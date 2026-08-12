"""Build the Utilization Report database design in the client's own workbook format.

Matches Database_Design_12.03_5.xlsx exactly:
  row 1  table name, merged across the sheet, bold on yellow
  row 2  what the table is for, merged, yellow, wrapped
  row 3  blank
  row 4  headers, bold, centred, wrapped
  row 5+ one row per column
  freeze panes below the header, Aptos Narrow 11 throughout
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from utilization_tables import TABLES, RELEASE, DECL, SOURCE

OUT = '/home/user/Jim/EDRMS_Utilization_Report_Database_Design_v1.xlsx'
FONT = 'Aptos Narrow'
YELLOW = PatternFill('solid', fgColor='FFFFFF00')  # ARGB, matching the client's sheets exactly

HEADERS = ['S/N', 'RELEASE EFFECTIVE', 'DECLARATION TYPE EFFECTIVE', 'COLUMN TITLE',
           'DESCRIPTION', 'REFERENCE TABLES', 'FIELD TYPE', 'FIELD VALUES/ CHOICES',
           'SAMPLE', 'WHERE THE VALUE COMES FROM', 'REMARKS', 'ADB Comments']
# Widths copied from the client's sheets so the two sit side by side comfortably.
WIDTHS = {'A': 6.9, 'B': 18, 'C': 22.9, 'D': 30, 'E': 61.1, 'F': 22.9,
          'G': 22.9, 'H': 34.4, 'I': 50.1, 'J': 24, 'K': 45, 'L': 22.9}
# A column nobody can fill is the one thing this workbook must not hide.
RED = PatternFill('solid', fgColor='FFFBE4E4')


def style_sheet(ws, title, blurb):
    n = len(HEADERS)
    last = get_column_letter(n)
    ws.merge_cells(f'A1:{last}1')
    ws['A1'] = title
    ws['A1'].font = Font(name=FONT, size=11, bold=True)
    ws['A1'].fill = YELLOW
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(f'A2:{last}2')
    ws['A2'] = blurb
    ws['A2'].font = Font(name=FONT, size=11)
    ws['A2'].fill = YELLOW
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 46

    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name=FONT, size=11, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[4].height = 32

    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'E5'


def write_table(wb, sheet, title, blurb, cols):
    ws = wb.create_sheet(sheet)
    style_sheet(ws, title, blurb)
    for n, (name, typ, desc, values, sample, ref, remarks) in enumerate(cols, start=1):
        row = 4 + n
        src = SOURCE.get(name)
        if src is None:
            raise SystemExit(f'{sheet}: {name} has no entry in SOURCE. '
                             'Every column must say where its value comes from.')
        blocked = src == 'NEEDS A DECISION'
        for i, v in enumerate([n, RELEASE, DECL, name, desc, ref, typ, values, sample,
                               src, remarks, ''], start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = Font(name=FONT, size=11, bold=(blocked and i == 10))
            c.alignment = Alignment(
                horizontal='center' if i in (1, 2, 3) else 'left',
                vertical='top',
                wrap_text=i in (5, 8, 9, 11))
            if blocked:
                c.fill = RED
    return len(cols)


wb = openpyxl.Workbook()
wb.remove(wb.active)

# ---- Structure sheet, the same role as theirs -----------------------------
ws = wb.create_sheet('Structure')
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 96

ws.merge_cells('A1:D1')
ws['A1'] = 'EDRMS UTILIZATION REPORT: REPORTING DATABASE DESIGN'
ws['A1'].font = Font(name=FONT, size=11, bold=True)
ws['A1'].fill = YELLOW
ws.merge_cells('A2:D2')
ws['A2'] = ('Four tables that feed the Utilization Report. These are reporting tables, separate from the tables the EDRMS '
            'application owns, and are intended to live in their own schema so nothing the application writes to is ever touched.')
ws['A2'].font = Font(name=FONT, size=11)
ws['A2'].fill = YELLOW
ws['A2'].alignment = Alignment(wrap_text=True, vertical='center')
ws.row_dimensions[2].height = 46

rows = [
    ['TABLE', 'ONE ROW PER', 'EST. ROWS', 'WHY IT MUST EXIST SEPARATELY'],
    ['Utilization Report Table', 'document', '3.47 million',
     'Holds declared and undeclared documents together. Without the undeclared there is no denominator and no declaration rate.'],
    ['Site Activity Table', 'SharePoint site', '1,057',
     'A site holding no documents would vanish from a document level count, and visit figures are measured per site so repeating them on every document row makes any total nonsense.'],
    ['User Activity Table', 'person', '9,400',
     'Total EDRMS Users counts people. Someone working in three sites appears in three rows of the Site Activity Table, so summing viewer counts overstates headcount.'],
    ['File Plan Table', 'term', 'a few hundred',
     'The file plan is the taxonomy that classifies content, so its terms are neither documents nor sites nor people.'],
]
for r, vals in enumerate(rows, start=4):
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(name=FONT, size=11, bold=(r == 4))
        c.alignment = Alignment(wrap_text=True, vertical='top',
                                horizontal='center' if r == 4 else 'left')
    ws.row_dimensions[r].height = 42

notes = [
    '',
    'HOW THESE DIFFER FROM THE APPLICATION TABLES',
    '',
    'No audit base class. The application tables carry CreatedBy, ModifiedBy, DeletedBy and their dates because a person edits those rows. '
    'Nobody edits a reporting row: a scheduled job writes the whole table and replaces it at the next refresh. SnapshotDate and RowLoadedDate '
    'do the equivalent job, recording which run a row came from and when it landed.',
    '',
    'SnapshotDate carries one value per run. It is what the Data as of line prints on every dashboard, and it is why every figure on a '
    'dashboard shares a single date rather than each panel being as current as its own source.',
    '',
    'Thresholds are not stored. LastActivityDate is a fact; "inactive after 90 days" is a judgement made in the report. Keeping judgements out '
    'of the database means changing 90 to 120 costs one edit in Power BI and no change to the job, the tables or the refresh.',
    '',
    'Nothing writes into these tables from Microsoft or AvePoint. Neither can push to PostgreSQL. A scheduled job pulls from Microsoft Graph, '
    'the Microsoft 365 usage reports and the term store, and writes the rows. Power BI then reads only these tables and never contacts SharePoint.',
    '',
    'TWO THINGS THAT ARE NOT SETTLED',
    '',
    'ADBDepartmentOwner has no source. The SharePoint column exists and is empty on every row. It arrives as a one time mapping of site to '
    'department from RAC, loaded onto the Site Activity Table, and every document inherits it through SiteUrl. Confirmed feasible by the '
    'development team on 10 August 2026, with no migration required.',
    '',
    'The File Plan Table is unverified. The five categories named in the requirement do not exist in the tenant term store, whose groups are '
    'ADB, ADB Exchange, ADB Test and similar. The Purview file plan holds 53 retention labels as a flat list with no such categories either. '
    'Where the institutional file plan actually lives is a question for the client before this table is built.',
]
r = 10
for line in notes:
    c = ws.cell(row=r, column=1, value=line)
    c.font = Font(name=FONT, size=11, bold=line.isupper() and bool(line))
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    if len(line) > 120:
        ws.row_dimensions[r].height = 46
    r += 1

total = 0
for sheet, title, blurb, cols in TABLES:
    total += write_table(wb, sheet, title, blurb, cols)

# ---- one sheet answering "can we actually get all of these?" --------------
from collections import Counter, defaultdict
tally = Counter()
where = defaultdict(list)
for sheet, _t, _b, cols in TABLES:
    for name, *_ in cols:
        src = SOURCE[name]
        tally[src] += 1
        where[src].append(f'{sheet.split()[1]}.{name}')

ws = wb.create_sheet('Where the values come from', 1)
for col, w in {'A': 26, 'B': 10, 'C': 62, 'D': 74}.items():
    ws.column_dimensions[col].width = w
ws.merge_cells('A1:D1')
ws['A1'] = 'CAN WE ACTUALLY GET EVERY COLUMN?'
ws['A1'].font = Font(name=FONT, size=11, bold=True)
ws['A1'].fill = YELLOW
ws.merge_cells('A2:D2')
ws['A2'] = (f'{total} columns across four tables. {total - tally["NEEDS A DECISION"]} have a source that exists today. '
            f'{tally["NEEDS A DECISION"]} do not, and those are shaded red on the table sheets. Nothing else is customised: '
            'every other column is either read from a system, computed from columns already present, or written by the job itself.')
ws['A2'].font = Font(name=FONT, size=11)
ws['A2'].fill = YELLOW
ws['A2'].alignment = Alignment(wrap_text=True, vertical='center')
ws.row_dimensions[2].height = 46

ORDER = ['EDRMS database', 'Microsoft Graph', 'M365 usage report', 'Site analytics',
         'SharePoint list', 'Term store', 'Derived at load', 'Job generated', 'NEEDS A DECISION']
HOW = {
 'EDRMS database': 'Read straight from public."Records" or ADBSites. Populated today for declared records; the scan supplies the same column for the rest.',
 'Microsoft Graph': 'One API call per object. GET /sites, /sites/{id}/drives, or the driveItem itself. All verified against the 7rkd12 tenant.',
 'M365 usage report': 'Downloadable as CSV from the Microsoft 365 admin centre, Reports, Usage. Two different reports: SharePoint site usage gives one row per site, SharePoint activity gives one row per user. Needs Reports.Read.All for the API version.',
 'Site analytics': 'GET /sites/{id}/analytics/{window}, the actorCount field. NOT in the usage export, which carries no unique viewer column at all. Only allTime and lastSevenDays exist as windows.',
 'SharePoint list': 'The Retention Label Mapping list, which RAC already maintains. RISK: it is unproven whether that list identifies libraries by ListId or by name. A name would be a fragile join.',
 'Term store': 'Walk the term store through Graph. UNVERIFIED: the categories in the requirement are not in this tenant.',
 'Derived at load': 'Computed by the job from columns it already has. No new source, no extra call.',
 'Job generated': 'The job writes it. Identity and run stamps.',
 'NEEDS A DECISION': 'Nothing supplies this. A person must provide a list, write a rule, or answer a question.',
}
r = 4
for k in ['SOURCE', *ORDER]:
    if k == 'SOURCE':
        vals = ['WHERE FROM', 'COLUMNS', 'HOW THE VALUE IS OBTAINED', 'WHICH COLUMNS']
    else:
        vals = [k, tally[k], HOW[k], ', '.join(where[k])]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(name=FONT, size=11, bold=(r == 4 or v == 'NEEDS A DECISION'))
        c.alignment = Alignment(wrap_text=True, vertical='top',
                                horizontal='center' if i == 2 else 'left')
        if k == 'NEEDS A DECISION':
            c.fill = RED
    ws.row_dimensions[r].height = 30 if r == 4 else 62
    r += 1

wb.save(OUT)
print(f'wrote {OUT}')
print(f'{len(TABLES)} tables, {total} columns')
for sheet, title, _, cols in TABLES:
    print(f'  {sheet:22} {len(cols):3} columns')
print()
for k in ORDER:
    print(f'  {k:20} {tally[k]:3}')
