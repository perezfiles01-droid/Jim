"""Build EDRMS_Utilization_Report_Requirements_2026-08-13.xlsx.

The client requirement assessment of 13 August 2026, as a workbook.

Source documents:
  EDRMS_Dashboard_requirements_1.pptx                 69 slides, the design deck
  R2026.4_Utilization_Report__Proposed_Metrics_6.docx  the proposed metric list

Never hand edit the workbook. Edit this file and rerun:

    python3 .claude/skills/edrms-utilization-report/scripts/build_requirements.py

Reference convention used throughout, the same one in REQUIREMENTS_2026-08-13.md:

    PPT s41              slide 41 of the deck
    DOC > Heading        that heading in the Word file, which has no section numbers
    STATUS s6            section 6 of STATUS.md
    T1 c27               table 1 column 27 of the database design
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "EDRMS_Utilization_Report_Requirements_2026-08-13.xlsx"
FONT = "Arial"

# ---------------------------------------------------------------- palette
# ADB palette, see BACKGROUND.md. Fills are the pale tints of it.
ADB_BLUE = "0067B1"
INK = "1F2933"
GREY = "6B7785"

FILL_HEAD = PatternFill("solid", fgColor=ADB_BLUE)
FILL_BAND = PatternFill("solid", fgColor="F2F6FA")
FILL_TITLE = PatternFill("solid", fgColor="E8EFF6")

# One fill per verdict, so the register can be read down the column at a glance.
VERDICT_FILL = {
    "Buildable now": PatternFill("solid", fgColor="D9EFD9"),
    "Needs department list": PatternFill("solid", fgColor="E4F0DA"),
    "Needs new column or join": PatternFill("solid", fgColor="FFF2CC"),
    "Needs application change": PatternFill("solid", fgColor="FFE4CC"),
    "Needs reference list": PatternFill("solid", fgColor="FFF7DA"),
    "Needs new data source": PatternFill("solid", fgColor="F8D7D7"),
    "Decision needed": PatternFill("solid", fgColor="E2DCF0"),
}

VERDICT_ORDER = list(VERDICT_FILL)

VERDICT_MEANING = {
    "Buildable now": "Sourceable from the 73 column design as it stands today",
    "Needs department list": "Buildable the moment RAC supply the site to department list",
    "Needs new column or join": "The design has to change: a new column, or a join that does not exist",
    "Needs application change": "Needs a new field in the EDRMS application. A development change request",
    "Needs reference list": "Not a measurement. Somebody has to maintain a list for it to exist",
    "Needs new data source": "A system we have not touched, or data that no API returns",
    "Decision needed": "The client has to answer before it can be scoped at all",
}

THIN = Side(style="thin", color="D4DCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------- sheet 2
# The six key views the client now wants, against the five we built.
VIEWS = [
    ("1", "Bank-wide Oversight",
     "PPT s13, s14, s15-s18, s34-s44; DOC > Bankwide oversight",
     "Overview, plus most of Records Management",
     "Rebuild and merge",
     "How is EDRMS performing across ADB? Adoption, usage, compliance and trends."),
    ("2", "Department Insights",
     "PPT s13, s14, s19-s28, s53-s66; DOC > Department insight",
     "Deleted on 10 August 2026",
     "Rebuild from nothing. The largest single piece of work in the deck",
     "How is a specific department managing its records?"),
    ("3", "Project Insights, Sovereign and Nonsovereign",
     "PPT s13, s14, s36, s37, s38",
     "Does not exist",
     "New. No data source exists to classify a project site",
     "Are project sites being used and maintained appropriately?"),
    ("4", "Institutional File Plan Insights",
     "PPT s13, s14, s29-s31, s47-s52; DOC > EDRMS Institutional File Plan insights",
     "A single panel inside Retention and File Plan",
     "Promote to a dashboard. Needs data that is in neither system",
     "Are users classifying records correctly?"),
    ("5", "Retention and Disposal Insights",
     "PPT s13, s14, s32, s44-s46; DOC > Retention and Disposition Metrics",
     "Retention and File Plan",
     "Closest match of the six. Extend it",
     "What records require action under the RRDS?"),
    ("6", "Records and Archive Holdings",
     "PPT s13, s14, s33, s67-s69; DOC > Records and Archives Holdings",
     "Does not exist",
     "New. A second data discovery exercise, recommend a separate workstream",
     "What records and archives are being managed by RAC?"),
    ("", "(no equivalent in the new design)",
     "",
     "Sites and Libraries",
     "Absorbed into Bank-wide Oversight and Department Insights",
     ""),
    ("", "(no equivalent in the new design)",
     "DOC > Format and Storage Analysis only, the deck drops it",
     "Format and Storage",
     "Demote to a panel on Bank-wide, per PPT s12",
     ""),
]


# ---------------------------------------------------------------- sheet 3
# Their review of the prototype we already built. Slides 4 to 12.
REVIEW = [
    ("PPT s4", "Overview, top panel",
     "Migration total is not a relevant top panel figure, and is a project concern rather than an ongoing one",
     "Remove any migration framing from the Overview top panel", "Small", "Yes"),
    ("PPT s4", "Overview, top panel",
     "Do not show one site total. Split three ways: Department / RM / Office, Nonsovereign projects, Sovereign projects",
     "Replace one KPI with three. Blocked, we cannot classify a site as sovereign or nonsovereign", "Blocked", "No"),
    ("PPT s4", "Overview",
     "We want users to know that information is consolidated", "A messaging point for the page copy", "Small", "Yes"),
    ("PPT s5", "Overview, total created sites",
     "Replace the total with an alphabetical table of departments, sortable by documents, records or users",
     "Replace the rollout treemap with a sortable department table", "Medium", "Needs department list"),
    ("PPT s5", "Overview",
     "Will there be EDRMS sites not created in Cloud Governance? If not, what is the value of this figure?",
     "Answer the question. If every site comes through Cloud Governance the figure is noise and goes", "Small", "Answer first"),
    ("PPT s6", "Overview, compare",
     "Comparison should be analytical, not obvious. Users against documents against records declared",
     "Rebuild Compare as three named ratios", "Medium", "Needs department list"),
    ("PPT s6", "Overview, compare",
     "Compare by library: how do you show hundreds of library options?",
     "Valid objection. Library comparison becomes a top N table, not a picker", "Small", "Yes"),
    ("PPT s7", "Overview, site rollout",
     "Site rollout is project only and will not be needed after go-live. Remove it",
     "Delete the Compliant site rollout panel", "Small, deletion", "Yes"),
    ("PPT s8", "Overview, active users per site",
     "Fold into the sites table, and also break it down on the Department dashboard",
     "Move it. Do not keep it as its own panel", "Small", "Yes"),
    ("PPT s9", "Records Management dashboard",
     "Adds no insight. Its stats belong on the Overview",
     "Dissolve the Records Management dashboard into Bank-wide Oversight", "Medium", "Yes"),
    ("PPT s10", "Records Management, declaration trend",
     "Genuinely useful, keep the date range. Wants it on the Overview, plus per department, office, RM and project. Asks what the default range is",
     "Move the trend to Bank-wide. Add a department cut. Answer the default range question, propose current year", "Medium", "Partly"),
    ("PPT s11", "Migration dashboard",
     "Not required at all", "Already removed on 10 August", "Done", "Done"),
    ("PPT s12", "Content volume dashboard",
     "Adds nothing except total size and storage growth. Amalgamate it",
     "Demote Format and Storage to a panel on Bank-wide", "Medium", "Yes"),
]


# ---------------------------------------------------------------- sheet 4
# The register. One row per requirement.
# (key view, panel, requirement, source, in prototype, in db design, verdict, note)
REG = []


def add(view, panel, req, src, proto, db, verdict, note):
    REG.append((view, panel, req, src, proto, db, verdict, note))


# --- 1. Bank-wide Oversight -------------------------------------------
V = "1. Bank-wide Oversight"

add(V, "Top panel", "Number of active EDRMS SharePoint sites, Department / RM / Office",
    "PPT s4, s15, s34, s35", "Partly, as one undivided site count",
    "T2 c6 IsEdrmsCompliant, T2 c14 LastActivityDate, T2 c7 ADBDepartmentOwner",
    "Needs department list",
    "Compliance rule is also still open, STATUS s6 gap 3b. Office and RM split needs a site type attribute we do not hold")
add(V, "Top panel", "Number of active EDRMS SharePoint sites for Nonsovereign projects",
    "PPT s4, s15, s34, s37", "No", "No", "Needs new data source",
    "Nothing in the term store, the usage export, Graph or Records distinguishes a project site. STATUS s6")
add(V, "Top panel", "Number of active EDRMS SharePoint sites for Sovereign projects",
    "PPT s4, s15, s34, s36", "No", "No", "Needs new data source", "Same as the row above")
add(V, "Top panel", "Total number of EDRMS users, monthly active",
    "PPT s15, s34, s39; DOC > Bankwide oversight", "Yes", "T3 c3 UserPrincipalName",
    "Buildable now", "Tenant reality per STATUS s5: 30 licensed users, 8 active")
add(V, "Top panel", "Total number of documents in EDRMS compliant sites",
    "PPT s15, s34, s40; DOC > Bankwide oversight", "Yes", "Table 1, all rows",
    "Buildable now",
    "Designed, but no data yet. Records holds declared only, STATUS s1. The undeclared denominator needs the weekly scan built")
add(V, "Top panel", "Total number of records declared in EDRMS",
    "PPT s15, s34, s41", "Yes", "T1 c23 IsDeclaredRecord", "Buildable now", "")
add(V, "Top panel", "Total number of physical counterparts identified",
    "PPT s15, s34, s42", "No", "T1 c27 HasPhysical", "Buildable now",
    "Designed, sourceable, and on no screen today. Appears on 24 of the 57 design slides. Quick win")
add(V, "Top panel", "Total number of records due for disposal",
    "PPT s34, s43", "Yes, on Retention", "T1 c32 EDRMSDueDateForDisposal", "Buildable now",
    "Moves up from the Retention dashboard to the top panel")
add(V, "Top panel", "Percentage of documents declared as records",
    "DOC > Bankwide oversight", "Yes", "T1 c23 over Table 1 row count", "Buildable now",
    "The declaration rate. Needs the undeclared denominator, same as total documents")

add(V, "Overview of EDRMS sites", "Alphabetical table: department, sites, documents, records declared, physical counterparts",
    "PPT s5, s16, s35", "No, we show a treemap", "T2 c7, T1 c23, T1 c27", "Needs department list",
    "The table replaces the treemap")
add(V, "Overview of EDRMS sites", "Each department name clickable, opening that department's dashboard",
    "PPT s16, s35, s16 note", "No", "n/a, a navigation behaviour", "Needs department list",
    "The single biggest build item in the deck. Depends on Department Insights existing")
add(V, "Overview of EDRMS sites", "Nonsovereign and Sovereign summary rows, clickable into Project Insights",
    "PPT s16, s35 note", "No", "No", "Needs new data source", "Blocked with the rest of Project Insights")
add(V, "Overview of EDRMS sites", "Indicator: number of sites created",
    "PPT s35", "Yes", "T2 c5 SiteCreatedDate", "Buildable now", "Confirmed from Graph createdDateTime, STATUS s5")
add(V, "Overview of EDRMS sites", "Indicator: number of sites deleted",
    "PPT s35", "No", "T2 c18 IsDeleted", "Needs new column or join",
    "Column exists but nothing populates it from the tenant yet")
add(V, "Overview of EDRMS sites", "Indicator: number of sites archived",
    "PPT s35; DOC > Site Trends", "No", "No", "Decision needed",
    "Archived has no definition. STATUS s6 already records this. Ask what archived means before building")
add(V, "Overview of EDRMS sites", "Indicator: sites inactive for 90 days",
    "PPT s35", "Yes", "T2 c14 LastActivityDate", "Buildable now",
    "Threshold conflict: PPT s35 says 90 days, DOC > Site Health says 300. Ask which")

add(V, "Comparison", "Active users against number of documents",
    "PPT s6, s17", "Partly", "T3 c3, Table 1 row count", "Needs department list", "")
add(V, "Comparison", "Active users against records declared",
    "PPT s6, s17", "Partly", "T3 c3, T1 c23", "Needs department list", "")
add(V, "Comparison", "Documents against records declared, the declaration rate",
    "PPT s6, s17", "Yes, on Records Management", "T1 c23", "Needs department list", "")

add(V, "Declaration trend", "Bank-wide record declaration trend over time with a date range",
    "PPT s10, s18", "Yes", "T1 c24 CreatedDate", "Buildable now",
    "Already built and the client likes it. Set an explicit default range, propose current year")
add(V, "Declaration trend", "The same trend cut by department, office, RM",
    "PPT s10, s18", "No", "T2 c7 ADBDepartmentOwner", "Needs department list", "")
add(V, "Declaration trend", "The same trend cut by project",
    "PPT s10", "No", "No", "Needs new data source", "")

add(V, "Users drill-down", "Users by department: percentage active, never accessed, not accessed in 90 days",
    "PPT s39", "No", "T3 c5 LastActivityDate", "Needs department list", "")
add(V, "Users drill-down", "Users split into staff, contractors and consultants",
    "PPT s21, s39, s54", "No", "No", "Needs new data source",
    "No such attribute in the design or the usage export. PPT s54 itself asks 'Source: database of end users? Owned and maintained by?'")
add(V, "Users drill-down", "Training completion and onboarded since go-live",
    "PPT s21, s39, s54", "No", "No", "Needs reference list", "An HR or LMS system, not SharePoint")

add(V, "Documents drill-down", "Documents created or uploaded per department, and size in GB",
    "PPT s40", "Partly", "T1 c9 FileCreatedDate, T1 c8 FileSize", "Needs department list",
    "FileSize is a new key that does not exist yet, STATUS s6 gap 2, closed as feasible via Graph size")
add(V, "Documents drill-down", "Number of users creating documents",
    "PPT s40", "No", "T1 c25 CreatedBy, declared rows only", "Needs new column or join",
    "We hold the declarer, not the author. The scan would have to capture Author for undeclared documents")
add(V, "Documents drill-down", "New documents month on month, and average monthly storage growth",
    "PPT s40", "Partly", "T1 c9, T1 c8 with SnapshotDate", "Needs department list",
    "Possible because history is kept on the small tables, STATUS s7, 12 Aug")

add(V, "Records declared drill-down", "Records declared and users declaring, per department",
    "PPT s41", "Yes, bank-wide", "T1 c23, T1 c25", "Needs department list", "")
add(V, "Records declared drill-down", "Records declared and users declaring, per division",
    "PPT s41, s42", "No", "No", "Decision needed",
    "Division was removed on 10 August, STATUS s7, because nothing populates it. Reopen or decline in writing")
add(V, "Records declared drill-down", "Highlight departments with zero record declarations",
    "PPT s41", "No", "T1 c23", "Needs department list", "")
add(V, "Records declared drill-down", "New records declared month on month per department",
    "PPT s41", "No", "T1 c24 CreatedDate", "Needs department list", "")

add(V, "Physical counterpart drill-down", "Physical records declared per department",
    "PPT s42", "No", "T1 c27 HasPhysical", "Needs department list", "")
add(V, "Physical counterpart drill-down", "Physical counterpart completion rate, records turned over to RAC against records declared with a counterpart",
    "PPT s42, s59", "No", "No", "Needs new data source",
    "Nothing records a transfer to RAC. Belongs to the holdings workstream, section 6 of this workbook")

add(V, "Disposal drill-down", "Records due for disposal per department, and next due date",
    "PPT s43", "Yes, bank-wide", "T1 c32 EDRMSDueDateForDisposal", "Needs department list", "")
add(V, "Disposal drill-down", "Disposal approver",
    "PPT s43, s60", "No", "No", "Needs application change", "")
add(V, "Disposal drill-down", "Status: Approved, Declined, Extended",
    "PPT s43, s60", "No", "No", "Needs application change",
    "DisposalStatus was removed from the design because it needs an application change. This requirement is the proof it is needed")
add(V, "Disposal drill-down", "Records disposed month on month, and disposal completion rate",
    "PPT s43; DOC > Disposition completed", "No", "No", "Needs application change", "")
add(V, "Disposal drill-down", "List of overdue disposal actions and pending approvals",
    "PPT s43", "Partly", "T1 c32 for overdue only", "Needs application change",
    "Overdue is computable today. Pending approvals is not")
add(V, "Disposal drill-down", "List of records due next quarter, for planning",
    "PPT s43", "Partly", "T1 c32", "Buildable now", "")
add(V, "Disposal drill-down", "Whether each due record has a physical counterpart",
    "PPT s43", "No", "T1 c27 HasPhysical", "Buildable now", "")

add(V, "Retention rollup", "Permanent and Temporary split with departments and libraries provisioned, records, counterparts, due and disposed",
    "PPT s44", "Partly", "T1 c30 EDRMSDuration", "Needs new column or join",
    "Split and counts are fine. Libraries provisioned per term needs the term join, and disposed needs the application change")

# --- 2. Department Insights -------------------------------------------
V = "2. Department Insights"

add(V, "Top panel", "Go-live date for the department",
    "PPT s19, s53", "No", "No", "Needs reference list", "Nobody has been named to maintain it")
add(V, "Top panel", "Total sites, users, documents, records declared, physical counterparts, records due for disposal",
    "PPT s19, s53", "No", "T2, T3, T1 as above", "Needs department list",
    "Every figure is one we can already produce bank-wide. It is the department filter that is missing")
add(V, "Top panel", "Every top panel stat clickable through to its detail",
    "PPT s19, s38, s53 notes", "No", "n/a, a navigation behaviour", "Needs department list", "")

add(V, "Sites list", "Site name, site owners, documents, records, physical counterparts, records due for disposal",
    "PPT s20, s55", "Site inventory exists, not per department", "T2 c4, T2 c16 SiteOwner", "Needs department list",
    "SiteOwner confirmed on 1,899 of 1,918 sites, STATUS s5. Note 19 sites have no owner at all")
add(V, "Sites list", "Number and list of sites with no activity in 90 days",
    "PPT s55", "Yes, bank-wide", "T2 c14 LastActivityDate", "Needs department list", "")
add(V, "Sites list", "Records declared month on month per site",
    "PPT s55", "Partly", "T1 c24 CreatedDate", "Needs department list", "")

add(V, "Users", "Users per site split into staff, contractors, consultants",
    "PPT s21, s54", "No", "No", "Needs new data source", "See the Bank-wide row. PPT s54 asks who owns this data")
add(V, "Users", "Training completion, and onboarded since go-live, per division",
    "PPT s21, s54", "No", "No", "Needs reference list",
    "PPT s54 sketches the source table itself: name, staff status, onboarded yes/no, training completed yes/no, date, notes")
add(V, "Users", "Active users month on month, users who have not accessed their site, new users to onboard",
    "PPT s54", "No", "T3 c5 LastActivityDate", "Needs department list",
    "PPT s54 asks 'is this possible?' of the never-accessed figure. Answer: yes for licensed users in the activity export")
add(V, "Users", "Total active users as a percentage of total staff",
    "PPT s54", "No", "T3 c3, denominator missing", "Needs reference list",
    "The numerator is ours. The staff headcount denominator is an HR figure")

add(V, "Visitors", "Total number of visitors per site",
    "PPT s22, s56", "No", "T2 c11 UniqueViewers7, T2 c12 UniqueViewersAllTime", "Buildable now",
    "Only two windows exist, 7 days and all time. Site analytics offers no 30 or 90 day window, STATUS s8")
add(V, "Visitors", "Visitors split internal and external",
    "PPT s22, s56", "No", "No", "Needs new data source", "Graph analytics returns a count with no internal or external split")
add(V, "Visitors", "Internal visitors by department",
    "PPT s22, s56", "No", "No", "Needs new data source", "")
add(V, "Visitors", "Access requests granted and denied",
    "PPT s56; DOC > Access Management", "No", "No", "Needs new data source",
    "A SharePoint permission workflow, not a reporting feed. PPT s56 asks 'Is this data available in SharePoint / Cloud Governance?' The answer is no, not through the reporting APIs")

add(V, "Documents", "Documents created or uploaded, users creating, size in GB, per site",
    "PPT s57", "No", "T1 c9, T1 c8, T1 c25", "Needs department list", "")
add(V, "Documents", "Documents migrated, users migrating, migrated size in GB",
    "PPT s57", "No", "No", "Decision needed",
    "Contradicts their own PPT s4 and s11, where migration was said not to be wanted. Query with the client")

add(V, "Record declaration", "Records declared, users declaring, declared record size in GB",
    "PPT s23, s58", "Yes, bank-wide", "T1 c23, T1 c25, T1 c8", "Needs department list", "")
add(V, "Record declaration", "The same figures per division, with collapsible division rows",
    "PPT s23, s58", "No", "No", "Decision needed", "The division decision again")
add(V, "Record declaration", "Sites and libraries with no declared record in 90 days",
    "PPT s58, s61", "Partly", "T1 c24, T1 c15 ListId", "Needs department list", "")
add(V, "Record declaration", "Declaration rate per site",
    "PPT s58", "Partly", "T1 c23 over Table 1", "Needs department list", "")

add(V, "Physical counterpart", "Records declared with a physical counterpart, per site and per division",
    "PPT s59", "No", "T1 c27 HasPhysical", "Needs department list", "Division tier excepted")
add(V, "Physical counterpart", "Physical counterpart completion rate against RAC turnover",
    "PPT s59", "No", "No", "Needs new data source", "")

add(V, "Declaration trend", "Record declaration trend for one department, office, RM or project",
    "PPT s24", "Yes, bank-wide only", "T1 c24 CreatedDate", "Needs department list", "")

add(V, "Library usage", "Library name, documents, records declared, physical counterparts, grouped by file plan category",
    "PPT s25, s61-s66", "Library table exists", "T1 c12 LibraryName, T1 c15 ListId", "Needs new column or join",
    "The figures are fine. The category grouping needs the term to library link that does not exist")
add(V, "Library usage", "Number of users per library",
    "PPT s25, s61-s66", "No", "No", "Needs new data source",
    "SharePoint reports viewers per site, never per library. Already recorded in STATUS s6")
add(V, "Library usage", "Libraries with no declared record in 90 days",
    "PPT s61", "Partly", "T1 c24, T1 c15", "Needs department list", "")

add(V, "Conventions", "Link to the approved site, library and folder convention, date of approval, last updated, version number",
    "PPT s26, s53; DOC > Conventions", "No", "No", "Needs reference list",
    "Not a measurement. Same conclusion already reached in STATUS s6")

add(V, "Disposal", "Library, records due for disposal, next due date",
    "PPT s27, s60", "Yes, bank-wide", "T1 c32, T1 c12", "Needs department list", "")
add(V, "Disposal", "Disposal approver, status Approved / Declined / Extended, records disposed, disposed size",
    "PPT s27, s60", "No", "No", "Needs application change", "")

add(V, "Programme dates", "CSIS-IR audit, convention review, physical records review, refresher training, focals CoP schedule, each with a date and status",
    "PPT s28, s53; DOC > Records management program dates", "No", "No", "Needs reference list",
    "Not measurements. A maintained list. STATUS s6 says the same of requirement 2.1.3 and 2.1.4")

# --- 3. Project Insights ----------------------------------------------
V = "3. Project Insights"

add(V, "Classification", "Identify which sites are project sites at all",
    "PPT s4, s13, s36, s37", "No", "No", "Needs new data source",
    "The one blocker. Everything else on this dashboard follows from it")
add(V, "Classification", "Split Sovereign from Nonsovereign",
    "PPT s15, s16, s34, s36, s37; DOC > Bankwide oversight", "No", "No", "Needs new data source",
    "STATUS s6 lists sovereign and nonsovereign under 'no source anywhere'")
add(V, "Project list", "Project number and project name, with sites, documents, records declared, physical counterparts",
    "PPT s36, s37", "No", "T2 c17 ProjectEndDate is the only project field held", "Needs new data source",
    "PPT s36 itself asks '(format?)' of the project identifier, so the client has not settled it either")
add(V, "Project profile", "Facility type, project number, modality, modality number, country or economy, project status, effectivity date, actual closing date",
    "PPT s38", "No", "No", "Needs new data source",
    "ADB operational data. It lives in ADB project systems, not in SharePoint")
add(V, "Project profile", "Sites, users, visitors, documents, records declared, physical counterparts, records due for disposal, for one project",
    "PPT s38", "No", "Yes, all of them", "Needs new data source",
    "Every figure here is one we can already produce. Purely the classification is missing")

# --- 4. Institutional File Plan Insights -------------------------------
V = "4. Institutional File Plan Insights"

add(V, "Top panel", "Total terms, and totals for Administration, People Management, Programs and Operations, Institutional Management, Other",
    "PPT s29, s47; DOC > EDRMS Institutional File Plan insights", "A flat panel",
    "T4 c6 CategoryName, T4 c4 TermName, T4 c7 Depth", "Decision needed",
    "Designed, but the source does not exist. STATUS s6: the term store holds six dropdown value sets, 16 terms, one level deep. Purview holds 53 flat retention labels. The institutional file plan is in neither")
add(V, "Term tables", "Per term: number of libraries provisioned",
    "PPT s30, s31, s47-s52", "No", "No", "Needs new column or join",
    "Table 4 has no key linking a term to a library. Nothing in Table 1 carries a TermId. This is a new column, plus a scan change")
add(V, "Term tables", "Per term: departments or offices or RMs provisioned",
    "PPT s47-s52", "No", "No", "Needs new column or join", "The same missing join, plus the department list")
add(V, "Term tables", "Per term: number of documents, records declared, physical counterparts",
    "PPT s30, s31, s48-s52", "No", "No", "Needs new column or join", "The same missing join")
add(V, "Indicators", "Top most-used terms by documents created and records declared",
    "PPT s48-s52 'Include indicators' note", "No", "Derivable once the join exists", "Needs new column or join", "")
add(V, "Indicators", "Least-used and unused terms, as candidates for review and deletion",
    "PPT s48-s52 'Include indicators' note", "No", "Derivable once the join exists", "Needs new column or join", "")
add(V, "Indicators", "New libraries created outside convention",
    "PPT s48-s52 'Include indicators' note", "No", "No", "Decision needed",
    "Convention compliance means comparing a library name against an approved naming convention. Nobody has given us the convention")
add(V, "Indicators", "Requests for new library creation or deletion, and requests for new terms",
    "PPT s48-s52 'Include indicators' note", "No", "No", "Needs new data source",
    "Not a report. A request workflow, most likely Cloud Governance")
add(V, "Records declaration", "Records declared by classification, and by business process",
    "DOC > Records Declaration", "No", "No", "Needs new column or join",
    "Classification means the file plan term, so the same join blocks it")

# --- 5. Retention and Disposal Insights --------------------------------
V = "5. Retention and Disposal"

add(V, "Structure", "Permanent retention screen and Temporary retention screen",
    "PPT s32, s44, s45, s46", "Yes, the retention profile already excludes Permanent",
    "T1 c30 EDRMSDuration, text so it can hold Permanent", "Buildable now", "")
add(V, "Retention dashboard", "Records due for disposition, within 30 days, within 90 days",
    "PPT s32, s44, s46; DOC > Retention Dashboard", "Yes", "T1 c32 EDRMSDueDateForDisposal", "Buildable now", "")
add(V, "Retention dashboard", "Retention label breakdown with durations, 5, 10, 20 years",
    "PPT s46", "Yes", "T1 c29 EDRMSRetentionLabel, T1 c30 EDRMSDuration", "Buildable now",
    "EDRMSDuration is text, cast it before arithmetic")
add(V, "Retention compliance", "Records with and without retention schedules",
    "DOC > Retention Compliance", "Partly", "Null test on T1 c29", "Buildable now", "")
add(V, "Retention compliance", "Libraries without mapped retention schedules",
    "DOC > Retention Compliance", "No", "Retention Label Mapping list, T1 c28", "Decision needed",
    "Held back deliberately. STATUS s11 next step 6: depends on whether the mapping list keys libraries by ListId or by name. A small check that releases two metrics")
add(V, "Term tables", "Departments and libraries provisioned per retention term",
    "PPT s44, s45, s46", "No", "No", "Needs new column or join", "The same term join as the File Plan dashboard")
add(V, "Disposal", "Number of records disposed",
    "PPT s32, s44, s46, s60; DOC > Disposition completed", "No", "No", "Needs application change",
    "There is no record anywhere of a disposal having happened")
add(V, "Disposal", "Records awaiting approval, disposition backlog, approval backlog",
    "DOC > Retention Dashboard, DOC > Disposition Risk Indicators", "No", "No", "Needs application change", "")
add(V, "Disposal", "Overdue dispositions, records beyond retention periods",
    "DOC > Disposition Risk Indicators", "Partly", "T1 c32", "Buildable now", "Computable from the due date alone")
add(V, "Disposal", "Physical records overdue for transfer",
    "DOC > Risk and Compliance", "No", "No", "Needs new data source", "Belongs to the holdings workstream")

# --- 6. Records and Archive Holdings -----------------------------------
V = "6. Records and Archive Holdings"

add(V, "Scope", "The client's own slide says 'RAC suggestions?', so the screen is unfinished on their side too",
    "PPT s33", "No", "No", "Decision needed", "Recommend a separate workstream with its own timeline")
add(V, "Storage", "By location: Archives Room, Records Center, Offsite Storage",
    "PPT s68; DOC > Storage Location Dashboard", "No", "No", "Needs new data source",
    "STATUS s6: PhysicalRecords was designed in the workbook and never built. No boxes, locations or facilities exist anywhere")
add(V, "Storage", "Requests, requestors by department and by RM, boxes stored, folders stored, month on month",
    "PPT s68", "No", "No", "Needs new data source", "")
add(V, "Storage", "Room capacity and percentage available storage capacity",
    "PPT s68, s69 closing notes", "No", "No", "Needs new data source", "Physical facility data")
add(V, "Retrieval", "Boxes and folders retrieved, by location, with status Loan, Return to owner, For Disposal",
    "PPT s69", "No", "No", "Needs new data source",
    "PPT s67 says retrieval is processed in eServe. A separate system with no established feed")
add(V, "Retrieval", "Outstanding and pending requests, boxes and folders disposed",
    "PPT s69", "No", "No", "Needs new data source", "")
add(V, "Inventory health", "Unverified physical files, missing files, files due for inventory verification, files scheduled for transfer",
    "DOC > Inventory Health", "No", "No", "Needs new data source", "Implies a physical inventory process we have no visibility of")
add(V, "Holdings", "Total physical files, total legacy records, total boxes, total storage locations",
    "DOC > Physical Records Holdings", "No", "No", "Needs new data source", "")
add(V, "Holdings", "Records by office location and by storage facility",
    "DOC > Physical Records Holdings", "No", "No", "Needs new data source",
    "Field office is already listed in STATUS s6 as having no source anywhere")
add(V, "Open questions", "Access to the IR Dashboard they screenshot",
    "PPT s67", "n/a", "n/a", "Decision needed", "It may already hold most of this. Ask for access")
add(V, "Open questions", "What is available in Opus and how its features could apply to the dashboard",
    "PPT s67", "n/a", "n/a", "Decision needed", "A direct question to answer, not a metric")

# --- 7. In the Word list only ------------------------------------------
V = "7. Word list only, not in the deck"

add(V, "Risk indicators, site health", "Active sites, inactive sites over 300 days, orphaned sites",
    "DOC > Risk and Compliance > Site Health", "Yes", "T2 c14 LastActivityDate, T2 c16 SiteOwner", "Buildable now",
    "Threshold conflict: PPT s35 says 90 days, this says 300. Ask which. 19 sites have no owner, STATUS s5")
add(V, "Site trends", "New sites created, site activity trend by month",
    "DOC > Site Trends", "Partly", "T2 c5 SiteCreatedDate, T2 c8 SiteVisits7", "Buildable now",
    "Possible now that history is kept, STATUS s7. A range sums SiteVisits7 only, never the 30 or 90 figures")
add(V, "Site trends", "Sites archived",
    "DOC > Site Trends", "No", "No", "Decision needed", "No definition")
add(V, "Library usage", "Most used libraries by views, downloads, uploads and edits",
    "DOC > Risk and Compliance > Library usage", "No", "No", "Needs new data source",
    "No per library activity feed exists, only per site. STATUS s6")
add(V, "Library health", "Largest libraries by record volume and by storage, library growth rate",
    "DOC > Risk and Compliance > Library usage", "Yes, already built", "T1 c15 ListId, T1 c8 FileSize", "Buildable now", "")
add(V, "Library health", "Libraries inactive at 90 or 180 days, libraries with no declared records",
    "DOC > Risk and Compliance", "Yes, already built", "T1 c14 LibraryLastActivityDate, T1 c23", "Buildable now", "")
add(V, "Library health", "Orphaned libraries, those without owners",
    "DOC > Risk and Compliance", "No", "No", "Needs new column or join", "We hold a site owner, not a library owner")
add(V, "Records quality", "Total duplicated records, by matching filenames",
    "DOC > Records Quality", "No", "T1 c4 Title", "Buildable now", "A self join on Title. Cheap and new")
add(V, "Records quality", "Orphaned records",
    "DOC > Records Quality", "No", "T1 c4, T1 c19", "Buildable now", "Needs a definition of orphaned, but the data is there")
add(V, "Format and storage", "Declared records by the eight format groups, with file counts and storage",
    "DOC > Format and Storage Analysis", "Yes, a whole dashboard", "T1 c7 FormatGroup, T1 c8 FileSize", "Buildable now",
    "Survives as a panel even though PPT s12 demotes it. FormatGroup still needs the extension mapping from RAC, STATUS s6")
add(V, "Information classification", "Records with and without sensitivity labels, records by classification level",
    "DOC > Information Classification", "No", "T1 c34 SensitivityLabelName", "Buildable now",
    "In the design and on no dashboard today. Another quick win")
add(V, "Access management", "Restricted and confidential records",
    "DOC > Access Management", "No", "T1 c34 SensitivityLabelName", "Buildable now",
    "Follows from the sensitivity label if the label set names those levels")
add(V, "Access management", "Access requests, external sharing instances, permission exceptions",
    "DOC > Access Management", "No", "No", "Needs new data source",
    "Audit log and permission data, not reporting APIs. A different Graph surface and a different permission set")
add(V, "Search analytics", "Searches performed, successful searches, frequently searched record categories",
    "DOC > Information Retrieval", "No", "No", "Needs new data source",
    "SharePoint search analytics are tenant level and are not exposed per record")
add(V, "Search analytics", "Records accessed per month, most viewed records, most downloaded records",
    "DOC > Top Content", "No", "No", "Needs new data source", "Not available at document grain")
add(V, "Search analytics", "Most accessed libraries",
    "DOC > Top Content", "No", "No", "Needs new data source", "Same as most used libraries above")


# ---------------------------------------------------------------- sheet 6
QUESTIONS = [
    (1, "Where does the Institutional File Plan actually live?",
     "It is in neither the term store nor Purview. The term store holds six dropdown value sets, 16 terms, one level deep. Purview holds 53 flat retention labels",
     "PPT s29-s31, s47-s52; STATUS s6",
     "A whole dashboard, plus records declared by classification",
     "Client",
     "Already outstanding before these documents arrived. STATUS s11 next step 2"),
    (2, "Is there a project site register, mapping site URL to project number and sovereign or nonsovereign?",
     "Nothing in SharePoint, Graph, the term store or Records classifies a project site",
     "PPT s36, s37, s38",
     "The entire Project Insights dashboard, plus three Bank-wide KPI tiles",
     "RAC, then AvePoint Cloud Governance",
     "Check Cloud Governance first. If it captures the project number at site request time this is one CSV and the dashboard unlocks"),
    (3, "Division: in or out?",
     "We removed the division tier on 10 August because nothing populates it. Four of their slides now ask for it",
     "PPT s41, s42, s54, s58; STATUS s7",
     "Four drill-down tables",
     "Client and RAC",
     "If it is in, who supplies the site or user to division mapping"),
    (4, "Do you want to raise a development change request for disposal status?",
     "Disposal approver, the three statuses and records disposed all need a new field in the EDRMS application",
     "PPT s43, s44, s46, s60; DOC > Retention Dashboard",
     "About twelve metrics across Bank-wide, Department and Retention",
     "Client, then Mihal Le",
     "We removed DisposalStatus from the design on these grounds. This requirement is the proof the change is needed"),
    (5, "Who maintains the three reference lists?",
     "Go-live dates, conventions, and records management programme dates. None of them are measurements and nothing can generate them",
     "PPT s19, s26, s28, s53",
     "Three Department Insights panels",
     "Client",
     "STATUS s6 reached the same conclusion about requirements 2.1.3 and 2.1.4"),
    (6, "Is an inactive site 90 days or 300 days?",
     "The deck says 90, the Word list says 300. The two documents contradict each other",
     "PPT s35; DOC > Risk and Compliance > Site Health",
     "Every inactivity figure on two dashboards",
     "Client",
     "Cheap to change either way, the threshold is not stored, STATUS s4"),
    (7, "Which system owns staff, contractor and consultant classification, and training completion?",
     "Nothing in SharePoint holds it. Their own slide asks the same question",
     "PPT s21, s39, s54",
     "The Users panels on Bank-wide and Department",
     "Client, likely HR or an LMS",
     "PPT s54 says 'Source: database of end users? Owned and maintained by?'"),
]


# ---------------------------------------------------------------- sheet 7
QUICKWINS = [
    (1, "Delete the Compliant site rollout panel", "PPT s7", "Small, deletion",
     "They said it is project only and will not be needed after go-live"),
    (2, "Dissolve Records Management into Bank-wide Oversight", "PPT s9", "Medium",
     "They said it adds no insight and its stats belong on the Overview"),
    (3, "Demote Format and Storage to a panel", "PPT s12", "Medium",
     "They said it adds nothing except total size and storage growth"),
    (4, "Move the declaration trend to Bank-wide and set an explicit default range", "PPT s10", "Medium",
     "They like the panel and asked what the default is. Propose current year"),
    (5, "Fold active users per site into the sites table", "PPT s8", "Small",
     "They asked for it to stop being its own panel"),
    (6, "Rebuild the department treemap as a sortable alphabetical table", "PPT s5, s16", "Medium",
     "Sortable by documents, records or users. Rows stay empty until the department list arrives"),
    (7, "Rebuild Compare as the three named ratios", "PPT s6, s17", "Medium",
     "Users to documents, users to records, documents to records"),
    (8, "Surface T1 c27 HasPhysical as a physical counterpart count", "PPT s15", "Small",
     "Designed, sourceable, and on no screen today. It appears on 24 of the 57 design slides"),
    (9, "Surface T1 c34 SensitivityLabelName", "DOC > Information Classification", "Small",
     "Same situation: in the design, on no dashboard"),
    (10, "Add duplicate and orphaned record counts", "DOC > Records Quality", "Small",
     "A self join on Title. A cheap new metric"),
    (11, "Check the Retention Label Mapping list's key, ListId or name", "STATUS s11 next step 6", "Small",
     "If it keys by ListId, two held-back metrics go straight in"),
]


# ---------------------------------------------------------------- helpers
def style_header(ws, row, ncols, height=30):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = FILL_HEAD
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = height


def title_block(ws, title, subtitle, ncols):
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=ADB_BLUE)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color=GREY)
    for r in (1, 2):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).fill = FILL_TITLE
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 16


def widths(ws, spec):
    for i, w in enumerate(spec, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def body(ws, first_row, last_row, ncols, band=True):
    for r in range(first_row, last_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if band and (r - first_row) % 2 == 1:
                cell.fill = FILL_BAND


# ---------------------------------------------------------------- build
wb = Workbook()

# ================= Sheet 1: Read me first =================
ws = wb.active
ws.title = "Read me first"
widths(ws, [3, 34, 96])

ws["B1"] = "EDRMS Utilization Report: client requirements assessment"
ws["B1"].font = Font(name=FONT, size=16, bold=True, color=ADB_BLUE)
ws["B2"] = "13 August 2026"
ws["B2"].font = Font(name=FONT, size=10, italic=True, color=GREY)

intro = [
    ("", ""),
    ("What this workbook is",
     "An assessment of two client documents against the prototype we have built and the database design "
     "behind it. Every requirement is recorded with the slide or heading it came from, whether it is already "
     "in the prototype, whether the design can source it, and where it cannot, why not."),
    ("", "No prototype code has been changed. This is the assessment that comes before that work."),
    ("Source documents",
     "EDRMS_Dashboard_requirements_1.pptx, 69 slides. Slides 1 to 12 are the client's annotated review of "
     "our existing prototype. Slides 13 to 69 are their own proposed design, screen by screen."),
    ("",
     "R2026.4_Utilization_Report__Proposed_Metrics_6.docx, a flat metric list with no section numbers."),
    ("Where the two disagree",
     "This assessment follows the deck. The Word file is a superset wish list containing three groups the "
     "deck never draws: Security and Classification, Search Analytics, and Risk Indicators. The deck is the "
     "one with layouts, click behaviour and the client's own comments on our work. Sheet 'Word list only' "
     "keeps everything that appears in the Word file alone, so nothing is lost."),
    ("", ""),
    ("How to read a reference", ""),
    ("PPT s41", "Slide 41 of the dashboard requirements deck"),
    ("DOC > Records Quality", "That heading in the proposed metrics Word file, which has no section numbers"),
    ("STATUS s6", "Section 6 of STATUS.md in this repository"),
    ("T1 c27", "Table 1, column 27 of the database design. The same numbering as "
               "EDRMS_Utilization_Report_Database_Design_v1.xlsx and utilizationdb.md"),
    ("", ""),
    ("The three headline findings", ""),
    ("1. Six key views, not five",
     "The client has redesigned the report around six key views (PPT s13, s14). Two of them do not exist for "
     "us at all, and one of them, Department Insights, we deleted on 10 August. Two of our five dashboards "
     "stop existing. This is not a set of tweaks."),
    ("2. Physical counterparts are everywhere",
     "'Number of physical counterparts' appears on 24 of the 57 design slides. T1 c27 HasPhysical is designed "
     "and sourceable, so the count is a quick win. The holdings world behind it, boxes, transfers and eServe "
     "retrievals, is a second data discovery exercise the size of the SharePoint one already done."),
    ("3. Division is back",
     "We removed the division tier on 10 August because nothing populated it (STATUS s7). PPT s41, s42, s54 "
     "and s58 all ask for it. That decision has to be reopened with the client, or the requirement declined "
     "in writing."),
    ("", ""),
    ("Sheets in this workbook", ""),
    ("Key views", "The six views they want, against the five we built"),
    ("Their review of ours", "Slides 4 to 12, their comments on our existing prototype. The cheapest changes in the document"),
    ("Requirements register", "The main sheet. One row per requirement, with its source, its verdict and why"),
    ("Word list only", "Included in the register, listed again here because it needs confirming as still wanted"),
    ("Questions to client", "The seven answers that unblock the most"),
    ("Start immediately", "Eleven items with no external dependency"),
    ("Summary", "Counts by verdict and by key view, computed from the register"),
]

r = 4
for left, right in intro:
    ws.cell(row=r, column=2, value=left).font = Font(name=FONT, size=10, bold=bool(left) and not right)
    ws.cell(row=r, column=3, value=right).font = Font(name=FONT, size=10)
    ws.cell(row=r, column=2).alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row=r, column=3).alignment = Alignment(vertical="top", wrap_text=True)
    if left and right:
        ws.cell(row=r, column=2).font = Font(name=FONT, size=10, bold=True)
    r += 1

ws.sheet_view.showGridLines = False


# ================= Sheet 2: Key views =================
ws = wb.create_sheet("Key views")
widths(ws, [5, 34, 40, 30, 40, 44])
title_block(ws, "The six key views the client now wants",
            "Both documents define this list. PPT s13 names them, PPT s14 gives each a question, an output and an action", 6)
hdr = ["#", "Their key view", "Where it is specified", "Our nearest equivalent today", "Status", "The question it answers, per PPT s14"]
for i, h in enumerate(hdr, 1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, 6)
for j, row in enumerate(VIEWS):
    for i, v in enumerate(row, 1):
        ws.cell(row=5 + j, column=i, value=v)
body(ws, 5, 4 + len(VIEWS), 6)
ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False


# ================= Sheet 3: Their review of ours =================
ws = wb.create_sheet("Their review of ours")
widths(ws, [12, 30, 62, 62, 16, 18])
title_block(ws, "Slides 4 to 12: the client's review of the prototype we built",
            "Mostly deletions. These are the cheapest changes in the whole document and should be done first", 6)
hdr = ["Slide", "Panel", "What they said", "What we do", "Effort", "Doable?"]
for i, h in enumerate(hdr, 1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, 6)
for j, row in enumerate(REVIEW):
    for i, v in enumerate(row, 1):
        ws.cell(row=5 + j, column=i, value=v)
body(ws, 5, 4 + len(REVIEW), 6)
ws.freeze_panes = "C5"
ws.sheet_view.showGridLines = False


# ================= Sheet 4: Requirements register =================
ws = wb.create_sheet("Requirements register")
widths(ws, [6, 26, 24, 58, 34, 26, 34, 24, 62])
title_block(ws, "Requirements register",
            "One row per requirement across both documents. Verdict column is colour coded, see the Summary sheet for what each verdict means", 9)
hdr = ["S/N", "Key view", "Panel or group", "Requirement", "Where it is asked",
       "In the prototype?", "In the database design?", "Verdict", "Why, and what unblocks it"]
for i, h in enumerate(hdr, 1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, 9, height=34)

for j, (view, panel, req, src, proto, db, verdict, note) in enumerate(REG):
    r = 5 + j
    ws.cell(row=r, column=1, value=j + 1)
    ws.cell(row=r, column=2, value=view)
    ws.cell(row=r, column=3, value=panel)
    ws.cell(row=r, column=4, value=req)
    ws.cell(row=r, column=5, value=src)
    ws.cell(row=r, column=6, value=proto)
    ws.cell(row=r, column=7, value=db)
    ws.cell(row=r, column=8, value=verdict)
    ws.cell(row=r, column=9, value=note)

last = 4 + len(REG)
body(ws, 5, last, 9, band=False)
for r in range(5, last + 1):
    ws.cell(row=r, column=8).fill = VERDICT_FILL[ws.cell(row=r, column=8).value]
    ws.cell(row=r, column=8).font = Font(name=FONT, size=9, bold=True)
    ws.cell(row=r, column=1).alignment = Alignment(vertical="top", horizontal="center")

ws.auto_filter.ref = f"A4:I{last}"
ws.freeze_panes = "D5"
ws.sheet_view.showGridLines = False


# ================= Sheet 5: Word list only =================
ws = wb.create_sheet("Word list only")
widths(ws, [6, 30, 58, 38, 26, 24, 58])
title_block(ws, "In the proposed metrics document but nowhere in the deck",
            "Not drawn on any slide. Confirm these are still wanted before costing them. All of these also appear in the register under key view 7", 7)
hdr = ["S/N", "Word heading", "Requirement", "Where it is asked", "In the database design?", "Verdict", "Note"]
for i, h in enumerate(hdr, 1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, 7)

wonly = [r for r in REG if r[0].startswith("7.")]
for j, (view, panel, req, src, proto, db, verdict, note) in enumerate(wonly):
    r = 5 + j
    ws.cell(row=r, column=1, value=j + 1)
    ws.cell(row=r, column=2, value=panel)
    ws.cell(row=r, column=3, value=req)
    ws.cell(row=r, column=4, value=src)
    ws.cell(row=r, column=5, value=db)
    ws.cell(row=r, column=6, value=verdict)
    ws.cell(row=r, column=7, value=note)
lastw = 4 + len(wonly)
body(ws, 5, lastw, 7, band=False)
for r in range(5, lastw + 1):
    ws.cell(row=r, column=6).fill = VERDICT_FILL[ws.cell(row=r, column=6).value]
    ws.cell(row=r, column=6).font = Font(name=FONT, size=9, bold=True)
ws.freeze_panes = "C5"
ws.sheet_view.showGridLines = False


# ================= Sheet 6: Questions to client =================
ws = wb.create_sheet("Questions to client")
widths(ws, [6, 52, 56, 32, 40, 24, 56])
title_block(ws, "The seven questions to send back, in priority order",
            "Ordered by how much each answer unblocks, not by how easy it is to ask", 7)
hdr = ["#", "Question", "Why we are asking", "Where it comes from", "What it unblocks", "Who answers", "Note"]
for i, h in enumerate(hdr, 1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, 7)
for j, row in enumerate(QUESTIONS):
    for i, v in enumerate(row, 1):
        ws.cell(row=5 + j, column=i, value=v)
body(ws, 5, 4 + len(QUESTIONS), 7)
for r in range(5, 5 + len(QUESTIONS)):
    ws.cell(row=r, column=2).font = Font(name=FONT, size=9, bold=True)
    ws.cell(row=r, column=1).alignment = Alignment(vertical="top", horizontal="center")
ws.freeze_panes = "C5"
ws.sheet_view.showGridLines = False


# ================= Sheet 7: Start immediately =================
ws = wb.create_sheet("Start immediately")
widths(ws, [6, 62, 34, 18, 72])
title_block(ws, "Work that can start now, with no external dependency",
            "None of these wait on the client, on RAC, or on a development change. Together they answer most of PPT s4 to s12", 5)
hdr = ["#", "Change", "Where it is asked", "Effort", "Why"]
for i, h in enumerate(hdr, 1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, 5)
for j, row in enumerate(QUICKWINS):
    for i, v in enumerate(row, 1):
        ws.cell(row=5 + j, column=i, value=v)
body(ws, 5, 4 + len(QUICKWINS), 5)
for r in range(5, 5 + len(QUICKWINS)):
    ws.cell(row=r, column=1).alignment = Alignment(vertical="top", horizontal="center")
ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False


# ================= Sheet 8: Summary =================
ws = wb.create_sheet("Summary")
widths(ws, [4, 32, 12, 12, 80])
title_block(ws, "Summary",
            "Every count on this sheet is a formula over the Requirements register, so it stays correct when a row is edited there", 5)

reg_last = 4 + len(REG)

ws["B4"] = "By verdict"
ws["B4"].font = Font(name=FONT, size=11, bold=True, color=ADB_BLUE)
for i, h in enumerate(["Verdict", "Count", "Share", "What it means"], 2):
    ws.cell(row=5, column=i, value=h)
style_header(ws, 5, 5)

r = 6
for v in VERDICT_ORDER:
    ws.cell(row=r, column=2, value=v).fill = VERDICT_FILL[v]
    ws.cell(row=r, column=2).font = Font(name=FONT, size=9, bold=True)
    ws.cell(row=r, column=3,
            value=f"=COUNTIF('Requirements register'!$H$5:$H${reg_last},$B{r})")
    ws.cell(row=r, column=4, value=f"=IF($C${6 + len(VERDICT_ORDER)}=0,0,C{r}/$C${6 + len(VERDICT_ORDER)})")
    ws.cell(row=r, column=5, value=VERDICT_MEANING[v])
    r += 1

total_row = r
ws.cell(row=total_row, column=2, value="Total requirements recorded")
ws.cell(row=total_row, column=3, value=f"=SUM(C6:C{total_row - 1})")
ws.cell(row=total_row, column=4, value=f"=IF(C{total_row}=0,0,1)")

body(ws, 6, total_row, 5, band=False)
for c in (2, 3, 4, 5):
    cell = ws.cell(row=total_row, column=c)
    cell.font = Font(name=FONT, size=9, bold=True)
    cell.fill = FILL_TITLE
for rr in range(6, total_row + 1):
    ws.cell(row=rr, column=3).number_format = "#,##0"
    ws.cell(row=rr, column=3).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=rr, column=4).number_format = "0.0%"
    ws.cell(row=rr, column=4).alignment = Alignment(horizontal="center", vertical="top")
for v in VERDICT_ORDER:
    pass

# By key view
start = total_row + 3
ws.cell(row=start - 1, column=2, value="By key view")
ws.cell(row=start - 1, column=2).font = Font(name=FONT, size=11, bold=True, color=ADB_BLUE)
for i, h in enumerate(["Key view", "Count", "Buildable now", "Of which blocked on something external"], 2):
    ws.cell(row=start, column=i, value=h)
style_header(ws, start, 5)

views_seen = []
for row in REG:
    if row[0] not in views_seen:
        views_seen.append(row[0])

r = start + 1
for v in views_seen:
    ws.cell(row=r, column=2, value=v)
    ws.cell(row=r, column=3, value=f"=COUNTIF('Requirements register'!$B$5:$B${reg_last},$B{r})")
    ws.cell(row=r, column=4,
            value=f"=COUNTIFS('Requirements register'!$B$5:$B${reg_last},$B{r},"
                  f"'Requirements register'!$H$5:$H${reg_last},\"Buildable now\")")
    ws.cell(row=r, column=5, value=f"=C{r}-D{r}")
    r += 1

vtot = r
ws.cell(row=vtot, column=2, value="Total")
ws.cell(row=vtot, column=3, value=f"=SUM(C{start + 1}:C{vtot - 1})")
ws.cell(row=vtot, column=4, value=f"=SUM(D{start + 1}:D{vtot - 1})")
ws.cell(row=vtot, column=5, value=f"=SUM(E{start + 1}:E{vtot - 1})")

body(ws, start + 1, vtot, 5, band=False)
for rr in range(start + 1, vtot + 1):
    for c in (3, 4, 5):
        ws.cell(row=rr, column=c).number_format = "#,##0"
        ws.cell(row=rr, column=c).alignment = Alignment(horizontal="center", vertical="top")
for c in (2, 3, 4, 5):
    cell = ws.cell(row=vtot, column=c)
    cell.font = Font(name=FONT, size=9, bold=True)
    cell.fill = FILL_TITLE

note_row = vtot + 2
ws.cell(row=note_row, column=2,
        value="Assumption behind these counts")
ws.cell(row=note_row, column=2).font = Font(name=FONT, size=10, bold=True)
ws.cell(row=note_row + 1, column=2,
        value="One register row is one requirement. Where a slide asks for the same figure at two levels, for example "
              "records declared bank-wide and again per department, it is recorded once at each level, because the two "
              "have different blockers. Where the deck and the Word file ask for the same thing, it is recorded once "
              "with both references. The register is therefore a count of work items, not a count of sentences in the "
              "two documents.")
ws.cell(row=note_row + 1, column=2).font = Font(name=FONT, size=9, color=GREY)
ws.cell(row=note_row + 1, column=2).alignment = Alignment(vertical="top", wrap_text=True)
ws.merge_cells(start_row=note_row + 1, start_column=2, end_row=note_row + 3, end_column=5)

# Every count on this sheet is a formula, which is what makes the sheet stay
# correct when a register row is edited. LibreOffice times out in the build
# environment (STATUS s12), so the formulas ship without cached values: Excel
# computes them on open, but a quick previewer will show these cells blank.
# The expected values are recorded here so a blank or a wrong one is obvious.
check_row = note_row + 5
ws.cell(row=check_row, column=2, value="Hand check, 13 August 2026")
ws.cell(row=check_row, column=2).font = Font(name=FONT, size=10, bold=True)
counts = {}
for row in REG:
    counts[row[6]] = counts.get(row[6], 0) + 1
expected = ", ".join(f"{v} {k}" for k, v in
                     sorted(counts.items(), key=lambda kv: -kv[1]))
ws.cell(row=check_row + 1, column=2,
        value="The counts above are formulas over the register, not typed numbers, so they follow any edit made "
              "there. LibreOffice cannot recalculate in the build environment (STATUS s12), so they ship with no "
              "cached value: Excel computes them the moment the file opens, but a lightweight previewer may show "
              "them blank. Counted by hand at build time, the register held "
              f"{len(REG)} requirements, being {expected}. If a figure above disagrees with that, the register has "
              "been edited since and the hand check is the stale one, not the formula.")
ws.cell(row=check_row + 1, column=2).font = Font(name=FONT, size=9, color=GREY)
ws.cell(row=check_row + 1, column=2).alignment = Alignment(vertical="top", wrap_text=True)
ws.merge_cells(start_row=check_row + 1, start_column=2, end_row=check_row + 4, end_column=5)

ws.sheet_view.showGridLines = False

wb.save(OUT)
print(f"wrote {OUT}: {len(REG)} register rows, {len(wb.sheetnames)} sheets")
