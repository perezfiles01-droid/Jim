import json, re, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
sys.path.insert(0, '/tmp/claude-0/-home-user-Jim/2ff9126d-9988-5801-be87-2b7c50c9e1d1/scratchpad')
from elements import ROWS, READY, SCAN, USAGE, BLOCK, DERIV
from tenant import (STATUS_OVERRIDE, STATUS_OVERRIDE_2, HOW_OVERRIDE_2, HOW_OVERRIDE_3,
                    tenant_note, FINDINGS, FINDINGS_2, FINDINGS_3, GAP1_PLAN)
FINDINGS = FINDINGS + FINDINGS_2 + FINDINGS_3

# Apply what the test tenant proved. Gap 2 is closed, so the elements blocked only by
# file size now wait on the scan like everything else.
ROWS = [r[:6] + [STATUS_OVERRIDE.get((r[0], r[2]), r[6])] + r[7:] for r in ROWS]
# Round 2: the real usage export moved three of them again.
ROWS = [r[:6] + [STATUS_OVERRIDE_2.get((r[0], r[2]), r[6])] + r[7:9]
        + [HOW_OVERRIDE_2.get((r[0], r[2]), r[9])] + r[10:] for r in ROWS]
# Round 3: the library screenshot traced Gap 1 to its cause. No status moves,
# but the fix is now specified rather than described.
ROWS = [r[:9] + [HOW_OVERRIDE_3.get((r[0], r[2]), r[9])] + r[10:] for r in ROWS]

OUT = '/home/user/Jim/EDRMS_Utilization_Report_Source_Data_v4.xlsx'

NAVY  = '0E1F35'
BLUE  = '0072BC'
LINE  = 'D9E2EC'
HEADF = PatternFill('solid', fgColor=NAVY)
SUBF  = PatternFill('solid', fgColor='F3F7FB')
BANDF = PatternFill('solid', fgColor='EAF1F8')
OKF   = PatternFill('solid', fgColor='E7F4E2')
AMBF  = PatternFill('solid', fgColor='FDF1E7')
REDF  = PatternFill('solid', fgColor='FBE9E9')
GRYF  = PatternFill('solid', fgColor='F0F2F5')
YELF  = PatternFill('solid', fgColor='FFFF00')

A     = 'Arial'
H     = Font(name=A, size=10, bold=True, color='FFFFFF')
B     = Font(name=A, size=10, bold=True)
N     = Font(name=A, size=10)
SM    = Font(name=A, size=9)
MONO  = Font(name='Consolas', size=9)
TITLE = Font(name=A, size=14, bold=True, color=NAVY)
SUBT  = Font(name=A, size=10, italic=True, color='5B6B7C')
thin  = Side(style='thin', color=LINE)
BOX   = Border(left=thin, right=thin, top=thin, bottom=thin)

STATUS_FILL = {READY: OKF, SCAN: AMBF, USAGE: AMBF, BLOCK: REDF, DERIV: GRYF}

wb = openpyxl.Workbook()

# =====================================================================
# Sheet 1: Read me
# =====================================================================
ws = wb.active
ws.title = 'Read me'
ws.sheet_view.showGridLines = False

def put(r, c, v, font=N, fill=None, align=None, wrap=False, border=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font
    if fill: cell.fill = fill
    cell.alignment = Alignment(horizontal=align or 'left', vertical='top', wrap_text=wrap)
    if border: cell.border = BOX
    return cell

ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 34
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 95

put(2, 2, 'EDRMS UTILIZATION REPORT', TITLE)
put(3, 2, 'Where every KPI and diagram gets its data, and what has to be built for the ones that cannot get it yet', SUBT)
put(4, 2, 'VERSION 4, 8 August 2026. Read against Database_Design_12.03, the deployed public."Records" table in drm-npr, '
          'the 7rkd12 test tenant, its five usage CSVs, and the Annual Meetings library with its metadata columns switched on.', SM)

r = 6
put(r, 2, 'THE ONE THING TO TAKE AWAY', B); r += 1
for line in [
 'public."Records" contains DECLARED RECORDS ONLY, about 1,990 in UAT. It holds no row at all for an undeclared document.',
 'So every figure about declared records can be produced today. Every figure that needs a denominator cannot, because the',
 'denominator does not exist in any system yet. That single fact explains most of the red rows in this workbook.',
]:
    put(r, 2, line, N); r += 1

r += 1
put(r, 2, 'THE TWO TABLES THE REPORT NEEDS', B); r += 1
put(r, 2, 'Table', B, SUBF, border=True)
put(r, 3, 'One row per', B, SUBF, border=True)
put(r, 4, 'Why it has to exist separately', B, SUBF, border=True); r += 1
for t, g, w in [
 ('rpt.utilization_report', 'document',
  'Holds declared AND undeclared documents together. Without the undeclared ones there is no denominator and no declaration rate.'),
 ('rpt.utilization_site_activity', 'SharePoint site',
  'A site with no documents in it has no document rows, so it would vanish from the site count. Visit counts are also per site, and repeating them on every document row makes any total nonsense.'),
]:
    put(r, 2, t, MONO, border=True)
    put(r, 3, g, N, border=True)
    put(r, 4, w, N, wrap=True, border=True)
    ws.row_dimensions[r].height = 30
    r += 1

r += 1
put(r, 2, 'STATUS VOCABULARY, used on every sheet', B); r += 1
put(r, 2, 'Status', B, SUBF, border=True)
put(r, 3, 'Count', B, SUBF, border=True)
put(r, 4, 'What it means and what unblocks it', B, SUBF, border=True); r += 1
STATUS_ROW = r
for st, meaning in [
 (READY, 'The data is in public."Records" today. Nothing to build. Point the query at it.'),
 (SCAN,  'Needs the weekly Microsoft Graph scan of every library in every compliant site. ONE job unblocks all of these at once.'),
 (USAGE, 'Needs the Microsoft 365 SharePoint site usage report, written weekly into the Site Activity Table. Independent of everything else.'),
 (BLOCK, 'Cannot be produced at all until a gap is closed. See the Gaps and actions sheet.'),
 (DERIV, 'Reads other elements rather than the database. Inherits their status.'),
]:
    put(r, 2, st, N, STATUS_FILL[st], border=True)
    c = ws.cell(row=r, column=3, value=f"=COUNTIF('KPIs and Diagrams'!$H:$H,$B{r})")
    c.font = B; c.border = BOX; c.alignment = Alignment(horizontal='center')
    put(r, 4, meaning, N, wrap=True, border=True)
    ws.row_dimensions[r].height = 26
    r += 1
put(r, 2, 'Total elements mapped', B, BANDF, border=True)
c = ws.cell(row=r, column=3, value=f'=SUM(C{STATUS_ROW}:C{r-1})')
c.font = B; c.fill = BANDF; c.border = BOX; c.alignment = Alignment(horizontal='center')
put(r, 4, 'Every KPI, chart, filter, drill level and label in the prototype.', N, BANDF, border=True)
r += 1
put(r, 2, 'The counts above are live formulas over the KPIs and Diagrams sheet, so they cannot drift from it. '
          'They populate the moment the file opens in Excel.', SM)
r += 2

put(r, 2, 'HOW TO USE THIS WORKBOOK', B); r += 1
for tab, use in [
 ('KPIs and Diagrams', 'The main sheet. One row per visual element, with the calculation in words, the status, the exact columns, and a runnable query.'),
 ('Database columns', 'All 58 columns of the two tables, whether each exists today, and exactly where in the workbook and the live database.'),
 ('Gaps and actions', 'The three gaps, what each one blocks, who owns it, and what closing it involves.'),
 ('Gap 1 action plan', 'The department and division fix, step by step, saying who does what and how long. Written to be forwarded.'),
 ('Decisions needed', 'Questions only RAC can answer. Each one changes a number, so leaving them open means the report cannot be signed off.'),
]:
    put(r, 2, tab, B, border=True)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    put(r, 3, use, N, wrap=True, border=True)
    ws.row_dimensions[r].height = 26
    r += 1

r += 1
put(r, 2, 'WHAT CHANGED IN VERSION 4, AFTER LOOKING INSIDE THE LIBRARY', B); r += 1
for line in [
 'THE STATUS COUNTS DO NOT MOVE IN THIS VERSION, and that is the honest result. What changed is that the biggest gap now has a',
 '   traced cause and a specified fix rather than a description. See the new Gap 1 action plan sheet.',
 'GAP 1 CAUSE FOUND. The SharePoint column ADBDepartmentOwner EXISTS and is EMPTY on every row. ADBMeta is empty because',
 '   nobody fills in the SharePoint column. The fix is upstream of the database, so no amount of database work solves it.',
 'DIVISION IS WORSE, and nobody had noticed. There is NO ADBDivisionOwner column anywhere. It was never created, not merely left empty.',
 'A SIMPLIFICATION: Item is a Record is populated and reliable, so IsDeclaredRecord can be read straight from SharePoint.',
 '   Keep the match against Records as a CROSS CHECK, which finally makes the label-without-declaring problem measurable.',
 'GAP 3a CLOSED. Graph returns createdDateTime with the site list. It unblocks no element on its own, because everything that',
 '   uses the created date also needs the compliance rule or department. It removes future work rather than moving the scoreboard.',
 'ROOT WEB TEMPLATE IS RULED OUT for the compliance rule: the EDRMS sites report Team Site, and so do 2,078 others.',
 'REASSURING: Retention Label Applied For Calculated matches the native compliance timestamp on every row, so it is derived, not typed.',
]:
    put(r, 2, '- ' + line, N)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
r += 1
put(r, 2, 'WHAT CHANGED IN VERSION 3, AFTER READING THE REAL USAGE EXPORT', B); r += 1
for line in [
 'ACTIVE USERS IS DISPROVEN. I said the usage report supplies unique viewers. It has 23 columns and none of them does.',
 '   Visited Page Count is unique PAGES, not people. Two elements move from Needs the usage feed to Blocked.',
 'TOTAL DOCUMENTS GOT MUCH CHEAPER. File Count in the usage report gives it per site, 26,660 across 977 sites in the test export.',
 '   That headline number no longer waits on the document scan, only on the Gap 3 compliance rule.',
 'A NEW BLOCKER: Site URL came back EMPTY on all 2,359 rows. Fixable by resolving Site Id through Graph, but nothing per site works until then.',
 'A LOAD RULE CONFIRMED: the detail file includes deleted sites, 657 of 2,359. Filter them or every count runs 28 percent high.',
 'A FRESHNESS LIMIT: Microsoft data ran three days behind on the day it was exported. The Data as of line must show THEIR date, not the job run time.',
 'A THIRD CANDIDATE for the compliance rule: Root Web Template is populated per site. Checkable once Site URL is resolved.',
]:
    put(r, 2, '- ' + line, N)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
r += 1
put(r, 2, 'WHAT CHANGED IN VERSION 2, AFTER INSPECTING THE TEST TENANT', B); r += 1
for line in [
 'GAP 2 IS CLOSED. Microsoft Graph returns file size on every file, verified on real items rather than taken from documentation.',
 '   Six elements moved from Blocked to Needs the scan. Gap 2 is no longer a separate problem, it is part of the scan.',
 'GAP 3 HAS A MUCH BETTER ANSWER. The Retention Label Mapping list is already the enforced registry of what EDRMS manages.',
 '   Declaration into an unmapped library already fails. The rule exists, it just has not been called a compliance rule.',
 'A NEW BUG WAS CAUGHT before the scan was written: folder sizes are cumulative, so summing every item double counts storage.',
 'A NEW DECISION SURFACED: documents in unmapped libraries are undeclarable, not undeclared. Counting them caps the rate below 100 percent.',
 'GAPS 1 AND 3a ARE UNCHANGED, and the usage feed still needs the real tenant.',
 'Every element now carries a column saying what the test tenant can and cannot settle for it.',
]:
    put(r, 2, '- ' + line, N)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
r += 1
put(r, 2, 'WHAT CHANGED FROM THE DRAFT source_data.xlsx', B); r += 1
for line in [
 'Yes / No for "data exists" was too blunt. Almost every honest answer is "yes for declared records, no for the rest", so a five value status replaced it.',
 'Added a column saying HOW to source anything missing, naming the system: Graph, the admin centre, the usage reports, AvePoint, or a RAC decision.',
 'Added the calculation in words beside each element, because no figure comes from a single column: it is a count, sliced by a grouping column, filtered by another.',
 'Added the elements the draft omitted: the four drill levels, both date filters, the sort controls, the Overview dashboard, and the Data as of line.',
 'Replaced the #VALUE! rows, which had lost their diagram names.',
 'Corrected four column types against the deployed CREATE TABLE: Id and ListId are uuid, EDRMSDuration is text, DeclarationType is already live.',
]:
    put(r, 2, '- ' + line, N)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1

# =====================================================================
# Sheet 2: KPIs and Diagrams
# =====================================================================
ws = wb.create_sheet('KPIs and Diagrams')
ws.sheet_view.showGridLines = False
HEADERS = ['#', 'Dashboard', 'Section', 'Element', 'Type', 'Shows in the prototype',
           'How the number is produced', 'Status', 'Source table', 'Column/s in the database',
           'If it is not there, how to source it', 'Can the TEST TENANT settle it?', 'Query sample']
WIDTHS  = [4, 20, 30, 34, 15, 26, 52, 24, 30, 30, 62, 62, 60]

ws.cell(row=1, column=1, value='EVERY KPI AND DIAGRAM IN THE PROTOTYPE, WHERE ITS DATA COMES FROM, AND WHAT THE TEST TENANT CAN SETTLE').font = TITLE
ws.cell(row=2, column=1,
        value='Version 2, revised 8 August 2026 after direct inspection of the 7rkd12 test tenant. Read Status first, then the test tenant column.').font = SUBT
for i, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=4, column=i, value=h)
    c.font = H; c.fill = HEADF; c.border = BOX
    c.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
ws.row_dimensions[4].height = 30

row = 5
last_dash = None
for n, e in enumerate(ROWS, start=1):
    dash, sec, el, typ, shows, calc, st, tbl, cols, how, sql = e
    band = dash != last_dash
    last_dash = dash
    vals = [n, dash, sec, el, typ, shows, calc, st, tbl, cols, how,
            tenant_note(dash, el, st, how), sql]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BOX
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        if i == 13:
            c.font = MONO
        elif i in (2, 4):
            c.font = B
        else:
            c.font = N
        if i == 8:
            c.fill = STATUS_FILL[st]
        elif band and i <= 3:
            c.fill = BANDF
    ws.row_dimensions[row].height = max(30, 11 * (sql.count('\n') + 1))
    row += 1

ws.freeze_panes = 'D5'
ws.auto_filter.ref = f'A4:M{row-1}'
LASTROW = row - 1

# =====================================================================
# Sheet 3: Database columns
# =====================================================================
d = json.load(open('/tmp/claude-0/-home-user-Jim/2ff9126d-9988-5801-be87-2b7c50c9e1d1/scratchpad/gen_out.js'))
def clean(t):
    t = re.sub(r'<[^>]+>', '', t)
    return t.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')

STLABEL = {'have': 'In the database', 'derived': 'Derived at load', 'new': 'New column',
           'planned': 'Designed, not built', 'gap': 'Missing, blocks a figure'}
STFILL  = {'have': OKF, 'derived': OKF, 'new': GRYF, 'planned': AMBF, 'gap': REDF}

ws = wb.create_sheet('Database columns')
ws.sheet_view.showGridLines = False
ws.cell(row=1, column=1, value='EVERY COLUMN OF THE TWO TABLES').font = TITLE
ws.cell(row=2, column=1,
        value='Columns keep the name they already have in the EDRMS database. JSON keys inside FileMeta, EDRMSMeta and ADBMeta '
              'become real columns keeping the key name. Workbook rows cite Database_Design_12.03, sheet "4 Records", the 2026.1 block at rows 56 to 82.').font = SUBT
CH = ['#', 'Table', 'Column', 'Type', 'What it holds', 'Which figure it produces', 'Status',
      'Where it is today', 'How to source it']
CW = [4, 26, 30, 14, 46, 46, 22, 52, 58]
for i, (h, w) in enumerate(zip(CH, CW), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=4, column=i, value=h)
    c.font = H; c.fill = HEADF; c.border = BOX
    c.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
ws.row_dimensions[4].height = 28

row = 5
for tname, cols, uses in [('Utilization Report Table', d['c1'], d['u1']),
                          ('Site Activity Table',      d['c2'], d['u2'])]:
    for i, (cdef, cuse) in enumerate(zip(cols, uses), start=1):
        name, typ, desc, flag = cdef
        _, fig, st, now, how = cuse
        vals = [i, tname, name, typ, clean(desc), clean(fig), STLABEL[st], clean(now), clean(how)]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=j, value=v)
            c.border = BOX
            c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            c.font = MONO if j == 3 else N
            if j == 7: c.fill = STFILL[st]
            if j == 2 and i == 1: c.fill = BANDF
        ws.row_dimensions[row].height = 28
        row += 1
ws.freeze_panes = 'D5'
ws.auto_filter.ref = f'A4:I{row-1}'

# =====================================================================
# Sheet 4: Gaps and actions
# =====================================================================
ws = wb.create_sheet('Gaps and actions')
ws.sheet_view.showGridLines = False
ws.cell(row=1, column=1, value='THE THREE GAPS, AND THE TWO FEEDS').font = TITLE
ws.cell(row=2, column=1, value='Nothing else in this workbook is blocked by anything but these five items.').font = SUBT
GH = ['#', 'What is missing', 'What it blocks', 'How many elements', 'Where to get it', 'Owner', 'Size of the job']
GW = [4, 34, 46, 16, 74, 16, 30]
for i, (h, w) in enumerate(zip(GH, GW), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=4, column=i, value=h)
    c.font = H; c.fill = HEADF; c.border = BOX
    c.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
ws.row_dimensions[4].height = 28

GAPS = [
 ('Gap 1', 'Department and division on anything',
  'Every department bar, the treemap grouping, all six department filters, both Overview top 5 lists, and drill levels 1 and 2.',
  '=COUNTIF(\'KPIs and Diagrams\'!$K:$K,"Gap 1*")',
  'The vocabularies ARE populated in the SharePoint term store. What is missing is the link from a site to its department. '
  'Attach it to the SITE, about 1,057 values, and let every document inherit it. Check AvePoint Cloud Governance first: it '
  'records the requesting department at provisioning, which would make this an export rather than data entry.',
  'IT, then RAC',
  'Large, but 1,057 rows not 3.47 million'),
 ('Gap 2', 'File size is not captured',
  'The 46.7 GB headline, storage by format, average file size, Largest Libraries, and the Overview storage donut.',
  '=COUNTIF(\'KPIs and Diagrams\'!$K:$K,"Gap 2*")',
  'CLOSED AS A GAP. Verified in the test tenant on 8 Aug 2026: Microsoft Graph returns size on every file unprompted, '
  'observed from 1,506 bytes up to 2,338,767. The scan supplies every storage figure at no extra cost. The FileMeta.FileSize '
  'key is now OPTIONAL, wanted only for per record detail at declaration time. WARNING: folders are returned as items with a '
  'CUMULATIVE size, so the scan must filter to files only or storage doubles.',
  'Development team',
  'Closed, folded into the scan'),
 ('Gap 3', 'No site created date, and no rule for EDRMS compliant',
  'The 1,057 compliant sites, the treemap, its date range, and the definition of which documents count toward the 3.47M.',
  '=COUNTIF(\'KPIs and Diagrams\'!$K:$K,"Gap 3*")',
  'THE RULE MAY ALREADY EXIST. The test tenant holds "No mapping library1", a real library full of real documents where '
  'declaration FAILS with "No Library and Retention Label Mapping found". The Retention Label Mapping list is therefore already '
  'the enforced registry of what EDRMS manages. PROPOSED RULE: a library is in scope if it appears in that list; a site is '
  'compliant if it holds at least one mapped library. No new list, no new process, and RAC already maintains it. '
  'The created date is still a separate read from the SharePoint admin centre, Active sites, Created column.',
  'RAC confirms the rule, IT exports the date',
  'Much smaller than it looked'),
 ('Feed 1', 'The weekly document scan',
  'Total Documents 3.47M, the declaration rate, every per library figure, and both date filters on the documents panel.',
  '=COUNTIF(\'KPIs and Diagrams\'!$H:$H,"Needs the document scan")',
  'A scheduled Microsoft Graph job walking every document library in every compliant site, matched to Records on ListId + ItemId. '
  'Roughly 3,500 API calls, comfortable overnight, impossible on demand. NOTE: if you only need the total count and not the list, '
  'SharePoint REST ItemCount per library gives it without enumerating a single item, which is far cheaper and unblocks both '
  'headline numbers on its own.',
  'Development team',
  'The largest single piece of work'),
 ('Feed 2', 'The Microsoft 365 usage feed',
  'Active Departmental Sites, Active Users, the 7 / 30 / 90 windows, and last activity.',
  '=COUNTIF(\'KPIs and Diagrams\'!$H:$H,"Needs the usage feed")',
  'EXPORTED AND READ, 8 Aug 2026. getSharePointSiteUsageDetail gives one row per site with 23 columns: Last Activity Date, '
  'Page View Count, Visited Page Count, Storage Used, Owner, Is Deleted, File Count and Root Web Template. THREE FINDINGS. '
  '1) There is NO unique viewer column, so Active Users is blocked, not merely waiting. 2) File Count supplies Total Documents, '
  'so that KPI leaves the scan. 3) Site URL was EMPTY on all 2,359 rows: resolve Site Id via Graph /sites/{siteId} and filter '
  'Is Deleted, which was 657 of 2,359.',
  'IT',
  'Small, but Site URL must be fixed first'),
]
row = 5
for g in GAPS:
    for j, v in enumerate(g, start=1):
        c = ws.cell(row=row, column=j, value=v)
        c.border = BOX
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        c.font = B if j in (1, 2) else N
        if j == 4: c.alignment = Alignment(horizontal='center', vertical='top'); c.font = B
        if j == 1: c.fill = REDF if v.startswith('Gap') else AMBF
    ws.row_dimensions[row].height = 74
    row += 1

row += 1
ws.cell(row=row, column=2, value='SUGGESTED ORDER OF WORK').font = B
row += 1
for n, (what, why) in enumerate([
 ('Export the SharePoint site usage report to CSV',
  'Ten minutes, no development, and it unblocks two KPI cards. Do it this week.'),
 ('Ask IT whether AvePoint Cloud Governance holds the department per site',
  'One question. If the answer is yes, Gap 1 stops being the largest item in the report.'),
 ('Get RAC to define EDRMS compliant, in writing',
  'It appears in two dashboard titles. Nothing that says "compliant" can be signed off until it exists.'),
 ('Run SharePoint REST ItemCount per library for a total document count',
  'Gives the 3.47M and the declaration rate without building the full scan. Days, not weeks.'),
 ('Build the weekly Graph scan',
  'Everything else. Do it after the definitions above are settled, or it gets built twice.'),
], start=1):
    ws.cell(row=row, column=2, value=f'{n}. {what}').font = B
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    c = ws.cell(row=row, column=3, value=why)
    c.font = N; c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 24
    row += 1

# =====================================================================
# Sheet: Gap 1 action plan
# =====================================================================
ws = wb.create_sheet('Gap 1 action plan')
ws.sheet_view.showGridLines = False
ws.cell(row=1, column=1, value='DEPARTMENT AND DIVISION: THE FIX, STEP BY STEP').font = TITLE
ws.cell(row=2, column=1,
        value='Written to be forwarded. The largest gap in the report, and the smallest version of the work that closes it.').font = SUBT

r = 4
ws.cell(row=r, column=1, value='WHY IT IS SMALLER THAN IT LOOKS').font = B
r += 1
for line in [
 'The obvious fix is to tag every document with a department. That is 3.47 million documents, and nobody is going to do it.',
 'But the report\'s own drill down goes Department to Division to Site to Library. A site sits BENEATH a division, which sits',
 'beneath a department. So the report never needed department on each document. It needs it on each SITE, about 1,057 of them,',
 'and every document inside inherits it. That is the whole trick: 1,057 rows instead of 3.47 million.',
]:
    c = ws.cell(row=r, column=1, value=line); c.font = N
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

r += 1
ws.cell(row=r, column=1, value='WHAT WE FOUND, 8 AUGUST 2026').font = B
r += 1
for line in [
 'DEPARTMENT: the SharePoint column ADBDepartmentOwner exists on the library, is named exactly like the database key, and is',
 '   EMPTY on every row. So the database is empty because the SharePoint column is empty. The fix is upstream of the database.',
 'DIVISION: there is NO column at all. Not empty, never created. So the site mapping is the only route short of building a new',
 '   managed metadata column across every library at ADB, which is a far bigger change.',
 'BOTH are solved by the SAME spreadsheet. One list, one extra column, both problems closed.',
]:
    c = ws.cell(row=r, column=1, value=line); c.font = N
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

r += 2
PH = ['Step', 'Who', 'What exactly', 'How long', 'Notes and warnings']
PW = [7, 26, 74, 22, 88]
for i, (h, w) in enumerate(zip(PH, PW), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=r, column=i, value=h)
    c.font = H; c.fill = HEADF; c.border = BOX
    c.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
ws.row_dimensions[r].height = 24
r += 1
for step, who, what, how, note in GAP1_PLAN:
    for j, v in enumerate([step, who, what, how, note], start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.border = BOX
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        c.font = B if j in (1, 2) else N
        if j == 1:
            c.fill = OKF if step in ('0', '2b') else BANDF
            c.alignment = Alignment(horizontal='center', vertical='top')
    ws.row_dimensions[r].height = 62
    r += 1

r += 1
ws.cell(row=r, column=1, value='THE ONE QUESTION THAT DECIDES EASY OR HARD').font = B
r += 1
c = ws.cell(row=r, column=1,
    value='"Is it acceptable that every document in a CWRD site counts as CWRD?"  If yes, everything above holds: one list, '
          'small job, done. If no, if a department site can hold documents belonging to other departments, the easy route '
          'breaks and documents must be tagged individually. That is the difference between a week and a quarter. The report\'s '
          'drill down already assumes the answer is yes, but it should be confirmed, not assumed.')
c.font = N; c.alignment = Alignment(wrap_text=True, vertical='top')
ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=5)
ws.row_dimensions[r].height = 30
r += 3

ws.cell(row=r, column=1, value='DRAFT NOTE TO THE CLIENT').font = B
r += 1
for line in [
 'To show the Utilization Report by department and division, we need one thing from you: a list of SharePoint sites with the',
 'department and division each one belongs to. We will send the site list with the URLs already filled in, so only two columns',
 'need completing. Roughly 1,000 rows, and most of them fill in bulk once sorted.',
 '',
 'BEFORE THAT: could you check whether AvePoint Cloud Governance recorded the requesting department when each site was',
 'provisioned? If it did, we may be able to export the list and skip the manual step entirely.',
 '',
 'ONE THING TO CONFIRM: is it acceptable for every document in a department\'s site to be counted under that department? The',
 'report\'s drill down assumes this, but we would rather have it confirmed.',
 '',
 'NOTE: Division does not currently exist as a field in SharePoint at all, so this list is the only way to report on it.',
]:
    c = ws.cell(row=r, column=1, value=line); c.font = N
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

# =====================================================================
# Sheet 5: Decisions needed
# =====================================================================
ws = wb.create_sheet('Decisions needed')
ws.sheet_view.showGridLines = False
ws.cell(row=1, column=1, value='DECISIONS ONLY RAC CAN MAKE').font = TITLE
ws.cell(row=2, column=1,
        value='Each one changes a number on the dashboard. Yellow cells are for you to fill in.').font = SUBT
DH = ['#', 'Question', 'Why it matters', 'Recommendation', 'RAC decision', 'Date agreed']
DW = [4, 44, 62, 62, 30, 14]
for i, (h, w) in enumerate(zip(DH, DW), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=4, column=i, value=h)
    c.font = H; c.fill = HEADF; c.border = BOX
    c.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
ws.row_dimensions[4].height = 28

DEC = [
 ('What makes a SharePoint site EDRMS compliant?',
  'Three figures cannot be produced without it and the phrase appears in two dashboard titles.',
  'UPDATED after the test tenant. Use the Retention Label Mapping list you already maintain: a library is in scope if it appears '
  'there, a site is compliant if it holds at least one mapped library. The application ALREADY enforces this, since declaration '
  'into an unmapped library fails. No new list and no new process, just a name for a rule that exists.'),
 ('Do documents in UNMAPPED libraries count in Total Documents?',
  'NEW, found in the test tenant. "No mapping library1" holds real documents that can never be declared, because the application '
  'rejects the attempt. They are undeclarable, not undeclared.',
  'No. Exclude them, and report them separately as a backlog figure. Counting them holds the declaration rate down permanently '
  'and means it can never reach 100 percent, which will be read as failure rather than as scope.'),
 ('Does a document declared twice count once or twice?',
  'UAT returns 1,990 rows against 1,984 distinct DocumentId. Six rows, but two reports using two rules is how one KPI ends up with two values.',
  'Once. Count distinct ListId + ItemId. Declaring the same file twice does not put two documents under records management.'),
 ('What counts as a document in Total Documents?',
  'Folders, version history and system files each move the 3.47M substantially, and the declaration rate with it.',
  'Compliant sites only, files only, exclude folders, exclude version history, current items only. Print the definition beside the KPI.'),
 ('Which file extensions map to which of the eight format groups?',
  'Without the list, FormatGroup has no rule and the Format and Storage dashboard cannot be built.',
  'A short list signed off once. The eight groups are already fixed in the prototype: PDF, Word, Excel, PowerPoint, Email, Image files, Video files, All other formats.'),
 ('How often should the report refresh?',
  'The document scan sets the cost. Daily is possible but roughly seven times the API load.',
  'Weekly. No decision in a utilization report turns on this morning versus last Tuesday.'),
 ('Is the 90 day Active Users fallback acceptable?',
  'Microsoft does not return unique viewers at 90 days. The dashboard currently shows site visits instead, with an amber note.',
  'Accept it and keep the note. The alternative is removing the 90 day option, which hides the limitation rather than stating it.'),
]
row = 5
for n, (q, why, rec) in enumerate(DEC, start=1):
    ws.cell(row=row, column=1, value=n).font = B
    for j, v in [(2, q), (3, why), (4, rec)]:
        c = ws.cell(row=row, column=j, value=v)
        c.font = B if j == 2 else N
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c.border = BOX
    for j in (1, 5, 6):
        c = ws.cell(row=row, column=j)
        c.border = BOX
        if j in (5, 6): c.fill = YELF
    ws.row_dimensions[row].height = 52
    row += 1
ws.cell(row=row + 1, column=2,
        value='Yellow cells are the ones to fill in. Everything else is a statement of fact or a recommendation.').font = SUBT

# =====================================================================
# Sheet: Test tenant findings
# =====================================================================
ws = wb.create_sheet('Test tenant findings')
ws.sheet_view.showGridLines = False
ws.cell(row=1, column=1, value='WHAT THE TEST TENANT SETTLED, WHAT IT DISPROVED, AND WHAT IT COULD NOT').font = TITLE
ws.cell(row=2, column=1,
        value='Checked 8 August 2026 as JimTest@7rkd12.onmicrosoft.com against org_csd_1.4testsite and its 22 libraries. '
              'Findings 1 to 9 come from browsing the tenant; 10 to 17 from reading the five usage CSVs exported from its '
              'admin centre. Everything below was observed directly, not inferred from documentation.').font = SUBT
FH = ['#', 'What was checked', 'What was found', 'What it changes', 'Outcome']
FW = [4, 40, 70, 78, 26]
for i, (h, w) in enumerate(zip(FH, FW), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=4, column=i, value=h)
    c.font = H; c.fill = HEADF; c.border = BOX
    c.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
ws.row_dimensions[4].height = 28

OUTFILL = {'Resolved': OKF, 'Confirmed': OKF, 'Proposal, needs RAC sign off': AMBF,
           'New risk, now known': AMBF, 'Decision needed': AMBF, 'Still blocked': REDF,
           'Disproven': REDF, 'New blocker, has a fix': REDF,
           'Promising, not yet checkable': AMBF, 'Constraint, now known': AMBF,
           'Opportunity': GRYF}
row = 5
for n, (chk, found, changes, outcome) in enumerate(FINDINGS, start=1):
    for j, v in enumerate([n, chk, found, changes, outcome], start=1):
        c = ws.cell(row=row, column=j, value=v)
        c.border = BOX
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        c.font = B if j == 2 else N
        if j == 5: c.fill = OUTFILL.get(outcome, GRYF)
    ws.row_dimensions[row].height = 78
    row += 1

row += 1
ws.cell(row=row, column=2, value='THE HEADLINE').font = B
row += 1
for line in [
 'The test tenant closed one gap outright, produced a far better answer to a second, and caught a bug before it was written.',
 'It cannot close the department link, the site created date, or the usage feed. Those three need the real tenant or a RAC decision.',
 'Its greatest value is as a rehearsal ground: the data there is adversarial, not tidy, so a scan that runs clean here will run clean at ADB.',
]:
    ws.cell(row=row, column=2, value='- ' + line).font = N
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    row += 1

# Put the findings right after the Read me, since they are the point of version 2.
wb.move_sheet('Test tenant findings', offset=-(len(wb.sheetnames) - 2))

wb.save(OUT)
print('written', OUT, 'elements', len(ROWS))
