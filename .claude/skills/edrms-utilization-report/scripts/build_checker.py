"""Build EDRMS_Utilization_Report_Checker_2026-08-14.xlsx.

A checker for the report: one sheet per dashboard, every figure on it, where
each figure comes from, which database column carries it, how it is computed,
whether it can be tested in the tenant today, and what to ask the client.

Never hand edit the workbook. Edit this file and rerun:

    python3 .claude/skills/edrms-utilization-report/scripts/build_checker.py

The dashboard inventory is kept in step with index.html by hand. If a panel is
added there, add its rows here. check_data.js guards the reverse direction, that
nothing in the requirements loses its home on the page.
"""

import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utilization_tables import TABLES, SOURCE

OUT = "EDRMS_Utilization_Report_Checker_2026-08-14.xlsx"
FONT = "Arial"
ADB_BLUE = "0067B1"
GREY = "6B7785"

FILL_HEAD = PatternFill("solid", fgColor=ADB_BLUE)
FILL_TITLE = PatternFill("solid", fgColor="E8EFF6")
FILL_SECT = PatternFill("solid", fgColor="F2F6FA")
THIN = Side(style="thin", color="D4DCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Status drives the colour, and is the column to filter on.
STATUS_FILL = {
    "Buildable now":            PatternFill("solid", fgColor="D9EFD9"),
    "Needs department list":    PatternFill("solid", fgColor="E4F0DA"),
    "Needs new column or join": PatternFill("solid", fgColor="FFF2CC"),
    "Needs application change": PatternFill("solid", fgColor="FFE4CC"),
    "Needs reference list":     PatternFill("solid", fgColor="FFF7DA"),
    "Needs new data source":    PatternFill("solid", fgColor="F8D7D7"),
    "Decision needed":          PatternFill("solid", fgColor="E2DCF0"),
}
STATUS_MEANING = {
    "Buildable now": "Sourceable from the 73 column design as it stands",
    "Needs department list": "Buildable the moment RAC supply the site to department list",
    "Needs new column or join": "The design must change: a new column, or a join that does not exist",
    "Needs application change": "Needs a new field in the EDRMS application, a development change request",
    "Needs reference list": "Not a measurement. Somebody has to maintain a list for it to exist",
    "Needs new data source": "A system we have not touched, or data no API returns",
    "Decision needed": "The client must answer before it can be scoped",
}

HDR = ["S/N", "Section", "Element", "Type", "What the figure is",
       "Source system", "Report or file to open",
       "Step by step: how to get that file",
       "Exact column heading in that file",
       "How the number is worked out",
       "Database table", "Database column",
       "In the design?", "In the tenant today?",
       "Effort", "Status", "Question to ask the client, and why",
       "Questions", "What we found", "FINAL REMARKS"]
# The last three are the client's own review columns, added to the workbook by
# hand on 14 Aug and folded back into the generator here so they survive a
# rebuild. Widths follow what he set, except "What we found", widened from 21
# because at that width a paragraph becomes a 400px tall row.
WID = [5, 24, 28, 9, 34, 18, 30, 62, 40, 60, 20, 28, 13, 15, 9, 24, 58,
       46, 40, 69]
COL_STATUS = 16          # P, the column the Summary counts and the filter uses

# Shorthands used a lot below.
T1, T2, T3, T4 = "1 Utilization Report", "2 Site Activity", "3 User Activity", "4 File Plan"
EDRMS, GRAPH, USAGE, ANALYTICS, TERMS, LIST, NONE = (
    "EDRMS database", "Microsoft Graph", "M365 usage report", "Site analytics",
    "Term store", "SharePoint list", "No source")
CG = "AvePoint Cloud Governance"
BUILD, DEPT, JOIN, APPCH, REFL, NEWSRC, DECIDE = (
    "Buildable now", "Needs department list", "Needs new column or join",
    "Needs application change", "Needs reference list", "Needs new data source",
    "Decision needed")

# ---------------------------------------------------------------------------
# RECIPES. One per real export or system. The steps are click by click, with
# the link, because the whole point of this workbook is that somebody can
# verify a number without asking anyone how.
#
# The column headings quoted below were read off the two evidence exports in
# this repo, not remembered:
#   evidence_SharePointSiteUsageDetail_2026-08-12.csv   2,575 rows, 23 columns
#   evidence_SharePointActivityUserDetail_2026-08-12.csv   30 rows, 12 columns
# ---------------------------------------------------------------------------
RC = {}
RC["usage"] = (USAGE, "SharePoint site usage report, exported as CSV",
 "1. Go to https://admin.microsoft.com/#/reportsUsage/SharePointSiteUsage\n"
 "   (tenant independent URL: signing in as an ADB reports reader gives ADB figures)\n"
 "2. Sign in as a Microsoft 365 reports reader or global reader\n"
 "3. Set the period selector to 30 days\n"
 "4. Click Export, top right. A CSV downloads, named like "
 "SharePointSiteUsageDetail_2026-08-12.csv\n"
 "5. Open it in Excel. It has 23 columns and one row per site, 2,575 rows in the test tenant\n"
 "6. WARNING: the Site URL column is EMPTY on every row. Use Site Id and match it to a "
 "site list from Graph if you need the address")
RC["user"] = (USAGE, "SharePoint activity user detail report, exported as CSV",
 "1. Go to https://admin.microsoft.com/#/reportsUsage/SharePointActivity\n"
 "   (tenant independent URL: signing in as an ADB reports reader gives ADB figures)\n"
 "2. Sign in as a Microsoft 365 reports reader or global reader\n"
 "3. Set the period selector to 30 days\n"
 "4. Click Export. A CSV downloads, named like "
 "SharePointActivityUserDetail_2026-08-12.csv\n"
 "5. Open it in Excel. It has 12 columns and one row per LICENSED user, 30 rows in the "
 "test tenant. It is NOT a list of active users, which is the mistake to avoid")
RC["db"] = (EDRMS, "The EDRMS database, drm-npr, table public.\"Records\"",
 "1. Connect to the drm-npr PostgreSQL database with any SQL client\n"
 "2. The table is public.\"Records\", quoted because of the capital R\n"
 "3. It holds DECLARED RECORDS ONLY, about 1,990 rows in UAT. There is no row for a "
 "document that was never declared, which is why every rate needs the scan built first")
RC["graph"] = (GRAPH, "Microsoft Graph, run interactively",
 "1. Go to https://aka.ms/ge, the Graph Explorer\n"
 "2. Sign in with a tenant account and consent to the scope named in the next column\n"
 "3. Paste the call from the Exact column heading cell into the address bar and click Run query\n"
 "4. Read the field named in that same cell out of the JSON response")
RC["cg"] = (CG, "Cloud Governance Workspace report, exported as CSV",
 "0. FOR ADB PRODUCTION: same product, same screen, same columns. Only the AvePoint Online "
 "Services instance differs. Ask Leah Bancale for access or for the export itself\n"
 "1. Open AvePoint Online Services, then Cloud Governance\n"
 "2. Left nav: Directory, then Workspace report\n"
 "3. Click Export report, top left. The file is queued, not downloaded immediately\n"
 "4. Collect it from Job monitor in the left nav when the job shows green\n"
 "5. The export carries 93 columns and one row per workspace, 1,209 in the test tenant. "
 "45 of those columns are empty on every row, so check population before relying on one\n"
 "6. The column that matters is 'EDRMS Site Type'. Filter to rows where it is not blank")

RC["spadmin"] = (USAGE, "SharePoint admin centre, Active sites",
 "1. Go to https://7rkd12-admin.sharepoint.com\n"
 "2. Open Sites, then Active sites\n"
 "3. Use the column chooser to add the columns you need, then Export to CSV")
RC["purview"] = (LIST, "Purview file plan",
 "1. Go to https://purview.microsoft.com\n"
 "2. Open Solutions, Records management, then File plan\n"
 "3. Click Export to download the label list. 53 labels in the test tenant, a flat list "
 "with no hierarchy")
RC["terms"] = (TERMS, "SharePoint term store",
 "1. Go to https://7rkd12-admin.sharepoint.com\n"
 "2. Open Content services, then Term store\n"
 "3. Expand the groups. 19 term groups were visible on 14 Aug 2026: ADB, ADB Exchange, "
 "ADB Test, ADB-Test, CPM Terms, Deleted Required, Emails, File Type, Finance, Group01, "
 "Group02, HR Tracking, LeilaTest, OneDrive, People, Search Dictionaries, SharePoint, "
 "Short Term Storage, System. NONE is named File Plan, Classification or Records.\n"
 "4. NOT RE-VERIFIED: the earlier note of 6 value sets and 16 terms one level deep was "
 "taken before those 19 groups were seen, and the groups were not expanded. Expand ADB "
 "and File Type and count depth before quoting either figure")
RC["maplist"] = (LIST, "Retention Label Mapping list, a SharePoint list",
 "1. Open the EDRMS site that holds the Retention Label Mapping list\n"
 "2. Open the list and check its columns first: does it key a library by ListId or by name?\n"
 "3. Export to CSV from the list toolbar")
RC["none"] = (NONE, "No file exists",
 "There is nothing to open. No system this report can reach holds this figure. "
 "See the question column for who to ask")
RC["derived"] = (EDRMS, "No single file. Computed from the tables above",
 "Produced by the refresh job from the sources above rather than read from one export. "
 "Verify it by checking the inputs named in the next two columns")

# Test recipes reused across rows, so a tester reads the same words each time.
TEST_DECL = ("Yes. Query public.\"Records\" in drm-npr and count rows. "
             "About 1,990 in UAT")
TEST_DOCS = ("Partly. The declared half is queryable today. The undeclared half "
             "needs the weekly SharePoint scan, which is not built")
TEST_USAGE = ("Yes. Export the SharePoint site usage report from "
              "admin.microsoft.com, Reports, Usage, SharePoint site usage")
TEST_USER = ("Yes. Export the SharePoint activity user detail report. "
             "Note only 8 of the 30 licensed users in UAT show any activity")
TEST_GRAPH = "Yes. Graph Explorer, aka.ms/ge, with the call in the formula column"
TEST_NO = "No. Nothing in the tenant holds this"
TEST_DEPT = ("Not until the department list exists. The measure itself is "
             "testable bank-wide today")

Q_DEPT = ("Ask RAC for the site to department list, roughly 1,057 rows, one "
          "department per site URL. Check AvePoint Cloud Governance first, it "
          "may already hold it from the site request. Why: this single list "
          "unlocks every departmental figure in the report")
Q_DIV = ("Ask who supplies division alongside department. Why: the tier was "
         "reinstated at the client's request but nothing in the tenant "
         "populates it, so the RAC list has to carry a division column too")
Q_PROJ = ("Ask whether a project site register exists: site URL to project "
          "number, and sovereign or nonsovereign. Check Cloud Governance "
          "first. Why: it is the only thing missing for the whole Project "
          "Insights dashboard, every measure on it is already producible")
Q_FILEPLAN = ("Ask where the institutional file plan actually lives. It is in "
              "neither the term store, which holds 6 dropdown value sets and "
              "16 terms one level deep, nor Purview, which holds 53 flat "
              "retention labels. Why: a whole dashboard waits on the answer")
Q_TERMJOIN = ("Ask the development team to add a term reference to the "
              "document row, or supply a term to library mapping list. Why: "
              "nothing links a term to a document, so no per term figure can "
              "be produced even once the file plan is located")
Q_DISPOSAL = ("Raise a change request for a disposal status field: status, "
              "approver, and date disposed. Why: nothing records that a "
              "disposal happened, so every completion figure is unanswerable")
# Superseded 14 Aug. The compliance list is not a technical detection problem,
# it is a business register, and Cloud Governance holds it. What remains is
# getting the PRODUCTION export and confirming the register matches reality.
Q_COMPLIANT = ("Ask Leah Bancale for the PRODUCTION Cloud Governance Workspace "
               "report export, and whether a converted site is recorded in "
               "Cloud Governance the same way a created one is. Why: 'EDRMS Site "
               "Type' in that export is the compliant site list, so this single "
               "file defines which sites are in the report at all")
Q_STAFF = ("Ask which system owns staff, contractor and consultant "
           "classification and training completion. The client asks the same "
           "question on their own slide 54. Why: nothing in SharePoint holds it")
Q_IDLE = ("Ask whether an inactive site is 90 days or 300. The deck says 90, "
          "the metrics document says 300. Why: it changes every inactivity "
          "figure, and it is a one line change either way")
Q_FORMAT = ("Ask RAC for the file extension to format group mapping, a short "
            "list. Why: without it every extension falls into All other formats")
Q_ARCHIVE = ("Ask what archived means for a site, and how it is marked. Why: "
             "there is no definition anywhere, so the figure cannot be counted")
Q_RAC = ("Ask RAC for the physical holdings system: boxes, folders, locations, "
         "transfers and retrievals. Slide 67 names eServe for retrievals and "
         "shows an IR Dashboard screenshot. Why: none of it is in any system "
         "this report reads, and access to that dashboard may answer most of it")
Q_ACCESS = ("Ask whether audit log access can be granted, and confirm the "
            "purpose. Why: access requests, external sharing and permission "
            "exceptions live in the audit log, a different Graph surface with "
            "a different permission set")
Q_REF = ("Ask who maintains this list and where it will live. Why: it is not a "
         "measurement, nothing can generate it, and without an owner the panel "
         "stays empty")


# =====================================================================
# THE CLIENT'S OWN REVIEW COLUMNS, 14 AUGUST
# ---------------------------------------------------------------------
# Three columns he added to the workbook by hand: his question, what
# checking returned, and the conclusion. His wording is kept as he wrote
# it. Where a question has since been answered by an actual measurement,
# the answer is added to "What we found" rather than folded into his text,
# so it stays visible which half is his and which half is the result.
# =====================================================================

Q_USERS_ASKED = (
    "Is it accurate to determine the total edrms users in the sharepoint export? "
    "Do we have any other report as basis to get this number? Why sharepoint "
    "specifically? It is kpi and when clicked, opens the list of departments and "
    "each department determines the active users? How to map this?")

F_USERS = (
    "M365 Activity Report: it does not contain any activity that guarantees a "
    "person used an EDRMS compliant site.\n\n"
    "This report only includes activities made to any SharePoint file anywhere, "
    "including a team site that has nothing to do with EDRMS. The figure "
    "overstates EDRMS adoption, because it counts non EDRMS SharePoint use.\n\n"
    "2nd trap: records only every licensed user, not every active one. 30 rows in "
    "the test tenant, 8 with activity above zero.\n\n"
    "Why SharePoint specifically: the EDRMS is not a separate application. It is "
    "app {B255A2AF-7F63-4A30-966A-5D5FD99F97D7} installed into SharePoint sites, "
    "so the only two traces of EDRMS use are SharePoint usage reporting, which is "
    "broad, and the Records table, which sees declarers only.\n\n"
    "Mapping users to departments: not from the RAC site list. That maps SITE to "
    "department. USER to department comes from the Entra ID 'department' "
    "attribute, joined on User Principal Name. Whether ADB populates it is NOT "
    "yet confirmed: four Graph Explorer attempts all returned the default "
    "property set, so the check has not actually run.")

RM_USERS = (
    "Kylee and Aldo have a conversation regarding how EDRMS user should be "
    "defined.\n\n"
    "CAVEATS:\n\n"
    "1. M365 Usage report: Activity only overstates. It counts anyone who touched "
    "any SharePoint file anywhere, including sites with nothing to do with "
    "records. Someone who has never declared anything appears as an EDRMS user.\n\n"
    "2. Users who declared a record: Declarers only understates. Not sure of the "
    "responsibility for each staff, but a records officer whose job involves "
    "searching, reviewing and retrieving records all day, and who declares "
    "nothing because declaring is not their job, would count as a non user.\n\n"
    "BUILDABLE:\n"
    "EDRMS Users is the headline: distinct licensed users with SharePoint "
    "activity in the last 90 days. Stated with the caveat that it covers all "
    "SharePoint activity, because the usage report cannot break activity down per "
    "site.\n\n"
    "Active Records Declarers sits beside it: distinct users who declared at "
    "least one record in the last 90 days.\n\n"
    "On 90 days: 30 is too tight for records work, 180 stops discriminating, and "
    "90 is the window the usage report offers natively and matches the site "
    "inactivity rule. On the declared side a rolling 3 months is exact. On the "
    "activity side it must mean the 90 day export, not a calendar quarter, or the "
    "two will not reconcile.\n\n"
    "The gap between the two figures is the most useful adoption number in the "
    "report. The definition is the client's to sign off.")

Q_DOCS_ASKED = (
    "Where would I be getting this: site list from Graph.\n\n"
    "In the export, does this File Count count what specific documents it "
    "includes? Does a folder get classified as a file?")

F_DOCS = (
    "FOLDERS: answered by measurement, not by reading documentation. Graph "
    "returns a facet on every item: an item carrying a 'folder' facet is a "
    "folder, one carrying a 'file' facet is a document.\n\n"
    "Measured on org_csd_1.4testsite, Documents library, 14 Aug 2026: 73 items "
    "returned, 19 of them folders, 54 documents. Counting rows instead of "
    "checking facets overstates by 35 percent.\n\n"
    "Storage is worse. A folder's 'size' is CUMULATIVE, so it already contains "
    "everything beneath it. Summing size over all 73 items gives 684.3 MB against "
    "a true 172.2 MB, four times over. The 54 file sizes reconcile to the root "
    "folder figure exactly.\n\n"
    "SYSTEM LIBRARIES: 21 libraries on that one site, 2 of them created by "
    "SharePoint itself and holding no business document, 'Apps for SharePoint' "
    "and 'Client Side Assets'. File Count in the usage export includes these. The "
    "scan excludes them, so the scan total will read LOWER than 32,833.\n\n"
    "SITE LIST: do NOT use GET /sites?search=* to enumerate. It is a search over "
    "the crawled index, not an enumeration. Measured 14 Aug: it returned 451 "
    "sites against 1,676 from Get-PnPTenantSite. A scan built on it silently "
    "misses most of the tenant.\n\n"
    "NOT VERIFIED: whether the plain GUID in the usage export's 'Site Id' is the "
    "same GUID that sits inside the Graph composite site id. If it is not, the "
    "usage report route needs a different join key.")

RM_DOCS = (
    "The exact number does not exist in any system today. Not in the database, "
    "not in any report. public.\"Records\" holds declared records only and has no "
    "concept of an undeclared document.\n\n"
    "TWO ROUTES, and the answer is to do both.\n\n"
    "Route A, the usage report: filter the site usage export to EDRMS compliant "
    "Site Ids, exclude rows where Is Deleted is True, sum File Count. Cheap, no "
    "development, gives a headline total as soon as RAC supplies the compliant "
    "site list. It gives a count per site and nothing else.\n\n"
    "Route B, the document scan: one row per document carrying ListId and "
    "ItemId. Only this can be joined to public.\"Records\" to say WHICH documents "
    "were declared, which is the declaration rate the whole report exists to "
    "show. Route A cannot answer that at all.\n\n"
    "THE FOUR QUESTIONS FOR DEVELOPMENT:\n\n"
    "1. Is it feasible to produce Total Documents in All EDRMS Compliant Sites as "
    "a KPI? We can produce declared records today, since they sit in "
    "public.\"Records\". We have no equivalent for documents that were never "
    "declared, and every percentage in the report needs that figure underneath "
    "it. Is this achievable, and roughly what does it cost?\n\n"
    "2. If it is feasible, where does the number come from? public.\"Records\" "
    "holds declared records only, about 1,990 rows in UAT, with no row at all for "
    "an undeclared document. So the count has to come from SharePoint rather than "
    "from the database. Which source would you use, and does it need a new "
    "scheduled job or can it come from an existing report?\n\n"
    "3. Can the SharePoint usage report do this without a document scan? The site "
    "usage export has Site Id, File Count and Is Deleted on every row. If we had "
    "the list of EDRMS compliant sites, could we filter that export to those Site "
    "Ids, exclude the deleted ones, and sum File Count? Two things we want your "
    "view on. First, File Count appears to include system libraries such as Apps "
    "for SharePoint and Client Side Assets, so we expect it to read high. Can "
    "that be corrected, or do we accept it and label it? Second, the Site Id in "
    "the usage export is a plain GUID, while the Graph site id is a composite of "
    "host, site GUID and web GUID. Are those the same site GUID, so the two can "
    "be joined reliably? And separately: we understand this route gives a count "
    "per site with no document level rows, so it cannot tell us which documents "
    "were declared. Is that your reading too?\n\n"
    "4. How do we identify EDRMS compliant sites across the whole tenant? A site "
    "is compliant when app {B255A2AF-7F63-4A30-966A-5D5FD99F97D7}, "
    "digital-records-management-system, is installed on it, and Site Contents "
    "shows it as its own row in a browser. What we do not know is which API "
    "returns that across all sites at once. Is there a query that does this, or "
    "should RAC maintain the list by hand? This one blocks BOTH routes equally.")

Q_DISP_ASKED = (
    "How do we determine the numbers here? Should the KPI change to a more "
    "specific title like Total Records Due for Disposal within 12 months or "
    "something similar?")

F_DISP = (
    "NOT VERIFIED: whether EDRMSDueDateForDisposal is actually populated. Run "
    "SELECT COUNT(*), COUNT(\"EDRMSDueDateForDisposal\") FROM public.\"Records\". "
    "If coverage is low, this KPI measures retention label coverage rather than "
    "disposal workload, which is a different thing wearing the same name.\n\n"
    "Records carrying no retention label have no due date and drop out of this "
    "figure SILENTLY. Nothing on screen says so. 'Records without a retention "
    "schedule' has to stay visible beside this KPI or the number flatters.\n\n"
    "Due for disposal does not mean awaiting action. Nothing records whether a "
    "disposal actually happened, because there is no DisposalStatus field. That "
    "is an application change for development, not a report change, and it is "
    "why S/N 36 to 38 on this sheet cannot be built.")

RM_DISP = (
    "DECIDED: change the KPI title to include the window. It now reads 'Total "
    "Records Due for Disposal, Next 12 Months'.\n\n"
    "This is consistent with the caption rule, which bans computed numbers in a "
    "label, not definitions. A window in the title is part of what the figure IS.\n\n"
    "The 30 day, 90 day and 12 month figures must NEST: the 30 cannot exceed the "
    "90, which cannot exceed the 12 month. If they do not nest, the query is "
    "wrong.\n\n"
    "'Records beyond retention period' stays visibly separate as the backlog. It "
    "is not part of what is falling due, it is what already fell due and was "
    "never actioned.")


Q_SITES_ASKED = (
    "How do we determine the list of all edrms compliant sites? Currently we "
    "know a site is an edrms compliant if it gets cg adopted. How are we going "
    "to know these list? Is it possible that all sites in the cloud governance "
    "be the count for this? Essentially when a site is cg created it is found "
    "in cloud governance and is it possible that all cg adopted sites are going "
    "to be move to the cloud governance then from there we can get the list?")

F_SITES = (
    "ANSWERED 14 Aug 2026 by exporting the Cloud Governance Workspace report.\n\n"
    "The export carries 93 columns over 1,209 workspaces. The column "
    "'EDRMS Site Type' is populated on 1,032 rows with a single value, "
    "'EDRMS Project Site'. It is a yes/no flag and it is the compliant site "
    "list. The 177 blanks are template and admin sites: edrmstemplate, "
    "template_drmdefault, app_edrms_data.\n\n"
    "1,032 against the 1,057 placeholder that has been in the prototype since "
    "before this project. Nobody could justify that number. It was close.\n\n"
    "'Department' is populated on 1,030 of the 1,032. That is the site to "
    "department mapping RAC has been asked for and 29 requirements have been "
    "waiting on. BUT 240 sites carry SEVERAL departments, semicolon separated, "
    "such as ADBI;BOD and ADBI;BOD;CSD;CWRD;SARD. That is 23 percent, and it "
    "contradicts the 10 August decision that a site has one department owner.\n\n"
    "Also arriving with it, all 100 percent populated over the 1,032: "
    "Created Time, Last Active Time, Primary Business Owner, Storage Used (GB), "
    "Storage Quota (GB), External Sharing for Site, and Status as Active, "
    "Locked or Archived.\n\n"
    "'Last Active Time' is the important one for THIS row. The M365 usage "
    "export fills Last Activity Date on only 381 of 1,918 live sites, so most "
    "sites cannot be judged active or inactive from it. Cloud Governance fills "
    "it on every row.\n\n"
    "NOT VERIFIED: whether every workspace carrying 'EDRMS Site Type' actually "
    "has the app installed. A register records intent. The app records reality. "
    "They can drift and nothing announces it.\n\n"
    "TEST TENANT. Every URL is 7rkd12 and the site names are generated, such as "
    "dane-Site-Title-ddos-125. No count here transfers to production.")

RM_SITES = (
    "Get the list from the cloud governance. In cloud gov, sites listed are "
    "both cg adopted and cg created which are the basis of the edrms compliant "
    "site number.\n\n"
    "Leah Bancale confirmed on 14 Aug that EDRMS sites exist which did not come "
    "through Cloud Governance originally, and that those get converted to become "
    "compliant. So the CG CREATED list alone would be incomplete: it would miss "
    "every converted site, and those are the older, established departmental "
    "sites most likely to hold the largest volume of declared records.\n\n"
    "This is the better definition, not merely the easier one. A site RAC "
    "designated as EDRMS where the app deployment failed disappears entirely "
    "under a technical test, and nobody ever learns it is broken. Under the "
    "register it appears with zero declared records and somebody asks why. That "
    "gap is a compliance finding worth reporting.\n\n"
    "STILL TO DO:\n"
    "1. Get the PRODUCTION export. Everything above is structure learned from "
    "the test tenant.\n"
    "2. Confirm a converted site is recorded in Cloud Governance the same way a "
    "created one is.\n"
    "3. Validate the register against the app once, on a sample of about twenty "
    "sites, in both directions.\n"
    "4. Put the multi-department question to RAC. Either a document counts to "
    "several departments, and departmental totals will not sum to the bank-wide "
    "figure, or one is picked as primary and the rest are dropped.")

def R(section, element, typ, measure, recipe, cols, compute,
      table, column, design, tenant, effort, status, question="",
      asked="", found="", remarks=""):
    """One checkable figure.

    recipe  a key into RC, which supplies the source system, the file to open
            and the click by click steps to get it
    cols    the exact column heading in that file, or the exact Graph field
    compute the arithmetic, written with those exact headings so a reader can
            reproduce the number in Excel without translating anything
    asked   his own question against this row, in his words
    found   what checking actually returned
    remarks the conclusion, once it is settled
    """
    src, report, steps = RC[recipe]
    return (section, element, typ, measure, src, report, steps, cols, compute,
            table, column, design, tenant, effort, status, question,
            asked, found, remarks)


# =====================================================================
# 01 BANK-WIDE OVERSIGHT
# =====================================================================
BW = [
 R("Top panel", "Total EDRMS Users", "KPI", "People who actually used SharePoint",
   "user", "Two columns: 'User Principal Name' and 'Viewed Or Edited File Count'",
   "Filter the sheet to rows where 'Viewed Or Edited File Count' is greater than 0, then count "
   "the remaining rows. In the test tenant the file has 30 rows but only 8 survive that filter, "
   "so the answer is 8, not 30. Do NOT just count rows: the export lists every licensed user",
   T3, "UserPrincipalName, ViewedOrEditedFileCount", "Yes", "Yes", "Easy", BUILD,
   "How should EDRMS User be defined? Aldo Enriquez raised this on 10 Aug and Kylee Williamson "
   "replied that it will need defining. Neither settled it, and every option is buildable today, "
   "so this is a definition to agree rather than work to do",
   asked=Q_USERS_ASKED, found=F_USERS, remarks=RM_USERS),
 R("Top panel", "Total Documents in EDRMS", "KPI", "Every document held, declared or not",
   "usage", "'File Count', and 'Site Id' to identify the site",
   "Sum the 'File Count' column across the sites that are EDRMS compliant. In the test tenant "
   "the full sum is 32,833 across 1,071 sites that hold anything. This is the interim answer: "
   "the real figure needs the weekly scan, because the usage export counts files rather than "
   "the documents the report defines",
   T1, "All rows WHERE IsEdrmsCompliant", "Yes", "Partly", "Medium", BUILD,
   "Four questions, in the FINAL REMARKS column, ready to send to development",
   asked=Q_DOCS_ASKED, found=F_DOCS, remarks=RM_DOCS),
 R("Top panel", "Total Records Declared", "KPI", "Documents formally declared as records",
   "db", "The whole table, and the IsDeleted flag",
   "Run: SELECT COUNT(*) FROM public.\"Records\" WHERE NOT \"IsDeleted\". About 1,990 in UAT. "
   "Every declared record is one row, so the count is the answer with no arithmetic",
   T1, "IsDeclaredRecord", "Yes", "Yes", "Easy", BUILD),
 R("Top panel", "Total Physical Counterparts", "KPI", "Declared records with a paper copy",
   "db", "The EDRMSMeta JSON column, key HasPhysical",
   "Run: SELECT COUNT(*) FROM public.\"Records\" WHERE \"EDRMSMeta\"->>'HasPhysical' = 'true'. "
   "The answer is a subset of Total Records Declared, so it must always be the smaller number",
   T1, "HasPhysical", "Yes", "Yes", "Easy", BUILD),
 R("Top panel", "Total Records Due for Disposal, Next 12 Months", "KPI",
   "Records reaching end of retention in 12 months",
   "db", "'EDRMSDueDateForDisposal'",
   "Run: SELECT COUNT(*) FROM public.\"Records\" WHERE \"EDRMSDueDateForDisposal\" BETWEEN "
   "CURRENT_DATE AND CURRENT_DATE + INTERVAL '12 months'. The date is already computed in the "
   "source database, so nothing needs adding up",
   T1, "EDRMSDueDateForDisposal", "Yes", "Yes", "Easy", BUILD,
   "Confirm the 12 month window is the one they want on the tile, and that the 30 day, "
   "90 day and 12 month figures should nest inside one another",
   asked=Q_DISP_ASKED, found=F_DISP, remarks=RM_DISP),
 R("Top panel", "Active EDRMS Sites, Department / RM / Office", "KPI", "Compliant sites still in use",
   "cg", "'EDRMS Site Type', 'Status', 'Last Active Time', 'Department'",
   "Filter to rows where 'EDRMS Site Type' is not blank, which is the compliant site list, then "
   "keep rows where 'Status' is Active, then count. In the test tenant that is 1,032 sites of "
   "1,209 workspaces. Break down by 'Department'. This REPLACES the earlier recipe over the M365 "
   "usage export, which could not apply a compliance filter at all and fills Last Activity Date "
   "on only 381 of 1,918 live sites. Cloud Governance fills it on every row",
   T2, "IsEdrmsCompliant, LastActivityDate, ADBDepartmentOwner", "Yes, 2 need a rule",
   "Yes, in test", "Easy", DEPT, Q_COMPLIANT + ". Also " + Q_IDLE,
   asked=Q_SITES_ASKED, found=F_SITES, remarks=RM_SITES),
 R("Top panel", "Active EDRMS Sites, Sovereign Projects", "KPI", "Sites belonging to sovereign projects",
   "none", "No column in any export identifies a project site",
   "Cannot be worked out. Once a project register exists, it would be: count the sites whose "
   "project number is a sovereign facility",
   T2, "No column exists", "No", "No", "Blocked", NEWSRC, Q_PROJ),
 R("Top panel", "Active EDRMS Sites, Nonsovereign Projects", "KPI", "Sites belonging to nonsovereign projects",
   "none", "No column in any export identifies a project site",
   "Cannot be worked out. Same as the row above, filtered to nonsovereign",
   T2, "No column exists", "No", "No", "Blocked", NEWSRC, Q_PROJ),

 R("Users drill", "Never accessed EDRMS", "Tile", "Licensed users who have never opened anything",
   "user", "'Last Activity Date'",
   "Count the rows where 'Last Activity Date' is BLANK. In the test tenant that is 5 of the 30 "
   "rows. A blank means the person has a licence but has never touched SharePoint",
   T3, "LastActivityDate", "Yes", "Yes", "Easy", BUILD),
 R("Users drill", "No access in 90 days", "Tile", "Users who have gone quiet",
   "user", "'Last Activity Date'",
   "Count rows where 'Last Activity Date' is not blank AND is older than 90 days from the "
   "'Report Refresh Date' at the top of the same row. Use the refresh date, not today, because "
   "Microsoft's figures lag by about two days",
   T3, "LastActivityDate", "Yes", "Yes", "Easy", BUILD),
 R("Users drill", "Staff, Contractors, Consultants", "Table column", "Employment class of each user",
   "none", "Nothing in the user export gives employment type. 'Assigned Products' gives the "
   "licence name only, such as MICROSOFT 365 E5 DEVELOPER",
   "Cannot be worked out from any Microsoft export. It would be a join from an HR or LMS extract "
   "on 'User Principal Name'",
   T3, "No column exists", "No", "No", "Blocked", NEWSRC, Q_STAFF),
 R("Users drill", "Training completion", "Tile", "Share of users who completed training",
   "none", "No column exists in any Microsoft export",
   "Cannot be worked out. Would be completed / total, both from a training system",
   "None", "No column exists", "No", "No", "Blocked", REFL, Q_STAFF),
 R("Users drill", "By department", "Table column", "Every user figure cut by department",
   "none", "No department column exists in the user export",
   "Cannot be worked out yet. Once the RAC list exists: match the user to their sites, then to "
   "the department that owns those sites",
   T3, "ADBDepartmentOwner", "Column exists, unpopulated", "No", "Easy", DEPT, Q_DEPT),

 R("Documents drill", "Documents held per department", "Table column", "Volume of content",
   "usage", "'File Count' and 'Site Id'",
   "Sum 'File Count' grouped by the department that owns each site. The department comes from "
   "the RAC list matched on site, it is not in the export",
   T1, "FileCreatedDate, SiteUrl", "Yes", "Partly", "Medium", DEPT, Q_DEPT),
 R("Documents drill", "Storage per department", "Table column", "Space consumed",
   "usage", "'Storage Used (Byte)'",
   "Sum 'Storage Used (Byte)' then divide by 1,073,741,824 to get GB. The test tenant totals "
   "142.4 GB. NOTE this includes version history, so it reads higher than the sum of file sizes, "
   "which is expected and not an error",
   T2, "StorageUsed", "Yes", "Yes", "Easy", DEPT, Q_DEPT),
 R("Documents drill", "Users creating documents", "Table column", "Distinct authors",
   "none", "No author column exists in any export, and the scan does not capture one",
   "Cannot be worked out for undeclared documents. For declared records only, it would be "
   "COUNT(DISTINCT \"CreatedBy\") on the Records table, which counts declarers not authors",
   T1, "CreatedBy covers declarers only", "Partly", "No", "Medium", JOIN,
   "Ask the development team to capture Author on the document scan. Why: without it, creation "
   "activity can only be measured for records, not for documents"),

 R("Records drill", "Records declared per department", "Table column", "Declaration volume",
   "db", "The whole table, plus SiteUrl to attach a department",
   "SELECT COUNT(*) FROM public.\"Records\" GROUP BY the department that owns \"SiteUrl\". "
   "The department itself comes from the RAC list, matched on site URL",
   T1, "IsDeclaredRecord, ADBDepartmentOwner", "Yes", "Partly", "Easy", DEPT, Q_DEPT),
 R("Records drill", "Users declaring records", "Table column", "How many people are declaring",
   "db", "'CreatedBy'",
   "SELECT COUNT(DISTINCT \"CreatedBy\") FROM public.\"Records\". Distinct, not a row count, "
   "or one person declaring 400 records counts 400 times",
   T1, "CreatedBy", "Yes", "Yes", "Easy", DEPT, Q_DEPT),
 R("Records drill", "Declaration rate", "Table column", "Share of content formally declared",
   "derived", "Two numbers: declared records, and total documents",
   "Records Declared divided by Total Documents, times 100. The numerator comes from the Records "
   "table, the denominator from the scan. Until the scan exists this rate CANNOT be produced, "
   "because the denominator does not exist anywhere",
   T1, "IsDeclaredRecord over all rows", "Yes", "No", "Medium", DEPT,
   "None, but be explicit with the client that every rate waits on the scan"),
 R("Records drill", "By division", "Drill tier", "The same figures one level below department",
   "none", "No division column exists anywhere",
   "Cannot be worked out. The RAC list would have to carry a division alongside department",
   T1, "No division column exists", "No", "No", "Easy once supplied", NEWSRC, Q_DIV),

 R("Counterparts drill", "Physical counterparts by department", "Table column",
   "Records with a paper copy", "db", "EDRMSMeta key HasPhysical, plus SiteUrl",
   "Count rows where the HasPhysical key is true, grouped by the department that owns the site",
   T1, "HasPhysical, ADBDepartmentOwner", "Yes", "Partly", "Easy", DEPT, Q_DEPT),
 R("Counterparts drill", "Turned over to RAC, Completion rate", "Table column",
   "Counterparts physically transferred", "none", "No column exists in any system",
   "Cannot be worked out. Would be transferred divided by flagged as having a counterpart",
   "None", "No column exists", "No", "No", "Blocked", NEWSRC, Q_RAC),

 R("Disposal drill", "Records due, Next due date", "Table column", "What falls due and when",
   "db", "'EDRMSDueDateForDisposal'",
   "COUNT(*) for the number due, and MIN(\"EDRMSDueDateForDisposal\") for the next date, "
   "grouped by department. Exclude anything whose duration is Permanent, which never falls due",
   T1, "EDRMSDueDateForDisposal", "Yes", "Yes", "Easy", DEPT, Q_DEPT),
 R("Disposal drill", "Approver, Status, Records disposed", "Table column",
   "Who approved and what happened", "none", "No column exists in the application or the database",
   "Cannot be worked out. Nothing anywhere records that a disposal was carried out",
   T1, "DisposalStatus was removed from the design", "No, deliberately", "No", "Blocked",
   APPCH, Q_DISPOSAL),

 R("Overview of sites", "Department table and treemap", "Table and chart",
   "Sites, documents, records and counterparts per department",
   "derived", "Combines the site usage export, the Records table and the RAC department list",
   "For each department: count its sites from the usage export, sum 'File Count' for documents, "
   "count Records rows for declared, count HasPhysical for counterparts. Join everything on the "
   "site, and the department onto the site from the RAC list",
   T1 + " and " + T2, "ADBDepartmentOwner and the measures above", "Yes", "Partly",
   "Medium", DEPT, Q_DEPT),
 R("Overview of sites", "Sites created, last 90 days", "Tile", "New compliant sites",
   "graph", "GET /sites?search=*&$select=id,displayName,webUrl,createdDateTime&$top=999, "
   "then the createdDateTime field",
   "Count the sites whose createdDateTime falls in the last 90 days. Graph returns this on "
   "every site, confirmed against the tenant",
   T2, "SiteCreatedDate", "Yes", "Yes", "Easy", BUILD),
 R("Overview of sites", "Inactive over 90 days", "Tile", "Sites with no recent activity",
   "usage", "'Last Activity Date' and 'Is Deleted'",
   "Filter out 'Is Deleted' True, then count rows where 'Last Activity Date' is older than 90 "
   "days before 'Report Refresh Date'. Only 381 of 1,918 live sites carry a date at all",
   T2, "LastActivityDate", "Yes", "Yes", "Easy", BUILD, Q_IDLE),
 R("Overview of sites", "Sites deleted", "Tile", "Sites removed since the last snapshot",
   "usage", "'Is Deleted'",
   "Count rows where 'Is Deleted' is True. To count deletions in a period instead, compare two "
   "weekly exports and count the Site Ids present in the older one and absent from the newer",
   T2, "IsDeleted", "Yes, unpopulated", "Partly", "Medium", JOIN,
   "Confirm that a deleted site should be counted from its disappearance between two exports. "
   "Why: absence cannot distinguish a deletion from a permissions change that hides the site"),
 R("Overview of sites", "Sites archived", "Tile", "Sites moved to an archived state",
   "none", "No column in any export marks a site as archived",
   "Cannot be worked out. No rule exists to compute it",
   T2, "No column exists", "No", "No", "Blocked", DECIDE, Q_ARCHIVE),

 R("Comparison", "Users against documents, users against records, documents against records",
   "Chart", "Three ratios per department", "derived",
   "Combines 'Viewed Or Edited File Count' from the user export, 'File Count' from the site "
   "export, and the Records row count",
   "Each pair is drawn side by side per department. The third pair, documents against records, "
   "is the declaration rate expressed as two bars rather than a percentage",
   T1 + ", " + T3, "UserPrincipalName, IsDeclaredRecord", "Yes", "Partly", "Easy", DEPT, Q_DEPT),

 R("Declaration trend", "Records declared by month", "Chart", "Declarations in each of 12 months",
   "db", "'CreatedDate'",
   "SELECT date_trunc('month', \"CreatedDate\"), COUNT(*) FROM public.\"Records\" GROUP BY 1 "
   "ORDER BY 1, then take the last 12 closed months. Use CreatedDate, which is when the record "
   "was declared, NOT EDRMSRetentionLabelApplied, which can be set on an undeclared document",
   T1, "CreatedDate", "Yes", "Yes", "Easy", BUILD),
 R("Declaration trend", "Records declared this month, Monthly average", "Tile",
   "Current month against its average", "db", "'CreatedDate'",
   "This month is the last bar of the same query. The average is the 12 month total divided by 12",
   T1, "CreatedDate", "Yes", "Yes", "Easy", BUILD),
 R("Declaration trend", "Records declared by year", "Chart", "Declarations per year",
   "db", "'CreatedDate'",
   "The same query grouped by year instead of month. The final bar is the last 12 months, so it "
   "must equal the total of the monthly chart above it",
   T1, "CreatedDate", "Yes", "Yes", "Easy", BUILD),
 R("Declaration trend", "Department filter", "Filter", "The same trend for one department",
   "none", "No department column exists in the Records table yet",
   "Add a filter on the department that owns the site in \"SiteUrl\", once the RAC list exists",
   T1, "ADBDepartmentOwner", "Column exists, unpopulated", "No", "Easy", DEPT, Q_DEPT),

 R("Site activity trend", "Site visits by month", "Chart", "Page views across the estate",
   "usage", "'Page View Count' and 'Report Refresh Date'",
   "Export the report weekly at the 7 day period. For each month, SUM 'Page View Count' across "
   "the weekly exports whose 'Report Refresh Date' falls in that month. CRITICAL: use the 7 DAY "
   "period only. Summing the 30, 90 or 180 day exports counts most days several times over, "
   "because those windows overlap while consecutive 7 day windows tile exactly",
   T2, "SiteVisits7, ReportRefreshDate", "Yes", "Partly", "Medium", BUILD,
   "Confirm the job start date with the client. Why: the trend can only begin when weekly "
   "capture begins, and history never captured cannot be recovered"),
 R("Site activity trend", "Department and period filters", "Filter", "The series cut two ways",
   "usage", "'Report Refresh Date', plus the RAC department list",
   "Period: keep only the exports whose 'Report Refresh Date' falls in the last N months. "
   "Department: keep only the sites the RAC list assigns to that department",
   T2, "ADBDepartmentOwner, ReportRefreshDate", "Yes", "Partly", "Easy", DEPT, Q_DEPT),

 R("Retention rollup", "Permanent and temporary split", "Chart and table",
   "How the holding divides by retention type", "db", "'EDRMSDuration' and 'EDRMSRetentionLabel'",
   "Permanent is WHERE \"EDRMSDuration\" = 'Permanent'. Temporary is everything else that has "
   "a label. The two must add up to the labelled record count. The column is TEXT, so it can "
   "hold Permanent alongside 10, and must be cast before any arithmetic",
   T1, "EDRMSDuration, EDRMSRetentionLabel", "Yes", "Yes", "Easy", BUILD),
 R("Retention rollup", "Departments and libraries provisioned per term", "Table column",
   "Reach of each retention term", "none", "No column links a term to a library or a document",
   "Cannot be worked out. The file plan table has no join key and the document row carries no term",
   T4, "No join exists", "No", "No", "Blocked", JOIN, Q_TERMJOIN),
 R("Retention rollup", "Disposed", "Table column", "Records actually disposed of",
   "none", "No column exists", "Cannot be worked out. Nothing records that a disposal happened",
   T1, "No column exists", "No", "No", "Blocked", APPCH, Q_DISPOSAL),

 R("Format and storage", "Files by format group", "Chart and table", "The eight format families",
   "db", "The FileMeta JSON column, key FileType",
   "Take the extension from the FileType key, map it to one of the eight groups using the RAC "
   "mapping list (doc and docx both become Word), then COUNT(*) per group. The eight groups must "
   "add up to Total Records Declared exactly, which is the check that the mapping missed nothing",
   T1, "FormatGroup, FileType", "Yes, mapping outstanding", "Partly", "Medium", DECIDE, Q_FORMAT),
 R("Format and storage", "Storage by format group, Average file size", "Chart and table",
   "Space each format consumes", "graph",
   "GET /sites/{siteId}/drive/items, then the size field on each item",
   "SUM the size field per format group for storage, and SUM(size)/COUNT(*) for the average. "
   "WARNING: a folder returns a CUMULATIVE size in Graph, so filter to files only or storage "
   "double counts",
   T1, "FileSize", "Yes, new key", "Partly", "Medium", BUILD),

 R("Risk and compliance", "Compliant, active and inactive sites", "Tile and chart",
   "Site health across the estate", "usage", "'Last Activity Date', 'Is Deleted'",
   "Live sites are rows where 'Is Deleted' is False, 1,918 in the test tenant. Active is those "
   "with 'Last Activity Date' inside 90 days, inactive is the rest. The compliant subset cannot "
   "be filtered yet",
   T2, "IsEdrmsCompliant, LastActivityDate", "Yes", "Partly", "Easy", DECIDE,
   Q_COMPLIANT + ". Also " + Q_IDLE),
 R("Risk and compliance", "Orphaned sites, no owner", "Tile", "Sites with no administrator",
   "usage", "'Owner Principal Name'",
   "Count rows where 'Owner Principal Name' is BLANK. In the test tenant that is exactly 19 of "
   "the 1,918 live sites, and those 19 are the ones worth chasing",
   T2, "SiteOwner", "Yes", "Yes", "Easy", BUILD),
 R("Risk and compliance", "Libraries across sites, active, dormant", "Tile",
   "Library counts and their activity", "graph",
   "GET /sites/{siteId}/drives, then count the drives and read lastModifiedDateTime on each",
   "Count the drives returned per site and sum across sites. A library is active if "
   "lastModifiedDateTime is inside 90 days, dormant at 180. The drive's list.id is the ListId "
   "that joins back to the records",
   T2 + " and " + T1, "LibraryCount, LibraryLastActivityDate", "Yes", "Yes", "Easy", BUILD),
 R("Risk and compliance", "Libraries with no declared records", "Tile", "Provisioned but unused",
   "derived", "The Graph drive list, and 'ListId' on the Records table",
   "List every drive from Graph, list the DISTINCT \"ListId\" values in Records, and count the "
   "drives that appear in the first list and not the second. Group on ListId, never on library "
   "name, because names repeat across sites and change on rename",
   T1, "ListId, IsDeclaredRecord", "Yes", "Yes", "Easy", BUILD),
 R("Risk and compliance", "Library growth rate", "Tile", "New records against the opening balance",
   "db", "'CreatedDate' and 'ListId'",
   "Records created inside the period, divided by records already held at the start of it, "
   "times 100",
   T1, "CreatedDate, ListId", "Yes", "Yes", "Easy", BUILD),
 R("Risk and compliance", "Library ranking, three measures", "Chart",
   "Largest by volume, by storage, and highest declaration rate", "derived",
   "'ListId' and 'LibraryName' from Records, the size field from Graph, and the site name",
   "Group on ListId. Volume is COUNT(*), storage is SUM of the Graph size field, rate is "
   "records divided by documents. Always display the library with its parent site, because a "
   "bare library name is ambiguous across sites",
   T1, "ListId, LibraryName, SiteName, FileSize", "Yes", "Partly", "Easy", BUILD),
 R("Risk and compliance", "Most used libraries", "Tile", "Libraries ranked by views and edits",
   "none", "No export reports activity below site level",
   "Cannot be worked out. The site usage export has 'Page View Count' per SITE only, and Graph "
   "analytics also reports per site. There is no per library activity feed at all",
   "None", "No column exists", "No", "No", "Blocked", NEWSRC,
   "Tell the client this cannot be produced and offer records created per library instead. "
   "Why: Microsoft exposes no per library activity anywhere"),
 R("Risk and compliance", "Orphaned libraries", "Tile", "Libraries with no owner",
   "none", "'Owner Principal Name' exists for SITES only",
   "Cannot be worked out. The tenant returns an owner for a site, never for a library inside it",
   "None", "No library owner column exists", "No", "No", "Blocked", JOIN,
   "Ask whether a library owner is maintained anywhere. Why: the site owner is the only one the "
   "tenant returns"),

 R("Records quality", "Duplicated records", "Tile", "Declared records sharing a filename",
   "db", "'Title'",
   "SELECT \"Title\", COUNT(*) FROM public.\"Records\" GROUP BY \"Title\" HAVING COUNT(*) > 1. "
   "Sum those counts for the total. It WILL overcount legitimate repeats such as a monthly "
   "report filed each month under the same name",
   T1, "Title", "Yes", "Yes", "Easy", BUILD,
   "Confirm the definition. Why: the client wrote 'with same filenames?' themselves, so they "
   "already suspect it is imprecise"),
 R("Records quality", "Orphaned records", "Tile", "Records whose parent no longer resolves",
   "derived", "'ListId' and 'SiteUrl' on Records, against the Graph site and drive lists",
   "Left join Records to the Graph lists on ListId and on site. Count the records that find no "
   "match, meaning their library or site no longer exists",
   T1, "ListId, SiteUrl", "Yes", "Yes", "Easy", BUILD,
   "Confirm what counts as orphaned. Why: a missing site, a missing library and a missing owner "
   "give three different numbers"),

 R("Classification", "Records by sensitivity label", "Chart and tile",
   "Protection applied to records", "db", "The FileMeta JSON column, key SensitivityLabelName",
   "COUNT(*) grouped by the SensitivityLabelName key, counting NULL as unlabelled. NOTE the site "
   "usage export also has a 'Site Sensitivity Label Id' column, but it is EMPTY on all 2,575 "
   "rows in the test tenant, so site level labels are not available and this must come from the "
   "record",
   T1, "SensitivityLabelName", "Yes", "Yes", "Easy", BUILD),
 R("Classification", "Restricted and Confidential records", "Tile", "Counts per level",
   "db", "The FileMeta key SensitivityLabelName",
   "COUNT(*) where the label equals each level name. The tile names must match the label names "
   "actually in use in the tenant",
   T1, "SensitivityLabelName", "Yes", "Yes", "Easy", BUILD,
   "Confirm the label names in use. Why: the tiles must carry the tenant's own wording"),

 R("Access management", "External sharing", "Tile", "Where content leaves its site",
   "usage", "'External Sharing', plus 'Anonymous Link Count', 'Company Link Count', "
   "'Secure Link For Guest Count' and 'Secure Link For Member Count'",
   "'External Sharing' is True or False per site and IS populated on all 2,575 rows, so count "
   "the True rows for sites where sharing is permitted. The four link count columns would give "
   "actual sharing instances, but every one of them is ZERO on every row in the test tenant, so "
   "confirm on production before promising a number",
   T2, "New column, not yet in the design", "No", "Yes, the policy flag", "Medium", JOIN,
   "Add External Sharing to the site table, and check the four link count columns on production. "
   "Why: they are all zero in the test tenant, which may mean no sharing or may mean not "
   "populated, and those are very different answers"),
 R("Access management", "Access requests granted and declined", "Tile", "Permission requests",
   "none", "No column in either export",
   "Cannot be worked out. These live in the audit log, a different Graph surface with a "
   "different permission set",
   "None", "No column exists", "No", "No", "Blocked", NEWSRC, Q_ACCESS),
 R("Access management", "Permission exceptions", "Tile", "Access outside the standard grants",
   "none", "No column in either export", "Cannot be worked out. Audit log data",
   "None", "No column exists", "No", "No", "Blocked", NEWSRC, Q_ACCESS),

 R("Search analytics", "Searches performed, Successful searches", "Tile", "How people find records",
   "none", "No column in any export",
   "Cannot be worked out. SharePoint search analytics are reported at tenant level and are never "
   "attributed to a record or a library",
   "None", "No column exists", "No", "No", "Blocked", NEWSRC,
   "Tell the client search analytics cannot be attributed to a record. Why: Microsoft reports "
   "them at tenant level only"),
 R("Search analytics", "Most viewed and most downloaded records", "Tile", "Top content",
   "none", "The nearest is 'Page View Count', which is per site",
   "Cannot be worked out at document level. Graph analytics reports a site, never a document",
   "None", "No column exists", "No", "No", "Blocked", NEWSRC,
   "Offer most active sites instead. Why: document grain analytics do not exist"),
]

# =====================================================================
# 02 DEPARTMENT INSIGHTS
# =====================================================================
DP = [
 R("Scope", "Department picker", "Filter", "Every panel follows one department",
   "none", "No department column exists in any Microsoft export",
   "Once RAC supply the list: match each site URL to its department, then every document "
   "inherits that department through its site. No per document lookup is needed",
   T2, "ADBDepartmentOwner", "Column exists, unpopulated", "No", "Easy", DEPT, Q_DEPT),
 R("Top panel", "Go-live date", "Label", "When the department started on EDRMS",
   "none", "No column exists", "Cannot be worked out. A maintained list, one date per department",
   "None", "No column exists", "No", "No", "Blocked", REFL, Q_REF),
 R("Top panel", "Sites, Users, Documents, Records, Counterparts, Due", "KPI",
   "The department's totals", "derived",
   "The same columns as Bank-wide: 'File Count', 'Storage Used (Byte)', the Records row count, "
   "the HasPhysical key and 'EDRMSDueDateForDisposal'",
   "Take each Bank-wide figure and add one filter: keep only the sites the RAC list assigns to "
   "this department. The arithmetic does not change, only the population",
   T1 + ", " + T2 + ", " + T3, "As Bank-wide, filtered", "Yes", "Partly", "Easy", DEPT, Q_DEPT),
 R("Top panel", "Sites Inactive over 90 Days", "KPI", "Idle sites in this department",
   "usage", "'Last Activity Date' and 'Is Deleted'",
   "Filter to this department's sites, drop 'Is Deleted' True, then count those whose "
   "'Last Activity Date' is more than 90 days before 'Report Refresh Date'",
   T2, "LastActivityDate", "Yes", "Yes", "Easy", DEPT, Q_IDLE),
 R("Users", "Active users by division", "Table", "Adoption inside the department",
   "none", "'User Principal Name' exists, but nothing carries a division",
   "The user count is COUNT(DISTINCT 'User Principal Name') where 'Viewed Or Edited File Count' "
   "is above zero. Splitting it by division cannot be done, because no division exists",
   T3, "UserPrincipalName", "Division does not exist", "No", "Easy once supplied", NEWSRC, Q_DIV),
 R("Users", "Staff, Contractors, Consultants, Training", "Table column", "Employment class",
   "none", "'Assigned Products' gives the licence name only, not an employment type",
   "Cannot be worked out from any Microsoft export. Needs an HR or LMS extract joined on "
   "'User Principal Name'",
   "None", "No column exists", "No", "No", "Blocked", NEWSRC, Q_STAFF),
 R("Visitors", "Visitors per site", "Table", "People opening content in each site",
   "graph", "GET /sites/{siteId}/analytics/allTime, then the actorCount field. Also "
   "/analytics/lastSevenDays for the 7 day window",
   "actorCount is the number of distinct people. Confirmed against the tenant, which returned "
   "actionCount 5535 and actorCount 12 on a test site. Only these TWO windows exist, all time "
   "and seven days. There is no 30 or 90 day window, so do not promise a monthly visitor figure",
   T2, "UniqueViewers7, UniqueViewersAllTime", "Yes", "Yes", "Easy", BUILD,
   "Warn the client only 7 day and all time windows exist. Why: their slide implies a monthly figure"),
 R("Visitors", "Internal against external visitors", "Table column", "Where visitors come from",
   "none", "Graph analytics returns actorCount with no internal or external split",
   "Cannot be worked out. The count is a single number with no breakdown",
   T2, "No column exists", "No", "No", "Blocked", NEWSRC, Q_ACCESS),
 R("Visitors", "Access requests granted and declined", "Table column", "Permission requests",
   "none", "No column in either export", "Cannot be worked out. Audit log data",
   "None", "No column exists", "No", "No", "Blocked", NEWSRC, Q_ACCESS),
 R("Sites", "Site list with owners", "Table", "Every site the department owns",
   "usage", "'Site Id', 'Owner Principal Name', 'Owner Display Name'",
   "List the department's sites and read the owner from 'Owner Principal Name'. It is filled on "
   "1,899 of the 1,918 live sites, so a few rows will legitimately show no owner. The 'Site URL' "
   "column is EMPTY on every row, so match 'Site Id' to a Graph site list to get the address",
   T2, "SiteName, SiteOwner, SiteUrl", "Yes", "Partly", "Easy", DEPT, Q_DEPT),
 R("Sites", "Documents, records, counterparts, due per site", "Table column", "What each site holds",
   "derived", "'File Count' from the usage export, and the Records table grouped by SiteUrl",
   "Documents come from 'File Count' on the site row. Records, counterparts and due dates come "
   "from the Records table grouped by SiteUrl. The site rows must sum to the department total",
   T1, "SiteUrl, IsDeclaredRecord, HasPhysical", "Yes", "Partly", "Easy", DEPT, Q_DEPT),
 R("Sites", "No activity in 90 days", "Table column", "Sites that have gone quiet",
   "usage", "'Last Activity Date'",
   "Mark the row if 'Last Activity Date' is more than 90 days before 'Report Refresh Date'. "
   "Derived in the report, never stored, so the threshold is one edit to change",
   T2, "LastActivityDate", "Yes", "Yes", "Easy", BUILD, Q_IDLE),
 R("Documents", "Documents, storage, users creating", "Table", "Content produced per site",
   "usage", "'File Count' and 'Storage Used (Byte)'",
   "Sum both per site. Divide bytes by 1,073,741,824 for GB. Users creating cannot be produced "
   "for undeclared documents, because no export carries an author",
   T1, "FileCreatedDate, FileSize, CreatedBy", "Partly", "Partly", "Medium", DEPT, Q_DEPT),
 R("Declaration", "Records declared and rate per site", "Table", "Declaration performance",
   "derived", "The Records row count per SiteUrl, over 'File Count' from the usage export",
   "Rate is records declared divided by documents held, times 100, per site. The numerator is "
   "the Records table, the denominator the usage export until the scan exists",
   T1, "IsDeclaredRecord, SiteUrl", "Yes", "Partly", "Easy", DEPT, Q_DEPT),
 R("Declaration", "Division rows under each site", "Drill tier", "The same figures per division",
   "none", "No division column exists anywhere",
   "Cannot be worked out until the RAC list carries a division",
   T1, "No division column exists", "No", "No", "Easy once supplied", NEWSRC, Q_DIV),
 R("Declaration", "Declaration donut and site ranking", "Chart", "Declared against everything held",
   "derived", "The Records count, and 'File Count'",
   "Two slices: declared records, and documents minus declared records. The ranking is the same "
   "counts ordered highest first",
   T1, "IsDeclaredRecord", "Yes", "Partly", "Easy", DEPT, Q_DEPT),
 R("Library usage", "Libraries by file plan category", "Table", "Content by classification",
   "none", "'LibraryName' and 'ListId' exist on Records, but nothing maps a library to a term",
   "The library counts are producible by grouping Records on ListId. Grouping those under a "
   "file plan category is not, because no column links a library to a term",
   T1 + " and " + T4, "LibraryName, ListId, CategoryName", "No join exists", "No", "Blocked",
   JOIN, Q_FILEPLAN + ". Also " + Q_TERMJOIN),
 R("Library usage", "Users per library", "Table column", "People working in each library",
   "none", "No export reports below site level",
   "Cannot be worked out. Viewer counts are per site, never per library",
   "None", "No column exists", "No", "No", "Blocked", NEWSRC,
   "Tell the client this is not available at library grain. Why: Microsoft exposes no per "
   "library viewer feed"),
 R("Trend", "Records declared by month", "Chart", "The department's declaration trend",
   "db", "'CreatedDate', filtered to the department's sites",
   "Group Records by month of CreatedDate, keeping only rows whose SiteUrl belongs to this "
   "department",
   T1, "CreatedDate, ADBDepartmentOwner", "Yes", "Partly", "Easy", DEPT, Q_DEPT),
 R("Disposal", "Records due by library, approver, status", "Table", "Disposal readiness",
   "db", "'EDRMSDueDateForDisposal' and 'LibraryName'",
   "COUNT and MIN of the due date grouped by library. Approver and status cannot be produced: "
   "no field records them",
   T1, "EDRMSDueDateForDisposal, LibraryName", "Partly", "Partly", "Medium", APPCH, Q_DISPOSAL),
 R("Conventions", "Convention link, approval date, version", "Panel", "The naming convention",
   "none", "No column exists", "Cannot be worked out. A maintained list, one row per department",
   "None", "No column exists", "No", "No", "Blocked", REFL, Q_REF),
 R("Programme dates", "Audit, review, training, CoP dates", "Panel", "The RM calendar",
   "none", "No column exists", "Cannot be worked out. A maintained list, five dates per department",
   "None", "No column exists", "No", "No", "Blocked", REFL, Q_REF),
]

# =====================================================================
# 03 PROJECT INSIGHTS
# =====================================================================
PJ = [
 R("Classification", "Which sites are project sites", "Filter", "The population this reports on",
   "none", "No column in any export identifies a project site",
   "Cannot be worked out. Once a register exists it loads onto the site row, exactly the way "
   "department will, and everything else here follows",
   T2, "No column exists", "No", "No", "Blocked", NEWSRC, Q_PROJ),
 R("Classification", "Sovereign against nonsovereign", "Filter", "Facility type",
   "none", "No column exists", "Cannot be worked out. Comes with the project register",
   T2, "No column exists", "No", "No", "Blocked", NEWSRC, Q_PROJ),
 R("Projects", "Project number and name", "Table", "The list of projects",
   "none", "No column exists in any Microsoft export",
   "Cannot be worked out. ADB operational data, held in ADB project systems",
   T2, "No column exists", "No", "No", "Blocked", NEWSRC, Q_PROJ),
 R("Projects", "Sites, documents, records, counterparts, due", "Table column",
   "What each project holds", "derived",
   "'File Count' and 'Storage Used (Byte)' from the usage export, and the Records table",
   "Identical arithmetic to the department figures, grouped by project instead. EVERY one of "
   "these is producible today for any site we can identify. Only the classification is missing",
   T1 + " and " + T2, "All the standard measures", "Yes", "Partly", "Easy", NEWSRC, Q_PROJ),
 R("Profile", "Facility type, modality, country, status, dates", "Panel", "Project attributes",
   "none", "No column exists. ProjectEndDate on ADBSites is the only project field held",
   "Cannot be worked out. Seven of the eight attributes exist in no system this report reads",
   T2, "ProjectEndDate only", "One field of eight", "No", "Blocked", NEWSRC, Q_PROJ),
 R("Profile", "Declaration and counterpart donuts", "Chart", "Recordkeeping for one project",
   "derived", "The Records count and the HasPhysical key, over 'File Count'",
   "Two slice donuts over the project's documents, the same arithmetic as the department view",
   T1, "IsDeclaredRecord, HasPhysical", "Yes", "Partly", "Easy", NEWSRC, Q_PROJ),
]

# =====================================================================
# 04 INSTITUTIONAL FILE PLAN
# =====================================================================
FP = [
 R("Structure", "The five categories", "KPI and table", "Top level groups of the plan",
   "terms", "The group names returned by GET /termStore/groups",
   "Walk the tree: groups, then sets inside each, then children one level at a time. The API "
   "returns ONE LEVEL PER CALL, so counting the terms under a category means looping. In the "
   "test tenant this returns 6 value sets and 16 terms one level deep, which is NOT the "
   "institutional file plan the client is describing",
   T4, "CategoryName, TermName, TermSetName, Depth", "Yes, unverified", "No", "Medium",
   DECIDE, Q_FILEPLAN),
 R("Structure", "Total terms, terms per category", "KPI", "Size of the plan",
   "terms", "The term names returned while walking the tree",
   "COUNT the terms found, grouped by the group they sit under. Whatever the term store holds "
   "today is countable, it is simply not the file plan being asked for",
   T4, "TermId, CategoryName", "Yes", "No", "Easy", DECIDE, Q_FILEPLAN),
 R("Term tables", "Documents, records, counterparts per term", "Table column",
   "How heavily each term is used", "none",
   "No column links a term to a document. Records has no term column, the file plan has no "
   "document key",
   "Cannot be worked out. Even once the file plan is located, this needs a new column on the "
   "document row or a term to library mapping list",
   T1 + " to " + T4, "No join key exists", "No", "No", "Blocked", JOIN, Q_TERMJOIN),
 R("Term tables", "Libraries and departments provisioned per term", "Table column",
   "Reach of each term", "none", "No join key, and no department list",
   "Cannot be worked out. Two blockers at once",
   T4, "No join key exists", "No", "No", "Blocked", JOIN, Q_TERMJOIN + ". Also " + Q_DEPT),
 R("Term usage", "Most used and least used terms", "Chart", "Usage pattern across the plan",
   "none", "No join key exists",
   "Cannot be worked out. Once the join exists it is COUNT per term, ordered",
   T1 + " to " + T4, "No join key exists", "No", "No", "Blocked", JOIN, Q_TERMJOIN),
 R("Classification", "Records by classification", "Tile", "Records by file plan term",
   "none", "No term column on the record",
   "Cannot be worked out. The term table read the other way round, and blocked by the same join",
   T1, "No term reference exists", "No", "No", "Blocked", JOIN, Q_TERMJOIN),
 R("Classification", "Records by business process", "Tile", "Records by business activity",
   "none", "No column exists, and no definition exists either",
   "Cannot be worked out, for two separate reasons: the term join is missing, and nobody has "
   "said what a business process is in ADB terms",
   "None", "No column exists", "No", "No", "Blocked", DECIDE,
   "Ask what a business process is and how it relates to a file plan term. Why: the two are not "
   "the same thing, and it cannot be counted until somebody says what it counts"),
]

# =====================================================================
# 05 RETENTION AND DISPOSAL
# =====================================================================
RD = [
 R("Structure", "Permanent and temporary screens", "Filter", "The two retention types",
   "db", "'EDRMSDuration'",
   "Permanent is WHERE EDRMSDuration = 'Permanent'. Temporary is every other labelled record. "
   "The column is TEXT so it can hold Permanent next to 10, and must be cast before any "
   "arithmetic on the number",
   T1, "EDRMSDuration", "Yes", "Yes", "Easy", BUILD),
 R("Retention", "Records due, 30, 90 and 12 month windows", "Tile and chart",
   "What falls due in each window", "db", "'EDRMSDueDateForDisposal'",
   "COUNT(*) where the due date falls inside each window from today. The three windows must "
   "NEST: the 30 day count cannot exceed the 90, which cannot exceed the 12 month. If they do "
   "not nest, the query is wrong",
   T1, "EDRMSDueDateForDisposal", "Yes", "Yes", "Easy", BUILD,
   "Confirm the 12 month window is the one they want on the tile, and that the 30 day, "
   "90 day and 12 month figures should nest inside one another",
   ),
 R("Retention", "Retention label and duration per term", "Table column",
   "How long each class is kept", "purview",
   "The label name and its retention period from the file plan export",
   "Group the records by EDRMSRetentionLabel and read the duration from EDRMSDuration. Purview "
   "holds 53 labels in the test tenant, as a flat list with no hierarchy",
   T1, "EDRMSRetentionLabel, EDRMSDuration", "Yes", "Yes", "Easy", BUILD),
 R("Retention", "Beyond retention period", "Tile", "Records past their disposal date",
   "db", "'EDRMSDueDateForDisposal'",
   "COUNT(*) where the due date is earlier than today. This means the date has PASSED, not that "
   "action is outstanding, because nothing records whether a disposal was carried out",
   T1, "EDRMSDueDateForDisposal", "Yes", "Yes", "Easy", BUILD,
   "Warn that this counts dates passed, not action outstanding. Why: the two read alike and are "
   "not the same"),
 R("Compliance", "Records with and without a retention schedule", "Tile",
   "Whether a record can be scheduled", "db", "'EDRMSRetentionLabel'",
   "COUNT where the label IS NULL against IS NOT NULL. The two must add up to the total declared "
   "record count. A record with no label has no due date and sits outside every disposal figure",
   T1, "EDRMSRetentionLabel", "Yes", "Yes", "Easy", BUILD),
 R("Compliance", "Libraries without a mapped retention schedule", "Tile", "Unmapped libraries",
   "maplist", "Whatever key the list uses for a library, either ListId or a name",
   "List every library from Graph, list the libraries present in the mapping list, and count the "
   "difference. DO THIS CHECK FIRST: open the list and see whether it keys by ListId or by name. "
   "By ListId the metric goes straight in. By name the join is too fragile to trust, because "
   "library names repeat across sites",
   T1, "ADBLibraryCategory, ListId", "Yes", "Unknown", "Easy", DECIDE,
   "Check the mapping list's key yourself before asking anyone. Why: it decides whether this "
   "metric is safe or fragile, and it is a five minute check"),
 R("Terms", "Departments and libraries provisioned per term", "Table column", "Reach of each term",
   "none", "No join key exists", "Cannot be worked out. Same missing link as the file plan",
   T4, "No join key exists", "No", "No", "Blocked", JOIN, Q_TERMJOIN),
 R("Disposal", "Records disposed, awaiting approval, backlog", "Tile", "What has happened",
   "none", "No column exists in the application or the database",
   "Cannot be worked out. Nothing records that a disposal was carried out",
   T1, "DisposalStatus was removed from the design", "No, deliberately", "No", "Blocked",
   APPCH, Q_DISPOSAL),
 R("Disposal", "Physical records overdue for transfer", "Tile", "Counterparts not yet with RAC",
   "none", "No column exists", "Cannot be worked out. Needs the RAC holdings system",
   "None", "No column exists", "No", "No", "Blocked", NEWSRC, Q_RAC),
]

# =====================================================================
# 06 RECORDS AND ARCHIVE HOLDINGS
# =====================================================================
RA = [
 R("Storage", "Boxes and folders by location", "Table", "Physical holdings per location",
   "none", "No file exists. A PhysicalRecords table was designed and never built",
   "Cannot be worked out. There is nothing to open and nothing to count",
   "None, designed not built", "No column exists", "Designed, not built", "No", "Blocked",
   NEWSRC, Q_RAC),
 R("Storage", "Requests and requestors", "Table column", "Who is depositing records",
   "none", "No file exists", "Cannot be worked out",
   "None", "No column exists", "No", "No", "Blocked", NEWSRC, Q_RAC),
 R("Storage", "Capacity used per location", "Chart", "How full each room is",
   "none", "No file exists", "Cannot be worked out. Facility data, held by whoever runs the rooms",
   "None", "No column exists", "No", "No", "Blocked", NEWSRC,
   Q_RAC + ". Answer their direct question: capacity can be displayed the moment somebody holds "
   "the figure"),
 R("Retrieval", "Boxes and folders retrieved, status", "Table", "Records taken out",
   "none", "No file exists. Slide 67 says retrieval is processed in eServe",
   "Cannot be worked out. eServe is a separate system with no established feed",
   "None", "No column exists", "No", "No", "Blocked", NEWSRC, Q_RAC),
 R("Inventory health", "Unverified, missing, due for verification", "Tile", "Inventory state",
   "none", "No file exists",
   "Cannot be worked out. Implies a physical inventory process with a system behind it",
   "None", "No column exists", "No", "No", "Blocked", NEWSRC, Q_RAC),
 R("Inventory health", "Total physical files, legacy records", "Tile", "Size of the holding",
   "none", "No file exists", "Cannot be worked out",
   "None", "No column exists", "No", "No", "Blocked", NEWSRC, Q_RAC),
 R("Whole dashboard", "Recommendation", "Note", "How to run this piece of work",
   "none", "n/a",
   "Run as its own workstream. It is a second data source discovery exercise of the same size as "
   "the SharePoint one already completed, and inside the utilization report it would hold up the "
   "other five dashboards",
   "n/a", "n/a", "n/a", "n/a", "n/a", DECIDE,
   "Ask for access to the IR Dashboard shown on slide 67, and ask what is available in Opus. "
   "Why: both are their own questions to us, and the IR Dashboard may already hold most of this"),
]

SHEETS = [
    ("01 Bank-wide Oversight", "Bank-wide Oversight", "PPT s13 to s18, s34 to s44", BW),
    ("02 Department Insights", "Department Insights", "PPT s19 to s28, s53 to s66", DP),
    ("03 Project Insights", "Project Insights", "PPT s36, s37, s38", PJ),
    ("04 Institutional File Plan", "Institutional File Plan Insights", "PPT s29 to s31, s47 to s52", FP),
    ("05 Retention and Disposal", "Retention and Disposal Insights", "PPT s32, s44 to s46", RD),
    ("06 Records and Archive Holdings", "Records and Archive Holdings", "PPT s33, s67 to s69", RA),
]


# ---------------------------------------------------------------- helpers
def style_header(ws, row, ncols, height=32):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, size=9.5, bold=True, color="FFFFFF")
        cell.fill = FILL_HEAD
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = height


def title_block(ws, title, subtitle, ncols):
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color=ADB_BLUE)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color=GREY)
    for r in (1, 2):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).fill = FILL_TITLE
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 16


def widths(ws, spec):
    for i, w in enumerate(spec, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def body(ws, first, last, ncols):
    for r in range(first, last + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER


# ---------------------------------------------------------------------------
# Build time check: every column heading quoted against a Microsoft export must
# actually exist in that export. The evidence CSVs in this repo are the real
# thing, 2,575 and 30 rows, so this is a check against reality rather than
# against a recollection. It is the single most valuable assertion in this
# file: a workbook that tells somebody to filter on a column that does not
# exist is worse than one that says nothing.
# ---------------------------------------------------------------------------
def _headings(path):
    import csv, io, os
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8-sig") as fh:
        return set(next(csv.reader(fh)))

_SITE = _headings("evidence_SharePointSiteUsageDetail_2026-08-12.csv")
_USER = _headings("evidence_SharePointActivityUserDetail_2026-08-12.csv")

if _SITE and _USER:
    import re as _re
    _known = _SITE | _USER
    _bad = []
    for _name, _t, _sl, _rows in [(n, t, sl, rws) for n, t, sl, rws in
                                  [("01", "", "", BW), ("02", "", "", DP), ("03", "", "", PJ),
                                   ("04", "", "", FP), ("05", "", "", RD), ("06", "", "", RA)]]:
        for _row in _rows:
            _report, _cols, _element = _row[5], _row[7], _row[1]
            if "usage report" in _report or "activity user detail" in _report:
                for _q in _re.findall(r"'([^']+)'", _cols):
                    if _q not in _known:
                        _bad.append("%s / %s: '%s' is not a column in that export"
                                    % (_name, _element, _q))
    if _bad:
        raise SystemExit("Refusing to build. Column headings that do not exist:\n  "
                         + "\n  ".join(_bad))
    print("checked every quoted export column heading against the evidence CSVs: all real")
else:
    print("WARNING: evidence CSVs not found, column headings were NOT checked")

wb = Workbook()

# ================= Read me first =================
ws = wb.active
ws.title = "Read me first"
widths(ws, [3, 30, 104])
ws["B1"] = "EDRMS Utilization Report: report checker"
ws["B1"].font = Font(name=FONT, size=16, bold=True, color=ADB_BLUE)
ws["B2"] = "13 August 2026. One sheet per dashboard, plus the database design behind them."
ws["B2"].font = Font(name=FONT, size=10, italic=True, color=GREY)

intro = [
    ("", ""),
    ("What this workbook is for",
     "Checking the report. Every figure on every dashboard is listed with where it comes from, "
     "which database column carries it, how it is computed, whether it can be tested in the tenant "
     "today, and what to ask the client about it."),
    ("How to use it",
     "Filter a dashboard sheet on the Status column. Green rows can be built now. Everything else "
     "names what is missing and, in the last column, exactly what to ask and why."),
    ("", ""),
    ("The columns on each dashboard sheet", ""),
    ("Section, Element, Type", "Where the figure sits on the dashboard, and what kind of thing it is"),
    ("What the figure is", "The measure in one line, so a reader does not have to open the report"),
    ("Source system", "EDRMS database, Microsoft Graph, M365 usage report, site analytics, term store, "
                      "SharePoint list, or No source"),
    ("Database table and column", "Which of the four tables carries it, and which column. "
                                  "Blank means the column does not exist"),
    ("How it is produced", "The formula or the direction to source it. Written to be handed to whoever "
                           "builds the Power BI measure"),
    ("In the design?", "Whether the 73 column design already carries it"),
    ("In the tenant today?", "Whether the data exists in the test tenant now"),
    ("Testable in the tenant, and how", "The actual check to run, with the export or the API call"),
    ("Effort", "Easy, Medium or Blocked"),
    ("Status", "The verdict. Colour coded, see the Summary sheet"),
    ("Question to ask the client", "What to ask, and why it matters. Blank means nothing to ask"),
    ("Questions", "His own question against that row, in his words"),
    ("What we found", "What checking actually returned. Kept separate from his question so it stays "
                      "visible which half is the ask and which half is the result"),
    ("FINAL REMARKS", "The conclusion, once it is settled"),
    ("", ""),
    ("THE TEST TENANT AND THE ADB TENANT", ""),
    ("The method transfers. The numbers do not.",
     "Everything in this workbook was proved against the 7rkd12 TEST tenant. Treat the METHOD as "
     "valid for ADB production: the same reports exist, under the same names, with the same column "
     "headings, and the same rules govern them. Microsoft and AvePoint do not ship different "
     "products to different tenants. What does NOT transfer is any figure. 1,032 compliant sites, "
     "1,676 sites, 32,833 files, 1,990 declared records: all test tenant. Re-run each export "
     "against production and read the real number off it."),
    ("What to change for ADB",
     "Two things only. Swap the tenant name in every admin URL, so 7rkd12-admin.sharepoint.com "
     "becomes ADB's equivalent. And point Cloud Governance at ADB's own AvePoint Online Services "
     "instance. Every click path, column name, Graph call and SQL statement is unchanged."),
    ("The four exports that reproduce this workbook in production", ""),
    ("1. Cloud Governance Workspace report",
     "AvePoint Online Services, Cloud Governance, Directory, Workspace report, Export report. "
     "Collect from Job monitor. Gives the EDRMS compliant site list via 'EDRMS Site Type', the "
     "site to department mapping via 'Department', plus owner, storage, activity and status. "
     "The single highest value file in the project. Ask Leah Bancale."),
    ("2. SharePoint site usage report",
     "admin.microsoft.com, Reports, Usage, SharePoint site usage, Export. Gives File Count per "
     "site. Join to export 1 on Site Id to get the interim document total."),
    ("3. SharePoint activity user detail report",
     "admin.microsoft.com, Reports, Usage, SharePoint activity, Export. One row per LICENSED "
     "user, not per active user. Filter on Viewed Or Edited File Count above zero."),
    ("4. The drm-npr database",
     "Any SQL client against public.\"Records\". Declared records only. Everything else in the "
     "report needs the weekly scan, which needs export 1 to know which sites to scan."),
    ("", ""),
    ("The single most important fact", ""),
    ("Records holds declared records only",
     "public.\"Records\" in drm-npr contains declared records only, about 1,990 in UAT, and no row at "
     "all for an undeclared document. So every figure about declared records can be produced today, "
     "and every figure needing a denominator waits on the weekly SharePoint scan being built."),
    ("", ""),
    ("The three query rules", "Getting any of these wrong produces a plausible wrong number, not an error"),
    ("1. Read the latest snapshot",
     "Or 1,057 compliant sites silently becomes 55,000 after a year of retained history"),
    ("2. A range sums SiteVisits7",
     "Never the 30, 90 or 180 day figures. Consecutive 7 day windows tile exactly, longer ones "
     "overlap and would count most days several times over"),
    ("3. Match on ReportRefreshDate, not SnapshotDate",
     "The first is what Microsoft measured, the second is when the job ran, and they differ. "
     "An export taken on 12 Aug carried figures as at 10 Aug"),
    ("", ""),
    ("Sheets", ""),
    ("01 to 06", "One per dashboard, in the client's own order from PPT s13"),
    ("DB 1 to DB 4", "The four tables, every column, its type, source and remarks"),
    ("Questions to client", "Every question from the dashboard sheets, deduplicated and prioritised"),
    ("Summary", "Counts by status and by dashboard, computed from the sheets"),
]
r = 4
for left, right in intro:
    ws.cell(row=r, column=2, value=left).font = Font(name=FONT, size=10, bold=bool(left))
    ws.cell(row=r, column=3, value=right).font = Font(name=FONT, size=10)
    for c in (2, 3):
        ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
    r += 1
ws.sheet_view.showGridLines = False

# ================= one sheet per dashboard =================
for sheet_name, title, slides, rows in SHEETS:
    ws = wb.create_sheet(sheet_name)
    widths(ws, WID)
    title_block(ws, title, "Client design: " + slides +
                ". Filter the Status column to see what can be built now.", len(HDR))
    for i, h in enumerate(HDR, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(HDR))
    for j, row in enumerate(rows):
        r = 5 + j
        ws.cell(row=r, column=1, value=j + 1)
        for i, v in enumerate(row, start=2):
            ws.cell(row=r, column=i, value=v)
    last = 4 + len(rows)
    body(ws, 5, last, len(HDR))
    for r in range(5, last + 1):
        st = ws.cell(row=r, column=COL_STATUS).value
        if st in STATUS_FILL:
            ws.cell(row=r, column=COL_STATUS).fill = STATUS_FILL[st]
            ws.cell(row=r, column=COL_STATUS).font = Font(name=FONT, size=9, bold=True)
        # the step by step cell holds newlines and must show them
        ws.cell(row=r, column=8).alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = None
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top", horizontal="center")
    ws.auto_filter.ref = f"A4:{get_column_letter(len(HDR))}{last}"
    ws.freeze_panes = "C5"
    ws.sheet_view.showGridLines = False

# ================= the four database tables =================
DBHDR = ["S/N", "Column", "Field type", "What it holds", "Values", "Sample",
         "Where the value comes from", "Joins to", "Remarks"]
DBWID = [5, 28, 24, 56, 30, 30, 20, 22, 70]

for idx, (key, title, blurb, cols) in enumerate(TABLES, start=1):
    ws = wb.create_sheet("DB %d %s" % (idx, key.split(" ", 1)[1][:22]))
    widths(ws, DBWID)
    title_block(ws, title, blurb, len(DBHDR))
    ws.row_dimensions[2].height = 42
    for i, h in enumerate(DBHDR, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(DBHDR))
    for j, (name, typ, desc, vals, sample, ref, rem) in enumerate(cols):
        r = 5 + j
        ws.cell(row=r, column=1, value=j + 1)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=typ)
        ws.cell(row=r, column=4, value=desc)
        ws.cell(row=r, column=5, value=vals)
        ws.cell(row=r, column=6, value=sample)
        ws.cell(row=r, column=7, value=SOURCE.get(name, ""))
        ws.cell(row=r, column=8, value=ref)
        ws.cell(row=r, column=9, value=rem)
    last = 4 + len(cols)
    body(ws, 5, last, len(DBHDR))
    for r in range(5, last + 1):
        ws.cell(row=r, column=2).font = Font(name=FONT, size=9, bold=True)
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top", horizontal="center")
        # a column nobody can fill is the one thing this sheet must not hide
        if ws.cell(row=r, column=7).value == "NEEDS A DECISION":
            ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor="F8D7D7")
            ws.cell(row=r, column=7).font = Font(name=FONT, size=9, bold=True)
    ws.auto_filter.ref = f"A4:{get_column_letter(len(DBHDR))}{last}"
    ws.freeze_panes = "C5"
    ws.sheet_view.showGridLines = False

# ================= questions =================
ws = wb.create_sheet("Questions to client")
widths(ws, [5, 26, 34, 92, 22])
title_block(ws, "Questions to ask, in priority order",
            "Ordered by how much each answer unblocks. Every one also appears against its "
            "figures on the dashboard sheets.", 5)
for i, h in enumerate(["#", "Topic", "Who answers", "What to ask, and why", "Unblocks"], 1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, 5)
QROWS = [
    ("Institutional file plan", "Client, records management", Q_FILEPLAN,
     "The whole File Plan dashboard, plus classification everywhere"),
    ("Site to department list", "RAC, check Cloud Governance first", Q_DEPT,
     "Every departmental figure, and all of Department Insights"),
    ("Project site register", "RAC, check Cloud Governance first", Q_PROJ,
     "The whole Project Insights dashboard"),
    ("Term to document join", "Development team", Q_TERMJOIN,
     "Every per term figure on two dashboards"),
    ("Disposal status field", "Client, then the development team", Q_DISPOSAL,
     "Disposal completion across three dashboards"),
    ("EDRMS app detection", "Mihal Le", Q_COMPLIANT,
     "Which sites are in the report at all, and 4 KPIs"),
    ("Division", "Client and RAC", Q_DIV, "Four panels across two dashboards"),
    ("Staff class and training", "Client, likely HR or an LMS", Q_STAFF,
     "The user panels on two dashboards"),
    ("Inactivity threshold", "Client", Q_IDLE, "Every inactivity figure"),
    ("Format group mapping", "RAC", Q_FORMAT, "The format and storage panel"),
    ("Reference lists", "Client", Q_REF,
     "Go-live dates, conventions and programme dates on Department Insights"),
    ("Archived sites", "Client", Q_ARCHIVE, "One tile, and the definition behind it"),
    ("Physical holdings", "RAC", Q_RAC, "The whole Records and Archive Holdings dashboard"),
    ("Audit log access", "Client and ITD", Q_ACCESS, "Access management and visitor split"),
]
for j, (topic, who, what, unb) in enumerate(QROWS):
    r = 5 + j
    ws.cell(row=r, column=1, value=j + 1)
    ws.cell(row=r, column=2, value=topic)
    ws.cell(row=r, column=3, value=who)
    ws.cell(row=r, column=4, value=what)
    ws.cell(row=r, column=5, value=unb)
last = 4 + len(QROWS)
body(ws, 5, last, 5)
for r in range(5, last + 1):
    ws.cell(row=r, column=2).font = Font(name=FONT, size=9, bold=True)
    ws.cell(row=r, column=1).alignment = Alignment(vertical="top", horizontal="center")
ws.freeze_panes = "B5"
ws.sheet_view.showGridLines = False

# ================= summary =================
ws = wb.create_sheet("Summary")
widths(ws, [4, 34, 12, 12, 78])
title_block(ws, "Summary",
            "Every count is a formula over the six dashboard sheets, so it follows any edit made there.", 5)
ws["B4"] = "By status"
ws["B4"].font = Font(name=FONT, size=11, bold=True, color=ADB_BLUE)
for i, h in enumerate(["Status", "Count", "Share", "What it means"], 2):
    ws.cell(row=5, column=i, value=h)
style_header(ws, 5, 5)

order = list(STATUS_FILL)
counts = {k: 0 for k in order}
for _, _, _, rows in SHEETS:
    for row in rows:
        counts[row[COL_STATUS - 1]] = counts.get(row[COL_STATUS - 1], 0) + 1

r = 6
for st in order:
    ws.cell(row=r, column=2, value=st).fill = STATUS_FILL[st]
    ws.cell(row=r, column=2).font = Font(name=FONT, size=9, bold=True)
    L = get_column_letter(COL_STATUS)
    parts = "+".join("COUNTIF('%s'!$%s:$%s,$B%d)" % (sh[0], L, L, r) for sh in SHEETS)
    ws.cell(row=r, column=3, value="=" + parts)
    ws.cell(row=r, column=4, value="=IF($C$%d=0,0,C%d/$C$%d)" % (6 + len(order), r, 6 + len(order)))
    ws.cell(row=r, column=5, value=STATUS_MEANING[st])
    r += 1
tot = r
ws.cell(row=tot, column=2, value="Total figures checked")
ws.cell(row=tot, column=3, value="=SUM(C6:C%d)" % (tot - 1))
body(ws, 6, tot, 5)
for rr in range(6, tot + 1):
    ws.cell(row=rr, column=3).number_format = "#,##0"
    ws.cell(row=rr, column=4).number_format = "0.0%"
    for c in (3, 4):
        ws.cell(row=rr, column=c).alignment = Alignment(horizontal="center", vertical="top")
for c in (2, 3, 4, 5):
    ws.cell(row=tot, column=c).font = Font(name=FONT, size=9, bold=True)
    ws.cell(row=tot, column=c).fill = FILL_TITLE

start = tot + 3
ws.cell(row=start - 1, column=2, value="By dashboard")
ws.cell(row=start - 1, column=2).font = Font(name=FONT, size=11, bold=True, color=ADB_BLUE)
for i, h in enumerate(["Dashboard", "Figures", "Buildable now", "Blocked on something"], 2):
    ws.cell(row=start, column=i, value=h)
style_header(ws, start, 5)
r = start + 1
for name, title, _, rows in SHEETS:
    ws.cell(row=r, column=2, value=title)
    L = get_column_letter(COL_STATUS)
    ws.cell(row=r, column=3, value="=COUNTA('%s'!$%s$5:$%s$200)" % (name, L, L))
    ws.cell(row=r, column=4,
            value="=COUNTIF('%s'!$%s$5:$%s$200,\"Buildable now\")" % (name, L, L))
    ws.cell(row=r, column=5, value="=C%d-D%d" % (r, r))
    r += 1
vtot = r
ws.cell(row=vtot, column=2, value="Total")
for c, col in ((3, "C"), (4, "D"), (5, "E")):
    ws.cell(row=vtot, column=c, value="=SUM(%s%d:%s%d)" % (col, start + 1, col, vtot - 1))
body(ws, start + 1, vtot, 5)
for rr in range(start + 1, vtot + 1):
    for c in (3, 4, 5):
        ws.cell(row=rr, column=c).number_format = "#,##0"
        ws.cell(row=rr, column=c).alignment = Alignment(horizontal="center", vertical="top")
for c in (2, 3, 4, 5):
    ws.cell(row=vtot, column=c).font = Font(name=FONT, size=9, bold=True)
    ws.cell(row=vtot, column=c).fill = FILL_TITLE

note = vtot + 2
ws.cell(row=note, column=2, value="Hand check at build time")
ws.cell(row=note, column=2).font = Font(name=FONT, size=10, bold=True)
expected = ", ".join("%d %s" % (v, k) for k, v in sorted(counts.items(), key=lambda kv: -kv[1]) if v)
ws.cell(row=note + 1, column=2,
        value="These counts are formulas over the six dashboard sheets, not typed numbers, so they "
              "follow any edit made there. LibreOffice cannot recalculate in the build environment, "
              "so they ship without cached values: Excel computes them on open, but a lightweight "
              "previewer may show them blank. Counted by hand at build time there were %d figures, "
              "being %s." % (sum(counts.values()), expected))
ws.cell(row=note + 1, column=2).font = Font(name=FONT, size=9, color=GREY)
ws.cell(row=note + 1, column=2).alignment = Alignment(vertical="top", wrap_text=True)
ws.merge_cells(start_row=note + 1, start_column=2, end_row=note + 4, end_column=5)
ws.sheet_view.showGridLines = False

wb.save(OUT)
print("wrote %s: %d sheets, %d figures checked, %d database columns"
      % (OUT, len(wb.sheetnames), sum(counts.values()),
         sum(len(c) for _, _, _, c in TABLES)))
