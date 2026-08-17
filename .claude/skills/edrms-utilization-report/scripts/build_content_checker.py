#!/usr/bin/env python3
"""
Build the dashboard content checker workbook.

One sheet per dashboard, one row per thing on the screen. For each row:
where it is, which slide asked for it, what it is for in plain English,
whether we can build it, the exact steps to prove it in the test tenant,
and which column the number actually comes from.

Written for a non technical reader. Every sentence in the output should be
readable by somebody who has never opened Graph Explorer.

    python3 build_content_checker.py

Never hand edit the workbook. Edit this file and rerun.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "..", "..", "..", "..",
                   "EDRMS_Dashboard_Content_Checker_2026-08-17.xlsx")

# ---------------------------------------------------------------- palette
# ADB blue, per BACKGROUND.md. Kept muted so the text stays the loud part.
BLUE     = "1B4F9C"
BLUEPALE = "E8EFF8"
GREEN    = "E3F1E4"
AMBER    = "FDF2DC"
RED      = "FAE3E3"
GREY     = "F2F4F7"
INK      = "1A2634"

THIN = Side(style="thin", color="C8D2DE")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------- verdicts
# The five words the Buildable column is allowed to say, and their colour.
READY = "Yes, today"
SCAN  = "Yes, after the weekly scan is built"
APP   = "No, needs a change to the EDRMS application"
ASK   = "No, needs something only the client has"
NONE  = "No, no source exists anywhere"

VERDICT_FILL = {READY: GREEN, SCAN: AMBER, APP: RED, ASK: RED, NONE: RED}


def proc(buildable, proven, where, link, steps, chain=None):
    """The process chain.

    Buildable? > Proven in the test tenant? > Where is the number? > Go to >
    the build chain > the specific steps.

    The build chain is the part to read first. It says, in one line per hop,
    which file the number comes out of and which key carries it to the next
    file, in the order you would actually do it. Everything below it is detail.
    """
    out = ["Buildable? " + buildable,
           "Proven in the test tenant? " + proven,
           "Where is the number? " + where]
    if link:
        out.append("Go to: " + link)
    if chain:
        out.append("")
        out.append("HOW THE NUMBER IS BUILT, one hop at a time:")
        for i, c in enumerate(chain, 1):
            out.append("   %d. %s" % (i, c))
    if steps:
        out.append("")
        out.append("STEP BY STEP:")
        for i, s in enumerate(steps, 1):
            out.append("   %d. %s" % (i, s))
    return "\n".join(out)


# ---------------------------------------------------------------- sources
# Every row on every sheet points at one of these. Defined once so that a
# correction to a click path corrects it on all six dashboards at once.

LINK_SITEUSAGE = "https://admin.microsoft.com/#/reportsUsage/SharePointSiteUsage"
LINK_ACTIVITY  = "https://admin.microsoft.com/#/reportsUsage/SharePointActivity"
LINK_PURVIEW   = "https://purview.microsoft.com  (Records management, then File plan)"
LINK_TERMSTORE = "https://7rkd12-admin.sharepoint.com  (Content services, then Term store)"
LINK_GRAPH     = "https://aka.ms/ge  (Graph Explorer)"
LINK_CG        = "AvePoint Online Services, then Cloud Governance, then Directory, then Workspace report"

CG_STEPS = [
    "Open AvePoint Online Services and sign in.",
    "Go to Cloud Governance, then Directory, then Workspace report.",
    "Click Export report. The export is queued, it does not download straight away.",
    "Go to Job monitor and download the CSV when the job finishes.",
    "Open the CSV. It has 93 columns and one row per workspace.",
]

SITEUSAGE_STEPS = [
    "Open the SharePoint site usage report at the link above.",
    "Set the period to 30 days.",
    "Click Export to get the CSV.",
    "Open the CSV and find the column named below.",
    "Warning: the Site URL column is blank on every row. Join on Site Id instead.",
]

RECORDS_STEPS = [
    "Open any SQL client and connect to the drm-npr PostgreSQL database.",
    "The table is public.\"Records\", in the public schema.",
    "Remember this table holds DECLARED RECORDS ONLY. There is no row for an undeclared document.",
    "Run the count described in the last column.",
]


SRC = {}

SRC["rec"] = dict(
    chain=[
        "In the drm-npr database, take the table public.\"Records\". One row per DECLARED record.",
        "Each row carries the SiteUrl it came from.",
        "Export the Cloud Governance Workspace report, which carries both \"URL\" and \"Department\".",
        "Join the two on the site URL.",
        "Group by Department and count the rows. That is records declared per department.",
    ],
    buildable=READY,
    proven="Yes. The table was queried directly and returned about 1,990 rows in the test tenant.",
    where="the drm-npr PostgreSQL database, table public.\"Records\".",
    link="Any SQL client. The database is drm-npr, schema public.",
    steps=RECORDS_STEPS,
)

SRC["phys"] = dict(
    buildable=READY,
    proven="Yes. The field is part of the record metadata and was seen populated.",
    where="the drm-npr database, the physical counterpart flag on each declared record.",
    link="Any SQL client. The database is drm-npr, schema public.",
    steps=RECORDS_STEPS[:3] + [
        "Count the declared records where the physical counterpart flag is true.",
    ],
)

SRC["due"] = dict(
    buildable=READY,
    proven="Yes. The disposal due date is held per record.",
    where="the drm-npr database, the disposal due date on each declared record.",
    link="Any SQL client. The database is drm-npr, schema public.",
    steps=RECORDS_STEPS[:3] + [
        "Count the declared records whose disposal due date falls inside the window you want.",
        "For the next due date, take the earliest due date that is still in the future.",
    ],
)

SRC["declarer"] = dict(
    buildable=READY,
    proven="No. The column exists and the query is one line, but it has never actually been run.",
    where="the drm-npr database, the created by field on each declared record.",
    link="Any SQL client. The database is drm-npr, schema public.",
    steps=RECORDS_STEPS[:3] + [
        "Run: SELECT COUNT(DISTINCT \"CreatedBy\") FROM public.\"Records\";",
        "This is one of the two cheap checks that need nobody else. Worth doing next.",
    ],
)

SRC["scan_docs"] = dict(
    chain=[
        "This needs the WEEKLY SCAN, which pulls two usage reports and joins them.",
        "Report one: the SharePoint site usage report. Take the column \"File Count\", which is keyed on \"Site Id\".",
        "Report two: the Cloud Governance Workspace report. It carries BOTH \"ID\" (the same GUID as Site Id) AND \"URL\".",
        "Join report one to report two on that GUID. Now every file count has a site URL.",
        "In the Workspace report, the same row also carries \"Department\".",
        "So: File Count -> Site Id -> Cloud Governance ID -> URL -> Department. Sum File Count per department.",
        "Cloud Governance is the bridge because it is the ONLY file carrying both the GUID and the URL.",
    ],
    buildable=SCAN,
    proven="No, and it cannot be until the scan exists. The interim figure below is real but is not the same measure.",
    where="nowhere yet. It needs the weekly SharePoint scan, which is designed and NOT BUILT.",
    link=LINK_SITEUSAGE + "  (for the interim figure only)",
    steps=[
        "There is no source for a true document count today, because the Records table holds declared records only.",
        "Interim substitute: export the SharePoint site usage report and sum the File Count column.",
        "That gave 32,833 files across 1,071 sites in the test tenant.",
        "Understand the difference: File Count is every file in the site, including system files, and it is per site not per document.",
        "The real fix is the weekly scan, which walks each compliant site and writes one row per document.",
        "This is the single biggest item we can fix ourselves. It needs no client input at all.",
    ],
)

SRC["scan_size"] = dict(
    buildable=SCAN,
    proven="Partly. Graph returns a size on every item, which was checked. Nothing collects it yet.",
    where="nowhere yet. Microsoft Graph returns it per file, but only the weekly scan would store it.",
    link=LINK_GRAPH,
    steps=[
        "In Graph Explorer, run: GET /sites/{siteId}/drives",
        "Then list the items in a drive. Each item returns a size in bytes.",
        "Sizes from 1,506 to 2,338,767 bytes were seen, so the field is real and populated.",
        "WARNING: a folder returns the size of everything inside it. Count files only, or storage is double counted.",
        "Interim substitute: the site usage report has Storage Used (Byte) per site. It summed to 142.4 GB in the test tenant.",
    ],
)

SRC["scan_creator"] = dict(
    buildable=SCAN,
    proven="No.",
    where="nowhere yet. It needs a NEW column that the weekly scan would fill.",
    link=LINK_GRAPH,
    steps=[
        "Do not confuse this with users declaring records. They are different people.",
        "The created by field on the Records table is the person who DECLARED the record.",
        "The person who CREATED or uploaded the document is only visible by scanning the document itself.",
        "The scan would read createdBy on each item from Graph and store it as a new column.",
    ],
)

SRC["cg_sites"] = dict(
    chain=[
        "Export the Cloud Governance Workspace report. One row per workspace, 1,209 rows.",
        "Keep only the rows where the column \"EDRMS Site Type\" is NOT blank. That leaves 1,032.",
        "Of those, keep only the rows where \"Status\" = Active AND \"Site Status\" = Active.",
        "Count the rows. That is the number of active EDRMS sites.",
    ],
    buildable=READY,
    proven="Yes for the EDRMS marker. The two Active filters are a NO-OP in the test tenant, which is the trap.",
    where="the Cloud Governance Workspace report. THREE columns, not one.",
    link=LINK_CG,
    steps=CG_STEPS + [
        "THE FILTER IS THREE CONDITIONS, and leaving two of them out still gives the right answer here:",
        "   EDRMS Site Type is not blank   (it is an EDRMS site)         -> 1,032 of 1,209",
        "   AND Status = Active            (not locked, not archived)",
        "   AND Site Status = Active       (not deleted)",
        "In the test tenant all 1,032 EDRMS sites are Active on both, so the last two change nothing.",
        "Across all 1,209 workspaces there are 59 Locked and 4 Archived, so in production they WILL matter.",
        "That is why this is written out: the short version passes every test here and is wrong at ADB.",
        "The 177 blanks are template and admin sites and are correctly excluded.",
        "SEPARATE OPEN QUESTION: the deck never defines \"active\". Lifecycle active gives 1,032.",
        "   Used in the last 7 days gives 681. Those are very different headline numbers.",
    ],
)

SRC["cg_dept"] = dict(
    chain=[
        "Export the Cloud Governance Workspace report.",
        "Filter to EDRMS sites, active, as above.",
        "Group the rows by the column \"Department\".",
        "Count the rows in each group. That is sites per department.",
        "WATCH OUT: 240 rows carry several departments in one cell, like CWRD;SARD.",
    ],
    buildable=READY,
    proven="Yes, but it is not clean. See the warning in the steps.",
    where="the Cloud Governance Workspace report, column \"Department\".",
    link=LINK_CG,
    steps=CG_STEPS + [
        "Read the Department column. It is filled on 1,030 of the 1,032 EDRMS sites.",
        "WARNING: 240 sites carry SEVERAL departments in one cell, separated by semicolons, for example CWRD;SARD.",
        "That is 23 percent of sites. Until the client gives a rule, departmental totals will not add up to the bank total.",
        "This is question 1 to the client and it should be settled before the tables are built.",
    ],
)

SRC["cg_owner"] = dict(
    buildable=READY,
    proven="Yes. Filled on all 1,032 EDRMS sites.",
    where="the Cloud Governance Workspace report, column \"Primary Business Owner\".",
    link=LINK_CG,
    steps=CG_STEPS + [
        "Read the Primary Business Owner column.",
        "Use this rather than the SharePoint export, where 19 sites have no owner at all.",
    ],
)

SRC["cg_active"] = dict(
    chain=[
        "Export the Cloud Governance Workspace report.",
        "Filter to active EDRMS sites.",
        "Read \"Last Active Time\" and compare it to today.",
        "Count the rows older than the threshold.",
    ],
    buildable=READY,
    proven="Yes. Filled on all 1,032 EDRMS sites.",
    where="the Cloud Governance Workspace report, column \"Last Active Time\".",
    link=LINK_CG,
    steps=CG_STEPS + [
        "Read the Last Active Time column and compare it to today.",
        "Use this rather than the SharePoint export, which fills its equivalent on only 381 of 1,918 sites.",
        "OPEN QUESTION: is an inactive site 90 days or 300? The deck says 90, the metrics document says 300.",
    ],
)

SRC["cg_storage"] = dict(
    buildable=READY,
    proven="Yes. Filled on all 1,032 EDRMS sites, and it carries the quota too.",
    where="the Cloud Governance Workspace report, columns \"Storage Used (GB)\" and \"Storage Quota (GB)\".",
    link=LINK_CG,
    steps=CG_STEPS + [
        "Read Storage Used (GB) per site and add up the sites you want.",
        "Storage Quota (GB) is the only place in any source that carries a quota.",
    ],
)

SRC["cg_archived"] = dict(
    buildable=READY,
    proven="Yes. The column exists with values Active, Locked and Archived.",
    where="the Cloud Governance Workspace report, column \"Status\".",
    link=LINK_CG,
    steps=CG_STEPS + [
        "Count the rows where Status is Archived.",
        "This closed a requirement that was blocked for weeks on the words no definition. Cloud Governance has one.",
        "The reporting table needs one new column to hold it.",
    ],
)

SRC["cg_deleted"] = dict(
    buildable=READY,
    proven="Yes.",
    where="the Cloud Governance Workspace report, column \"Site Status\".",
    link=LINK_CG,
    steps=CG_STEPS + ["Count the rows where Site Status is Deleted."],
)

SRC["cg_created"] = dict(
    buildable=READY,
    proven="Yes, but read the warning. It is NOT the go-live date.",
    where="the Cloud Governance Workspace report, column \"Created Time\".",
    link=LINK_CG,
    steps=CG_STEPS + [
        "Read Created Time. This is when the SHAREPOINT SITE was created.",
        "WARNING: that is not when the site became EDRMS compliant. A site can exist first and be converted later.",
        "Every date column in the export was checked and none records the conversion. See the Go-Live row.",
    ],
)

SRC["m365_activity"] = dict(
    chain=[
        "Export the SharePoint activity user detail report. One row per LICENSED user.",
        "Keep the rows where \"Viewed Or Edited File Count\" is above zero. Those are the people with activity.",
        "Count them. That is users with recorded activity.",
        "IT IS NOT total EDRMS users. The export has no site column, so it cannot be narrowed to EDRMS at all.",
    ],
    buildable=READY,
    proven="Yes, and the result was a surprise worth knowing. See the steps.",
    where="the SharePoint activity user detail report, column \"Viewed Or Edited File Count\".",
    link=LINK_ACTIVITY,
    steps=[
        "Open the SharePoint activity report at the link above.",
        "Set the period to 30 days and click Export.",
        "IMPORTANT: the export has one row per LICENSED user, not per active user. Counting rows is wrong.",
        "Filter to rows where Viewed Or Edited File Count is above zero. Those are the active users.",
        "In the test tenant that is 8 active users out of 30 licensed rows.",
    ],
)

SRC["m365_never"] = dict(
    buildable=READY,
    proven="Yes. 25 of 30 rows carried a last activity date, so 5 were blank.",
    where="the SharePoint activity user detail report, column \"Last Activity Date\".",
    link=LINK_ACTIVITY,
    steps=[
        "Export the SharePoint activity report as above.",
        "A blank Last Activity Date means that person has never used SharePoint.",
        "A date older than your threshold means they have not used it recently.",
    ],
)

SRC["m365_visits"] = dict(
    chain=[
        "Export the SharePoint site usage report, period 7 days.",
        "Take \"Page View Count\", keyed on \"Site Id\".",
        "Join to the Cloud Governance Workspace report on ID to get the URL and the Department.",
        "Sum per department. To cover a longer range, ADD CONSECUTIVE 7 DAY EXPORTS.",
        "Never add the 30, 90 and 180 day figures together: those windows overlap.",
    ],
    buildable=READY,
    proven="Yes. Page View Count is populated in the export.",
    where="the SharePoint site usage report, column \"Page View Count\".",
    link=LINK_SITEUSAGE,
    steps=SITEUSAGE_STEPS + [
        "Read Page View Count for the 7 day period.",
        "THE RULE THAT MATTERS: to cover a range, add up consecutive 7 day figures. Never add the 30, 90 or 180 day figures.",
        "Consecutive 7 day windows fit together exactly. Longer windows overlap and would count most days several times.",
        "Getting this wrong gives a plausible wrong number, not an error message, so it is easy to miss.",
    ],
)

SRC["graph_libs"] = dict(
    buildable=READY,
    proven="Yes. The call returns and list.id is the key the Records table uses.",
    where="Microsoft Graph, the drives call on each site.",
    link=LINK_GRAPH,
    steps=[
        "In Graph Explorer, run: GET /sites/{siteId}/drives",
        "Each drive returned is one document library.",
        "list.id in the response is the ListId, which is what joins a library to the Records table.",
    ],
)

SRC["purview"] = dict(
    buildable=READY,
    proven="Yes, the labels were seen. The permanent versus temporary split was NOT.",
    where="Purview, the file plan, which holds 53 retention labels.",
    link=LINK_PURVIEW,
    steps=[
        "Open Purview, go to Records management, then File plan.",
        "There are 53 retention labels in the test tenant, as a FLAT LIST with no hierarchy.",
        "The label on each record is available, so you can group records by label today.",
        "WHAT IS MISSING: nothing says which of the 53 labels are permanent and which are temporary.",
        "The design has a retention status field fed from EDRMSMasters. Confirm whether that table already carries the split.",
        "If it does, this closes with no work at all. That is question 11.",
    ],
)

SRC["purview_split"] = dict(
    buildable=ASK,
    proven="No. The labels were seen. The rule that sorts them into two groups was NOT.",
    where="nowhere confirmed. We hold the label on every record. We do not hold the rule.",
    link=LINK_PURVIEW,
    steps=[
        "Open Purview, Records management, File plan. There are 53 retention labels, as a flat list.",
        "Nothing in that list says which labels mean keep forever and which mean keep until a date.",
        "CHECK THIS FIRST, BEFORE ASKING THE CLIENT: the design has a retention status field fed from the EDRMSMasters table.",
        "If EDRMSMasters already carries the permanent versus temporary split, this question closes with NO WORK AT ALL.",
        "Only if it does not, ask the client for the mapping of 53 labels to the two groups. That is question 11.",
        "Slides 32, 44, 45 and 46 all split the holding this way, so a whole dashboard rests on the answer.",
    ],
)

SRC["ask_project"] = dict(
    buildable=ASK,
    proven="No, and it cannot be. Nothing in any system we can reach records it.",
    where="NOWHERE. No system says which SharePoint site belongs to which project.",
    link="",
    steps=[
        "This was searched for in SharePoint, in Cloud Governance and in the drm-npr database. It is in none of them.",
        "ASK THE CLIENT FOR: a list of project number, project name, sovereign or nonsovereign, and the site URL or URLs.",
        "This is question 2, and it is the cheapest unblock in the whole deck.",
        "One spreadsheet turns three blank screens into working ones.",
        "Everything else about these sites we can already count. We simply cannot tell which sites are project sites.",
    ],
)

SRC["ask_projattr"] = dict(
    buildable=ASK,
    proven="No.",
    where="NOWHERE in EDRMS. These come from an ADB project system that has never been named in this work.",
    link="",
    steps=[
        "Facility type, project number, modality, country, status, effectivity and closing dates are not EDRMS data.",
        "ASK THE CLIENT: which system holds these, and can we read it? That is question 16.",
        "This is SEPARATE from the project register. Question 2 tells us which sites are project sites.",
        "This one fills the top third of the screen. Both are needed before Project Insights shows anything.",
    ],
)

SRC["ask_usertotal"] = dict(
    buildable=ASK,
    proven="No. What was proven is a DIFFERENT measure, users with activity.",
    where="NOWHERE. No source knows who has ACCESS to EDRMS.",
    link=LINK_ACTIVITY,
    chain=[
        "There is no chain, and that is the finding. Here is why each candidate fails.",
        "Candidate 1: count the rows in the M365 activity export. That is every LICENSED user in the tenant.",
        "   At ADB that is close to the whole Bank, roughly 9,400 people. It is not EDRMS users.",
        "Candidate 2: count the rows with activity above zero. That is ACTIVE users, not total users.",
        "   The client's own s39 asks for \"% of active users\" as an indicator UNDERNEATH this figure,",
        "   so the headline must be the population and active users must be a subset of it.",
        "   It also asks for \"users who have NEVER accessed EDRMS\", which cannot be counted",
        "   from a list that excludes them by definition.",
        "Candidate 3: read the members of each EDRMS site's Microsoft 365 group.",
        "   1,028 of the 1,032 EDRMS sites are \"Team site (no Microsoft 365 group)\", so there is no group to read.",
        "UNTESTED AND WORTH TRYING: every site has PERMISSIONS whether or not it has a group.",
        "   Reading role assignments per site would give users per EDRMS site directly.",
        "   That is scan-shaped work, about the size of the document scan, and nobody has tried it.",
        "WHAT WE OFFER INSTEAD: users with recorded activity, which is measurable today and honest.",
    ],
    steps=[
        "Do not publish the licensed user count as EDRMS users. At ADB it would say the whole Bank uses EDRMS.",
        "The prototype tile is renamed to what the number really is.",
        "Ask the client for the user register, question 3, which also solves staff versus contractor.",
    ],
)

SRC["ask_userreg"] = dict(
    buildable=ASK,
    proven="No.",
    where="NOWHERE. Microsoft 365 knows who used SharePoint. It cannot know what kind of person they are.",
    link="",
    steps=[
        "Microsoft 365 tells us who signed in and when. It does not know staff from contractor from consultant.",
        "It has no idea whether anyone completed training.",
        "GOOD NEWS: the client's own slide 54 already drew the table they would need to give us.",
        "Their columns: Name, Staff status, Onboarded since go-live, Training completed, Date completed, Notes.",
        "Their own slide even asks Source: Database of end users? Owned and maintained by?",
        "ASK THE CLIENT: who owns that list, can we have it, and how often. That is question 3.",
    ],
)

SRC["ask_userdept"] = dict(
    chain=[
        "GOOD NEWS, and this replaces the Graph call previously suggested here.",
        "Export the Cloud Governance USERS export, not the Workspace one.",
        "It carries a \"Department\" per person (column 37), filled on 228 of 286 people in the test tenant.",
        "It also carries \"User principal name\", which is the same key the M365 activity report uses.",
        "So: activity report -> User principal name -> Users export -> Department.",
        "The test tenant values are Microsoft demo data (R&D, HR, Retail), so they mean nothing.",
        "What they PROVE is that the field flows through. The only open question is whether ADB fills it in.",
        "Pull this export from production and the answer is immediate. No client involvement needed.",
    ],
    buildable=ASK,
    proven="No, but there is one cheap check nobody has run yet. See the steps.",
    where="NOWHERE yet. We can put SITES in departments. We cannot put PEOPLE in departments.",
    link=LINK_GRAPH,
    steps=[
        "Cloud Governance closed the department gap for SITES on 14 August. That does nothing for PEOPLE.",
        "A person does not live in a site, so site to department cannot carry them.",
        "CHEAP CHECK WORTH RUNNING: in Graph Explorer, run",
        "GET /users?$select=userPrincipalName,department,jobTitle",
        "If ADB fills in the department field on each person's account, this gap closes for free.",
        "Nobody has checked whether they do. It needs no client involvement at all.",
        "If they do not fill it in, the user register from question 3 must carry a department per person.",
    ],
)

SRC["ask_division"] = dict(
    buildable=ASK,
    proven="Proven ABSENT, which is a stronger answer than not found.",
    where="NOWHERE. Cloud Governance HAS a Division column and it is EMPTY on all 1,032 rows.",
    link=LINK_CG,
    steps=CG_STEPS + [
        "Look at the Division column. It is blank on every one of the 1,032 EDRMS sites.",
        "This moves division from we have not found the source to no source has it, which is worth telling the client.",
        "ASK THE CLIENT: either a list of divisions with the sites or people in each, or a decision to drop the division tier.",
        "That is question 5. Six columns across six screens depend on the answer.",
        "Better to remove six columns than to ship them empty.",
    ],
)

SRC["ask_golive"] = dict(
    buildable=ASK,
    proven="Proven absent from the export. There is one place left to look.",
    where="NOWHERE confirmed. Every date column in the Cloud Governance export was checked.",
    link=LINK_CG,
    steps=[
        "Created Time is the SharePoint site creation date, which is a different thing.",
        "Leah Bancale confirmed that sites not created in Cloud Governance are converted afterwards.",
        "For a converted site the site existed first and became EDRMS later. Nothing records that second date.",
        "ONE PLACE LEFT TO LOOK: Cloud Governance Job monitor lists job types including Site manual import and Apply profile for site.",
        "If those jobs are dated and exportable, that IS the go-live date and we need nothing from the client.",
        "ASK: can we have a Job monitor export to check? That is question 4.",
    ],
)

SRC["ask_fileplan"] = dict(
    buildable=ASK,
    proven="Proven absent from both candidate systems. This is not a guess.",
    where="NOWHERE. It is in neither the SharePoint term store nor Purview.",
    link=LINK_TERMSTORE,
    steps=[
        "The term store was opened and read. It holds 19 term groups.",
        "NONE of them is called Institutional Management, Administration, People Management, Programs and Operations or Other.",
        "What it does hold are dropdown value lists, one level deep, which is not a classification scheme.",
        "Purview was opened and read. It holds 53 retention labels as a flat list.",
        "A retention schedule is not a file plan. Neither system contains what slides 47 to 52 draw.",
        "ASK THE CLIENT: where does the Institutional File Plan actually live? That is question 12.",
        "It is the whole of one dashboard, seven slides, plus six more on Department Insights.",
    ],
)

SRC["ask_termjoin"] = dict(
    buildable=ASK,
    proven="No. This is a design gap, not a missing export.",
    where="NOWHERE. Nothing links a file plan term to a library or to a document.",
    link="",
    steps=[
        "THIS IS THE DEEPER HALF OF THE FILE PLAN PROBLEM, and it survives whatever answer we get about where the plan lives.",
        "Even holding the file plan, there is no field on a document, and no field on a library, naming its term.",
        "The file plan table in the design has no key that reaches the document table.",
        "ASK THE CLIENT: either a rule we can apply, for example the library name IS the term, which their own slides 48 and 61 hint at,",
        "or a field added to each library recording its term. That is question 13.",
        "THIS IS THE LONGEST LEAD ITEM IN THE DECK. It needs a design change, then an application or convention change, then a scan change.",
        "It will not get shorter by waiting, so it is worth raising now.",
    ],
)

SRC["ask_visitors"] = dict(
    buildable=ASK,
    proven="Partly, and the honest split is in the steps.",
    where="partly the SharePoint site usage report, partly Cloud Governance, partly nowhere confirmed.",
    link=LINK_CG,
    steps=[
        "The client's own slide 56 asks us: is this data available in SharePoint or Cloud Governance?",
        "TOTAL VISITORS: yes. Page View Count per site in the usage report.",
        "INTERNAL VERSUS EXTERNAL: probably. Cloud Governance publishes a guest user report, but it has not been joined to sites yet.",
        "ACCESS REQUESTS GRANTED AND DENIED: a Cloud Governance request report exists. It has NOT been confirmed to hold approvals and refusals.",
        "We should check both and come back rather than putting this to the client as a blocker. That is question 15.",
    ],
)

SRC["app_disposal"] = dict(
    buildable=APP,
    proven="No, and it cannot be proven until the field exists.",
    where="NOWHERE. The EDRMS application does not capture it.",
    link="",
    steps=[
        "THIS IS NOT A QUESTION FOR THE CLIENT. It is a development change request for Mihal Le.",
        "The client's own slides 43 and 60 specify it precisely enough to write.",
        "The application needs to capture, per record: disposal status with values Pending, Approved, Declined and Extended;",
        "the disposal approver, as a person; the disposal date; and for an extension, the new due date and the reason.",
        "An extension that does not record what it moved to, and why, cannot be audited.",
        "WHAT IT UNBLOCKS: five of the eight columns on slide 43, four indicators, records disposed on slide 44, and all of slide 60.",
        "UNTIL IT EXISTS, every disposal figure this report can produce is what is DUE. None is what was DONE.",
    ],
)

SRC["none_rac"] = dict(
    buildable=NONE,
    proven="No. It is a physical event, not a database field.",
    where="NOWHERE. Handing a box to RAC is a physical custody event and no system we can reach records it.",
    link="",
    steps=[
        "The slide defines this as records turned over for RAC storage against records declared with a counterpart.",
        "WE CAN PRODUCE THE BOTTOM HALF today. The top half does not exist anywhere.",
        "The prototype prints Not captured here. It previously printed 58 percent, which was invented.",
        "ASK THE CLIENT: what system or register records a physical counterpart actually reaching RAC custody? That is question 6.",
    ],
)

SRC["none_holdings"] = dict(
    buildable=NONE,
    proven="No.",
    where="NOWHERE. No system we can reach holds boxes, folders, locations, capacity, transfers or retrievals.",
    link="",
    steps=[
        "THE CLIENT HAS NOT SPECIFIED THIS DASHBOARD. Their slide 33 is literally the words RAC suggestions?",
        "Their slide 67 says: we would also like to learn what is available in Opus and how we can apply the features.",
        "The client has since confirmed that NOTHING from Opus goes in.",
        "Their slide 68 asks US: can room capacity and percentage of available storage capacity be included?",
        "The prototype used to answer that open question with an invented chart. It no longer does.",
        "Slides 68 and 69 ARE drawn with real column headings, so the SHAPE is a requirement even though nothing fills it.",
        "ASK THE CLIENT: is this dashboard in scope for this delivery, or a later phase? That is question 17.",
    ],
)

SRC["reflist"] = dict(
    buildable=ASK,
    proven="No, and it never can be. This is not a measurement.",
    where="NOWHERE, and correctly so. It is a list somebody maintains, not a number a system calculates.",
    link="",
    steps=[
        "Nothing in SharePoint or the database can generate this, and nothing should.",
        "It is a reference list: a document link, an approval date, a version, or a set of planned dates.",
        "ASK THE CLIENT: who maintains this list, where does it live, and how do we read it?",
        "The honest answer may be a small spreadsheet that somebody updates. That is fine, as long as it has an owner.",
    ],
)

SRC["derived"] = dict(
    buildable=READY,
    proven="Yes, once both halves are available. It is arithmetic, not a new source.",
    where="calculated from two figures that already exist. See the last column for the sum.",
    link="",
    steps=[
        "This is not read from anywhere. It is worked out from two other numbers on this list.",
        "Find both halves first, then apply the sum shown in the last column.",
        "If either half is not available yet, this is not available either.",
    ],
)

SRC["history"] = dict(
    buildable=ASK,
    proven="No. It is a decision, not a lookup.",
    where="nowhere yet. Month on month needs history to be KEPT, which is a decision nobody has taken.",
    link="",
    steps=[
        "Nine slides ask for a measure month on month.",
        "The main reporting table as designed is REPLACED every week, so it knows today and nothing else.",
        "Month on month is not a query. It is a decision to keep old snapshots.",
        "IT IS CHEAP IF DECIDED NOW AND EXPENSIVE LATER, because history that was never kept cannot be recovered.",
        "ASK THE CLIENT: how far back should the dashboard look? Twelve months and three years cost different amounts.",
        "That is question 8.",
    ],
)


def S(key, basis, buildable=None):
    """Turn a source key plus a basis sentence into the two output columns."""
    s = SRC[key]
    b = buildable or s["buildable"]
    return b, proc(b, s["proven"], s["where"], s["link"], s["steps"],
                   s.get("chain")), basis


# ---------------------------------------------------------------- the rows
# Each row: (what it is, where on the screen, slide, purpose, source key, basis)

BW = [
 ("Number of active EDRMS SharePoint sites for Department, RM, office",
  "Top panel, tile 1 (opens a table)", "s34, s15",
  "The headline count of how many SharePoint sites are actually running EDRMS. Everything else on this dashboard is a breakdown of these sites.",
  "cg_sites",
  'Found in report? Yes. Cloud Governance Workspace report, column "EDRMS Site Type". Count the rows where it is not blank. Test tenant answer: 1,032.'),

 ("Total number of active EDRMS SharePoint sites for Sovereign projects",
  "Top panel, tile 2 (opens a table)", "s34, s36",
  "The same count, but only for sites belonging to sovereign projects, so project work can be seen separately from departmental work.",
  "ask_project",
  "Cannot be produced. We can count sites perfectly well. We cannot tell which of them are sovereign project sites, because nothing records it."),

 ("Total number of active EDRMS SharePoint sites for Nonsovereign projects",
  "Top panel, tile 3 (opens a table)", "s34, s37",
  "The same again for nonsovereign projects.",
  "ask_project",
  "Cannot be produced, for exactly the same reason as the sovereign tile."),

 ("Total number of EDRMS users (WHAT THE CLIENT ASKED FOR)",
  "Top panel, tile 4. Renamed on the page", "s34, s39",
  "The client wants the population of people who hold EDRMS access. No source knows that. See the next row for what the page shows instead.",
  "ask_usertotal",
  'CANNOT BE PRODUCED. Every candidate source counts something else. The row below is the alternative that is on the page.'),

 ("EDRMS users with recorded activity (WHAT WE GIVE INSTEAD)",
  "Top panel, tile 4", "s39 indicators",
  "How many people have actually used EDRMS. It answers the adoption question the client is really asking, and unlike the population it is measurable today.",
  "m365_activity",
  'Found in report? Yes. SharePoint activity user detail, column "Viewed Or Edited File Count". Count the rows above zero. Test tenant: 8 of 30 rows.'),

 ("Total number of documents in EDRMS",
  "Top panel, tile 5 (opens a table)", "s34, s40",
  "Everything held in EDRMS, declared or not. This is the denominator for the declaration rate, which is why it matters more than it looks.",
  "scan_docs",
  'Needs the scan. Interim substitute: SharePoint site usage report, column "File Count", summed across sites. Test tenant: 32,833 files over 1,071 sites. Be clear that File Count is not the same measure.'),

 ("Total number of records declared in EDRMS",
  "Top panel, tile 6 (opens a table)", "s34, s41",
  "How many documents have been formally declared as records. This is the core compliance number for the whole report.",
  "rec",
  'Found in report? No, this is a database query, not a report. Count the rows in public."Records". Test tenant: about 1,990 rows, 1,984 distinct documents.'),

 ("Total number of physical counterparts identified",
  "Top panel, tile 7 (opens a table)", "s34, s42",
  "Declared records that also exist as paper. It tells RAC how much physical material sits behind the digital holding.",
  "phys",
  'Found in report? No, a database query. Count the declared records where the physical counterpart flag is true.'),

 ("Total number of records due for disposal",
  "Top panel, tile 8 (opens a table)", "s34, s43",
  "Records whose retention period is ending. This is the work queue for records management.",
  "due",
  'Found in report? No, a database query. Count the declared records whose disposal due date falls in the window.'),

 ("Retention and disposal insights",
  "Top panel, tile 9 (NAVIGATES to another dashboard)", "s34, s44",
  "Not a breakdown. Clicking it leaves this screen and opens the Retention and Disposal dashboard. This is why slide 44 carries the Bank-wide banner: the banner marks the way in, not the ownership.",
  "due",
  "The tile's number is read from the Retention dashboard's own total, so the two can never disagree."),

 ("Institutional File Plan insights",
  "Top panel, tile 10 (NAVIGATES to another dashboard)", "s34, s47",
  "The same idea: clicking it opens the Institutional File Plan dashboard rather than a table.",
  "ask_fileplan",
  "The tile's number is read from the File Plan dashboard's own total. That total is itself unsourced, see the File Plan sheet."),

 ("Department / office / RM (column)",
  "Sites table (tile 1 drill)", "s35, also drawn at s16",
  "Names each department so every other figure on the row can be attributed to somebody.",
  "cg_dept",
  'Found in report? Yes. Cloud Governance Workspace report, column "Department". WARNING: 240 sites carry several departments separated by semicolons.'),

 ("Number of EDRMS SharePoint sites (column)",
  "Sites table (tile 1 drill)", "s35",
  "How many EDRMS sites that department runs.",
  "cg_sites",
  'Found in report? Yes. Count the Workspace report rows where "EDRMS Site Type" is filled, grouped by "Department".'),

 ("Total number of documents (column)",
  "Sites table (tile 1 drill)", "s35",
  "How much content that department holds.",
  "scan_docs",
  'Needs the scan. Interim: sum "File Count" from the site usage report, joined to the department via Site Id.'),

 ("Total number of records declared (column)",
  "Sites table (tile 1 drill)", "s35",
  "How much of that content has been formally declared. The compliance figure per department.",
  "rec",
  'Count rows in public."Records", grouped by the site\'s department.'),

 ("Total number of physical counterparts (column)",
  "Sites table (tile 1 drill)", "s35",
  "How much paper sits behind that department's declared records.",
  "phys",
  "Count declared records with the physical flag true, grouped by department."),

 ("Sites created, last 90 days (indicator)",
  "Sites table, small stat above the table", "s35",
  "Shows whether the estate is still growing.",
  "cg_created",
  'Found in report? Yes. Workspace report, column "Created Time". Count rows newer than 90 days. This is SITE creation, not EDRMS go-live.'),

 ("Sites deleted (indicator)",
  "Sites table, small stat above the table", "s35",
  "Shows sites removed, so the count going down can be explained.",
  "cg_deleted",
  'Found in report? Yes. Workspace report, column "Site Status". Count rows where it reads Deleted.'),

 ("Sites archived (indicator)",
  "Sites table, small stat above the table", "s35",
  "Sites closed but retained. Previously blocked because nobody had a definition of archived. Cloud Governance has one.",
  "cg_archived",
  'Found in report? Yes. Workspace report, column "Status". Count rows where it reads Archived. Needs one new column in the reporting table.'),

 ("Inactive over 90 days (indicator)",
  "Sites table, small stat above the table", "s35",
  "Sites nobody is using, which are candidates for cleanup or a conversation with the owner.",
  "cg_active",
  'Found in report? Yes. Workspace report, column "Last Active Time". Count rows older than the threshold. OPEN: is the threshold 90 days or 300?'),

 ("Project number and name, and all four measures (whole table)",
  "Sovereign projects table (tile 2 drill)", "s36",
  "The same measures as the department table, cut by project instead. Lets a project team see its own records position.",
  "ask_project",
  "NOT ONE FIGURE on this table can be produced today. Every measure is individually available. The missing piece is only the link from a site to a project."),

 ("Project number and name, and all four measures (whole table)",
  "Nonsovereign projects table (tile 3 drill)", "s37",
  "The same table for nonsovereign projects.",
  "ask_project",
  "Same as sovereign. Blocked entirely on the project register, which is question 2."),

 ("Total number of users, staff (column)",
  "Users table (tile 4 drill)", "s39",
  "How many EDRMS users are Bank staff, as opposed to contractors or consultants.",
  "ask_userreg",
  "Cannot be produced. Microsoft 365 knows who signed in. It does not know their employment type. Needs the user register, question 3."),

 ("Total number of users, Contractors (column)",
  "Users table (tile 4 drill)", "s39",
  "The contractor share of users, which matters for access reviews when contracts end.",
  "ask_userreg",
  "Cannot be produced. Same reason. Their own slide 54 drew the table that would fix it."),

 ("Total number of users, Consultants (column)",
  "Users table (tile 4 drill)", "s39",
  "The consultant share, for the same reason.",
  "ask_userreg",
  "Cannot be produced. Same reason."),

 ("Completion of training (column)",
  "Users table (tile 4 drill)", "s39",
  "Whether users have been trained, so gaps can be targeted rather than guessed at.",
  "ask_userreg",
  "Cannot be produced. No system we can reach records training. This is one of the four columns the user register would fill at once."),

 ("Onboarded since go-live (column)",
  "Users table (tile 4 drill)", "s39",
  "New users since the department went live, which is the onboarding queue.",
  "ask_golive",
  "Cannot be produced, and it is blocked TWICE: it needs the user register AND the go-live date. Questions 3 and 4."),

 ("Users by department (the grouping itself)",
  "Users table (tile 4 drill)", "s39",
  "Every column on this screen is per department, so without this the whole table has no rows.",
  "ask_userdept",
  "A GAP NOBODY HAD NAMED until the deck was read. Closing the department gap on 14 August fixed SITES, not PEOPLE. Run the Graph users check first, it is free."),

 ("Percentage of active users (indicator)",
  "Users table, small stat", "s39",
  "Adoption. How many of the people who could use EDRMS actually do.",
  "m365_activity",
  'Two figures. Active users (Viewed Or Edited File Count above zero) divided by all rows in the activity export, as a percentage. Test tenant: 8 / 30 = 27%.'),

 ("Never accessed EDRMS (indicator)",
  "Users table, small stat", "s39",
  "People who have never used it at all. The retraining list.",
  "m365_never",
  'Found in report? Yes. SharePoint activity report, column "Last Activity Date". Count the rows where it is BLANK.'),

 ("Not accessed in 90 days (indicator)",
  "Users table, small stat", "s39",
  "People who have stopped using it. Candidates for cleanup, or people who left the Bank.",
  "m365_never",
  'Found in report? Yes. Same column. Count rows where the date is older than 90 days.'),

 ("Number of documents created / uploaded (column)",
  "Documents table (tile 5 drill)", "s40",
  "How much content each department is putting into EDRMS.",
  "scan_docs",
  'Needs the scan. Interim: sum "File Count" per site from the usage report and group by department.'),

 ("Number of users creating documents (column)",
  "Documents table (tile 5 drill)", "s40",
  "How many people are actually contributing, as opposed to how many have access.",
  "scan_creator",
  "Needs the scan AND A NEW COLUMN. Do not use the Records created by field: that is who DECLARED the record, a different person and a different question."),

 ("Documents size in GB (column)",
  "Documents table (tile 5 drill)", "s40",
  "Storage consumed per department, which drives cost and capacity planning.",
  "scan_size",
  'Needs the scan. Interim: site usage report, column "Storage Used (Byte)", summed and converted to GB. Test tenant: 142.4 GB.'),

 ("New documents month on month (indicator)",
  "Documents table, small stat", "s40",
  "Whether the system is growing, flat or stalling.",
  "history",
  "Needs a decision, not a source. The reporting table is replaced weekly, so it knows today only. Question 8."),

 ("Average monthly storage growth (indicator)",
  "Documents table, small stat", "s40",
  "Capacity planning. This is one of the two things slide 12 said were worth keeping when it demoted the content volume screen.",
  "history",
  "Needs the same history decision, plus the file size from the scan."),

 ("Total number of records declared (column)",
  "Records declared table (tile 6 drill)", "s41",
  "The core compliance number, per department.",
  "rec",
  'Count rows in public."Records" grouped by department.'),

 ("Total number of users declaring records (column)",
  "Records declared table (tile 6 drill)", "s41",
  "How many people are doing the declaring. A high record count from one person is a different story from the same count from fifty.",
  "declarer",
  'SELECT COUNT(DISTINCT "CreatedBy") FROM public."Records". Never actually run. One of the two cheap checks worth doing next.'),

 ("Number of records declared per division (column)",
  "Records declared table (tile 6 drill)", "s41",
  "The same figure one level further down the organisation.",
  "ask_division",
  "Cannot be produced. Cloud Governance HAS a Division column and it is empty on all 1,032 rows. Question 5."),

 ("Number of users declaring records per division (column)",
  "Records declared table (tile 6 drill)", "s41",
  "The same again for people.",
  "ask_division",
  "Cannot be produced. TWO OF THE FOUR COLUMNS on this screen are division columns, so half of it prints nothing today."),

 ("Sites with zero record declarations (indicator)",
  "Records table, small stat", "s41",
  "The most valuable compliance figure in the deck: sites that are supposed to be declaring records and are not.",
  "rec",
  'Two sources. Take the compliant site list (Cloud Governance) and subtract the sites that appear in public."Records". THIS IS THE EXACT CASE the business register was chosen for, because a site where the app failed still shows up here and gets asked about.'),

 ("Declaration rate against documents created (indicator)",
  "Records table, small stat", "s41",
  "What share of content is actually being declared. The single best measure of whether EDRMS is working.",
  "derived",
  "TWO COLUMNS MAKE THIS ONE NUMBER: records declared (from the Records table) divided by documents (from the weekly scan), times 100. The top half exists today. The bottom half waits on the scan."),

 ("Total number of physical records declared (column)",
  "Physical counterparts table (tile 7 drill)", "s42",
  "Paper behind the digital records, per department.",
  "phys",
  "Count declared records with the physical flag true, grouped by department."),

 ("Share with counterpart (column)",
  "Physical counterparts table (tile 7 drill)", "s42",
  "What proportion of declared records also exist on paper.",
  "derived",
  "TWO FIGURES: physical counterparts divided by records declared, as a percentage. Both halves are available today, so this one works now."),

 ("Turned over to RAC (column)",
  "Physical counterparts table (tile 7 drill)", "s42",
  "Whether the paper actually reached RAC custody. Currently reads Not captured on the screen, deliberately.",
  "none_rac",
  "NOTHING SOURCES THIS. It printed 58 percent until 17 August, which was invented. Question 6."),

 ("Columns 3, 4 and 5 of this table",
  "Physical counterparts table (tile 7 drill)", "s42",
  "Worth flagging: these three columns are copied word for word from slide 41 and only column 2 was edited to say physical. It reads like a slide that was partly updated.",
  "phys",
  "ASK BEFORE BUILDING LITERALLY: should these be the physical equivalents rather than repeats of the records columns? That is question 10."),

 ("Number of records due for disposal (column)",
  "Disposal table (tile 8 drill)", "s43",
  "The disposal work queue per department.",
  "due",
  "Count declared records whose disposal due date is inside the window."),

 ("Next due date for disposal (column)",
  "Disposal table (tile 8 drill)", "s43",
  "When the next action falls due, so departments can plan.",
  "due",
  "The earliest disposal due date still in the future, per department."),

 ("With physical counterpart (column)",
  "Disposal table (tile 8 drill)", "s43",
  "Which of the records due for disposal also have paper to deal with. A physical disposal is a different job.",
  "phys",
  "Count records due for disposal where the physical flag is also true. Both halves exist today."),

 ("Disposal approver (column)",
  "Disposal table (tile 8 drill)", "s43",
  "Who signed off the disposal. Required for audit. Reads Not captured on the screen.",
  "app_disposal",
  "The EDRMS application does not have this field. Change request, not a client question."),

 ("Status: Approved (column)",
  "Disposal table (tile 8 drill)", "s43",
  "Disposals that were approved. Reads Not captured.",
  "app_disposal",
  "Same change request. The client's own slide names the exact values, which is why this is writable as a specification rather than a question."),

 ("Status: Declined (column)",
  "Disposal table (tile 8 drill)", "s43",
  "Disposals that were refused, and therefore still held.",
  "app_disposal",
  "Same change request."),

 ("Status: Extended (column)",
  "Disposal table (tile 8 drill)", "s43",
  "Disposals postponed. An extension also needs to record the new date and the reason, or it cannot be audited.",
  "app_disposal",
  "Same change request. Note the extension needs TWO more fields beyond the status."),

 ("Records disposed month on month (indicator)",
  "Disposal table, small stat", "s43",
  "What was actually destroyed or transferred, as opposed to what was due.",
  "app_disposal",
  "Needs a disposal DATE field, which does not exist. Until then every disposal figure is what is due and none is what was done."),

 ("Due within 30 days, 90 days, 12 months, next quarter (indicators)",
  "Disposal table, small stats", "s43",
  "The near term workload, so records management can staff for it.",
  "due",
  "All four are the same query with a different date window on the disposal due date. Producible today."),

 ("Comparison: users against documents, users against records, documents against records",
  "Comparison panel", "s17, from s6",
  "Three ratios that put the raw counts in context. A department with many users and few records is a different problem from one with few users and many records.",
  "derived",
  "THREE PAIRS OF FIGURES. Users comes from the activity report, records from the Records table, documents from the scan. Two of the three need the scan before they work."),

 ("Records Declaration Trend (cumulative curve)",
  "Trend panel, with a date range filter", "s18, from s10",
  "Shows declared records building up over time, so you can see whether adoption is accelerating or stalling.",
  "rec",
  'The declaration date on each record, counted per month and then ADDED UP RUNNING TOTAL. WATCH OUT: a cumulative curve ENDS at the total. It does not sum to it. Checking the wrong one passes happily on a chart that is wrong by a factor of six.'),

 ("Trend: default range, and a per department cut",
  "Trend panel", "s10",
  "The client asked two things here that are still unsettled.",
  "rec",
  "OPEN QUESTION. Slide 10 asks what the default range should be (they suggest current year) and asks for a trend per department. The 16 August instruction REMOVED the department filter. Those conflict. Question 14."),
]


DP = [
 ("Go-Live date",
  "Above the top panel", "s53",
  "When this department started using EDRMS. Everything else on the screen should be read against it: a department live for one month is not failing.",
  "ask_golive",
  "Cannot be produced. Reads Not captured on the screen. Check the Cloud Governance Job monitor before asking the client."),

 ("Total number of sites created",
  "Top panel, tile 1", "s53",
  "How many EDRMS sites this department runs.",
  "cg_sites",
  'Found in report? Yes. Workspace report, count rows where "EDRMS Site Type" is filled and "Department" matches.'),

 ("Total number of EDRMS users",
  "Top panel, tile 2", "s53",
  "How many people in the department are using EDRMS.",
  "ask_userdept",
  "Blocked on user to department. We can count active users bank-wide. We cannot yet say which department a person belongs to."),

 ("Total number of site visitors",
  "Top panel, tile 3", "s53, s56",
  "How much the department's sites are being looked at, as opposed to contributed to.",
  "m365_visits",
  'Found in report? Yes. Site usage report, column "Page View Count". THE RULE: add consecutive 7 day figures. Never add the 30, 90 or 180 day figures, they overlap.'),

 ("Total number of documents in EDRMS",
  "Top panel, tile 4", "s53, s57",
  "How much content the department holds.",
  "scan_docs",
  'Needs the scan. Interim: sum "File Count" for that department\'s sites.'),

 ("Total number of records declared in EDRMS",
  "Top panel, tile 5", "s53, s58",
  "The department's compliance number.",
  "rec",
  'Count rows in public."Records" for that department\'s sites.'),

 ("Total number of physical counterparts identified",
  "Top panel, tile 6", "s53, s59",
  "Paper behind the department's declared records.",
  "phys",
  "Count declared records with the physical flag, for that department."),

 ("Total number of records due for disposal",
  "Top panel, tile 7", "s53, s60",
  "The department's disposal queue.",
  "due",
  "Count declared records due for disposal, for that department."),

 ("Name of sites, and site owners (columns)",
  "Sites table (tile 1 drill), sorted and paged", "s55",
  "Lists the department's sites and who owns each one, so a question about a site has a name attached.",
  "cg_owner",
  'Found in report? Yes. Workspace report, columns "Site URL" and "Primary Business Owner". Use Cloud Governance, not the SharePoint export, where 19 sites have no owner.'),

 ("Documents, records, counterparts, records due (columns)",
  "Sites table (tile 1 drill)", "s55",
  "The same four measures again, this time per site rather than per department.",
  "rec",
  "THE SAME FOUR MEASURES appear on nine screens in this deck. Build them ONCE at document level and group differently per screen, and every screen agrees by construction."),

 ("Sites with no activity in 90 days (indicator)",
  "Sites table, small stat", "s55",
  "Sites nobody is touching, which the department focal should be asked about.",
  "cg_active",
  'Found in report? Yes. Workspace report, column "Last Active Time".'),

 ("Division (column)",
  "Users table (tile 2 drill)", "s54",
  "Splits the department's users one level further down.",
  "ask_division",
  "Cannot be produced. The Division column exists in Cloud Governance and is empty on all 1,032 rows."),

 ("Users by staff, contractor and consultant (columns)",
  "Users table (tile 2 drill)", "s54",
  "Who the department's users actually are.",
  "ask_userreg",
  "Cannot be produced. THE CLIENT ALREADY DREW THE TABLE THAT WOULD FIX THIS, on this very slide, and asked themselves who owns it."),

 ("Training completed, and date completed (columns)",
  "Users table (tile 2 drill)", "s54",
  "Whether the department's users were trained, and when.",
  "ask_userreg",
  "Cannot be produced. Same table, same question 3."),

 ("Active users, users who have not accessed, new users (indicators)",
  "Users table, small stats", "s54",
  "The client's note asks for these month on month, to spot people who left the Bank and people who need onboarding.",
  "m365_activity",
  "The activity half is producible. The month on month half needs the history decision, question 8. The department half needs user to department."),

 ("Total number of visitors (column)",
  "Visitors table (tile 3 drill)", "s56",
  "How much the sites are being read.",
  "m365_visits",
  'Found in report? Yes. Site usage report, column "Page View Count".'),

 ("Visitors external and internal (columns)",
  "Visitors table (tile 3 drill)", "s56",
  "Whether the readers are Bank staff or outside parties, which is an access risk question.",
  "ask_visitors",
  "PROBABLY producible. Cloud Governance publishes a guest user report but it has not been joined to sites yet. We should check rather than ask."),

 ("Access requests granted and denied (columns)",
  "Visitors table (tile 3 drill)", "s56",
  "Who asked for access and what was decided.",
  "ask_visitors",
  "UNCONFIRMED. A Cloud Governance request report exists. Nobody has confirmed it holds approvals and refusals. Check before answering the client."),

 ("Documents created / uploaded, users creating, documents size (columns)",
  "Documents table (tile 4 drill)", "s57",
  "The department's content, who is producing it and how much space it takes.",
  "scan_docs",
  "Needs the scan for all three. The users creating column additionally needs a new column, and is NOT the same as users declaring."),

 ("Documents migrated, users migrating, migrated size (columns)",
  "Documents table (tile 4 drill)", "s57",
  "Content brought in from the old system.",
  "scan_docs",
  "THE DECK CONTRADICTS ITSELF HERE. Slide 11 says of the migration dashboard: I do not believe this is required at all. Slide 57 then asks for three migration figures. Question 9. Nothing marks a document as migrated today in any case."),

 ("Records declared, and users declaring, with collapsible division rows",
  "Record declaration table (tile 5 drill)", "s58",
  "The department's declaration performance, openable to division level. The client specifically asked for collapsible division rows.",
  "rec",
  "The department level works today. THE DIVISION ROWS THE CLIENT ASKED FOR cannot be filled, because nothing populates division."),

 ("Sites with no declared record in 90 days (indicator)",
  "Declaration table, small stat", "s58",
  "Sites that have gone quiet on declaring, which is the compliance early warning.",
  "rec",
  'Compare the compliant site list against the latest declaration date per site in public."Records".'),

 ("Declaration rate per site (indicator)",
  "Declaration table, small stat", "s58",
  "What share of each site's content is being declared.",
  "derived",
  "TWO COLUMNS: records declared divided by documents, per site, as a percentage. Waits on the scan for the bottom half."),

 ("Physical records declared, per division (columns)",
  "Physical counterpart table (tile 6 drill)", "s59",
  "Paper counterparts for the department.",
  "phys",
  "The department total works today. The division split does not. Note this slide repeats slide 58's columns, the same probable copy as slide 42."),

 ("Library name, records due, next due date (columns)",
  "Disposal table (tile 7 drill)", "s60",
  "The disposal queue broken down by library, which is how a records officer actually works through it.",
  "due",
  'Producible. Join the due records to their library via ListId, which Graph returns as list.id on the drives call.'),

 ("Disposal approver and status columns",
  "Disposal table (tile 7 drill)", "s60",
  "Who approved what. Reads Not captured.",
  "app_disposal",
  "The whole of slide 60 waits on the same change request as slide 43."),

 ("Library usage by file plan category",
  "Library usage panel, behind a category picker", "s61 to s66",
  "Shows which parts of the file plan the department actually uses, one screen per category in the deck.",
  "ask_termjoin",
  "BLOCKED TWICE: we do not have the file plan (question 12) and nothing links a term to a library or document (question 13). Six slides depend on it."),

 ("Record declaration trend (cumulative)",
  "Trend panel", "s24",
  "The department's declaration building up over time.",
  "rec",
  "Same as bank-wide: count per month, then running total. A cumulative curve ENDS at the total rather than summing to it."),

 ("Conventions: approved document, approval date, version",
  "Conventions and programme dates panel", "s26",
  "The department's agreed naming and structure rules, so a reviewer can check what they are being held to.",
  "reflist",
  "NOT A MEASUREMENT. It is a document link and some dates that somebody maintains. Needs an owner, not a query."),

 ("Records management programme dates",
  "Conventions and programme dates panel", "s28",
  "Planned activities: audits, reviews, training, focal group meetings.",
  "reflist",
  "NOT A MEASUREMENT either. Same answer: who maintains this list and where does it live."),
]


PJ = [
 ("Project name and the eight profile fields",
  "Project profile panel", "s38",
  "Facility type, project number, type and modality, modality number, country, status, effectivity date and closing date. The context for everything below.",
  "ask_projattr",
  "NONE of these is EDRMS data. They come from an ADB project system that has never been named in this work. Question 16."),

 ("Total number of sites created",
  "Top panel, tile 1", "s38",
  "How many EDRMS sites this project runs.",
  "ask_project",
  "Blocked on the project register. We can count sites. We cannot tell which belong to this project."),

 ("Total number of EDRMS users",
  "Top panel, tile 2", "s38",
  "How many people work in the project's sites.",
  "ask_project",
  "Blocked on the project register, and separately on user to department."),

 ("Total number of site visitors",
  "Top panel, tile 3", "s38",
  "How much the project's sites are read.",
  "ask_project",
  'The measure itself is available ("Page View Count"). The grouping is not.'),

 ("Total number of documents in EDRMS",
  "Top panel, tile 4", "s38",
  "The project's content volume.",
  "ask_project",
  "Blocked on the project register, AND on the scan. Two separate blockers on one tile."),

 ("Total number of records declared in EDRMS",
  "Top panel, tile 5", "s38",
  "The project's compliance number.",
  "ask_project",
  "The measure works today. Only the project grouping is missing."),

 ("Total number of physical counterparts identified",
  "Top panel, tile 6", "s38",
  "Paper behind the project's records.",
  "ask_project",
  "The measure works today. Only the project grouping is missing."),

 ("Total number of records due for disposal",
  "Top panel, tile 7", "s38",
  "The project's disposal queue. Especially relevant for closed projects.",
  "ask_project",
  "The measure works today. Only the project grouping is missing."),

 ("Site level drill tables behind each tile",
  "Table below the tiles", "s38",
  "Each tile opens the same figure site by site, so a project team can see which of its sites is behind.",
  "ask_project",
  "Same blocker throughout. THE WHOLE DASHBOARD turns on one spreadsheet: question 2."),

 ("Number of users chart, declaration trend, documents against records",
  "Three charts at the foot", "s38",
  "The project's user mix, its declaration over time, and how much of its content is declared.",
  "ask_project",
  "Same blocker, plus the user mix additionally needs the user register (question 3) and the documents half needs the scan."),
]


FP = [
 ("Total terms",
  "Top stats", "s47",
  "How big the file plan is. The one figure here that would work the moment we are handed the plan.",
  "ask_fileplan",
  "Blocked on the file plan itself. Count the terms once we have it. Question 12."),

 ("Total terms per category, all five categories",
  "Top stats", "s47",
  "How the plan is distributed across Institutional Management, Administration, People Management, Programs and Operations, and Other.",
  "ask_fileplan",
  "Blocked on the file plan. NOTE A CONFLICT IN THE DECK: slide 29 lists FOUR categories and omits Institutional Management and Other. Slide 47 lists FIVE. The prototype follows slide 47 because the detailed design beats the outline, but this needs confirming. That is question 18."),

 ("Category rollup table: documents, records, counterparts, records due per category",
  "Rollup table", "s47",
  "The same four measures again, grouped by file plan category this time.",
  "ask_termjoin",
  "BLOCKED ON THE JOIN, not just the plan. Even holding the file plan, nothing links a term to a document. This survives the answer to question 12."),

 ("Term table for a category: each term with its measures",
  "One screen per category", "s48 to s52",
  "Which individual terms are being used and how heavily.",
  "ask_termjoin",
  "Same. Question 13, the longest lead item in the deck."),

 ("Most used terms, and least used terms",
  "Two charts per category screen", "s48 to s52",
  "Underused terms mean the plan does not match how people actually work, which is a reason to revise it.",
  "ask_termjoin",
  "Same blocker. THIS IS THE MOST USEFUL THING ON THE DASHBOARD and it is the furthest from being buildable."),

 ("The five indicators repeated on every category screen",
  "Below each category table", "s48 to s52",
  "The client repeats the same five indicators on all five category screens, which is a strong signal they matter.",
  "ask_termjoin",
  "Same blocker throughout."),

 ("A possible shortcut worth testing",
  "Applies to the whole dashboard", "s48, s61",
  "The client uses Annual Meetings as BOTH a library name and a file plan term, on two different slides.",
  "ask_termjoin",
  "IF THE LIBRARY NAME IS THE TERM, this whole dashboard unblocks without any application change. Worth putting to the client as the first option in question 13, because it is by far the cheapest answer."),
]


RD = [
 ("Permanent and temporary rollup: departments, libraries, records, counterparts, records due, records disposed",
  "Landing screen", "s44",
  "The whole holding split into records kept forever and records with an end date. This is the landing screen reached from the Bank-wide tile, which is why slide 44 carries the Bank-wide banner.",
  "purview_split",
  "FIVE OF THE SIX COLUMNS work today. Records disposed does not, it needs the change request. But see the row below: the split itself is unproven."),

 ("The permanent versus temporary split itself",
  "Applies to the whole dashboard", "s44, s45, s46",
  "Everything on this dashboard depends on sorting 53 retention labels into two groups.",
  "purview_split",
  "WE HOLD THE LABEL ON EVERY RECORD. WE DO NOT HOLD THE RULE that turns 53 labels into two groups. The design has a retention status field fed from EDRMSMasters. CHECK THAT FIRST: if it already carries the split, this closes with no work. Question 11."),

 ("Number of libraries provisioned",
  "Rollup table", "s44",
  "How many libraries are under each retention class.",
  "graph_libs",
  "Found in report? No, a Graph call. GET /sites/{siteId}/drives, and count the drives. list.id joins a library to the Records table."),

 ("Departments, offices and RMs provisioned",
  "Rollup table", "s44",
  "How widely each retention class is used across the Bank.",
  "cg_dept",
  'Join the declared records to their site, then to the Workspace report column "Department", and count distinct departments.'),

 ("Number of records disposed",
  "Rollup table", "s44",
  "What was actually destroyed. Reads Not captured.",
  "app_disposal",
  "Needs the disposal fields. Question 7. Every other figure on this screen is what is DUE."),

 ("Permanent retention table: top level terms with document counts",
  "Permanent retention panel", "s45",
  "Which categories of record are kept forever, and how much of each.",
  "purview_split",
  "Drawn as its OWN table rather than sharing one with temporary, because the client's two slides carry different columns: permanent has a document count and NO retention label."),

 ("Temporary retention table: terms, retention label, due date, disposed count",
  "Temporary retention panel", "s46",
  "Records with an end date, when they fall due, and what has been done.",
  "purview_split",
  "The label and the due date work today. The disposed count does not, it needs the change request. Note this table carries the retention label where the permanent one does not."),
]


RA = [
 ("The whole dashboard",
  "A notice at the top of the screen", "s33, s67",
  "THE CLIENT HAS NOT SPECIFIED THIS DASHBOARD. Slide 33 is literally the words RAC suggestions? and slide 67 asks what Opus offers. It is kept because slide 13 names it as one of the six key views and slides 68 and 69 are drawn with real headings, so the SHAPE is a requirement even though nothing fills it.",
  "none_holdings",
  "EVERY MEASURE ON THIS DASHBOARD READS Not captured. That is deliberate. Until 17 August it printed 1,840 boxes, 9,260 folders and 78 percent capacity used, none of which had any source at all."),

 ("Total boxes stored, total folders stored, storage requests, retrieval requests",
  "Top stats", "s68, s69",
  "The headline physical holding figures.",
  "none_holdings",
  "No system we can reach holds boxes or folders. All four read Not captured."),

 ("Storage table: location, requests, requestors, boxes, folders, Remarks",
  "Storage panel", "s68",
  "The client's own column headings, including their Remarks column, kept verbatim so they can see we read the slide.",
  "none_holdings",
  "The COLUMN NAMES are a requirement. The CONTENT has no source. Every cell reads Not captured."),

 ("Retrieval table: location, requests, requestors, boxes, folders, Status, Remarks",
  "Retrieval panel", "s69",
  "Their retrieval table, including their own status vocabulary: Loan, Return to owner, For Disposal.",
  "none_holdings",
  "Same. Slide 67 puts retrieval in eServe, which is a lead we can follow IF the client confirms it. Nothing from Opus goes in, by client instruction."),

 ("Room capacity and percentage of storage capacity available",
  "Listed as an indicator, deliberately not drawn", "s68",
  "Whether there is physical space left.",
  "none_holdings",
  "THE SHARPEST CASE IN THE PROJECT. Slide 68 asks US: can room capacity and percentage of available storage capacity be included? We were answering their open question with an invented chart. Now it is listed as a question, not drawn as a fact."),

 ("The other indicators each slide asks for",
  "Listed below each table, not drawn", "s68, s69",
  "Month on month movements, outstanding requests, counterparts not yet transferred, disposal of boxes and folders.",
  "none_holdings",
  "LISTED RATHER THAN DRAWN, because drawing them would mean inventing them a second time."),
]


SHEETS = [
    ("Bank-wide Oversight", "s15, s34 to s44 and s47", BW),
    ("Department Insights", "s19 to s28, s53 to s66", DP),
    ("Project Insights WITHDRAWN", "s36 to s38",        PJ),
    ("Institutional File Plan", "s29 to s31, s47 to s52", FP),
    ("Retention and Disposal",  "s32, s44 to s46",     RD),
    ("Records and Archive Holdings", "s33, s67 to s69", RA),
]

COLS = [
    ("#", 5),
    ("What is on the dashboard", 42),
    ("Where to find it on the screen", 30),
    ("Which slide asked for it", 16),
    ("What it is for, in plain English", 52),
    ("Can we build it?", 22),
    ("How to prove it, step by step", 78),
    ("Which column the number comes from", 60),
]


def style_header(ws, row, text, fill, size=11, height=None):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLS))
    c = ws.cell(row=row, column=1)
    c.font = Font(bold=True, size=size, color="FFFFFF" if fill == BLUE else INK)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    if height:
        ws.row_dimensions[row].height = height


def build_sheet(wb, name, slides, rows):
    ws = wb.create_sheet(name[:31])
    ws.sheet_view.showGridLines = False

    for i, (h, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    style_header(ws, 1, name, BLUE, size=14, height=26)
    withdrawn = "WITHDRAWN" in name
    style_header(ws, 2, ("The client's slides for this dashboard: " + slides + ".   " +
                 ("THIS DASHBOARD IS NOT ON THE PAGE. It was withdrawn on 17 August because not one figure "
                  "on it could be produced. It is kept whole and working in "
                  "EDRMS_Prototype_with_project_sites_2026-08-17.html and goes straight back when the client "
                  "supplies the project register. The rows below say what each screen would need."
                  if withdrawn else
                  "Every row below is something actually on the screen today.")),
                 AMBER if withdrawn else BLUEPALE, size=10, height=32 if withdrawn else 20)

    # count the verdicts so the reader gets the shape before the detail
    tally = {}
    for r in rows:
        b = SRC[r[4]]["buildable"]
        tally[b] = tally.get(b, 0) + 1
    order = [READY, SCAN, APP, ASK, NONE]
    summ = ";   ".join("%s: %d" % (k, tally[k]) for k in order if k in tally)
    style_header(ws, 3, "Of the %d things on this screen, %s" % (len(rows), summ),
                 GREY, size=10, height=20)

    hr = 5
    for i, (h, _) in enumerate(COLS, 1):
        c = ws.cell(row=hr, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        c.border = BOX
    ws.row_dimensions[hr].height = 34

    r = hr + 1
    for n, (what, where, slide, purpose, key, basis) in enumerate(rows, 1):
        buildable, process, basis_txt = S(key, basis)
        vals = [n, what, where, slide, purpose, buildable, process, basis_txt]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=True,
                                    horizontal="center" if i == 1 else "left")
            c.border = BOX
            c.font = Font(size=10)
            if i == 2:
                c.font = Font(size=10, bold=True)
            if i == 6:
                c.fill = PatternFill("solid", fgColor=VERDICT_FILL[buildable])
                c.font = Font(size=10, bold=True)
        # tall enough for the process column, which is the long one
        ws.row_dimensions[r].height = max(60, 11 * (process.count("\n") + 1))
        r += 1

    ws.freeze_panes = ws.cell(row=hr + 1, column=3)
    return tally


def build_readme(wb, tallies):
    ws = wb.create_sheet("Read me first", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 104

    r = [1]

    def head(t, fill=BLUE, size=13, h=26):
        ws.cell(row=r[0], column=2, value=t)
        ws.merge_cells(start_row=r[0], start_column=2, end_row=r[0], end_column=3)
        c = ws.cell(row=r[0], column=2)
        c.font = Font(bold=True, size=size, color="FFFFFF" if fill == BLUE else INK)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[r[0]].height = h
        r[0] += 1

    def pair(a, b, fill=None):
        ws.cell(row=r[0], column=2, value=a).font = Font(bold=True, size=10)
        ws.cell(row=r[0], column=3, value=b).font = Font(size=10)
        ws.cell(row=r[0], column=2).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=r[0], column=3).alignment = Alignment(vertical="top", wrap_text=True)
        if fill:
            ws.cell(row=r[0], column=2).fill = PatternFill("solid", fgColor=fill)
        ws.row_dimensions[r[0]].height = max(15, 13 * (1 + len(b) // 100))
        r[0] += 1

    def gap():
        r[0] += 1

    head("EDRMS Utilization Report: what is on each dashboard, and how to prove it", BLUE, 14, 30)
    gap()
    pair("What this workbook is",
         "One sheet per dashboard. One row per thing on the screen. It answers four questions for every "
         "single item: which slide asked for it, what it is for, whether we can build it, and exactly how "
         "to prove the number is real.")
    pair("Why it exists",
         "So you can see the whole delivery in one place: what is ready, what we can fix ourselves, and "
         "what we have to ask the client for. Nothing here is guessed. Every row traces to a slide in "
         "EDRMS_Dashboard_requirements_1.pptx and to a source that was actually opened and read.")
    pair("The habit behind it",
         "Do not trust an expected column name or a plausible figure. Export the real file and read it. "
         "Every significant error in this project was caught that way, and several had survived for weeks "
         "looking perfectly correct.")
    gap()

    head("Department / office / RM: what that heading means", BLUEPALE, 12, 22)
    gap()
    pair("It is ONE column, not three",
         "Department / office / RM is three kinds of organisational unit, and a site might be owned by any "
         "of them. A department (CSD, Corporate Services Department), an office (BIOC, Office of Business "
         "Intelligence and Operations Coordination), or a resident mission. The client needed one heading "
         "that covers all three, because the answer has to go in one cell. Read it as: whoever owns this site.")
    pair("The test tenant behaves this way, confirmed",
         "All offices and resident missions sit in the SAME column, the one called Department. Its NAME is "
         "just narrower than its CONTENTS. Proof from the test tenant: of the five codes in that column, BOD "
         "is the Board of Directors and ADBI is the ADB Institute, and neither is a department. The client's "
         "own slide 39 does the same thing, listing BIOC (an office) and CCSD (a department) as rows of one column.")
    pair("THE GAP: we cannot tell which TYPE a code is",
         "Nothing in any export says BIOC is an office and CCSD is a department. So \"show me the site count per "
         "unit\" works today, and \"show me only the resident missions\", or any subtotal by unit type, does not. "
         "No slide asks for a subtotal by type, so this blocks nothing that is drawn, but it is worth asking the "
         "client for their unit codes with a type against each. It would also give us the full list of valid codes.")
    pair("ALSO UNRESOLVED: what does RM stand for?",
         "Probably Resident Mission, which fits ADB's structure of departments, offices, resident missions and "
         "projects. But the Cloud Governance export has a Records Manager column, and in a records management deck "
         "RM is at least as commonly Records Manager. The test tenant cannot settle it: Records Manager is filled "
         "on 2 of 1,032 rows with the values Alex Wilber and Admin, which is test noise. It matters: a resident "
         "mission is another organisational unit and lands in the same column, while a records manager is a PERSON "
         "and would be a different requirement needing a different source. One line to ask, and it changes what gets built.")
    gap()
    pair("The filter behind every site count",
         "Three conditions, and two of them look unnecessary here:   "
         "EDRMS Site Type is not blank (it is an EDRMS site)   AND   Status = Active (not locked, not archived)   "
         "AND   Site Status = Active (not deleted).   In the test tenant all 1,032 EDRMS sites are Active on both, "
         "so the last two change nothing and the short version passes every test. Across all 1,209 workspaces there "
         "are 59 Locked and 4 Archived, so at ADB they will matter. This is exactly the kind of filter that is "
         "silently wrong in production and correct in test.", AMBER)
    gap()

    head("How to read the columns", BLUEPALE, 12, 22)
    gap()
    pair("What is on the dashboard", "The item itself: a tile, a table column, a chart or a panel.")
    pair("Where to find it on the screen", "So you can point at it while reading the row.")
    pair("Which slide asked for it", "The slide number in the client's own deck. If a row has no slide, it should not be on the screen.")
    pair("What it is for", "Why anyone would want this number. Written for a reader who has never seen the system.")
    pair("Can we build it?", "One of five answers, colour coded. See the next section.")
    pair("How to prove it, step by step", "The chain: buildable, then whether it is proven in the test tenant, then where the number lives, then the link, then the clicks. Follow it top to bottom and you will end up looking at the real number.")
    pair("Which column the number comes from", "The exact column name. Where a figure needs two columns, the sum is written out, for example: records declared divided by documents, times 100, equals the declaration rate.")
    gap()

    head("The five answers in the Can we build it column", BLUEPALE, 12, 22)
    gap()
    pair(READY, "The data exists in something we already hold. Building the report is the only work left.", GREEN)
    pair(SCAN, "It needs the weekly SharePoint scan, which is designed and not yet built. NEEDS NOBODY OUTSIDE THE TEAM. This is the biggest thing we can fix ourselves.", AMBER)
    pair(APP, "The EDRMS application does not capture the field. A development change request for Mihal Le. This is all disposal status, approver and dates.", RED)
    pair(ASK, "Only the client can supply it: the project register, the user register, divisions, go-live dates, the file plan.", RED)
    pair(NONE, "No system anywhere holds it. Mostly the physical holdings dashboard, which the client has not specified.", RED)
    gap()

    head("What the whole thing adds up to", BLUEPALE, 12, 22)
    gap()

    total = {}
    for t in tallies:
        for k, v in t.items():
            total[k] = total.get(k, 0) + v
    n = sum(total.values())
    for k in [READY, SCAN, APP, ASK, NONE]:
        if k in total:
            pair(k, "%d of %d items, %d percent" % (total[k], n, round(100 * total[k] / n)),
                 VERDICT_FILL[k])
    gap()
    pair("The three things worth doing first",
         "1. BUILD THE WEEKLY SCAN. Every document figure, declaration rate and storage figure waits on "
         "it, and it needs no client input at all. "
         "2. ASK FOR THE PROJECT REGISTER (question 2). One spreadsheet turns three blank screens into "
         "working ones. "
         "3. RAISE THE FILE PLAN JOIN (question 13). It is the longest lead item in the deck and it will "
         "not get shorter by waiting.")
    pair("Two free checks nobody has run",
         "Neither needs another person. (a) In Graph Explorer: GET /users?$select=userPrincipalName,"
         "department,jobTitle. If ADB fills the department field in, the user to department gap closes "
         "for nothing. (b) In the database: SELECT COUNT(DISTINCT \"CreatedBy\") FROM public.\"Records\". "
         "That is the users declaring records figure, and it has never actually been run.")
    gap()

    head("Two warnings that have already caught us out", AMBER, 12, 22)
    gap()
    pair("Test tenant, not production",
         "Every figure quoted in this workbook was measured against the TEST tenant 7rkd12. THE METHOD "
         "transfers to ADB exactly: same reports, same names, same columns. NO FIGURE transfers. Re-run "
         "each export against production and read the real number.")
    pair("A number that reconciles is not a proven number",
         "In August a rounding fault made a department show 103 records due against a real 75, and every "
         "automatic check passed, because the checks were comparing two figures that were both wrong. "
         "Where a figure can be checked against something outside the report, check it there.")
    gap()

    pair("Built by", "build_content_checker.py. Never hand edit this workbook: edit the script and rerun, "
                     "or the next rebuild will silently discard your changes.")
    pair("Sources behind it", "EDRMS_Dashboard_requirements_1.pptx (69 slides), "
                              "REQUIREMENTS_AUDIT_2026-08-17.md, STATUS.md, and the evidence CSV exports in the repo.")


def main():
    wb = Workbook()
    wb.remove(wb.active)
    tallies = [build_sheet(wb, n, s, r) for n, s, r in SHEETS]
    build_readme(wb, tallies)
    wb.active = 0
    path = os.path.abspath(OUT)
    wb.save(path)
    n = sum(len(r) for _, _, r in SHEETS)
    print("wrote %s" % path)
    print("%d sheets, %d content rows" % (len(SHEETS) + 1, n))
    for (name, _, rows), t in zip(SHEETS, tallies):
        short = {READY: "ready", SCAN: "scan", APP: "app", ASK: "ask", NONE: "none"}
        print("  %-32s %3d rows   %s" % (name, len(rows),
              "  ".join("%s=%d" % (short[k], t[k])
                        for k in [READY, SCAN, APP, ASK, NONE] if k in t)))


if __name__ == "__main__":
    main()
