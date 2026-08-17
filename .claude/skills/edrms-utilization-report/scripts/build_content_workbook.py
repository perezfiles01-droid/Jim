#!/usr/bin/env python3
"""Build the dashboard content workbook from the prototype itself.

    node extract_contents.js /home/user/Jim/index.html > /tmp/contents.json
    python3 build_content_workbook.py /tmp/contents.json <out.xlsx>

The client's instruction was "do it similar to how exactly the prototype is
named so I will not get confused". So the NAME column is never typed here: it
comes from extract_contents.js, which reads the labels off the rendered page.
The workbook cannot disagree with the screen because it holds no second copy
of the names.

Everything else, what a row of the measure counts, whether it is buildable,
which file and column carry it and how the number is built, comes from the
SOURCES table below. That table is hand maintained and deliberately partial:
where nothing here has been proved against a real export, the cells are left
BLANK and the verdict reads "To confirm". An expectation written into a source
column is indistinguishable from a verified fact three weeks later, which is
the single habit that has caught every real error in this project.

NEVER HAND EDIT THE WORKBOOK. Edit this file, or the prototype, and rerun.
"""
import json, sys, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY   = "0E1F35"
BLUE   = "0072BC"
HEADBG = "10243E"
GREEN  = "DDEEDA"   # ready
AMBER  = "FBE7D5"   # needs an application change or a scan we have not built
RED    = "F7D9D9"   # needs the client
GREY   = "ECEFF3"   # to confirm

# ---------------------------------------------------------------------------
# What we know, per measure name as the prototype shows it.
#
#   counts    : what ONE unit of this measure is. The question that caught the
#               visitors error (a page view is an event, a visitor is a person)
#   verdict   : Ready | Needs the weekly scan | Needs an app change | Needs the
#               client | No source | To confirm
#   file      : the export or table that carries it
#   column    : the column on that export
#   build     : how the number is built, one hop at a time
#   blocked   : what it waits on, and which audit question
# ---------------------------------------------------------------------------
R  = "Ready"
SC = "Needs the weekly scan"
AC = "Needs an application change"
CL = "Needs the client"
NO = "No source anywhere"
TC = "To confirm"
NM = "Not a measure"      # a heading or a title, nothing to source

CG   = "Cloud Governance Workspace report"
USE  = "SharePoint site usage detail (M365)"
ACT  = "SharePoint activity user detail (M365)"
DB   = 'drm-npr PostgreSQL, public."Records"'
CGU  = "Cloud Governance Users export"

SOURCES = {
 "Number of active EDRMS SharePoint sites for Department, RM, office": dict(
   counts="One SharePoint site that RAC has designated as EDRMS and that is still live.",
   verdict=R, file=CG, column="EDRMS Site Type, Status, Site Status",
   build=("1. Export the Workspace report from Cloud Governance, Directory.\n"
          "2. Keep rows where EDRMS Site Type is not blank.\n"
          "3. AND Status = Active.\n"
          "4. AND Site Status = Active.\n"
          "5. Count the rows."),
   blocked=("All three filters are needed. In the test tenant all 1,032 are Active on both, "
            "so the short version passes every test here and is wrong at ADB, where 59 are "
            "Locked and 4 Archived.")),

 "EDRMS users with recorded activity, last 180 days": dict(
   counts="One licensed person who opened a SharePoint site in the window.",
   verdict=CL, file=ACT, column="User Principal Name, Last Activity Date",
   build=("1. Export SharePoint activity user detail at period D180.\n"
          "2. Keep rows with a Last Activity Date inside the window.\n"
          "3. Count distinct User Principal Name."),
   blocked=("TWO problems. (a) The export has NO SITE DIMENSION, so it cannot be narrowed to "
            "EDRMS: at ADB it returns every licensed user, about 9,400. (b) 180 days is the "
            "longest period Microsoft serves, so a user last active 200 days ago is "
            "indistinguishable from one who never used EDRMS. Audit questions 3 and 16.")),

 "Total number of documents in EDRMS": dict(
   counts="One document held in an EDRMS site, declared or not.",
   verdict=SC, file="The weekly scan, not yet built", column="-",
   build=("1. Take the compliant site list from the Workspace report.\n"
          "2. For each site, walk its libraries through Microsoft Graph.\n"
          "3. Count items that are files, not folders.\n"
          "INTERIM: sum File Count from the site usage export, joined on Site Id."),
   blocked="The weekly scan is the single biggest self-help item left, and needs nobody else."),

 "Total number of records declared in EDRMS": dict(
   counts="One declared record. The table holds NO ROW for an undeclared document.",
   verdict=R, file=DB, column="ListId, ItemId",
   build='SELECT COUNT(*) FROM public."Records"',
   blocked=""),

 "Total number of physical counterparts identified": dict(
   counts="One declared record that also exists on paper.",
   verdict=R, file=DB, column="HasPhysical",
   build='SELECT COUNT(*) FROM public."Records" WHERE "HasPhysical" IS TRUE',
   blocked=("Turned over to RAC is a DIFFERENT measure and has no source: no system records a "
            "physical custody event. Audit question 6.")),

 "Total number of records due for disposal, next 12 months": dict(
   counts="One declared record whose disposal date falls inside the next 12 months.",
   verdict=R, file=DB, column="EDRMSDueDateForDisposal",
   build=('SELECT COUNT(*) FROM public."Records"\n'
          "WHERE \"EDRMSDueDateForDisposal\" BETWEEN now() AND now() + interval '12 months'"),
   blocked="Records DISPOSED needs a disposal status the application does not have. Audit question 7."),

 "Records due for disposal, next 12 months": dict(
   counts="The same measure as the tile above. This card navigates rather than drilling.",
   verdict=R, file=DB, column="EDRMSDueDateForDisposal",
   build="As above. The card opens the Retention and Disposal dashboard.", blocked=""),

 "File plan terms in use": dict(
   counts="One institutional file plan term that at least one library is classified under.",
   verdict=CL, file="", column="",
   build="",
   blocked=("The file plan exists in NEITHER system. The term store holds dropdown value lists, "
            "Purview holds 53 retention labels as a flat list. And nothing joins a term to a "
            "document: table 4 has no join key and table 1 carries no TermId. Audit questions "
            "12 and 13, the longest lead item in the deck.")),

 "Sites created, last 90 days": dict(
   counts="One EDRMS site whose SharePoint site was created in the window.",
   verdict=R, file=CG, column="Created Time",
   build="Count compliant sites where Created Time is inside the last 90 days.",
   blocked=("Created Time is when the SITE was created, NOT the EDRMS go-live date. For a "
            "converted site those differ and no column holds the second. Audit question 4.")),

 "Sites deleted": dict(counts="One EDRMS site in the deleted state.", verdict=R,
   file=CG, column="Site Status", build="Count compliant sites where Site Status = Deleted.",
   blocked=""),

 "Sites archived": dict(counts="One EDRMS site in the archived state.", verdict=R,
   file=CG, column="Status", build="Count compliant sites where Status = Archived.",
   blocked="Previously blocked on 'no definition'. Cloud Governance supplies one."),

 "Inactive over 90 days": dict(
   counts="One EDRMS site with no recorded activity for more than 90 days.",
   verdict=R, file=CG, column="Last Active Time",
   build="Count compliant sites where Last Active Time is older than 90 days.",
   blocked=("The threshold is unsettled: the deck says 90 days, the metrics document says 300. "
            "It is a report-level judgement, never stored, so changing it is one edit.")),

 "Department / office / RM": dict(
   counts=("One organisational unit: a department (CSD), an office (BIOC) or a resident "
           "mission. One column holds all three and nothing says which type a code is."),
   verdict=R, file=CG, column="Department",
   build=("Group the compliant site list by Department.\n"
          "Do NOT use DRM Department: it is filled on all 1,032 rows and every row reads "
          "'CGAdoption'. B-Department, Department22 and Division are empty on every row."),
   blocked=("240 of 1,032 sites (23%) name SEVERAL departments in one cell, e.g. 'CWRD;SARD'. "
            "Counting under every department overstates the total by 26%; counting under the "
            "first named makes SARD vanish. RAC's call. Audit question 1. Separately, RM is "
            "unconfirmed: Resident Mission or Records Manager. Audit question 14.")),

 "Number of EDRMS SharePoint sites": dict(
   counts="One EDRMS site owned by this unit.", verdict=R, file=CG,
   column="EDRMS Site Type, Department",
   build="Count compliant sites grouped by Department.",
   blocked="Same multi-department problem as above. Audit question 1."),

 "Total number of documents": dict(counts="One document in this unit's sites.", verdict=SC,
   file="The weekly scan", column="-", build="Sum the document count over the unit's sites.",
   blocked="Waits on the scan."),

 "Total number of records declared": dict(counts="One declared record in this unit's sites.",
   verdict=R, file=DB, column="SiteUrl",
   build=('1. SELECT "SiteUrl", COUNT(*) FROM public."Records" GROUP BY "SiteUrl"\n'
          "2. Join SiteUrl to the Workspace report's URL.\n"
          "3. Group by that report's Department."),
   blocked="Same multi-department problem. Audit question 1."),

 "Total number of physical counterparts": dict(counts="One declared record with a paper original.",
   verdict=R, file=DB, column="HasPhysical, SiteUrl",
   build="As above, filtered to HasPhysical.", blocked=""),

 "Users with recorded activity": dict(
   counts="One person with activity in the window. NOT a site and NOT a visit.",
   verdict=CL, file=ACT, column="User Principal Name",
   build="Count distinct User Principal Name with activity inside the window.",
   blocked="No site dimension, so it cannot be narrowed to EDRMS. Audit question 3."),

 "Never accessed EDRMS": dict(
   counts="One licensed person with no recorded activity at all.",
   verdict=CL, file=ACT, column="Last Activity Date",
   build="Count rows where Last Activity Date is blank.",
   blocked=("Indistinguishable from 'last active more than 180 days ago', because 180 is the "
            "longest period the report serves. Audit question 16.")),

 "No access in 90 days": dict(
   counts="One person whose last activity is older than 90 days.",
   verdict=CL, file=ACT, column="Last Activity Date",
   build="Count rows where Last Activity Date is older than 90 days.",
   blocked="Same as above, plus the EDRMS narrowing problem. Audit question 3."),

 "Users with activity": dict(counts="One person with activity in the window.", verdict=CL,
   file=ACT, column="User Principal Name, Last Activity Date",
   build="Count distinct users with activity, grouped by the user's department.",
   blocked=("Needs a USER to department mapping, which is different from the SITE to department "
            "mapping already found. Cloud Governance Users export carries Department on 228 of "
            "286 people, keyed on User principal name, the same key the activity report uses. "
            "Untested against production. Audit question 3.")),

 "Never accessed": dict(counts="One person with no recorded activity at all.", verdict=CL,
   file=ACT, column="Last Activity Date", build="As 'Never accessed EDRMS', per unit.",
   blocked="Needs the user to department mapping. Audit question 3."),

 "Departments in the list": dict(counts="One organisational unit appearing in the table.",
   verdict=R, file=CG, column="Department",
   build="Count distinct Department over the compliant site list.", blocked=""),

 "Documents held": dict(counts="One document.", verdict=SC, file="The weekly scan", column="-",
   build="Sum the document count over every compliant site.", blocked="Waits on the scan."),

 "Storage consumed": dict(counts="Bytes held across EDRMS sites.", verdict=R, file=CG,
   column="Storage Used (GB)",
   build=("Sum Storage Used (GB) over the compliant sites.\n"
          "Never infer storage from a document count: that was the 3400 divisor removed on "
          "17 Aug 2026."),
   blocked="Graph returns a CUMULATIVE size on folders, so a scan must filter to files only."),

 "Storage": dict(counts="Bytes held by this unit's sites.", verdict=R, file=CG,
   column="Storage Used (GB)", build="Sum Storage Used (GB) grouped by Department.", blocked=""),

 "People who declared a record": dict(
   counts="One person who declared at least one record. NOT a record and NOT a document.",
   verdict=R, file=DB, column="CreatedBy",
   build='SELECT COUNT(DISTINCT "CreatedBy") FROM public."Records"',
   blocked=("One line of SQL that has never been run. Until it is, the prototype shows demo "
            "data here. This is the cheapest real number left in the project.")),

 "Records declared": dict(counts="One declared record.", verdict=R, file=DB, column="ListId, ItemId",
   build='COUNT(*) from public."Records", grouped by the site\'s department.', blocked=""),

 "Documents declared": dict(
   counts="A RATE, not a count: declared records over documents held, for the same unit.",
   verdict=SC, file=DB + " and the weekly scan", column="-",
   build="Records declared divided by documents held, per unit.",
   blocked=("The denominator does not exist in any system yet. public.\"Records\" holds NO ROW "
            "for an undeclared document, so every rate waits on the scan.")),

 "Physical counterparts": dict(counts="One declared record with a paper original.", verdict=R,
   file=DB, column="HasPhysical", build="Count HasPhysical, grouped by department.", blocked=""),

 "Records due": dict(counts="One record whose disposal date is inside the window.", verdict=R,
   file=DB, column="EDRMSDueDateForDisposal", build="Count by department.", blocked=""),

 "Next due date": dict(counts="The earliest disposal date still ahead, for this unit.",
   verdict=R, file=DB, column="EDRMSDueDateForDisposal",
   build="MIN(EDRMSDueDateForDisposal) where the date is in the future, per unit.", blocked=""),

 "With counterpart": dict(counts="One record due for disposal that also has a paper original.",
   verdict=R, file=DB, column="EDRMSDueDateForDisposal, HasPhysical",
   build="Count records due where HasPhysical is true.", blocked=""),

 "Total number of visitors": dict(
   counts="One PERSON who opened a site. A person visiting fifty times is ONE visitor.",
   verdict=SC, file="Microsoft Graph site analytics", column="actorCount",
   build="GET /sites/{siteId}/analytics/allTime, read actorCount, per site.",
   blocked=("Graph offers allTime and lastSevenDays ONLY. There is no 30 day unique viewer "
            "figure, however often it is asked for.")),

 "Page views": dict(
   counts="One EVENT. Fifty visits by one person are fifty page views.",
   verdict=R, file=USE, column="Page View Count",
   build="Sum Page View Count per site, then group by the site's department.",
   blocked="Page views and visitors are two measures. They must never be shown as one."),

 "Views per visitor": dict(counts="A RATE: page views over visitors, for the same site.",
   verdict=SC, file=USE + " and Graph analytics", column="Page View Count, actorCount",
   build="Page views divided by visitors, per site.",
   blocked="Waits on the visitor figure, which needs a per site Graph call."),

 "Name of sites": dict(counts="One EDRMS SharePoint site.", verdict=R, file=CG,
   column="Workspace Name, URL", build="List the compliant sites for the selected unit.", blocked=""),

 "Site owners": dict(counts="The person accountable for the site.", verdict=R, file=CG,
   column="Primary Business Owner",
   build="Read Primary Business Owner per site.",
   blocked="Populated on 100% of Cloud Governance rows. The M365 export had 19 sites with no owner."),
}

VERDICT_FILL = {R:GREEN, SC:AMBER, AC:AMBER, CL:RED, NO:RED, TC:GREY, NM:"FFFFFF"}

HEAD = ["#", "Where on screen", "What it is",
        "Name, exactly as the prototype shows it",
        "What ONE of these counts",
        "Can we build it?",
        "Which file or table", "Which column",
        "HOW THE NUMBER IS BUILT, one hop at a time",
        "Blocked on, and the question to ask"]
WIDTH = [5, 34, 15, 46, 44, 22, 30, 28, 60, 56]


def style_header(ws, row=1):
    for c in range(1, len(HEAD) + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=HEADBG)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 34
    for i, w in enumerate(WIDTH, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Set by reference, not by ws.cell(): calling ws.cell() to get the anchor
    # MATERIALISES that cell, so the next append() lands a row lower and every
    # sheet starts with a blank row.
    ws.freeze_panes = f"A{row + 1}"


def build(rows, out):
    wb = Workbook()
    thin = Side(style="thin", color="C9D3DF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- Read me first -----------------------------------------------------
    ws = wb.active
    ws.title = "Read me first"
    ws.column_dimensions["A"].width = 118
    intro = [
        ("EDRMS Utilization Report: what is on every dashboard, and whether we can build it", "h1"),
        ("", ""),
        ("Every NAME in this workbook is read off the prototype itself, not typed here.", "b"),
        ("The build script drives the live page, opens every drill on every dashboard and records "
         "each label exactly as it appears. So a name in column D and a name on screen cannot "
         "disagree. If a label changes in the prototype, rerun the script and this workbook "
         "follows it.", ""),
        ("", ""),
        ("How to read a row", "h2"),
        ("Column D is what you see on screen. Column E is the question that has caught more errors "
         "in this project than any other: what is ONE of these actually counting? A visitor is a "
         "person; a page view is an event. Getting that wrong produces a plausible wrong number, "
         "not an error.", ""),
        ("Column F is the verdict, and it is colour coded:", ""),
        ("    Ready                        we can produce this today from an export we already pull", "ready"),
        ("    Needs the weekly scan        buildable, but waits on a job nobody has built yet. Needs NO ONE else", "scan"),
        ("    Needs an application change  the EDRMS application does not carry the field", "scan"),
        ("    Needs the client            we cannot proceed without an answer or a list from ADB", "client"),
        ("    No source anywhere          no system we can reach holds this", "client"),
        ("    To confirm                  not yet checked against a real export. NOT a claim either way", "tc"),
        ("", ""),
        ("Columns G, H and I are the proof chain: which file, which column, and the steps to get "
         "from the export to the number. Column J says what it waits on and which audit question "
         "covers it.", ""),
        ("", ""),
        ("Where a cell is BLANK, nothing has been verified.", "b"),
        ("That is deliberate. An expectation written into a source column is indistinguishable "
         "from a verified fact three weeks later, and that single habit has caught every real "
         "error in this project. A blank cell is honest; a plausible guess is not.", ""),
        ("", ""),
        ("The figures in the prototype are DEMO DATA.", "b"),
        ("Nothing on the page is a measured ADB figure. The prototype is a specification for the "
         "future Power BI report, agreed before anyone opens Power BI Desktop. Where a real "
         "test-tenant figure exists it is recorded in STATUS.md section 5, not here.", ""),
        ("", ""),
        ("Two things about 'Department / office / RM'", "h2"),
        ("One column holds three kinds of organisational unit: a department (CSD), an office "
         "(BIOC) or a resident mission. Nothing in any export says which type a code is, so "
         "'sites per unit' works and 'only the resident missions' does not.", ""),
        ("And 240 of 1,032 sites name several departments in one cell. Counting a site under "
         "every department it names overstates the estate by 26 percent; counting it under the "
         "first named makes SARD vanish entirely. That is RAC's decision, audit question 1.", ""),
        ("", ""),
        ("Never hand edit this workbook.", "b"),
        ("It is generated by build_content_workbook.py from extract_contents.js. A hand edit is "
         "lost on the next run, and worse, it makes the workbook disagree with the screen.", ""),
    ]
    r = 1
    for text, kind in intro:
        c = ws.cell(row=r, column=1, value=text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if kind == "h1":
            c.font = Font(bold=True, size=14, color=NAVY)
        elif kind == "h2":
            c.font = Font(bold=True, size=11, color=BLUE)
        elif kind == "b":
            c.font = Font(bold=True, size=10)
        else:
            c.font = Font(size=10)
        if kind in ("ready", "scan", "client", "tc"):
            c.font = Font(size=10, name="Consolas")
            c.fill = PatternFill("solid", fgColor={"ready": GREEN, "scan": AMBER,
                                                   "client": RED, "tc": GREY}[kind])
        ws.row_dimensions[r].height = 15 if len(text) < 95 else 15 * (len(text) // 95 + 1)
        r += 1

    # ---- one sheet per dashboard ------------------------------------------
    order, seen = [], set()
    for row in rows:
        if row["dashboard"] not in seen:
            seen.add(row["dashboard"]); order.append(row["dashboard"])

    tally = {}
    for dash in order:
        title = re.sub(r"[\\/*?:\[\]]", "", dash)[:31]
        ws = wb.create_sheet(title)
        ws.append(HEAD)
        style_header(ws)
        n = 0
        for row in [x for x in rows if x["dashboard"] == dash]:
            n += 1
            s = SOURCES.get(row["name"], {})
            # A band heading or a drill title is navigation, not a figure. It
            # has nothing to source, so calling it "To confirm" would bury the
            # measures that genuinely are unconfirmed under structural rows.
            structural = row["kind"] in ("Band", "Drill title", "Panel title")
            verdict = s.get("verdict", NM if structural else TC)
            tally[verdict] = tally.get(verdict, 0) + 1
            ws.append([n, row["where"], row["kind"], row["name"],
                       s.get("counts", ""), verdict, s.get("file", ""),
                       s.get("column", ""), s.get("build", ""), s.get("blocked", "")])
            rr = ws.max_row
            for c in range(1, len(HEAD) + 1):
                cell = ws.cell(row=rr, column=c)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = Font(size=9.5)
                cell.border = border
            ws.cell(row=rr, column=4).font = Font(size=9.5, bold=True)
            ws.cell(row=rr, column=6).fill = PatternFill(
                "solid", fgColor=VERDICT_FILL.get(verdict, GREY))
            body = s.get("build", "")
            ws.row_dimensions[rr].height = max(30, 12 * (body.count("\n") + 1))
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEAD))}{ws.max_row}"

    # ---- summary -----------------------------------------------------------
    ws = wb.create_sheet("Verdict tally")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12
    ws.append(["Verdict", "Rows"]); style_header(ws)
    for c in range(1, 3):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=HEADBG)
    for k in [R, SC, AC, CL, NO, TC, NM]:
        if k in tally:
            ws.append([k, tally[k]])
            ws.cell(row=ws.max_row, column=1).fill = PatternFill(
                "solid", fgColor=VERDICT_FILL[k])
    ws.append(["Total", sum(tally.values())])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)

    wb.save(out)
    print(f"wrote {out}: {len(rows)} rows across {len(order)} dashboards")
    for k, v in sorted(tally.items(), key=lambda x: -x[1]):
        print(f"  {k:<30} {v}")


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "/tmp/contents.json"
    out = sys.argv[2] if len(sys.argv) > 2 else "EDRMS_Dashboard_Contents.xlsx"
    build(json.load(open(src)), out)
