#!/usr/bin/env python3
"""
Build EDRMS_RA_Gap_Checker_2026-08-24.xlsx

One sheet, laid out exactly like sheet 6 of the gap checker: the client's eight
columns, the autofilter on row 4, the section bands, amber for what is in the
prototype and red on pink for what is not.

What differs from the 21 August cut is the content, because the dashboard was
recut on 24 August. Drawn is not sourced, and this sheet says so on every row:
the "In the prototype?" column answers whether the client can see it, and the
two right hand columns say what it still needs before it carries a real figure
rather than a split of the client's own s67 totals.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"

H_FILL = PatternFill("solid", fgColor="10243E")
H_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BAND = PatternFill("solid", fgColor="EEF2F6")
BAND_FT = Font(name=FONT, size=10, bold=True, color="10243E")

IN_FILL = PatternFill("solid", fgColor="FDEFDC")          # in the prototype: amber
IN_FONT = Font(name=FONT, size=10)
OUT_FILL = PatternFill("solid", fgColor="FCE4E4")         # not in it: pink
OUT_FONT = Font(name=FONT, size=10, bold=True, color="C00000")

TITLE_FT = Font(name=FONT, size=14, bold=True, color="10243E")
NOTE_FT = Font(name=FONT, size=9, italic=True, color="6B7A8C")
LINK_FT = Font(name=FONT, size=9, color="0563C1", underline="single")

THIN = Side(style="thin", color="D9E1EA")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP = Alignment(vertical="top", wrap_text=True)

COLS = ["#", "Requirement Items", "Type", "In the prototype?", "Slide",
        "Why it is not there", "What it needs before it can be built",
        "Question to the client"]
WIDTHS = [5, 52, 14, 15, 12, 44, 40, 15]

# The deck link as far as the client's own sheet shows it. The rest of the URL
# is not recoverable from the screenshot and is nowhere in this repo, so it is
# written as text rather than wired as a hyperlink that would go nowhere.
DECK_URL = ("https://avepointcrm.sharepoint.com/:p:/r/sites/Gen_ADB_DRM_Internal/"
            "Shared%20Documents/General/04.%20Requirement%20")

ILL = ("Drawn. The figure is illustrative, a split of the client's own s67 "
       "total, not a measurement. ")

ROWS = [
    ("BAND", "THE DASHBOARD ITSELF"),

    ("Records and Archive Holdings, as a key view", "Dashboard", "Yes", "s13, s14",
     "", "Present, and since 24 Aug it is a built screen rather than a stub. Every "
     "measure below still needs its source before it carries a real figure.", "Q17"),
    ("The two tables drawn with the client's own column headings", "Panel", "Yes",
     "s68, s69", "", "Shape and column names are the client's own. Content is illustrative.", "Q17"),
    ("Physical counterparts identified", "Indicator", "Yes", "s68",
     "", "SOURCED. Declared records in drm-npr flagged as having a paper counterpart. "
     "The one figure on this dashboard with a real parent.", ""),

    ("BAND", "SLIDE 67, THE OVERVIEW SCREEN"),

    ("Total storage and retrieval requests", "Tile", "Yes", "s67",
     "", "The client's own published total. Needs the eServe request feed to become live.", "Q17"),
    ("Total storage activities", "Tile", "Yes", "s67",
     "", "The client's own published total. Needs RAC's holdings register to become live.", "Q17"),
    ("Total staff supported", "Tile", "Yes", "s67",
     "", "The client's own published total. Needs the requestor on the eServe request.", "Q17"),
    ("Total retrieval activities", "Tile", "Yes", "s67",
     "", "The client's own published total. Needs RAC's retrieval record.", "Q17"),
    ("Storage activities donut: boxes and folders stored", "Chart", "Yes", "s67",
     "", "The 92 / 8 split the client printed. Needs RAC's holdings register.", "Q17"),
    ("Retrieval activities donut: archived and records material", "Chart", "Yes", "s67",
     "", "The 91 / 9 split the client printed. Needs RAC's retrieval record.", "Q17"),
    ("Storage and retrieval requests, by month", "Chart", "Yes", "s67",
     "", ILL + "Needs a dated eServe request extract.", "Q17"),
    ("Storage and retrieval activities, by month", "Chart", "Yes", "s67",
     "", ILL + "Needs a dated holdings register.", "Q17"),
    ("Retrieval counts material removed from the holdings, as processed from "
     "eServe, not requests raised", "Definition", "Yes", "s67",
     "", "The client's own note on s67, now stated on the screen so the definition "
     "travels with the figure.", "Q17"),

    ("BAND", "SLIDE 68, STORAGE"),

    ("Location, as Archives Room / Records Center / Offsite Storage", "Table column",
     "Yes", "s68", "", "The client's three locations. No system carries which one a box "
     "is in. Check the Opus Locations export, and confirm it is a picklist not free text.", "Q17"),
    ("No. of requests", "Table column", "Yes", "s68",
     "", ILL + "Needs the eServe request extract. Opus reports physical records requests.", "Q17"),
    ("Total number of requestors (departments)", "Table column", "Yes", "s68",
     "", ILL + "Check whether the Opus requestor carries a department.", "Q17"),
    ("Total number of requestors (RMs)", "Table column", "Yes", "s68",
     "", ILL + "Check the Opus requests export.", "Q17"),
    ("Total number of boxes stored", "Table column", "Yes", "s68",
     "", ILL + "Check the Opus Physical records Explorer export for a box type.", "Q17"),
    ("Total number of folders stored", "Table column", "Yes", "s68",
     "", ILL + "Check the same Opus Explorer export.", "Q17"),
    ("Remarks", "Table column", "Yes", "s68",
     "", "Carries text, not a figure, because it is a note somebody maintains. "
     "ASK whether Remarks is captured anywhere.", "Q17"),
    ("Indicator: new boxes and folders month on month per department", "Indicator",
     "Yes", "s68", "", ILL + "Drawn as two cuts, by month and by department. Needs an "
     "accession date and a department per item.", "Q17"),
    ("No. and list of outstanding / pending requests", "Indicator", "Yes", "s68",
     "", "The count is drawn. THE LIST IS NOT. Needs the eServe request extract carrying "
     "request state; Opus does show pending and overdue loan counts.", "Q17"),
    ("Physical counterparts identified but not transferred", "Indicator", "Yes", "s68",
     "", "Half sourced. The counterparts identified are real, from EDRMS. The transfer "
     "event is recorded nowhere, so the difference cannot yet be computed.", "Q6, Q17"),
    ("Room capacity and % available storage capacity", "Chart", "No", "s68",
     "REMOVED ON 24 AUGUST, deliberately. A 65 / 80 / 45 per cent capacity chart stood "
     "here and was taken out: the slide asks US whether capacity CAN be included, and "
     "drawing a bar answered the client's own question with an invented number.",
     "A capacity figure per location from RAC, and a definition of what full means. "
     "Check the Opus Locations export for a capacity field, then answer the question.", "Q17"),

    ("BAND", "SLIDE 69, RETRIEVAL"),

    ("Location", "Table column", "Yes", "s69", "", "As s68. Opus Locations.", "Q17"),
    ("No. of requests", "Table column", "Yes", "s69",
     "", ILL + "Opus requests export. A retrieval is a loan in Opus terms.", "Q17"),
    ("Total number of requestors (departments / RMs)", "Table column", "Yes", "s69",
     "", ILL + "Opus requests export.", "Q17"),
    ("Total number of boxes retrieved", "Table column", "Yes", "s69",
     "", ILL + "Opus requests export, loan type.", "Q17"),
    ("Total number of folders retrieved", "Table column", "Yes", "s69",
     "", ILL + "Opus requests export.", "Q17"),
    ("Status: Loan, Return to owner, For Disposal", "Table column", "Yes", "s69",
     "", "Carried as the MIX per location, not one value per row, because the client's "
     "column names a vocabulary. Check the Opus status values against their three words.", "Q17"),
    ("Remarks", "Table column", "Yes", "s69", "", "As s68. ASK.", "Q17"),
    ("Indicator: retrieval requests and boxes / folders retrieved month on month "
     "per department", "Indicator", "Yes", "s69",
     "", ILL + "Drawn as two cuts, by month and by department. Needs the Opus requests "
     "export with dates and department.", "Q17"),
    ("No. and list of outstanding / pending requests", "Indicator", "No", "s69",
     "Not drawn on the retrieval screen. The s68 pending panel covers storage requests "
     "and records awaiting transfer, not outstanding retrievals.",
     "The eServe request extract carrying retrieval request state. The figure is already "
     "held in the data as RA_PENDING_RETR and can be drawn in one edit.", "Q17"),
    ("No. of boxes / folders disposed", "Indicator", "Yes", "s69",
     "", ILL + "Drawn as a table by location. Needs RAC's disposal register; Opus shows "
     "pending destruction.", "Q17"),
    ("Can room activities update capacity and freed up storage?", "Question", "No", "s69",
     "An open question the client put to US, not a requirement to build. It is NOT "
     "answered. The freed capacity column that once sat beside the disposal table went "
     "with the capacity chart on 24 August, for the same reason.",
     "The client's decision on the definition, then the capacity source. Answer it once "
     "the Opus Locations export is read.", "Q17"),

    ("BAND", "SLIDE 67, WHAT THE CLIENT SAID ABOUT THIS DASHBOARD"),

    ("Retrieval as processed from eServe", "Source", "No", "s67",
     "eServe has never been connected to this project and nobody has named an owner. "
     "The DEFINITION it implies is now honoured on the screen; the FEED is not.",
     "ASK. Is eServe the retrieval system of record, and who owns it.", "Q17"),
    ("What is available in Opus and how we can apply it", "Source", "No", "s67",
     "The client asked this on s67. It was closed on 17 Aug as 'nothing from Opus', and "
     "nothing on this dashboard is taken from Opus or from the IR Dashboard.",
     "Run the seven exports in OPUS_EXPORT_INSTRUCTIONS.md, then put it back to the client.", "Q17"),

    ("BAND", "TAKEN OFF THE DASHBOARD ON 24 AUGUST, RECORDED SO THE REMOVAL IS NOT SILENT"),

    ("Unverified holdings", "Tile", "No", "none",
     "Removed. It appears nowhere in s67, s68 or s69. The standing rule since 17 August "
     "is that undrawn is enough on all six dashboards.",
     "A client requirement asking for it. None exists.", ""),
    ("Missing items", "Tile", "No", "none",
     "Removed, same reason.", "A client requirement asking for it. None exists.", ""),
    ("Due for verification", "Tile", "No", "none",
     "Removed, same reason.", "A client requirement asking for it. None exists.", ""),
    ("Legacy files", "Tile", "No", "none",
     "Removed, same reason.", "A client requirement asking for it. None exists.", ""),
    ("1,840 boxes, 9,260 folders and 342 requests as headline figures", "Figure", "No",
     "none",
     "Removed. Invented outright, and reconciling with nothing. The dashboard is now "
     "re-based on the eight totals the client themselves printed on s67.",
     "Nothing. These should not come back.", ""),
]


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Records and Archive Holdings"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Records and Archive Holdings"
    ws["A1"].font = TITLE_FT
    ws.merge_cells("A1:H1")

    ws["A2"] = "REFERENCE DECK:"
    ws["A2"].font = NOTE_FT
    ws.merge_cells("A2:H2")

    ws["A3"] = DECK_URL
    ws["A3"].font = LINK_FT
    ws.merge_cells("A3:H3")

    hr = 4
    for i, c in enumerate(COLS, start=1):
        cell = ws.cell(row=hr, column=i, value=c)
        cell.fill, cell.font, cell.border = H_FILL, H_FONT, BORD
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS[i - 1]
    ws.row_dimensions[hr].height = 30

    r = hr + 1
    n = 0
    for row in ROWS:
        if row[0] == "BAND":
            ws.cell(row=r, column=1, value=row[1])
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            for i in range(1, 9):
                ws.cell(row=r, column=i).fill = BAND
                ws.cell(row=r, column=i).border = BORD
            ws.cell(row=r, column=1).font = BAND_FT
            ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
            r += 1
            continue

        item, typ, yn, slide, why, needs, q = row
        n += 1
        missing = (yn == "No")
        for i, v in enumerate([n, item, typ, yn, slide, why, needs, q], start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = OUT_FONT if missing else IN_FONT
            cell.fill = OUT_FILL if missing else IN_FILL
            cell.border = BORD
            cell.alignment = TOP
        r += 1

    ws.auto_filter.ref = f"A{hr}:H{r - 1}"
    ws.freeze_panes = "A5"

    out = "EDRMS_RA_Gap_Checker_2026-08-24.xlsx"
    wb.save(out)
    print("wrote", out, "rows:", n, "last row:", r - 1)


if __name__ == "__main__":
    build()
