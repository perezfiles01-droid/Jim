# -*- coding: utf-8 -*-
"""Per dashboard: every gap in the RAC requirements, and whether the merged
site CSV solves it. Rows and wording are read from the checker workbook so
the two files cannot drift apart."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC='/root/.claude/uploads/dff1b875-2cc7-5f42-9522-30eacbb8d668/658b240f-EDRMS_Utilization_Dashboard_Checker_1.xlsx'
OUT='/home/user/Jim/EDRMS_What_the_Lookup_Solves_by_Dashboard.xlsx'
F="Arial"; NAVY="1F3864"; HDR="1F3864"; ALT="F7F9FC"; BAND="EDF1F7"
GREEN="1F6B3B"; RED="9C2B2B"; AMBER="8A5A00"
GFILL="E3F2E8"; AFILL="FDF3DF"

DEPT="Department per site, for 92 sites"
DIV="Division per site, 27 values, for 92 sites"
LIB="Library names per site, 148 likely EDRMS libraries"
PROD="Production Yes/No, excluding 54 demo and UAT sites"
NONE="Nothing. This file carries no such data."

# verdict, what the CSV gives, what is still needed
V={
 '1 Bank-wide Oversight':{
  '1':("PARTLY",DEPT,"Only 92 sites of the full estate. Good news: every site here carries ONE department, so the semicolon multi department problem does not appear. Confirm that holds at full scale."),
  '2':("NO",NONE,"A user register and an agreed definition of an EDRMS user."),
  '3':("NO",NONE,"A document level scan. This file counts nothing."),
  '6':("NO",NONE,"Due is not disposed. Needs the disposal system."),
  '9':("NO",NONE,"The site to project register. There is no project column in this file."),
  '10':("NO",NONE,"The site to project register."),
  '12':("PARTLY",PROD,"Demo and UAT sites can now be excluded. But the Production flag is NOT the created versus adopted split, which is still unresolved."),
  '13':("NO",NONE,"A document level scan."),
  '24':("NO",NONE,"A user register."),
  '26':("NO",NONE,"A source that looks back further than 180 days."),
  '28':("NO",NONE,"A user register carrying employment type."),
  '29':("NO",NONE,"The same user register."),
  '30':("NO",NONE,"The same user register."),
  '31':("NO",NONE,"A training system."),
  '32':("NO",NONE,"The user register and a go-live date per site."),
  '34':("NO",NONE,"A document level scan carrying file size."),
  '35':("NO",NONE,"A document level scan carrying createdBy."),
  '36':("NO",NONE,"A document level scan carrying FileCreatedDate."),
  '37':("NO",NONE,"A document level scan carrying size and created date."),
  '38':("SOLVED",DIV,"Nothing further for these 92 sites. Extend the file to the full estate."),
  '42':("NO",NONE,"The denominator still needs a document level scan."),
  '43':("SOLVED",DIV,"Nothing further for these 92 sites."),
  '44':("PARTLY",DIV,"Division is solved. The user half still needs the user register."),
  '48':("SOLVED",DIV,"Nothing further for these 92 sites."),
  '49':("NO",NONE,"A register recording physical handover to RAC."),
  '50':("NO",NONE,"Due is not disposed."),
  '54':("NO",NONE,"The disposal system."),'55':("NO",NONE,"The disposal system."),
  '56':("NO",NONE,"The disposal system."),'57':("NO",NONE,"The disposal system."),
  '58':("NO",NONE,"The disposal system."),'59':("NO",NONE,"The disposal system."),
  '60':("NO",NONE,"The disposal system."),
  '61':("NO",NONE,"A document level scan for an accurate denominator."),
  '61b':("NO",NONE,"Active users cannot be scoped to EDRMS. Needs a list of users entitled to EDRMS and activity attributed per site."),
  '61c':("NO",NONE,"The same. No available Microsoft report attributes activity to EDRMS sites."),
 },
 '2 Department Insights':{
  '2':("PARTLY",PROD,"Demo sites can be excluded. Created versus adopted is still not distinguishable."),
  '3':("NO",NONE,"A user register and a definition of EDRMS user."),
  '8':("NO",NONE,"Due is not disposed."),
  '10':("PARTLY",LIB,"Real library names now exist so the convention can be tested against them. Still needs the approved convention document itself."),
  '11':("NO",NONE,"A go-live date per site. There is no date column in this file."),
  '12':("NO",NONE,"Activity attributed to EDRMS sites."),
  '13':("NO",NONE,"Activity attributed to EDRMS sites."),
  '14':("NO",NONE,"Activity attributed to EDRMS sites."),
  '15':("NO",NONE,"A window longer than 180 days."),
  '16':("NO",NONE,"A user register carrying employment type."),
  '17':("NO",NONE,"The same user register."),'18':("NO",NONE,"The same user register."),
  '19':("NO",NONE,"A training system."),
  '20':("PARTLY",PROD,"Demo sites excluded. Created versus adopted still unresolved."),
  '21':("NO",NONE,"A document level scan."),
  '24':("NO",NONE,"Due is not disposed."),
  '25':("PARTLY",DEPT,"A site can no longer appear under several departments, because this file gives one department per site. Coverage is 92 sites."),
  '27':("NO",NONE,"A document level scan."),
  '30':("NO",NONE,"Due is not disposed."),
  '41':("NO",NONE,"A report splitting visitors internal and external."),
  '42':("NO",NONE,"The same."),
  '43':("NO",NONE,"A source recording access requests and outcomes."),
  '44':("NO",NONE,"The same."),
  '47':("NO",NONE,"A document level scan carrying file size."),
  '50':("NO",NONE,"A document level scan carrying file size."),
  '51':("NO",NONE,"A document level scan carrying createdBy."),
  '57':("NO",NONE,"A document level scan carrying file size."),
  '58':("SOLVED",DIV,"Nothing further for these 92 sites."),
  '59':("PARTLY",DIV,"Division solved. The user half needs the user register."),
  '65':("SOLVED",DIV,"Nothing further for these 92 sites."),
  '66':("NO",NONE,"A register recording physical handover to RAC."),
  '73':("NO",NONE,"The disposal system."),'74':("NO",NONE,"The disposal system."),
  '75':("NO",NONE,"The disposal system."),'76':("NO",NONE,"The disposal system."),
  '77':("NO",NONE,"The disposal system."),'78':("NO",NONE,"The disposal system."),
  '79':("SOLVED",LIB,"Filter out the 110 Lists/* entries and 6 system libraries first. Raw 264 is not the library count."),
  '80':("NO",NONE,"A document level scan. This file names libraries but counts nothing inside them."),
  '83':("NO",NONE,"The file plan category list."),
  '84':("NO",NONE,"Nothing will ever fill this. Activity is never reported per library."),
  '85':("PARTLY",DIV+" and "+LIB,"Division and library names both arrive. The document counts per library still need the scan."),
  '86':("NO",NONE,"A programme schedule from the client."),
 },
 '3 Project Insights':{str(i):("NO",NONE,"The site to project register, and for the eight profile fields the ADB project system. This file has no project column of any kind.") for i in range(1,21)},
 '4 Institutional File Plan':{
  **{str(i):("NO",NONE,"The Institutional File Plan term list, with one row per term and its category.") for i in range(1,25)},
  '9':("PARTLY",DEPT,"Departments per site now exist. Still needs the term list to attribute a department to a term."),
  '10':("PARTLY",LIB,"Library names now exist. Still needs the term list to know which library belongs to which term."),
  '15':("PARTLY",DEPT,"Same as row 9."),
  '16':("PARTLY",LIB,"Same as row 10."),
  '22':("PARTLY",LIB,"Real names to test: 99 contain a space, 13 a double space, 1 carries a '(1)' suffix. Still needs the approved convention."),
 },
 '5 Retention and Disposal':{
  '7':("SOLVED",DEPT,"Nothing further for these 92 sites. Note the row is off the mockup by an earlier design decision, not by sourcing."),
  '8':("SOLVED",LIB,"Filter Lists/* and system libraries first."),
  '9':("N/A",'Not a sourcing gap',"Feasible from the records database already. Off the mockup by design."),
  '10':("N/A",'Not a sourcing gap',"Feasible from the records database already. Off the mockup by design."),
  '11':("NO",NONE,"The disposal system."),
  '12':("NO",NONE,"The disposal system."),
  '13':("PARTLY",LIB,"Library names arrive, which is one half of linking a library to a term. Still needs the file plan term list."),
  '16':("SOLVED",LIB,"Filter Lists/* and system libraries first."),
  '19':("PARTLY",LIB,"Same as row 13."),
  '22':("SOLVED",LIB,"Filter Lists/* and system libraries first."),
  '26':("NO",NONE,"The disposal system."),
 },
 '6 Records and Archive Holdings':{str(i):("NO",NONE,"A system that records physical archive activity. This file touches none of it.") for i in range(1,29)},
}

def gapof(why,needs):
    t=(why+' '+needs).lower()
    if 'not feasible' in t: return 'Not feasible as specified'
    if 'need source' in t or 'no source' in t or 'need to identify' in t or 'needs:' in t: return 'Needs a source'
    if 'limitation' in t: return 'Built, with a limitation'
    if 'removed' in t: return 'Removed from the mockup'
    return 'Removed from the mockup'

src=openpyxl.load_workbook(SRC,data_only=True)
wb=openpyxl.Workbook(); wb.remove(wb.active)
thin=Side(style="thin",color="C6CEDC"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
tally={}

def hdr(ws,row,cols,widths):
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    for i,c in enumerate(cols,1):
        cell=ws.cell(row,i,c); cell.font=Font(name=F,size=9,bold=True,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor=HDR)
        cell.alignment=Alignment(wrap_text=True,vertical="center"); cell.border=BORD
    ws.row_dimensions[row].height=30

for ws_src in src.worksheets:
    name=ws_src.title
    ws=wb.create_sheet(name[:31])
    ws["A1"]=name; ws["A1"].font=Font(name=F,size=14,bold=True,color=NAVY)
    ws["A2"]="Every requirement item on this dashboard that is not fully deliverable today, and whether the merged site CSV solves it."
    ws["A2"].font=Font(name=F,size=9,italic=True,color="595959")
    ws.freeze_panes="A5"
    hdr(ws,4,["Checker row","Requirement item","In the mockup?","What was missing",
              "Does the CSV solve it?","What the CSV gives","What is still needed"],
             [11,34,12,24,15,32,46])
    r=5; c=dict(SOLVED=0,PARTLY=0,NO=0); lastnum=None; suffix=0
    for row in ws_src.iter_rows(min_row=5,max_row=ws_src.max_row,max_col=7):
        num,item,typ,mock,slide,why,needs=[('' if x.value is None else str(x.value)) for x in row]
        if not item.strip(): continue
        num=num.strip()
        if num:
            lastnum=num; suffix=0
        else:
            # a continuation row of the requirement above it, e.g. the three
            # Comparison ratios which share one number in the checker
            suffix+=1; num=f"{lastnum}{chr(ord('a')+suffix)}"
        vd=V.get(name,{}).get(num)
        if not vd: continue
        verdict,gives,still=vd
        put=[num,item.strip().replace('\n',' '),mock.strip() or '-',gapof(why,needs),verdict,gives,still]
        for i,v in enumerate(put,1):
            cell=ws.cell(r,i,v); cell.font=Font(name=F,size=9)
            cell.alignment=Alignment(wrap_text=True,vertical="top"); cell.border=BORD
        col={'SOLVED':GREEN,'PARTLY':AMBER,'NO':RED}.get(verdict,"595959")
        ws.cell(r,5).font=Font(name=F,size=9,bold=True,color=col)
        if verdict in ('SOLVED','PARTLY'):
            fill=GFILL if verdict=='SOLVED' else AFILL
            for i in range(1,8): ws.cell(r,i).fill=PatternFill("solid",fgColor=fill)
        if verdict in c: c[verdict]+=1
        r+=1
    tally[name]=(c,r-5)
    # counts by formula so they follow the rows
    ws.cell(r+1,4,"Solved").font=Font(name=F,size=9,bold=True,color=GREEN)
    ws.cell(r+1,5,f'=COUNTIF(E5:E{r-1},"SOLVED")').font=Font(name=F,size=9,bold=True)
    ws.cell(r+2,4,"Partly").font=Font(name=F,size=9,bold=True,color=AMBER)
    ws.cell(r+2,5,f'=COUNTIF(E5:E{r-1},"PARTLY")').font=Font(name=F,size=9,bold=True)
    ws.cell(r+3,4,"Not solved").font=Font(name=F,size=9,bold=True,color=RED)
    ws.cell(r+3,5,f'=COUNTIF(E5:E{r-1},"NO")').font=Font(name=F,size=9,bold=True)

# ------------------------------------------------------------- SUMMARY FIRST
ws=wb.create_sheet("0 Summary",0)
for i,w in enumerate([30,12,12,12,12,62],1): ws.column_dimensions[get_column_letter(i)].width=w
ws["A1"]="What the lookup CSV solves, by dashboard"
ws["A1"].font=Font(name=F,size=14,bold=True,color=NAVY)
ws["A2"]="Gap items are requirement rows that are not fully deliverable today: removed from the mockup, needing a source, not feasible as specified, or built with a stated limitation."
ws["A2"].font=Font(name=F,size=9,italic=True,color="595959")
hdr(ws,4,["Dashboard","Gap items","Solved","Partly","Not solved","What the CSV changes on this dashboard"],
        [30,12,12,12,12,62])
note={
 '1 Bank-wide Oversight':"The three per division rows come back. Department grouping becomes runnable. Demo sites can be excluded. Everything about users, documents, projects and disposal is untouched.",
 '2 Department Insights':"The biggest gain. Division returns, library names arrive so the library table can be built, and a site no longer appears under several departments. Documents per library, users, disposal and go-live are untouched.",
 '3 Project Insights':"Nothing. There is no project column anywhere in this file. All 20 rows stand exactly where they were.",
 '4 Institutional File Plan':"Almost nothing on its own. Five rows get one of their two ingredients, but every row still waits on the file plan term list. The 148 library names are NOT that list.",
 '5 Retention and Disposal':"The retention half improves: departments and libraries provisioned become countable. The disposal half is untouched, because due is still not disposed.",
 '6 Records and Archive Holdings':"Nothing. This dashboard needs a physical archive system and the CSV has no bearing on it.",
}
r=5
for k,(c,n) in tally.items():
    vals=[k,n,c['SOLVED'],c['PARTLY'],c['NO'],note.get(k,'')]
    for i,v in enumerate(vals,1):
        cell=ws.cell(r,i,v); cell.font=Font(name=F,size=9,bold=(i==1))
        cell.alignment=Alignment(wrap_text=True,vertical="top"); cell.border=BORD
    ws.cell(r,3).font=Font(name=F,size=9,bold=True,color=GREEN)
    ws.cell(r,4).font=Font(name=F,size=9,bold=True,color=AMBER)
    ws.cell(r,5).font=Font(name=F,size=9,bold=True,color=RED)
    r+=1
for i,lab in [(1,"TOTAL")]:
    ws.cell(r,1,lab).font=Font(name=F,size=10,bold=True,color=NAVY)
for col in range(2,6):
    L=get_column_letter(col)
    ws.cell(r,col,f'=SUM({L}5:{L}{r-1})').font=Font(name=F,size=10,bold=True,color=NAVY)
    ws.cell(r,col).border=BORD
ws.cell(r,1).border=BORD; ws.cell(r,6).border=BORD
for i in range(1,7): ws.cell(r,i).fill=PatternFill("solid",fgColor=BAND)

r+=2
ws.cell(r,1,"HOW TO READ THIS").font=Font(name=F,size=10,bold=True,color=NAVY); r+=1
for t,col in [
 ("SOLVED means the CSV supplies everything that was missing for that row, for the 92 sites it covers.",GREEN),
 ("PARTLY means it supplies one ingredient of two or more. The row still cannot be built.",AMBER),
 ("NO means the row stands exactly where it was before this file arrived.",RED),
 ("Coverage caveat on every SOLVED and PARTLY: 92 sites, of which only 38 are production. The estate is far larger.",NAVY),
 ("The CSV contains no counts. It maps sites to departments, divisions and libraries. Every number still comes from the records database and the M365 reports, joined through this mapping.",NAVY),
]:
    ws.cell(r,1,t).font=Font(name=F,size=9,color=col,bold=(col!=NAVY))
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); r+=1

wb.save(OUT); print("saved",OUT)
for k,(c,n) in tally.items(): print(f'  {k:32} gaps={n:3} solved={c["SOLVED"]:2} partly={c["PARTLY"]:2} no={c["NO"]:2}')
