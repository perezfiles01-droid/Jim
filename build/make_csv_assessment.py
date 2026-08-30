# -*- coding: utf-8 -*-
"""Assessment of EDRMS_Declared_Records_MergedVLOOKUP_SITE.csv against the checker."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1F3864"; HDR="1F3864"; BAND="EDF1F7"; ALT="F7F9FC"
GREEN="1F6B3B"; RED="9C2B2B"; AMBER="8A5A00"
F="Arial"

wb=openpyxl.Workbook(); wb.remove(wb.active)
thin=Side(style="thin",color="C6CEDC")
BORD=Border(left=thin,right=thin,top=thin,bottom=thin)

def sheet(name,widths,title,sub):
    ws=wb.create_sheet(name)
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws["A1"]=title; ws["A1"].font=Font(name=F,size=14,bold=True,color=NAVY)
    ws["A2"]=sub;   ws["A2"].font=Font(name=F,size=9,italic=True,color="595959")
    ws.freeze_panes="A5"
    return ws

def header(ws,row,cols):
    for i,c in enumerate(cols,1):
        cell=ws.cell(row,i,c)
        cell.font=Font(name=F,size=9,bold=True,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor=HDR)
        cell.alignment=Alignment(wrap_text=True,vertical="center")
        cell.border=BORD
    ws.row_dimensions[row].height=30

def band(ws,row,text,ncols):
    ws.cell(row,1,text).font=Font(name=F,size=9,bold=True,color=NAVY)
    for i in range(1,ncols+1):
        ws.cell(row,i).fill=PatternFill("solid",fgColor=BAND); ws.cell(row,i).border=BORD

def put(ws,row,vals,alt=False,colors=None):
    for i,v in enumerate(vals,1):
        cell=ws.cell(row,i,v)
        cell.font=Font(name=F,size=9,color=(colors or {}).get(i,"000000"),
                       bold=(colors or {}).get("bold")==i)
        cell.alignment=Alignment(wrap_text=True,vertical="top")
        cell.border=BORD
        if alt: cell.fill=PatternFill("solid",fgColor=ALT)

# =============================================================== 1 SUMMARY
ws=sheet("1 Summary",[34,16,86],
  "Assessment: EDRMS_Declared_Records_MergedVLOOKUP_SITE.csv",
  "What this file fixes in the requirement checker, and what it does not. Every figure below was measured from the CSV itself.")

r=4
ws.cell(r,1,"WHAT THIS FILE IS").font=Font(name=F,size=10,bold=True,color=NAVY); r+=1
header(ws,r,["Item","Value","Note"]); r+=1
rows=[
 ("Source tenant","asiandevbank","PRODUCTION. This is the first ADB production data in this project. Everything proven before now came from the 7rkd12 test tenant."),
 ("Rows in file","7,146","Only 943 carry data. The other 6,203 are empty Excel padding rows and should be deleted before anyone opens it."),
 ("Populated rows","943",""),
 ("Distinct sites","92","Identified by the URL column."),
 ("Production sites","38","Production = Yes."),
 ("Non-production sites","54","Production = No. Demo and UAT sites, which must be excluded from every reported figure."),
 ("Departments","16","BPMSD, CSD, CWRD, ITD, NARO, OCRP, OOMP, OPEC, ORM, OSPF, SARD, SD1, SLRM, SPD, URM, VPAC."),
 ("Divisions","27","Genuinely populated, not a copy of Department: 518 of 943 rows carry a Division different from the Department."),
 ("Path segments","264","The Site column is a LIBRARY path, not a site. Splits into 148 likely EDRMS libraries, 110 SharePoint Lists, 6 system libraries."),
 ("Likely EDRMS libraries","148","After removing Lists/* and the six SharePoint system libraries."),
 ("Count columns","0","There is no document count, record count, counterpart count, date or user anywhere in this file."),
]
for i,(a,b,c) in enumerate(rows):
    put(ws,r,[a,b,c],alt=i%2==1); ws.cell(r,1).font=Font(name=F,size=9,bold=True); r+=1

r+=1
ws.cell(r,1,"THE HEADLINE").font=Font(name=F,size=10,bold=True,color=NAVY); r+=1
for t,col in [
 ("This file is a MAPPING file, not a data file. Despite the name, it contains no declared record figures at all.",RED),
 ("What it does is tell us, for 92 sites, which department and division owns them and what libraries sit inside them.",NAVY),
 ("That is genuinely valuable, because two of those three things were hard blockers. But it produces no number on its own.",NAVY),
 ("It closes the DIVISION blocker outright, which is 9 requirement rows that were removed from the mockup.",GREEN),
 ("It supplies the site to department mapping that 30 requirement rows depend on for their grouping.",GREEN),
 ("It gives library names per site, which partly opens 4 more rows.",GREEN),
 ("It does NOT touch the site to project register, the file plan term list, the user register, disposal, or the physical archive.",RED),
 ("Coverage is the main limit: 92 sites, and only 38 of them production. The EDRMS estate is far larger.",AMBER),
]:
    ws.cell(r,1,t).font=Font(name=F,size=9,color=col,bold=(col!=NAVY))
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3); r+=1

# ============================================================ 2 WHAT IT FIXES
ws=sheet("2 What it fixes",[5,26,13,13,34,26,26,13,40],
  "What this CSV fixes in the checker",
  "Keyed to the requirement rows in EDRMS_Utilization_Dashboard_Checker_1.xlsx.")
r=4
header(ws,r,["#","What the CSV supplies","Checker sheet","Checker rows","Requirement affected",
             "Status before","Status with this file","Fixed?","What still limits it"]); r+=1

data=[
 ("BAND","DIVISION. The clearest win: these rows were taken off the mockup for one reason, and that reason is gone."),
 (1,"Division per site, 27 values","1 Bank-wide","38","Per division breakdown","Removed from mockup. Division empty on every EDRMS site.","Buildable for the 92 sites in this file","Yes","Only these 92 sites. Every other site still has no division."),
 (2,"Division per site","1 Bank-wide","43","Number of records declared per division","Removed. Division is undefined.","Buildable","Yes","Coverage only."),
 (3,"Division per site","1 Bank-wide","44","Number of users declaring records per division","Removed. Division is undefined.","Partly buildable","Partly","Division is solved; the user side still needs the user register."),
 (4,"Division per site","1 Bank-wide","48","Per division, two columns","Removed. Division is undefined.","Buildable","Yes","Coverage only."),
 (5,"Division per site","2 Department","58","Number of records declared per division","Removed from the mockup.","Buildable","Yes","Coverage only."),
 (6,"Division per site","2 Department","59","Number of users declaring records per division","Removed from the mockup.","Partly buildable","Partly","Needs the user register for the user half."),
 (7,"Division per site","2 Department","65","Per division, two columns","Removed from the mockup.","Buildable","Yes","Coverage only."),
 (8,"Division per site","2 Department","66","Physical counterpart completion rate","Removed. Division undefined.","Still blocked","No","Division was never the real blocker here. This needs a RAC custody event, which nothing records."),
 (9,"Division per site","2 Department","85","Indicators per library, per division","Removed. Division undefined and nothing below site level.","Partly unblocked","Partly","Division and library names both arrive. Document counts per library still need the Graph scan."),

 ("BAND","DEPARTMENT. Not a new capability, but it closes the gap that 30 rows quietly depend on."),
 (10,"Site to department mapping, 16 departments","1 Bank-wide","1, 11","Sites by Department, RM, office and the Department column","Department column blank in the database and in SharePoint. Approach agreed, list outstanding.","The list now exists for 92 sites","Partly","Only 92 sites. Also note this file gives ONE department per site, so the semicolon multi-department problem does not appear here. Confirm that holds across the full estate."),
 (11,"Site to department mapping","All sheets","30 rows in total","Every figure whose method reads 'grouped by Department' or 'joined for department'","The join was designed but had no mapping table to join to.","The join can now be made","Partly","Coverage. These rows were never marked No in the mockup, so this does not change any Yes or No, it makes the stated method actually runnable."),

 ("BAND","LIBRARY LEVEL. Partial. Library NAMES arrive; library CONTENTS do not."),
 (12,"Library names per site, 148 likely EDRMS libraries","2 Department","79","Library names","Not feasible. No source today.","Buildable for these sites","Yes","Needs the Lists/* and system libraries filtered out first. Raw count of 264 is not the library count."),
 (13,"Library names per site","5 Retention","16, 22","Number of libraries provisioned","Need source. Nothing reports libraries below site level.","Countable for these sites","Yes","Coverage, and the same filtering caveat."),
 (14,"Library names per site","5 Retention","8","Number of libraries provisioned on the rollup","Removed from mockup, no source.","Countable","Yes","Row is still off the mockup by an earlier design decision, not by sourcing."),
 (15,"Library names only","2 Department","80","Number of documents per library","Not feasible. No source.","STILL BLOCKED","No","This file names libraries but counts nothing inside them. Still needs the document level Graph scan."),
 (16,"Nothing relevant","2 Department","84","Number of users per library","Removed. No source, ever.","STILL BLOCKED","No","Unchanged and unchangeable. Microsoft reports activity per user and per site, never per library."),

 ("BAND","NAMING CONVENTION. A capability nobody asked this file for, but it is there."),
 (17,"Actual library names, so the convention can be tested","4 File Plan","22","Indicator: new libraries created outside convention","Need source. Term list not available.","Testable once the convention is supplied","Partly","We can now list real library names and flag outliers: 99 contain spaces, 13 contain double spaces, 1 carries a '(1)' duplicate suffix. Still needs the approved convention document to test against."),

 ("BAND","DATA QUALITY. Not a requirement row, but it improves every figure in the report."),
 (18,"Production Yes/No flag","All sheets","Every figure","Excluding demo and UAT sites","No way to tell a demo site from a real one, so demo sites inflated every total.","54 of 92 sites can now be excluded","Yes","This is a real gain that was never written as a requirement. Confirm the flag is trustworthy."),
]
i=0
for row in data:
    if row[0]=="BAND":
        band(ws,r,row[1],9); r+=1; continue
    put(ws,r,list(row),alt=i%2==1)
    v=str(row[7]).lower()
    ws.cell(r,8).font=Font(name=F,size=9,bold=True,
        color=GREEN if v=="yes" else (RED if v=="no" else AMBER))
    i+=1; r+=1

# ========================================================== 3 STILL MISSING
ws=sheet("3 Still missing",[5,30,40,13,16,44],
  "What this CSV does NOT fix",
  "Everything below is exactly where it was before this file arrived.")
r=4
header(ws,r,["#","Still missing","What it blocks","Checker rows","Owner","What done looks like"]); r+=1
miss=[
 (1,"Site to project register","All 20 Project Insights rows, Bank-wide tiles 9 and 10, and the s36 and s37 project lists","20 + 2","RAC","One row per EDRMS site carrying its project number and the sovereign or nonsovereign identifier, keyed on Site Id. This CSV carries no project column of any kind."),
 (2,"The ADB project system","The eight profile fields on s38: facility type, modality, country, status, effectivity date, closing date, project number, project name","8","ITD to name it","A named system and a contact. Still never named in this work by anyone."),
 (3,"Institutional File Plan term list","All 24 Institutional File Plan rows, and the Term column on both retention screens","24 + 2","RAC","One row per term with its category. NOTE: the 148 library names in this CSV are NOT the term list. They carry no category, and 148 libraries against 23 terms is the wrong granularity. They are the usage side, not the vocabulary."),
 (4,"Document level Graph scan","Accurate document counts, document size, monthly growth, users creating documents, the declaration rate denominator, and documents per library","about 10","ITD","One row per document carrying Site Id, ListId, ItemId, LibraryId, LibraryName, size, created date, created by. This CSV names libraries but counts nothing inside them."),
 (5,"EDRMS user register","Staff, contractor and consultant splits, training completion, onboarded since go-live, the users pie","8","RAC","A file keyed on User Principal Name carrying employment type. Nothing user related is in this CSV."),
 (6,"Definition of EDRMS user","Every user tile and two of the three comparison ratios","several","RAC to define","One written definition. Untouched by this file."),
 (7,"Activity attributed to EDRMS sites","Active users, no access counts, never accessed","several","ITD","A source giving activity per user per site. Untouched."),
 (8,"Disposal decision system","Approver, Approved, Declined, Extended, records disposed, disposed size, completion rate, overdue and pending","13","ITD change request","A scheduled release or a written deferral. Untouched."),
 (9,"Physical custody event to RAC","Physical counterpart completion rate on s42 and s59","2","RAC","A register recording handover. Untouched."),
 (10,"Go-live date per site","Go-Live date on s53, onboarded since go-live on s39","2","RAC","A date per site. This CSV has no date column at all."),
 (11,"Access request and visitor origin source","Visitors internal and external, access requests granted and denied","4","ITD","A source recording requests and outcomes. Untouched."),
 (12,"Physical archive activity source","27 of the 28 Records and Archive Holdings rows","27","RAC","Any register of storage and retrieval requests. Untouched."),
 (13,"Training system","Completion of training, training completion rate","2","Client","The system where training completion is recorded. Untouched."),
 (14,"CG created versus adopted","'Sites created' being a true created count","5","ITD","A distinguishing column. The Production flag is NOT this: it separates demo from real, not newly created from converted."),
 (15,"Coverage of the full estate","Everything this file does fix, beyond the 92 sites it names","all of the above fixes","RAC and ITD","The same six columns for every EDRMS compliant site in production. Until then, every gain from this file applies to 92 sites only, 38 of them production."),
 (16,"Any count at all","Every figure on every dashboard","all","n/a","This file maps sites to owners and libraries. It produces no number. The counts still come from the records database and the M365 reports, joined through this mapping."),
]
for i,row in enumerate(miss):
    put(ws,r,list(row),alt=i%2==1); ws.cell(r,2).font=Font(name=F,size=9,bold=True); r+=1

# ========================================================== 4 DATA QUALITY
ws=sheet("4 Fix before use",[5,30,30,40,40],
  "Things to fix in the file itself before anyone builds on it",
  "None of these is fatal. All of them will cause a wrong number if they are not handled.")
r=4
header(ws,r,["#","Issue","Detail","Why it matters","What to do"]); r+=1
dq=[
 (1,"6,203 empty rows","Rows 945 to 7,147 are all commas. The file is 13 percent data.","Anyone opening it in Excel sees a mostly empty sheet and may think the lookup failed. Any naive row count is wrong by a factor of seven.","Delete the trailing rows before circulating it."),
 (2,"The Site column is not a site","Site is a LIBRARY path. URL is the site. 934 distinct Site values against 92 distinct sites.","Reading Site as the site key gives 934 sites instead of 92, overstating the estate tenfold. The column name invites exactly that mistake.","Rename the columns: Site to LibraryPath, URL to SiteUrl. Or split into two files, one per site and one per library."),
 (3,"System libraries counted as libraries","264 path segments include 110 SharePoint Lists (Lists/*) and 6 system libraries: Style Library, SitePages, SiteAssets, FormServerTemplates, AppCatalog, Shared Documents.","Any 'number of libraries provisioned' figure is inflated by 116 out of 264, so it would read 264 where the real answer is about 148.","Filter Lists/* and the six system names before counting. Confirm with ITD whether more system libraries exist."),
 (4,"Duplicate rows","6 sites appear more than once with identical values. edrms_demo-uat8 appears 4 times, edrms_demo-uat29 3 times.","A join on a duplicated key multiplies the fact rows, which silently inflates every count that passes through it. This is the single most dangerous item on this sheet.","De-duplicate on the site and library pair before using it as a lookup table."),
 (5,"Join key is a URL string","There is no Site Id or GUID. The only key is a URL.","URL joins break on renames, trailing slashes and case differences, and they break silently by returning no match rather than an error.","Add the SharePoint Site Id GUID as a column. Every other export in this project carries it."),
 (6,"Naming inconsistency","99 library names contain a space, 13 contain a double space, 1 is 'Events (1)'.","Some of this is real off-convention creation, which is a finding. Some is likely a data entry slip, such as the double spaces.","Worth reviewing with RAC. It is evidence for the 'libraries created outside convention' indicator, once the convention is supplied."),
 (7,"Mostly non-production","54 of 92 sites are Production = No.","If the flag is ignored, demo and UAT sites are counted as real, and the majority of the sites in this file are not real.","Filter on Production = Yes for every reported figure. Keep the rest only for testing."),
]
for i,row in enumerate(dq):
    put(ws,r,list(row),alt=i%2==1); ws.cell(r,2).font=Font(name=F,size=9,bold=True); r+=1

# =========================================================== 5 QUESTIONS
ws=sheet("5 Questions",[5,54,12,52],
  "Questions to put back, to get the most out of this file",
  "Short list. Each one either widens what the file covers or removes a caveat from it.")
r=4
header(ws,r,["#","Question","Ask","Why it matters"]); r+=1
qs=[
 (1,"Can we have these same six columns for every EDRMS compliant site in production, not just these 92?","RAC","This is the single question that multiplies the value of the file. Every gain here currently applies to 92 sites."),
 (2,"Is the Production flag reliable, and is Production = Yes the right filter for reporting?","RAC","38 of 92 sites hang on this answer. If the flag is wrong, the report counts demo sites as real."),
 (3,"Does the 'P' suffix on a library name mean permanent retention?","RAC","There are 18 pairs where both X and XP exist, for example AnnualMeet and AnnualMeetP. If P means permanent, this file feeds the permanent and temporary split on the Retention and Disposal dashboard directly. If it means something else, we must not assume it."),
 (4,"Can the SharePoint Site Id be added as a column?","ITD","It removes the fragile URL join, and it matches every other export in this project."),
 (5,"Is one department and one division per site the rule across the whole estate?","RAC","In this file no site carries two departments, which is good news. The Cloud Governance export does show semicolon separated multi department sites, so we need to know which is true at full scale."),
 (6,"Is the division list of 27 values complete and current?","RAC","CSD shows 5 divisions and SARD shows 6. If any are missing, the per division tables will under-report."),
 (7,"Where does this file come from, and can it be produced on a schedule?","ITD","A one time extract goes stale. The report needs this refreshed on the same cadence as the other sources."),
 (8,"Are the 148 library names the full list, or only libraries that currently hold declared records?","RAC and ITD","If it is only libraries with declarations, then empty libraries are invisible, and 'libraries provisioned' would under-report."),
]
for i,row in enumerate(qs):
    put(ws,r,list(row),alt=i%2==1); ws.cell(r,2).font=Font(name=F,size=9,bold=True); r+=1

wb.save("/home/user/Jim/EDRMS_Merged_Site_CSV_Assessment.xlsx")
print("saved")
