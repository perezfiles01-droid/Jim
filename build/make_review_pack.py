# -*- coding: utf-8 -*-
"""RAC review pack for the two dashboards agreed in scope on 28 Aug 2026.

Adds the two columns RAC asked for that the checker does not have: where each
figure is sourced from, and who keeps it current after the project ends.
Requirement wording and row numbers are read from the checker at build time.
"""
import openpyxl, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC='/root/.claude/uploads/dff1b875-2cc7-5f42-9522-30eacbb8d668/658b240f-EDRMS_Utilization_Dashboard_Checker_1.xlsx'
OUT='/home/user/Jim/EDRMS_RAC_Review_Pack_BankWide_and_Department.xlsx'
IN_SCOPE=['1 Bank-wide Oversight','2 Department Insights']

F="Arial"; NAVY="1F3864"; HDR="1F3864"; ALT="F7F9FC"; BAND="EDF1F7"
GREEN="1F6B3B"; RED="9C2B2B"; AMBER="8A5A00"; YELLOW="FFF6D5"

# ---- source and maintenance classification -------------------------------
# Decide BLOCKED first. The checker's prose often names a source only in order
# to rule it out ("the CG Groups export cannot help either"), so scanning for
# source names before establishing the verdict credits a row with a source it
# was explicitly denied. That reads as authoritative and is wrong.
MISSING=[
 ('user register',      'NO SOURCE. Needs a register of EDRMS users carrying employment type',
                        'RAC, or HR','Somebody must keep the register current as people join and leave.'),
 ('training system',    'NO SOURCE. No training system is connected to this project','n/a','n/a'),
 ('disposal',           'NO SOURCE. Disposal is a separate future release','n/a until that release','n/a'),
 ('go live',            'NO SOURCE. Needs a go-live date per site','RAC','A one off list, then updated as new sites go live.'),
 ('go-live',            'NO SOURCE. Needs a go-live date per site','RAC','A one off list, then updated as new sites go live.'),
 ('site to project',    'NO SOURCE. Needs a site to project register','RAC','Out of scope for now. Projects were deferred on 28 Aug.'),
 ('access request',     'NO SOURCE. Nothing available records site access requests','n/a','n/a'),
 ('custody',            'NO SOURCE. No system records physical handover to RAC','RAC, if a register is created',
                        'Manual, unless it becomes part of the physical records workflow.'),
 ('turned over',        'NO SOURCE. No system records physical handover to RAC','RAC, if a register is created',
                        'Manual, unless it becomes part of the physical records workflow.'),
 ('file plan term list','NO SOURCE TODAY. Agreed on 28 Aug to move the file plan into the Postgres database',
                        'RAC','RAC maintains the terms through the admin view being built.'),
 ('division',           'Division is now available from Cloud Governance, for the sites we have',
                        'Whoever provisions the site',
                        'Division must be set at provisioning. It is empty on most sites today.'),
 ('graph scan',         'NOT BUILT. Needs a scheduled Microsoft Graph scan over the documents',
                        'ITD, once built','ITD runs the job. No RAC effort once it exists.'),
 ('weekly scan',        'NOT BUILT. Needs a scheduled Microsoft Graph scan over the documents',
                        'ITD, once built','ITD runs the job. No RAC effort once it exists.'),
 ('below site level',   'NOT BUILT. Nothing available reports below site level',
                        'ITD, once a scan exists','None once it exists.'),
 ('programme',          'NO SOURCE. Nothing records EDRMS programme or rollout dates','RAC','RAC would maintain the schedule.'),
 ('per user per site',  'NO SOURCE. Activity cannot be attributed to EDRMS sites specifically',
                        'n/a','n/a. This is how Microsoft reports it and no join produces it.'),
]
def classify(why,needs,mock):
    t=(why+' '+needs).lower()
    blocked=(mock.strip().lower()=='no' or 'not feasible' in t or 'no source' in t
             or 'need source' in t or 'needs:' in t or 'need to identify' in t)
    if blocked:
        for key,srctxt,who,eff in MISSING:
            if key in t:
                return srctxt,who,eff
        return ('NO SOURCE identified','To be confirmed','To be confirmed')
    # not blocked: name the sources it genuinely reads
    src=[]; who=None; eff=None
    if 'public."records"' in t or 'existing in the database' in t or 'drm-npr' in t:
        src.append('EDRMS records database (drm-npr)')
        who='Nobody. EDRMS writes it.'
        eff='None. The figure appears because a record was declared.'
    if 'cloud governance' in t or 'cg workspace' in t or 'cg export' in t or 'cg compliant' in t:
        src.append('AvePoint Cloud Governance, Workspace report')
        who=who or 'Whoever provisions the site'
        eff=('The department and division must be set correctly when a site is provisioned. '
             'If they are left blank the site drops out of every departmental figure, silently.')
    if 'sharepoint site usage' in t or 'file count' in t or 'page view count' in t or 'visited page' in t:
        src.append('Microsoft 365, SharePoint site usage report')
        who=who or 'Nobody. Microsoft generates it.'
        eff=eff or 'None. It is a point in time count per site and cannot be split below site level.'
    if 'activity user detail' in t or 'activity export' in t or 'last activity date' in t:
        src.append('Microsoft 365, SharePoint activity user detail')
        who=who or 'Nobody. Microsoft generates it.'
        eff=eff or 'None, but the window is capped at 180 days and cannot be widened.'
    if not src:
        return ('Derived by arithmetic from other columns on the same row',
                'Nobody. It is arithmetic.','None.')
    return ' + '.join(src), who, eff

DEFN={('1 Bank-wide Oversight','2'),('1 Bank-wide Oversight','24'),
      ('2 Department Insights','3'),('2 Department Insights','12')}

def status(why,needs,mock):
    t=(why+' '+needs).lower()
    if mock.strip().lower()=='no': return 'Not in the mockup'
    if 'not feasible' in t: return 'Built, but not as you specified'
    if 'relabel' in t or 'replaced with' in t: return 'Built, under a different label'
    if 'limitation' in t: return 'Built, with a limitation'
    if 'feasible' in t: return 'Built and sourceable today'
    return 'Built'

# targeted questions, keyed (sheet, row)
Q={
 ('1 Bank-wide Oversight','1'):"Confirmed on 28 Aug that all EDRMS sites are active today. You asked to track inactive and deleted EDRMS sites in future, and to exclude the roughly 3,000 non EDRMS SharePoint sites. Can you confirm we build the tile as active sites now, with an inactive count added later?",
 ('1 Bank-wide Oversight','2'):"We still have no definition of an EDRMS user. Our proposal is someone who has declared a record, because that is the only one countable today. Do you accept that, or do you want a different definition?",
 ('1 Bank-wide Oversight','3'):"The document count is per site and cannot be split by library, format or created date without a new scan being built. Does that requirement still stand?",
 ('1 Bank-wide Oversight','11'):"Is Department, office and RM one field or three? It changes how every table on this dashboard groups.",
 ('1 Bank-wide Oversight','12'):"The export cannot separate a newly created site from an adopted one. Do you need that distinction, or is 'EDRMS compliant sites' acceptable?",
 ('1 Bank-wide Oversight','17'):"Does 'number of sites created' mean all SharePoint sites or only EDRMS compliant ones? On 28 Aug you said only EDRMS sites are tracked, so we have assumed compliant only.",
 ('1 Bank-wide Oversight','20'):"You asked on 28 Aug for two separate measures: a site not accessed in X days, and a site with no record declared in X days. You suggested 90 days may be better than 180 for the declaration one. Please confirm both thresholds and we will colour code at 30, 90 and 180.",
 ('1 Bank-wide Oversight','26'):"The activity report only looks back 180 days, so 'never accessed' cannot be proven, only 'not in 180 days'. Do you accept that wording?",
 ('1 Bank-wide Oversight','28'):"This needs a register of EDRMS users carrying staff, contractor or consultant. Does one exist, and who would keep it current?",
 ('1 Bank-wide Oversight','38'):"Division is now available from Cloud Governance for the sites we have. Do you want the per division columns brought back?",
 ('1 Bank-wide Oversight','41'):"We can list departments with zero declarations. Is that the alert you want, or would you rather have it as a flag on the site table?",
 ('1 Bank-wide Oversight','54'):"Disposal has its own release and no system records an approval today. Do you agree these columns wait for that release?",
 ('2 Department Insights','1'):"On 28 Aug you asked to be able to generate the report per department, for example SARD showing HQ plus its RMs and divisions, rather than only a flat bank-wide file. Does the department picker here answer that, or do you need a separate export?",
 ('2 Department Insights','3'):"Same question as Bank-wide: we need your definition of an EDRMS user.",
 ('2 Department Insights','4'):"We can count site visits, meaning page views and visited pages, but not distinct visitors. Is 'site visits' acceptable?",
 ('2 Department Insights','10'):"We now have real library names, so we can flag libraries that do not follow the convention. Can you send us the approved site, library and folder naming convention to test against?",
 ('2 Department Insights','11'):"Nothing records when a site went live on EDRMS. Can RAC supply a go-live date per site, or do you accept the site creation date as a stand in? Note these can be years apart for an adopted site.",
 ('2 Department Insights','16'):"Same user register question. Without it, the staff, contractor and consultant split cannot be built.",
 ('2 Department Insights','25'):"A site can carry more than one department. On 28 Aug the view was that one department is ultimately the owner. Can RAC nominate a primary department per site so figures do not double count?",
 ('2 Department Insights','31'):"Confirming again that we report site visits rather than visitors.",
 ('2 Department Insights','41'):"No available report splits visitors into internal and external. Is this something you still need, and do you know of a source?",
 ('2 Department Insights','43'):"Nothing records site access requests or their outcomes. Same question: still needed, and is there a source?",
 ('2 Department Insights','58'):"Division is now available. Do you want the per division breakdown restored?",
 ('2 Department Insights','79'):"We can now list library names per site. Please confirm the library table is what you expected.",
 ('2 Department Insights','80'):"Documents per library needs a scan that does not exist yet. Is a document count per library a must have, or is records declared per library enough for now?",
 ('2 Department Insights','84'):"Users per library cannot be produced by any source, now or later. Microsoft reports activity per person and per site, never per library. Can this be dropped?",
 ('2 Department Insights','86'):"Programme dates are not recorded anywhere we can reach. Do you have a programme schedule we could read?",
}

src=openpyxl.load_workbook(SRC,data_only=True)
wb=openpyxl.Workbook(); wb.remove(wb.active)
thin=Side(style="thin",color="C6CEDC"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)

# --------------------------------------------------------------- HOW TO USE
ws=wb.create_sheet("0 How to use this")
for i,w in enumerate([4,104],1): ws.column_dimensions[get_column_letter(i)].width=w
ws["A1"]="EDRMS Utilization Report: review pack for Bank-wide Oversight and Department Insights"
ws["A1"].font=Font(name=F,size=14,bold=True,color=NAVY)
r=3
for t,bold,col in [
 ("Why only two dashboards",True,NAVY),
 ("On 28 August it was agreed to start with Bank-wide Oversight and Department Insights, because they are the interactive ones. Project Insights is deferred while sub-OPS and NSO are still being designed. The other three follow once these two are settled.",False,"000000"),
 ("",False,"000000"),
 ("What is new in this pack",True,NAVY),
 ("Two columns the earlier assessment file did not have, added because they are what you asked for in the workshop:",False,"000000"),
 ("   'Where the figure comes from' names the actual source for every row.",False,"000000"),
 ("   'Who keeps it current after go-live' and 'What it costs RAC to maintain' answer the question of who runs this once the project ends.",False,"000000"),
 ("",False,"000000"),
 ("How to review it",True,NAVY),
 ("Work down each sheet. Where we have a question, it is in the 'Our question for you' column. Not every row has one: most rows just need you to confirm the source and the maintenance are acceptable.",False,"000000"),
 ("Type your response in the two yellow columns. 'Your answer' for the question, 'Your comments' for anything else, including wording you would prefer.",False,"000000"),
 ("",False,"000000"),
 ("The test to apply to every row",True,NAVY),
 ("For each component, ask the three questions you set out in the workshop: where is this sourced from, how is it collated, and what would a records and archives team member have to do to keep it current. If the answer to the third is more than you can sustain, tell us and we will take it off rather than build something that goes stale.",False,"000000"),
 ("",False,"000000"),
 ("Sheet 3 is new",True,NAVY),
 ("It lists the requirements that came out of the 28 August workshop itself and are not in the original deck. They need your confirmation before we build them.",False,"000000"),
]:
    c=ws.cell(r,2,t); c.font=Font(name=F,size=10,bold=bold,color=col)
    c.alignment=Alignment(wrap_text=True,vertical="top"); r+=1

COLS=["#","Requirement item","Type","In the mockup?","Status today",
      "Where the figure comes from","Who keeps it current after go-live",
      "What it costs RAC to maintain","Our question for you","Your answer","Your comments"]
W=[5,32,11,11,20,30,22,34,44,26,26]

for name in IN_SCOPE:
    s=src[name]; ws=wb.create_sheet(name[:31])
    for i,w in enumerate(W,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws["A1"]=name; ws["A1"].font=Font(name=F,size=14,bold=True,color=NAVY)
    ws["A2"]="Please complete the two yellow columns. Rows with no question still need you to confirm the source and the maintenance are acceptable."
    ws["A2"].font=Font(name=F,size=9,italic=True,color="595959")
    for i,c in enumerate(COLS,1):
        cell=ws.cell(4,i,c); cell.font=Font(name=F,size=9,bold=True,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor=HDR)
        cell.alignment=Alignment(wrap_text=True,vertical="center"); cell.border=BORD
    ws.row_dimensions[4].height=34; ws.freeze_panes="A5"
    r=5; last=None; sfx=0
    for row in s.iter_rows(min_row=5,max_row=s.max_row,max_col=7):
        num,item,typ,mock,slide,why,needs=[('' if x.value is None else str(x.value)) for x in row]
        if not item.strip(): continue
        num=num.strip()
        if num: last=num; sfx=0
        else: sfx+=1; num=f"{last}{chr(ord('a')+sfx)}"
        source,who,effort=classify(why,needs,mock)
        if (name,num) in DEFN:
            source=('NO AGREED DEFINITION. There is no register of who is entitled to EDRMS. '
                    'What is countable today is activity, or people who declared a record.')
            who='RAC must define it first'
            effort='Depends entirely on the definition chosen. Declaring-a-record needs nothing; entitlement needs a register somebody keeps current.'
        vals=[num,item.strip().replace('\n',' '),typ.strip(),mock.strip() or '-',
              status(why,needs,mock),source,who,effort,Q.get((name,num),''),'','']
        for i,v in enumerate(vals,1):
            cell=ws.cell(r,i,v); cell.font=Font(name=F,size=9)
            cell.alignment=Alignment(wrap_text=True,vertical="top"); cell.border=BORD
            if i in (10,11): cell.fill=PatternFill("solid",fgColor=YELLOW)
            elif r%2==0: cell.fill=PatternFill("solid",fgColor=ALT)
        st=vals[4]
        ws.cell(r,5).font=Font(name=F,size=9,bold=True,
            color=GREEN if st.startswith('Built and') else (RED if st.startswith('Not in') else AMBER))
        if 'NO SOURCE' in source or 'NOT BUILT' in source:
            ws.cell(r,6).font=Font(name=F,size=9,bold=True,color=RED)
        if vals[8]: ws.cell(r,9).font=Font(name=F,size=9,bold=True,color=NAVY)
        r+=1

# ------------------------------------------------- NEW FROM THE WORKSHOP
ws=wb.create_sheet("3 New from 28 Aug workshop")
W2=[5,34,44,26,26,26]
for i,w in enumerate(W2,1): ws.column_dimensions[get_column_letter(i)].width=w
ws["A1"]="New requirements raised in the 28 August workshop"
ws["A1"].font=Font(name=F,size=14,bold=True,color=NAVY)
ws["A2"]="These are not in the original requirements deck. They came out of the discussion and need your confirmation before we build them."
ws["A2"].font=Font(name=F,size=9,italic=True,color="595959")
for i,c in enumerate(["#","What was asked for","What we understood","Where it would come from","Your confirmation","Your comments"],1):
    cell=ws.cell(4,i,c); cell.font=Font(name=F,size=9,bold=True,color="FFFFFF")
    cell.fill=PatternFill("solid",fgColor=HDR); cell.alignment=Alignment(wrap_text=True,vertical="center"); cell.border=BORD
ws.row_dimensions[4].height=30; ws.freeze_panes="A5"
NEW=[
 (1,"Track inactive EDRMS sites","All EDRMS sites are active today because the programme is new. CSD is about to make five inactive. You want inactive EDRMS sites counted and visible once that happens.","AvePoint Cloud Governance, site status"),
 (2,"Track deleted EDRMS sites","A deleted EDRMS site should still be visible as a record of what existed.","AvePoint Cloud Governance, site status"),
 (3,"Do not track non EDRMS SharePoint sites","The roughly 3,000 ordinary SharePoint sites stay out of this dashboard. Only EDRMS provisioned sites are counted.","Scope rule, no source needed"),
 (4,"A notes field on the site record","So that a site being decommissioned, reused or split can carry an explanation.","New field in the database"),
 (5,"Two separate inactivity measures","One: the site has not been accessed in X days. Two: no record has been declared in X days. You want BOTH, not one standing in for the other.","M365 site usage for the first, the records database for the second"),
 (6,"Colour coded inactivity thresholds","30, 90 and 180 days, turning red beyond 180. You said 180 days is likely too long for no declaration and suggested 90.","Derived from the two measures above"),
 (7,"Generate the report per department","Not only a flat bank-wide file. SARD should be able to see HQ plus its RMs and divisions.","Department and division from Cloud Governance"),
 (8,"A filter to exclude sub-OPS and NSO sites","Those sites are out of scope for this release and should not distort the figures.","A flag on the site record"),
 (9,"Project closing date","The one project attribute that matters for records: it triggers making the site read only and starts the 12 year retention clock.","The project system, once named"),
 (10,"Shorter labels on cards and tables","'Declared records' rather than 'Total number of records declared'. You noted the long wording came from the requirements deck and offered to send preferred wording.","Wording only, no source needed"),
 (11,"Correct the dashboard name","Records and Archives Holdings, with an s on Archives. Our nav currently reads Archive.","Wording only"),
 (12,"Read only sites, later","Not for this release. Currently tracked in a spreadsheet. To be revisited once the EDRMS rollout is complete.","To be decided"),
]
r=5
for i,(n,ask,und,srcs) in enumerate(NEW):
    for j,v in enumerate([n,ask,und,srcs,'',''],1):
        cell=ws.cell(r,j,v); cell.font=Font(name=F,size=9,bold=(j==2))
        cell.alignment=Alignment(wrap_text=True,vertical="top"); cell.border=BORD
        if j in (5,6): cell.fill=PatternFill("solid",fgColor=YELLOW)
        elif i%2==1: cell.fill=PatternFill("solid",fgColor=ALT)
    r+=1

wb.save(OUT); print("saved",OUT)
for w in wb.worksheets:
    n=sum(1 for rr in w.iter_rows(min_row=5) if rr[1].value)
    print(f'  {w.title:32} rows={n}')
