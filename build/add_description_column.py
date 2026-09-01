# -*- coding: utf-8 -*-
"""Insert a Description column into the two RAC decision sheets.

openpyxl's insert_cols moves cell VALUES but leaves merged ranges and data
validation sqrefs pointing at the old columns. The section bands are merged
A:F and the RAC Decision dropdown lives on column E, so both are rebuilt by
hand after the insert or the sheet arrives with a broken dropdown and bands
that stop one column short.
"""
import openpyxl, copy
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

SRC='/root/.claude/uploads/dff1b875-2cc7-5f42-9522-30eacbb8d668/028d7904-EDRMS_Utilization_Dashboard_Checker_2.xlsx'
OUT='/home/user/Jim/EDRMS_Utilization_Dashboard_Checker_with_Description.xlsx'
NEWCOL=4   # insert Description as column D, right after Requirement Item

BW={
 7:"How many EDRMS sites exist across the whole bank, counted once per site, with the department, RM or office that owns each one.",
 8:"How many people count as EDRMS users across the whole bank. What makes someone an EDRMS user is not yet agreed.",
 9:"Every document sitting in every EDRMS compliant site across the bank, whether it has been declared as a record or not.",
 10:"Every document across the bank that somebody has formally declared as a record.",
 11:"Of the records declared bank-wide, how many were flagged as also existing on paper.",
 12:"Records across the bank whose disposal due date falls inside a chosen window. Which window is not yet agreed.",
 13:"Not a figure. A button that takes the reader to the Retention and Disposal dashboard.",
 14:"Not a figure. A button that takes the reader to the Institutional File Plan dashboard.",
 15:"How many EDRMS sites across the bank belong to a sovereign project.",
 16:"How many EDRMS sites across the bank belong to a nonsovereign project.",
 18:"The first column of the site table. One row per owning unit, so units can be compared side by side.",
 19:"For the one unit on that row: how many EDRMS sites it owns.",
 20:"For the one unit on that row: every document held across that unit's EDRMS sites.",
 21:"For the one unit on that row: how many of its documents have been declared as records.",
 22:"For the one unit on that row: how many of its declared records also exist on paper.",
 23:"How many EDRMS sites have been created, counted across the whole bank.",
 24:"How many EDRMS sites have been deleted, counted across the whole bank.",
 25:"How many EDRMS sites have been archived, counted across the whole bank.",
 26:"How many EDRMS sites have had no activity at all in the last 90 days, bank-wide.",
 28:"A table listing every sovereign project, with the EDRMS sites, documents, records and counterparts belonging to it.",
 29:"The same table for nonsovereign projects.",
 30:"Not a figure. Clicking a project row opens Project Insights for that project.",
 32:"How many people count as EDRMS users bank-wide, as the headline above the user breakdown.",
 33:"How many distinct departments, offices and RMs appear in that user breakdown.",
 34:"How many people who have access to an EDRMS site have never once opened one.",
 35:"How many people with access have not opened an EDRMS site in the last 90 days.",
 36:"Of the EDRMS users bank-wide, how many are ADB staff.",
 37:"Of the EDRMS users bank-wide, how many are contractors.",
 38:"Of the EDRMS users bank-wide, how many are consultants.",
 39:"How many EDRMS users have completed EDRMS training.",
 40:"How many users were added to EDRMS after the system went live.",
 42:"How many documents have been put into EDRMS across the whole bank.",
 43:"How much storage space those documents occupy, bank-wide.",
 44:"How many distinct people have created a document in EDRMS.",
 45:"How many documents were added in each month, so growth can be seen over time.",
 46:"How much the stored volume grows in an average month.",
 47:"The same document figures split by division, the level below department.",
 49:"How many records have been declared bank-wide, as the headline above the records breakdown.",
 50:"How many distinct people have declared at least one record.",
 51:"Which departments, offices and RMs have declared no records at all.",
 52:"What share of the documents held have been declared as records.",
 53:"Records declared, split by division rather than by department.",
 54:"People declaring records, split by division rather than by department.",
 56:"How many declared records were flagged as also existing on paper, bank-wide.",
 57:"How many distinct people declared those records that have a paper counterpart.",
 58:"What share of all declared records also exist on paper.",
 59:"The counterpart figures split by division.",
 60:"Of the paper counterparts identified, how many have actually been handed over to RAC for storage.",
 62:"How many records reach their disposal due date, counted across the bank.",
 63:"The earliest disposal due date coming up.",
 64:"Of the records falling due for disposal, which ones also have paper to deal with.",
 65:"How many records fall due for disposal in the next quarter.",
 66:"Who approved each disposal.",
 67:"How many disposal requests were approved.",
 68:"How many disposal requests were declined.",
 69:"How many had their retention extended instead of being disposed.",
 70:"How many records were actually disposed in each month.",
 71:"How many disposals are past their date or still waiting on a decision.",
 72:"Of the records that fell due, what share were actually disposed.",
 74:"A comparison showing, for each unit, how many documents it holds against how many it has declared.",
 75:"A comparison showing how many documents there are for each active user.",
 76:"A comparison showing how many declared records there are for each active user.",
 77:"A running total of records declared over time, bank-wide, so the curve shows adoption.",
 78:"How many records were declared inside the date range the reader picks.",
 79:"The average number of records declared per month inside that same range.",
}

DI={
 7:"Not a figure. The control that chooses which department, office or RM the whole screen is about.",
 8:"How many EDRMS sites the selected unit owns.",
 9:"How many people count as EDRMS users for the selected unit.",
 10:"How many people visited the selected unit's EDRMS sites.",
 11:"Every document held across the selected unit's EDRMS sites, declared or not.",
 12:"How many of the selected unit's documents have been declared as records.",
 13:"How many of the selected unit's declared records also exist on paper.",
 14:"How many of the selected unit's records reach their disposal due date.",
 15:"Not a figure. Each tile opens a table underneath showing the detail behind it.",
 16:"The naming rules the selected unit's sites, libraries and folders are meant to follow.",
 17:"The date the selected unit started using EDRMS.",
 19:"How many of the selected unit's people have used EDRMS recently.",
 20:"How many of the selected unit's people opened EDRMS in the last 180 days.",
 21:"How many had access but did not open EDRMS in the last 180 days.",
 22:"How many of the selected unit's people have access to an EDRMS site but have never once opened it.",
 23:"Of the selected unit's EDRMS users, how many are ADB staff.",
 24:"Of the selected unit's EDRMS users, how many are contractors.",
 25:"Of the selected unit's EDRMS users, how many are consultants.",
 26:"What share of the selected unit's EDRMS users have completed EDRMS training.",
 28:"How many EDRMS sites the selected unit owns, shown above its site table.",
 29:"Every document across the selected unit's sites, shown above its site table.",
 30:"Records declared across the selected unit's sites.",
 31:"Paper counterparts across the selected unit's sites.",
 32:"Records in the selected unit falling due for disposal within the next 12 months.",
 33:"One row per site. The name of each EDRMS site the selected unit owns.",
 34:"For the one site on that row: who owns it.",
 35:"For the one site on that row: how many documents it holds.",
 36:"For the one site on that row: how many records have been declared in it.",
 37:"For the one site on that row: how many of its records also exist on paper.",
 38:"For the one site on that row: how many of its records fall due for disposal.",
 40:"How many people visited the selected unit's EDRMS sites.",
 41:"A table showing visit activity for each of the selected unit's sites.",
 42:"How many pages were viewed in total across the selected unit's sites.",
 43:"How many distinct pages were opened at least once.",
 44:"Not a figure. Buttons that switch the visit figures between the last 7, 30, 90 or 180 days.",
 45:"Not a figure. A control to look at one particular month or year.",
 46:"One row per site in the visits table.",
 47:"For the one site on that row: how many pages were viewed.",
 48:"For the one site on that row: how many distinct pages were opened.",
 49:"For the one site on that row: what kind of EDRMS site it is.",
 50:"Of the people who visited the selected unit's sites, how many came from outside ADB.",
 51:"Of the people who visited the selected unit's sites, how many came from inside ADB.",
 52:"How many requests to join the selected unit's sites were approved.",
 53:"How many requests to join the selected unit's sites were refused.",
 55:"How many documents have been put into the selected unit's sites.",
 56:"How many of the selected unit's EDRMS sites are currently active.",
 57:"How much storage space the selected unit's documents occupy.",
 58:"One row per site in the documents table.",
 59:"For the one site on that row: how many documents were put into it.",
 60:"For the one site on that row: how much space its documents occupy.",
 61:"How many distinct people created documents in the selected unit's sites.",
 63:"Records declared across the selected unit's sites.",
 64:"How many distinct people in the selected unit declared at least one record.",
 65:"Which of the selected unit's sites have had no record declared in the last 180 days.",
 66:"One row per site: how many records were declared in that site.",
 67:"One row per site: how many people declared records in that site.",
 68:"How much storage space the selected unit's declared records occupy.",
 69:"Records declared, split by the divisions inside the selected department.",
 70:"People declaring records, split by the divisions inside the selected department.",
 72:"How many of the selected unit's declared records also exist on paper.",
 73:"The selected unit's total declared records, shown alongside for comparison.",
 74:"One row per site: how many of that site's records also exist on paper.",
 75:"One row per site: how many records were declared in it, shown for comparison.",
 76:"For the one site on that row: what share of its records also exist on paper.",
 77:"The counterpart figures split by the divisions inside the selected department.",
 78:"Of the selected unit's paper counterparts, how many have been handed over to RAC.",
 80:"Records in the selected unit falling due for disposal within the next 3 months.",
 81:"Records in the selected unit falling due for disposal within the next 6 months.",
 82:"Records in the selected unit falling due for disposal within the next 12 months.",
 83:"One row per library: how many of its records fall due for disposal.",
 84:"One row per library: the earliest date a record in it falls due.",
 85:"One row per library: whether those records also exist on paper.",
 86:"Who approved each disposal in the selected unit.",
 87:"How many disposal requests in the selected unit were approved.",
 88:"How many disposal requests in the selected unit were declined.",
 89:"How many had their retention extended instead of being disposed.",
 90:"How many of the selected unit's records were actually disposed, by month and year.",
 91:"How much storage space the disposed records had occupied.",
 93:"One row per library. The name of each library inside the selected unit's sites.",
 94:"For the one library on that row: how many documents it holds.",
 95:"For the one library on that row: how many records have been declared in it.",
 96:"For the one library on that row: how many of its records also exist on paper.",
 97:"Not a figure. A control to switch between the file plan categories.",
 98:"For the one library on that row: how many people use it.",
 99:"The library figures split by the divisions inside the selected department.",
 101:"The key dates in the EDRMS rollout programme for the selected unit.",
}

wb = openpyxl.load_workbook(SRC)
report = {}

for name, DESC in [('Bank-wide (RAC)', BW), ('Department Insights (RAC)', DI)]:
    ws = wb[name]

    # 1. remember the merged bands and the dropdowns, then clear both
    bands = [(r.min_row, r.min_col, r.max_col) for r in list(ws.merged_cells.ranges)]
    for r in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(r))
    saved_dv = []
    for dv in list(ws.data_validations.dataValidation):
        cells = [str(c) for c in dv.sqref.ranges]
        saved_dv.append((dv.type, dv.formula1, dv.allow_blank, cells))
    ws.data_validations.dataValidation = []

    # 2. the insert itself
    ws.insert_cols(NEWCOL)

    # 3. put the bands back one column wider
    for row, c0, c1 in bands:
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1 + 1)

    # 4. put the dropdowns back, shifted one column right
    for typ, f1, blank, cells in saved_dv:
        dv = DataValidation(type=typ, formula1=f1, allowBlank=blank)
        ws.add_data_validation(dv)
        for ref in cells:
            shifted = []
            for part in ref.split(':'):
                col = ''.join(ch for ch in part if ch.isalpha())
                rownum = ''.join(ch for ch in part if ch.isdigit())
                idx = openpyxl.utils.column_index_from_string(col)
                shifted.append(f"{get_column_letter(idx+1 if idx>=NEWCOL else idx)}{rownum}")
            dv.add(':'.join(shifted))

    # 5. header, styled like the one beside it
    src = ws.cell(5, 3)
    h = ws.cell(5, NEWCOL, "Description")
    h.font = copy.copy(src.font); h.fill = copy.copy(src.fill)
    h.border = copy.copy(src.border); h.alignment = copy.copy(src.alignment)
    ws.column_dimensions[get_column_letter(NEWCOL)].width = 62
    # insert_cols does not shift the column widths, so restore the three that
    # moved right by hand or Feasibility Note arrives 22 wide and RAC Comments
    # falls back to the default width.
    for letter, w in (('E', 60), ('F', 22), ('G', 40)):
        ws.column_dimensions[letter].width = w

    # 6. the descriptions
    body = ws.cell(7, 3)
    n = 0
    for row, text in DESC.items():
        c = ws.cell(row, NEWCOL, text)
        c.font = copy.copy(body.font); c.border = copy.copy(body.border)
        c.alignment = copy.copy(body.alignment)
        # the sheet is zebra striped, so take the fill from the cell beside it
        c.fill = copy.copy(ws.cell(row, 3).fill)
        n += 1
    # every requirement row must have one
    missing = [r for r in range(7, ws.max_row+1)
               if ws.cell(r, 3).value and not ws.cell(r, NEWCOL).value]
    report[name] = (n, missing, len(bands), sum(len(c) for _,_,_,c in saved_dv))

wb.save(OUT)
for k,(n,missing,bands,dvs) in report.items():
    print(f'{k}: {n} descriptions | bands restored {bands} | dv ranges restored {dvs}')
    print('   rows left without a description:', missing or 'none')
print('saved', OUT)
