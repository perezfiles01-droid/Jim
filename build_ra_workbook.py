#!/usr/bin/env python3
"""
Build EDRMS_Records_and_Archive_Holdings_2026-08-24.xlsx

One workbook for one dashboard: Records and Archive Holdings, and nothing else.
Every figure is read out of the live prototype by extract_ra.js, which mounts
index.html in a browser and computes the splits with the page's own split(),
weights() and DATA. Nothing here is retyped by hand, because every significant
error in this project came from a plausible figure somebody typed.

    node extract_ra.js "$(pwd)/index.html" "$(pwd)/ra_data.json"
    python3 build_ra_workbook.py
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
NAVY = "10243E"

H_FILL = PatternFill("solid", fgColor=NAVY)
H_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BAND = PatternFill("solid", fgColor="EEF2F6")
BAND_FT = Font(name=FONT, size=10, bold=True, color=NAVY)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
TITLE_FT = Font(name=FONT, size=14, bold=True, color=NAVY)
NOTE_FT = Font(name=FONT, size=9, italic=True, color="6B7A8C")
WARN_FT = Font(name=FONT, size=10, bold=True, color="C00000")

THIN = Side(style="thin", color="D9E1EA")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP = Alignment(vertical="top", wrap_text=True)
NUM = "#,##0"

D = json.load(open("ra_data.json"))
T = D["totals"]
YEAR = T["RA_YEAR"]


def new_sheet(wb, tab, title, subtitle, span=6):
    ws = wb.create_sheet(tab)
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_FT
    ws["A2"] = subtitle
    ws["A2"].font = NOTE_FT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    ws.row_dimensions[2].height = 26
    ws["A2"].alignment = TOP
    return ws


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def header(ws, row, cols):
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.fill, cell.font, cell.border = H_FILL, H_FONT, BORD
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 42
    return row + 1


def band(ws, row, text, span):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    for i in range(1, span + 1):
        ws.cell(row=row, column=i).fill = BAND
        ws.cell(row=row, column=i).border = BORD
    ws.cell(row=row, column=1).font = BAND_FT
    return row + 1


def data_row(ws, row, values, bold=False):
    for i, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = BOLD if bold else BODY
        cell.border = BORD
        cell.alignment = TOP
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            cell.number_format = NUM
            cell.alignment = Alignment(vertical="top", horizontal="right")
    return row + 1


ILLUSTRATIVE = ("Illustrative. A distribution of the client's own published total, "
                "not a measurement.")
PUBLISHED = "Published by the client on slide 67."

wb = Workbook()
wb.remove(wb.active)

# ---------------------------------------------------------------- 1. Read me
ws = new_sheet(wb, "Read me", "Records and Archive Holdings",
               "The one dashboard, and nothing else. Generated from the prototype as it "
               "stands on 24 August 2026.", span=3)
widths(ws, [42, 22, 78])
r = 4
for para in [
    "This workbook covers the sixth dashboard of the EDRMS Utilization Report prototype, "
    "Records and Archive Holdings, and no part of the other five.",
    "Every figure was read out of the running prototype rather than retyped. extract_ra.js "
    "mounts index.html in a browser, opens the dashboard, and computes each split with the "
    "page's own split(), weights() and DATA. build_ra_workbook.py then writes this file. Re-run "
    "both when the dashboard changes.",
    "The reporting period is calendar %d, because that is the year the client's own slide 67 "
    "screenshot covers." % YEAR,
    "WHAT IS THE CLIENT'S AND WHAT IS OURS. The eight totals on the next table are the client's, "
    "printed by them on slide 67. Every location, month and department figure in this workbook is "
    "a distribution of one of those totals, drawn so the shape of the report can be seen. Those "
    "distributions are illustrative and are marked as such on every sheet. Nothing is taken from "
    "Opus or from the IR Dashboard, which the client ruled out on 17 August.",
    "Sheet 7 reproduces the nine reconciliation checks the dashboard asserts at runtime, each with "
    "a live formula, so this workbook proves its own arithmetic rather than asking to be trusted.",
    "Sheet 8 carries the two questions slides 68 and 69 put back to this project. They are open and "
    "they are not answered here. A capacity chart that answered them with an invented percentage "
    "was removed from the dashboard on 24 August.",
]:
    ws.cell(row=r, column=1, value=para).font = BODY
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(row=r, column=1).alignment = TOP
    ws.row_dimensions[r].height = 15 * (1 + len(para) // 105)
    r += 2

r = band(ws, r, "The eight totals published by the client on slide 67", 3)
r = header(ws, r, ["Measure", "Value", "Note"])
for label, key, note in [
    ("Total storage and retrieval requests", "RA_REQUESTS", PUBLISHED),
    ("Total storage activities", "RA_STORE_ACTS", PUBLISHED),
    ("Boxes stored", "RA_BOXES_STORED", "The 92 per cent slice of the client's storage donut."),
    ("Folders stored", "RA_FOLDERS_STORED", "The 8 per cent slice of the client's storage donut."),
    ("Total staff supported", "RA_STAFF", PUBLISHED),
    ("Total retrieval activities", "RA_RETR_ACTS", PUBLISHED),
    ("Archived material retrieved", "RA_ARCHIVED_RETRIEVED", "The 91 per cent slice of the client's retrieval donut."),
    ("Records material retrieved", "RA_RECORDS_RETRIEVED", "The 9 per cent slice of the client's retrieval donut."),
]:
    r = data_row(ws, r, [label, T[key], note])

r += 1
r = band(ws, r, "Figures set by this project, within the client's totals", 3)
r = header(ws, r, ["Measure", "Value", "Note"])
for label, key, note in [
    ("Storage requests", "RA_STORE_REQUESTS", "Splits the client's 259 requests into storage and retrieval."),
    ("Retrieval requests", "RA_RETR_REQUESTS", "Splits the client's 259 requests into storage and retrieval."),
    ("Requestors, departments", "RA_REQUESTORS_DEPTS", "Splits the client's 283 staff supported."),
    ("Requestors, RMs", "RA_REQUESTORS_RMS", "Splits the client's 283 staff supported."),
    ("Boxes retrieved", "RA_BOXES_RETRIEVED", "Splits the client's 541 retrieval activities."),
    ("Folders retrieved", "RA_FOLDERS_RETRIEVED", "Splits the client's 541 retrieval activities."),
    ("Boxes disposed", "RA_BOXES_DISPOSED", "Slide 69 asks for this. No source yet."),
    ("Folders disposed", "RA_FOLDERS_DISPOSED", "Slide 69 asks for this. No source yet."),
    ("Storage requests outstanding", "RA_PENDING_STORE", "Slide 68 asks for this. No source yet."),
    ("Retrieval requests outstanding", "RA_PENDING_RETR", "Held in the data, not drawn on the screen."),
    ("Records awaiting transfer to RAC", "RA_AWAITING_TRANSFER", "Of the counterparts identified below. The transfer event has no source."),
    ("Physical counterparts identified", "WITH_PHYSICAL",
     "SOURCED. Declared records recorded in EDRMS as having a paper counterpart. "
     "The one figure on this dashboard with a real parent."),
]:
    r = data_row(ws, r, [label, T[key], note])

# ------------------------------------------------------------- 2. Overview
ws = new_sheet(wb, "Overview (s67)", "Overview, slide 67",
               "The shape the client screenshotted, filled with the totals they themselves "
               "printed. Retrieval counts material removed from the holdings, as processed from "
               "eServe, rather than requests raised. That is the client's own third note on s67.",
               span=4)
widths(ws, [46, 16, 16, 62])
r = 4
r = band(ws, r, "The four tiles", 4)
r = header(ws, r, ["Tile", "Value", "", "Caption on the dashboard"])
for label, key, sub in [
    ("Total storage and retrieval requests", "RA_REQUESTS", "Requests raised across both activities"),
    ("Total storage activities", "RA_STORE_ACTS", "Boxes and folders taken into the holdings"),
    ("Total staff supported", "RA_STAFF", "People served across departments and RMs"),
    ("Total retrieval activities", "RA_RETR_ACTS", "Material removed from the holdings"),
]:
    r = data_row(ws, r, [label, T[key], "", sub])

r += 1
r = band(ws, r, "Storage activities donut", 4)
r = header(ws, r, ["Slice", "Value", "Share", "Note"])
for label, key in [("Boxes stored", "RA_BOXES_STORED"), ("Folders stored", "RA_FOLDERS_STORED")]:
    r = data_row(ws, r, [label, T[key], round(T[key] / T["RA_STORE_ACTS"], 4), PUBLISHED])
    ws.cell(row=r - 1, column=3).number_format = "0.0%"
r = data_row(ws, r, ["Total activities", T["RA_STORE_ACTS"], 1.0, ""], bold=True)
ws.cell(row=r - 1, column=3).number_format = "0.0%"

r += 1
r = band(ws, r, "Retrieval activities donut", 4)
r = header(ws, r, ["Slice", "Value", "Share", "Note"])
for label, key in [("Archived material retrieved", "RA_ARCHIVED_RETRIEVED"),
                   ("Records material retrieved", "RA_RECORDS_RETRIEVED")]:
    r = data_row(ws, r, [label, T[key], round(T[key] / T["RA_RETR_ACTS"], 4), PUBLISHED])
    ws.cell(row=r - 1, column=3).number_format = "0.0%"
r = data_row(ws, r, ["Total activities", T["RA_RETR_ACTS"], 1.0, ""], bold=True)
ws.cell(row=r - 1, column=3).number_format = "0.0%"

M = D["M"]
MONTHS = D["MONTHS"]
r += 1
r = band(ws, r, "Storage and retrieval requests, by month. " + ILLUSTRATIVE, 4)
r = header(ws, r, ["Month", "Storage requests", "Retrieval requests", "Total requests raised"])
for i, m in enumerate(MONTHS):
    r = data_row(ws, r, [m, M["storeReq"][i], M["retrReq"][i], M["storeReq"][i] + M["retrReq"][i]])
r = data_row(ws, r, ["Total", T["RA_STORE_REQUESTS"], T["RA_RETR_REQUESTS"], T["RA_REQUESTS"]], bold=True)

r += 1
r = band(ws, r, "Storage and retrieval activities, by month. " + ILLUSTRATIVE, 4)
r = header(ws, r, ["Month", "Boxes stored", "Folders stored", "Storage activities"])
for i, m in enumerate(MONTHS):
    r = data_row(ws, r, [m, M["boxesStored"][i], M["foldStored"][i],
                         M["boxesStored"][i] + M["foldStored"][i]])
r = data_row(ws, r, ["Total", T["RA_BOXES_STORED"], T["RA_FOLDERS_STORED"], T["RA_STORE_ACTS"]], bold=True)

# -------------------------------------------------------------- 3. Storage
ws = new_sheet(wb, "Storage (s68)", "Storage, slide 68",
               "The client's own table and the client's own column names, including Remarks. "
               "Location columns total the published figures exactly. " + ILLUSTRATIVE, span=7)
widths(ws, [22, 14, 20, 16, 18, 18, 28])
r = 4
# The column names are READ OFF THE RENDERED TABLE, never typed here. A
# workbook that keeps its own copy of a name disagrees with the screen within
# the week: this block used to say "(departments and RMs)" where the slide and
# the screen both say "(departments / RMs)", and dropped the "(year, month)"
# qualifier the client puts on ten of the fourteen columns.
cols = D["storeCols"]
r = header(ws, r, cols)
first = r
for d in D["storage"]:
    r = data_row(ws, r, [d["location"], d["requests"], d["requestors_depts"], d["requestors_rms"],
                         d["boxes"], d["folders"], d["remarks"]])
last = r - 1
r = data_row(ws, r, ["Total"] + [f"=SUM({get_column_letter(c)}{first}:{get_column_letter(c)}{last})"
                                 for c in range(2, 7)] + [""], bold=True)
for c in range(2, 7):
    ws.cell(row=r - 1, column=c).number_format = NUM

r += 1
ws.cell(row=r, column=1, value="Remarks is the client's own column. It is a note somebody "
        "maintains, not a measurement, so it carries text rather than a figure.").font = NOTE_FT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

# ------------------------------------------------------------ 4. Retrieval
ws = new_sheet(wb, "Retrieval (s69)", "Retrieval, slide 69",
               "The client's Status column names a vocabulary, Loan / Return to owner / For "
               "Disposal, not one value per row. The mix is therefore carried per location and "
               "printed as the mix, rather than picking one and implying Archives Room is always "
               "a loan. " + ILLUSTRATIVE, span=10)
widths(ws, [22, 14, 22, 18, 18, 10, 16, 12, 28, 12])
r = 4
# As on s68, read off the page. The client's own Status column is one column
# naming a vocabulary, so the workbook keeps it as printed AND breaks the mix
# into three countable columns beside it, which a spreadsheet can total and a
# sentence cannot.
_rc = D["retrCols"]
_status_i = next(i for i, c in enumerate(_rc) if c.startswith("Status"))
cols = (_rc[:_status_i]
        + ["Status: Loan", "Status: Return to owner", "Status: For Disposal"]
        + _rc[_status_i + 1:]
        + [_rc[_status_i] + ", as printed"])
r = header(ws, r, cols)
first = r
for d in D["retrieval"]:
    r = data_row(ws, r, [d["location"], d["requests"], d["requestors"], d["boxes_retrieved"],
                         d["folders_retrieved"], d["mix"][0], d["mix"][1], d["mix"][2],
                         d["remarks"], d["status"]])
last = r - 1
r = data_row(ws, r, ["Total"] + [f"=SUM({get_column_letter(c)}{first}:{get_column_letter(c)}{last})"
                                 for c in range(2, 9)] + ["", ""], bold=True)
for c in range(2, 9):
    ws.cell(row=r - 1, column=c).number_format = NUM

r += 1
ws.cell(row=r, column=1, value="The three status columns on a location total that location's "
        "retrieval requests. The dashboard asserts this at runtime; sheet 7 checks it here.").font = NOTE_FT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

# ---------------------------------------------- 5. Month and department cuts
ws = new_sheet(wb, "Month and dept cuts", "Month and department cuts",
               "Slides 68 and 69 both ask for the month on month cut per department. The month "
               "vector and the department vector are independent, which is what stops every "
               "department showing the same shape across the year. " + ILLUSTRATIVE, span=5)
widths(ws, [30, 20, 20, 8, 20])
r = 4
r = band(ws, r, "Boxes stored and boxes retrieved, month on month", 5)
r = header(ws, r, ["Month", "New boxes stored", "Boxes retrieved", "", "Folders retrieved"])
for i, m in enumerate(MONTHS):
    r = data_row(ws, r, [m, M["boxesStored"][i], M["boxesRetr"][i], "", M["foldRetr"][i]])
r = data_row(ws, r, ["Total", T["RA_BOXES_STORED"], T["RA_BOXES_RETRIEVED"], "",
                     T["RA_FOLDERS_RETRIEVED"]], bold=True)

r += 1
r = band(ws, r, "Boxes stored and boxes retrieved, by department, office or RM", 5)
r = header(ws, r, ["Department, office or RM", "New boxes stored", "Boxes retrieved", "", "Code"])
rows = sorted(zip(D["depts"], D["storageByDept"], D["retrievalByDept"]),
              key=lambda x: -x[1])
for dept, s, rt in rows:
    name = dept["name"] or dept["code"]
    r = data_row(ws, r, [f'{name} ({dept["code"]})', s, rt, "", dept["code"]])
r = data_row(ws, r, ["Total", T["RA_BOXES_STORED"], T["RA_BOXES_RETRIEVED"], "", ""], bold=True)

r += 1
ws.cell(row=r, column=1, value="The department cut is sized by each unit's declared holding "
        "rather than drawn flat. A flat draw saturates the small units, which is the fault that "
        "gave one office 270 people declaring out of 270 users on 17 August.").font = NOTE_FT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

# ------------------------------------------------- 6. Pending and disposal
ws = new_sheet(wb, "Pending and disposal", "Pending requests, transfers and disposal",
               "Slide 68's second callout and slide 69's third. " + ILLUSTRATIVE, span=4)
widths(ws, [40, 16, 70, 20])
r = 4
r = band(ws, r, "Pending requests and transfers (slide 68)", 4)
r = header(ws, r, ["Measure", "Value", "What it means", "Sourced?"])
r = data_row(ws, r, ["Storage requests outstanding", T["RA_PENDING_STORE"],
                     "Requests raised and not yet fulfilled", "No"])
r = data_row(ws, r, ["Records awaiting transfer to RAC", T["RA_AWAITING_TRANSFER"],
                     "Counterparts identified in EDRMS, not yet recorded as transferred, of %s identified"
                     % f'{T["WITH_PHYSICAL"]:,}', "Partly"])
r = data_row(ws, r, ["Physical counterparts identified", T["WITH_PHYSICAL"],
                     "Declared records recorded as having a paper counterpart",
                     "Yes"])
ws.cell(row=r - 1, column=4).font = Font(name=FONT, size=10, bold=True, color="1F7A4D")
r = data_row(ws, r, ["Retrieval requests outstanding", T["RA_PENDING_RETR"],
                     "Held in the data. Not drawn on the dashboard.", "No"])

r += 1
r = band(ws, r, "Boxes and folders disposed, by location (slide 69)", 4)
r = header(ws, r, ["Location", "No. of boxes disposed", "No. of folders disposed", ""])
first = r
for d in D["retrieval"]:
    r = data_row(ws, r, [d["location"], d["boxes_disposed"], d["folders_disposed"], ""])
last = r - 1
r = data_row(ws, r, ["Total", f"=SUM(B{first}:B{last})", f"=SUM(C{first}:C{last})", ""], bold=True)
ws.cell(row=r - 1, column=2).number_format = NUM
ws.cell(row=r - 1, column=3).number_format = NUM

r += 1
ws.cell(row=r, column=1, value="A freed capacity column used to sit beside this table. It is gone "
        "with the capacity chart, for the reason on the Open items sheet.").font = NOTE_FT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

# -------------------------------------------------------- 7. Reconciliation
ws = new_sheet(wb, "Reconciliation", "Reconciliation",
               "The nine checks the dashboard asserts at runtime, reproduced here with live "
               "formulas. Every Verdict cell must read Pass. A split that quietly loses a unit "
               "is the bug that showed a department 103 records due against a real 75 for four "
               "days in August.", span=5)
widths(ws, [58, 16, 16, 12, 46])
r = 4
r = header(ws, r, ["Check", "Left side", "Right side", "Verdict", "Why it matters"])

S = D["storage"]
R = D["retrieval"]
checks = [
    ("Boxes stored by location total the client's published figure",
     sum(x["boxes"] for x in S), T["RA_BOXES_STORED"],
     "The location split may not lose or gain a box."),
    ("Folders stored by location total the client's published figure",
     sum(x["folders"] for x in S), T["RA_FOLDERS_STORED"],
     "The location split may not lose or gain a folder."),
    ("Storage requests by location total the published request split",
     sum(x["requests"] for x in S), T["RA_STORE_REQUESTS"], "As above, for requests."),
    ("Retrieval requests by location total the published request split",
     sum(x["requests"] for x in R), T["RA_RETR_REQUESTS"], "As above, for requests."),
    ("Requestors by location total the published staff supported",
     sum(x["requestors_depts"] + x["requestors_rms"] for x in S), T["RA_STAFF"],
     "Departments plus RMs is the client's 283."),
    ("Boxes retrieved by location total the published retrieval figure",
     sum(x["boxes_retrieved"] for x in R), T["RA_BOXES_RETRIEVED"], "As above, for retrieval."),
    ("Folders retrieved by location total the published retrieval figure",
     sum(x["folders_retrieved"] for x in R), T["RA_FOLDERS_RETRIEVED"], "As above, for retrieval."),
    ("Nothing is retrieved that was never stored (boxes)",
     sum(x["boxes_retrieved"] for x in R), sum(x["boxes"] for x in S),
     "Left must be the smaller. A holding cannot give out what it never took in."),
    ("Nothing is disposed that was never retrieved (boxes)",
     sum(x["boxes_disposed"] for x in R), sum(x["boxes_retrieved"] for x in R),
     "Left must be the smaller."),
    ("The status mix totals retrieval requests, all locations",
     sum(sum(x["mix"]) for x in R), sum(x["requests"] for x in R),
     "Loan plus Return to owner plus For Disposal is the request count."),
    ("Each monthly series sums to its annual total: boxes stored",
     sum(M["boxesStored"]), T["RA_BOXES_STORED"], "A month chart may not drift from its tile."),
    ("Each monthly series sums to its annual total: boxes retrieved",
     sum(M["boxesRetr"]), T["RA_BOXES_RETRIEVED"], "As above."),
    ("Each monthly series sums to its annual total: storage requests",
     sum(M["storeReq"]), T["RA_STORE_REQUESTS"], "As above."),
    ("Each monthly series sums to its annual total: retrieval requests",
     sum(M["retrReq"]), T["RA_RETR_REQUESTS"], "As above."),
    ("The department cut totals the same boxes as the location cut: stored",
     sum(D["storageByDept"]), T["RA_BOXES_STORED"],
     "Two cuts of one total must agree."),
    ("The department cut totals the same boxes as the location cut: retrieved",
     sum(D["retrievalByDept"]), T["RA_BOXES_RETRIEVED"], "Two cuts of one total must agree."),
    ("Storage plus retrieval requests equal the published total",
     T["RA_STORE_REQUESTS"] + T["RA_RETR_REQUESTS"], T["RA_REQUESTS"],
     "The client's 259."),
    ("Boxes plus folders stored equal the published storage activities",
     T["RA_BOXES_STORED"] + T["RA_FOLDERS_STORED"], T["RA_STORE_ACTS"], "The client's 7,305."),
    ("Boxes plus folders retrieved equal the published retrieval activities",
     T["RA_BOXES_RETRIEVED"] + T["RA_FOLDERS_RETRIEVED"], T["RA_RETR_ACTS"], "The client's 541."),
    ("Records awaiting transfer do not exceed the counterparts identified",
     T["RA_AWAITING_TRANSFER"], T["WITH_PHYSICAL"], "Left must be the smaller."),
]
LTE = {7, 8, 19}  # checks where left must be less than or equal, not equal
for n, (name, left, right, why) in enumerate(checks):
    op = "<=" if n in LTE else "="
    r = data_row(ws, r, [name, left, right,
                         f'=IF(B{r}{op}C{r},"Pass","FAIL")', why])
    ws.cell(row=r - 1, column=4).font = BOLD
    ws.cell(row=r - 1, column=4).alignment = Alignment(vertical="top", horizontal="center")

r += 1
ws.cell(row=r, column=1, value="Rows 8, 9 and 20 are less than or equal checks. All others are "
        "equalities.").font = NOTE_FT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

# ------------------------------------------------------------ 8. Open items
ws = new_sheet(wb, "Open items", "Open items and what is not sourced",
               "Two questions the client put to this project and this project has not answered, "
               "and every measure on the dashboard that has no source yet.", span=4)
widths(ws, [46, 16, 62, 46])
r = 4
r = band(ws, r, "Questions the client asked us, still open", 4)
r = header(ws, r, ["Question, in the client's words", "Slide", "Status", "What answering it needs"])
for q, s, st, need in [
    ("Can room capacity and % available storage capacity be included?", "68",
     "OPEN. Not answered. A 65 / 80 / 45 per cent capacity chart was drawn here and was removed "
     "on 24 August, because it answered the client's own question with an invention.",
     "A capacity figure per location from RAC, and a definition of what full means."),
    ("Does retrieval free up storage capacity, and should the capacity figure update?", "69",
     "OPEN. Not answered. This is a definition the client owns, not one to assume.",
     "The client's decision, then the capacity source above."),
]:
    r = data_row(ws, r, [q, s, st, need])
    ws.cell(row=r - 1, column=3).font = WARN_FT

r += 1
r = band(ws, r, "Measures on the dashboard with no source yet", 4)
r = header(ws, r, ["Measure", "Slide", "Why it is not sourced", "What it needs"])
for m, s, why, need in [
    ("Storage requests, retrieval requests", "68, 69",
     "Requests are raised in eServe. There is no feed from eServe into the reporting database.",
     "An eServe extract, or a manual return from RAC."),
    ("Requestors, departments and RMs", "68, 69",
     "Same as above. The requestor is on the eServe request.", "An eServe extract."),
    ("Boxes and folders stored, by location", "68",
     "The physical holdings are not in any system this project can read.",
     "RAC's own holdings register, per location."),
    ("Boxes and folders retrieved, by location", "69", "As above.", "RAC's holdings register."),
    ("Status: Loan, Return to owner, For Disposal", "69",
     "The vocabulary is the client's. Nothing records which applies to a retrieval.",
     "The status field on RAC's retrieval record."),
    ("Boxes and folders disposed", "69", "No disposal event is recorded anywhere readable.",
     "RAC's disposal register."),
    ("Storage requests outstanding", "68", "Requires the request lifecycle from eServe.",
     "An eServe extract carrying request state."),
    ("Records awaiting transfer to RAC", "68",
     "Half sourced. The counterparts identified are real, from EDRMS. The transfer event is not "
     "recorded, so the difference cannot be computed.",
     "A transfer event, either in EDRMS or in RAC's register."),
    ("Location of a holding: Archives Room, Records Center, Offsite Storage", "68, 69",
     "The three locations are the client's. No system carries which one a box is in.",
     "RAC's holdings register."),
    ("Month on month, per department", "68, 69",
     "Depends on every source above carrying a date and a department.",
     "A dated, department stamped holdings register."),
]:
    r = data_row(ws, r, [m, s, why, need])

r += 1
r = band(ws, r, "What IS sourced", 4)
r = header(ws, r, ["Measure", "Value", "Source", "Note"])
r = data_row(ws, r, ["Physical counterparts identified", T["WITH_PHYSICAL"],
                     'public."Records" in drm-npr, declared records flagged as having a paper counterpart',
                     "The one figure on this dashboard with a real parent, and the bridge between "
                     "this report and the physical estate."])
ws.cell(row=r - 1, column=2).font = Font(name=FONT, size=10, bold=True, color="1F7A4D")

OUT = "EDRMS_Records_and_Archive_Holdings_2026-08-24.xlsx"
wb.save(OUT)
print("wrote", OUT, "sheets:", wb.sheetnames)
