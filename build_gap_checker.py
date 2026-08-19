#!/usr/bin/env python3
"""
Build EDRMS_Dashboard_Gap_Checker_2026-08-19.xlsx

One sheet per dashboard. Every row is one thing the client's deck asks for.
Rows already in the prototype are left in plain font. Rows the client asked for
that are NOT in the prototype are highlighted, carry their slide number, and say
what they are blocked on so the question can go back to the client.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"

# ---- styling -------------------------------------------------------------
H_FILL   = PatternFill("solid", fgColor="10243E")   # ADB deep navy
H_FONT   = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BAND     = PatternFill("solid", fgColor="EEF2F6")   # section band
BAND_FT  = Font(name=FONT, size=10, bold=True, color="10243E")

IN_FONT  = Font(name=FONT, size=10)                 # in the prototype: plain
OUT_FONT = Font(name=FONT, size=10, bold=True, color="C00000")
OUT_FILL = PatternFill("solid", fgColor="FCE4E4")

TITLE_FT = Font(name=FONT, size=14, bold=True, color="10243E")
NOTE_FT  = Font(name=FONT, size=9, italic=True, color="6B7A8C")

THIN = Side(style="thin", color="D9E1EA")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TOP = Alignment(vertical="top", wrap_text=True)

COLS = ["#", "Item, in the client's words where possible", "Type",
        "In the prototype?", "Slide", "Why it is not there",
        "What it needs before it can be built", "Question to the client"]
WIDTHS = [5, 52, 14, 15, 12, 44, 40, 15]


def sheet(wb, tab, title, subtitle, rows):
    """rows: list of ('BAND', text) or (item, type, yes_no, slide, why, needs, q)"""
    ws = wb.create_sheet(tab)
    ws.sheet_view.showGridLines = False

    ws["A1"] = title
    ws["A1"].font = TITLE_FT
    ws["A2"] = subtitle
    ws["A2"].font = NOTE_FT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    hr = 4
    for i, c in enumerate(COLS, start=1):
        cell = ws.cell(row=hr, column=i, value=c)
        cell.fill, cell.font, cell.border = H_FILL, H_FONT, BORD
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS[i - 1]
    ws.row_dimensions[hr].height = 30

    r = hr + 1
    n = 0
    for row in rows:
        if row[0] == "BAND":
            ws.cell(row=r, column=1, value=row[1])
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            for i in range(1, 9):
                ws.cell(row=r, column=i).fill = BAND
                ws.cell(row=r, column=i).border = BORD
            ws.cell(row=r, column=1).font = BAND_FT
            ws.cell(row=r, column=1).alignment = Alignment(vertical="center")
            r += 1
            continue

        item, typ, yn, slide, why, needs, q = row
        n += 1
        vals = [n, item, typ, yn, slide, why, needs, q]
        missing = (yn == "No")
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = OUT_FONT if missing else IN_FONT
            cell.alignment = TOP
            cell.border = BORD
            if missing:
                cell.fill = OUT_FILL
        r += 1

    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    ws.auto_filter.ref = f"A{hr}:H{r-1}"
    return ws, hr, r - 1


# =========================================================================
wb = Workbook()
wb.remove(wb.active)

# ---------------------------------------------------------------- README
rd = wb.create_sheet("README")
rd.sheet_view.showGridLines = False
rd["A1"] = "EDRMS Utilization Report: dashboard gap checker"
rd["A1"].font = Font(name=FONT, size=16, bold=True, color="10243E")
rd["A2"] = "Generated 19 August 2026 from index.html as it stands, checked against EDRMS_Dashboard_requirements_1.pptx"
rd["A2"].font = NOTE_FT

readme = [
    ("", ""),
    ("HOW TO READ THIS WORKBOOK", ""),
    ("One sheet per dashboard.", "Every row is one thing the client's deck asks for on that dashboard."),
    ("Plain black rows", "Already in the prototype. Nothing to do, nothing to ask."),
    ("Red highlighted rows", "The client asked for it and it is NOT in the prototype. Each carries the slide it "
                             "is drawn on, why it is not there, and what it needs. These are the rows to take back "
                             "to the client."),
    ("", ""),
    ("THE HEADLINE FINDING", ""),
    ("The client requires SIX dashboards.", "Slide 13 and slide 14 both name six key views. The prototype has FIVE."),
    ("Project Insights is missing entirely.", "It has its own sheet in this workbook, sheet '3 Project Insights', "
                                              "listing everything on it as not built. Slides 36, 37 and 38."),
    ("", ""),
    ("WHY THINGS ARE MISSING, IN FOUR KINDS", ""),
    ("ASK", "Only the client can supply it. A register, a list, a definition."),
    ("APP", "Needs a new field in the EDRMS application. A change request."),
    ("SCAN", "Needs the weekly SharePoint scan, which is designed and not built."),
    ("NONE", "No source exists anywhere today, in any system."),
    ("", ""),
    ("WHAT THIS WORKBOOK IS NOT", ""),
    ("It is not a figure checker.", "It does not verify any number. It checks presence against the deck only."),
    ("It does not carry Opus.", "AvePoint Opus was opened on 19 August and may close several of these gaps. "
                                "See OPUS_EXPORT_INSTRUCTIONS.md. Nothing here assumes it will."),
    ("Numbers in the prototype are test tenant.", "Presence is what this workbook tracks, not values."),
    ("", ""),
    ("ONE THING FOUND WHILE BUILDING THIS", ""),
    ("Two tiles are on screen but not sourced.", "'Sites deleted' prints a hardcoded 12 and 'Sites archived' a "
                                                 "hardcoded 31, while the named constants SITES_DELETED_90 and "
                                                 "SITES_ARCHIVED_90 are both null and unused. The code's own comment "
                                                 "says these should print as a dash because a zero would be a claim. "
                                                 "They are counted as present on sheet 1, flagged in the 'why' column."),
]
r = 4
for a, b in readme:
    rd.cell(row=r, column=1, value=a).font = Font(name=FONT, size=10, bold=bool(a and not b))
    rd.cell(row=r, column=2, value=b).font = Font(name=FONT, size=10)
    rd.cell(row=r, column=1).alignment = TOP
    rd.cell(row=r, column=2).alignment = TOP
    r += 1
rd.column_dimensions["A"].width = 42
rd.column_dimensions["B"].width = 95

# legend swatches
r += 1
rd.cell(row=r, column=1, value="LEGEND").font = Font(name=FONT, size=10, bold=True)
r += 1
c = rd.cell(row=r, column=1, value="In the prototype")
c.font = IN_FONT
rd.cell(row=r, column=2, value="Plain font, no fill. Nothing to action.").font = IN_FONT
r += 1
c = rd.cell(row=r, column=1, value="Not in the prototype")
c.font, c.fill = OUT_FONT, OUT_FILL
c2 = rd.cell(row=r, column=2, value="Red on pink. Take these to the client.")
c2.font, c2.fill = OUT_FONT, OUT_FILL

# =========================================================================
# 1. BANK-WIDE OVERSIGHT
bw = [
    ("BAND", "TOP PANEL, THE TEN TILES THE CLIENT DREW ON SLIDE 34"),
    ("Number of active EDRMS SharePoint sites for Department, RM, office", "Tile", "Yes", "s34, s35", "", "", ""),
    ("Total number of EDRMS users", "Tile", "Yes", "s34, s39",
     "Built, but relabelled 'EDRMS users with recorded activity, last 180 days'. "
     "The population who hold access is not knowable from any source we reach.", "", "Q3"),
    ("Total number of documents in EDRMS", "Tile", "Yes", "s34, s40", "", "", ""),
    ("Total number of records declared in EDRMS", "Tile", "Yes", "s34, s41", "", "", ""),
    ("Total number of physical counterparts identified", "Tile", "Yes", "s34, s42", "", "", ""),
    ("Total number of records due for disposal", "Tile", "Yes", "s34, s43", "", "", ""),
    ("Retention and disposal insights (navigation tile)", "Tile", "Yes", "s34, s44", "", "", ""),
    ("Institutional File Plan insights (navigation tile)", "Tile", "Yes", "s34, s47", "", "", ""),
    ("Total number of active EDRMS SharePoint sites for Sovereign projects",
     "Tile", "No", "s34",
     "Removed with Project Insights. Nothing says which site belongs to which project.",
     "ASK. A site to project register, roughly one row per project site.", "Q2"),
    ("Total number of active EDRMS SharePoint sites for Nonsovereign projects",
     "Tile", "No", "s34",
     "Removed with Project Insights. Same blocker.",
     "ASK. The same register.", "Q2"),

    ("BAND", "SLIDE 35, OVERVIEW OF EDRMS SITES. Also drawn at s16, the same table"),
    ("Department / office / RM", "Table column", "Yes", "s16, s35", "", "", ""),
    ("Number of EDRMS SharePoint sites", "Table column", "Yes", "s16, s35", "", "", ""),
    ("Total number of documents", "Table column", "Yes", "s16, s35", "", "", ""),
    ("Total number of records declared", "Table column", "Yes", "s16, s35", "", "", ""),
    ("Total number of physical counterparts", "Table column", "Yes", "s16, s35", "", "", ""),
    ("Each name clickable into that unit's dashboard", "Behaviour", "Yes", "s35", "", "", ""),
    ("Indicator: no. of sites created", "Indicator", "Yes", "s35",
     "Built as a tile on the sites drill. Sourceable from CG Created Time. "
     "Note this is site CREATION, not EDRMS go-live, which is a different date.", "", "Q4"),
    ("Indicator: no. of sites deleted", "Indicator", "Yes", "s35",
     "ON SCREEN, BUT THE FIGURE IS A HARDCODED 12. The named constant "
     "SITES_DELETED_90 is null and unused. Displayed, not sourced.",
     "Wire the tile to the constant, or source it from CG Site Status.", ""),
    ("Indicator: no. of sites archived", "Indicator", "Yes", "s35",
     "ON SCREEN, BUT THE FIGURE IS A HARDCODED 31. SITES_ARCHIVED_90 is null "
     "and unused, and the code's own comment says archived has no definition "
     "at all. Displayed, not sourced, and arguably a claim we cannot make.",
     "ASK. What archived means. Then wire the tile or drop the figure.", ""),
    ("Indicator: inactive site for 90 days", "Indicator", "Yes", "s35",
     "Built as a tile. The 90 against 300 day threshold is still unsettled.",
     "", ""),

    ("BAND", "SLIDE 36 AND 37, SOVEREIGN AND NONSOVEREIGN PROJECT SITE LISTS"),
    ("Sovereign project site list, all columns", "Table", "No", "s36",
     "Removed with Project Insights. Every row is blocked on the same missing register.",
     "ASK. A site to project register.", "Q2"),
    ("Nonsovereign project site list, all columns", "Table", "No", "s37",
     "Removed with Project Insights. Same blocker.", "ASK. The same register.", "Q2"),
    ("Project rows clickable into Project Insights", "Behaviour", "No", "s35, s36, s37",
     "The destination dashboard does not exist.", "Build Project Insights first.", "Q2"),

    ("BAND", "SLIDE 39, USERS"),
    ("Total number of users, count", "Drill", "Yes", "s39", "", "", ""),
    ("Indicator: % of active users", "Indicator", "Yes", "s39", "", "", ""),
    ("Indicator: never accessed EDRMS", "Indicator", "Yes", "s39", "", "", ""),
    ("Indicator: not accessed in 90 days", "Indicator", "Yes", "s39", "", "", ""),
    ("Total number of users (staff)", "Table column", "No", "s39",
     "Removed 17 Aug. No register of EDRMS users exists that carries employment type.",
     "ASK. A user register with staff / contractor / consultant.", "Q3"),
    ("Total number of users (Contractors)", "Table column", "No", "s39",
     "Removed 17 Aug. Same register.", "ASK. The same register.", "Q3"),
    ("Total number of users (Consultants)", "Table column", "No", "s39",
     "Removed 17 Aug. Same register.", "ASK. The same register.", "Q3"),
    ("Completion of training", "Table column", "No", "s39",
     "Removed 17 Aug. No training system is connected to this project.",
     "ASK. Where training completion is recorded.", "Q3"),
    ("Onboarded since go-live", "Table column", "No", "s39",
     "Removed 17 Aug. Needs both the user register and a go-live date per site.",
     "ASK. The register, and the go-live date.", "Q3, Q4"),
    ("User to department mapping", "Join", "No", "s39",
     "Gap 1 was closed for SITES on 14 Aug using Cloud Governance Department. "
     "That does nothing for PEOPLE. A user does not live in a site.",
     "ASK. A user to department source.", "Q3"),

    ("BAND", "SLIDE 40, DOCUMENTS"),
    ("Number of documents created / uploaded", "Table column", "Yes", "s40",
     "Present, but the figure waits on the weekly scan for a real value.", "", ""),
    ("Documents size (in GB)", "Table column", "Yes", "s40",
     "Present as Storage. Waits on the scan for a real value.", "", ""),
    ("Number of users creating documents", "Table column", "No", "s40",
     "Never built. CreatedBy on the document table is the DECLARER, not the creator. "
     "A distinct creator count needs a second column.",
     "SCAN, plus a new column on the document table.", ""),
    ("Indicator: new documents month on month", "Indicator", "No", "s40",
     "Never built. Needs snapshot history and a decision on how far back.",
     "Decision needed on the month on month window.", "Q8"),
    ("Indicator: average monthly storage growth", "Indicator", "No", "s40",
     "Never built. Same decision.", "Decision needed on the window.", "Q8"),
    ("Per division breakdown", "Table column", "No", "s40",
     "Removed 17 Aug. Nobody has defined what a division is or what populates it.",
     "ASK. A definition of division and a source for it.", "Q5"),

    ("BAND", "SLIDE 41, RECORDS DECLARED"),
    ("Total number of records declared", "Table column", "Yes", "s41", "", "", ""),
    ("Total number of users declaring records", "Table column", "Yes", "s41", "", "", ""),
    ("Indicator: zero record declarations", "Indicator", "Yes", "s41", "", "", ""),
    ("Indicator: declaration rate against documents", "Indicator", "Yes", "s41",
     "Present as a bank-wide tile. The per row version waits on the scan.", "", ""),
    ("Number of records declared per division", "Table column", "No", "s41",
     "Removed 17 Aug. Division is undefined.", "ASK. Division.", "Q5"),
    ("Number of users declaring records per division", "Table column", "No", "s41",
     "Removed 17 Aug. Division is undefined.", "ASK. Division.", "Q5"),

    ("BAND", "SLIDE 42, PHYSICAL COUNTERPARTS"),
    ("Total number of physical records declared", "Table column", "Yes", "s42", "", "", ""),
    ("Total number of users declaring records", "Table column", "Yes", "s42", "", "", ""),
    ("Share of records with a counterpart", "Table column", "Yes", "s42",
     "Built as a derived column in place of the removed completion rate.", "", ""),
    ("Per division, two columns", "Table column", "No", "s42",
     "Removed 17 Aug. Division is undefined. Note these columns look copied "
     "verbatim from s41 and may be a slip in the client's own slide.",
     "ASK. Division, and confirm the slide was intended.", "Q5, Q10"),
    ("Indicator: physical counterpart completion rate (turned over to RAC)",
     "Indicator", "No", "s42",
     "Removed 17 Aug. 'Turned over for RAC storage' is a physical custody event. "
     "No system we can reach records it. AvePoint Opus may.",
     "NONE today. Check Opus physical records for a custody or transfer event.", "Q6"),

    ("BAND", "SLIDE 43, RECORDS DUE FOR DISPOSAL"),
    ("Number of records due for disposal", "Table column", "Yes", "s43", "", "", ""),
    ("Next due date for disposal", "Table column", "Yes", "s43", "", "", ""),
    ("With physical counterpart?", "Table column", "Yes", "s43", "", "", ""),
    ("Indicator: upcoming next quarter", "Indicator", "Yes", "s43",
     "Present as the due within 30 and 90 day tiles.", "", ""),
    ("Disposal approver", "Table column", "No", "s43",
     "Removed 17 Aug. The EDRMS application has no approver field. "
     "Opus records an approver on its manual approval workflow.",
     "APP, or Opus. Check the Opus approval history export.", "Q7"),
    ("Status (Approved)", "Table column", "No", "s43",
     "Removed 17 Aug. No status field. Opus does record Approved.",
     "APP, or Opus. Opus confirmed to carry Approved.", "Q7"),
    ("Status (Declined)", "Table column", "No", "s43",
     "Removed 17 Aug. No status field. Opus records 'Rejected', which is the "
     "same outcome under a different word.",
     "APP, or Opus. A word mapping decision for RAC.", "Q7"),
    ("Status (Extended)", "Table column", "No", "s43",
     "Removed 17 Aug. No status field, and Opus does NOT carry Extended either. "
     "Confirmed on the Opus Explorer screen, 19 Aug.",
     "APP. Or confirm Extended is not a process at ADB and drop the column.", "Q7"),
    ("Indicator: records disposed month on month", "Indicator", "No", "s43",
     "Removed 17 Aug. Needs a disposal date, which no field holds.",
     "APP, or Opus. Opus reports destroyed records.", "Q7"),
    ("Indicator: overdue actions, pending approvals", "Indicator", "No", "s43",
     "Removed 17 Aug. Needs the status field. Opus shows pending and overdue counts.",
     "APP, or Opus.", "Q7"),
    ("Indicator: disposal completion rate", "Indicator", "No", "s43",
     "Removed 17 Aug. Needs the status field.", "APP, or Opus.", "Q7"),

    ("BAND", "THE REST OF BANK-WIDE"),
    ("Comparison, three named ratios", "Panel", "Yes", "s17, s6", "", "", ""),
    ("Bank-wide records declaration trend, cumulative", "Panel", "Yes", "s18, s10", "", "", ""),
    ("Records declared this month", "Tile", "Yes", "Metrics doc",
     "Kept on the trend panel when its original panel was removed.", "", ""),
    ("Trend per department / office / RM / Projects", "Filter", "No", "s10",
     "The 16 Aug instruction removed the department filter. s10 asks for it. "
     "The two conflict and the client has not been asked which wins.",
     "Decision needed. Also the default range, s10 suggests current year.", "Q14"),
]

# =========================================================================
# 2. DEPARTMENT INSIGHTS
dp = [
    ("BAND", "SLIDE 53, THE DEPARTMENT PROFILE AND TOP PANEL"),
    ("Department picker driving the whole screen", "Behaviour", "Yes", "s53", "", "", ""),
    ("Total number of sites created", "Tile", "Yes", "s53", "", "", ""),
    ("Total number of EDRMS users", "Tile", "Yes", "s53", "", "", ""),
    ("Total number of site visitors", "Tile", "Yes", "s53", "", "", ""),
    ("Total number of documents in EDRMS", "Tile", "Yes", "s53", "", "", ""),
    ("Total number of records declared in EDRMS", "Tile", "Yes", "s53", "", "", ""),
    ("Total number of physical counterparts identified", "Tile", "Yes", "s53", "", "", ""),
    ("Total number of records due for disposal", "Tile", "Yes", "s53", "", "", ""),
    ("Clickable top panel stats", "Behaviour", "Yes", "s53", "", "", ""),
    ("Approved SharePoint site, library and folder convention", "Panel", "Yes", "s53, s26, s28", "", "", ""),
    ("Go-Live date", "Field", "No", "s53",
     "Removed 17 Aug. Cloud Governance carries site CREATION date, which is a "
     "different thing. Nobody holds the date a site became EDRMS compliant.",
     "ASK. A go-live date per site, or agreement that creation date stands in.", "Q4"),

    ("BAND", "SLIDE 54, USERS AND ACTIVITY"),
    ("Total number of active users", "Tile", "Yes", "s54", "", "", ""),
    ("Total number of staff", "Table column", "No", "s54",
     "Removed 17 Aug. No user register.", "ASK. The user register.", "Q3"),
    ("Total number of contractors (#, %)", "Table column", "No", "s54",
     "Removed 17 Aug. No user register.", "ASK. The user register.", "Q3"),
    ("Total number of consultants (#, %)", "Table column", "No", "s54",
     "Removed 17 Aug. No user register.", "ASK. The user register.", "Q3"),
    ("Training completion rate (%)", "Table column", "No", "s54",
     "Removed 17 Aug. No training system connected.",
     "ASK. Where training completion lives.", "Q3"),

    ("BAND", "SLIDE 55, LIST OF SITES"),
    ("Name of sites", "Table column", "Yes", "s55", "", "", ""),
    ("Site owners", "Table column", "Yes", "s55", "", "", ""),
    ("Number of documents", "Table column", "Yes", "s55", "", "", ""),
    ("Number of records", "Table column", "Yes", "s55", "", "", ""),
    ("Number of physical counterparts", "Table column", "Yes", "s55", "", "", ""),
    ("Number of records due for disposal", "Table column", "Yes", "s55", "", "", ""),

    ("BAND", "SLIDE 56, SITE VISITORS"),
    ("Total number of visitors", "Table column", "Yes", "s56", "", "", ""),
    ("Site visits table, by site", "Table", "Yes", "s56", "", "", ""),
    ("Total number of visitors (external) (#, %)", "Table column", "No", "s56",
     "Removed 17 Aug. M365 usage reporting does not split a visitor "
     "internal from external.",
     "ASK. Confirm what external means at ADB, then find a source.", "Q15"),
    ("Total number of visitors (internal) (#, %)", "Table column", "No", "s56",
     "Removed 17 Aug. Same.", "ASK. Same.", "Q15"),
    ("Total number of access requests granted (#, %)", "Table column", "No", "s56",
     "Removed 17 Aug. Access requests are a SharePoint governance event "
     "no source in this project captures.",
     "ASK, and a source check we owe.", "Q15"),
    ("Total number of access requests denied (#, %)", "Table column", "No", "s56",
     "Removed 17 Aug. Same.", "ASK. Same.", "Q15"),

    ("BAND", "SLIDE 57, DOCUMENTS"),
    ("Number of documents created / uploaded", "Table column", "Yes", "s57", "", "", ""),
    ("Documents size (in GB)", "Table column", "Yes", "s57", "", "", ""),
    ("Number of users creating documents", "Table column", "No", "s57",
     "Never built. CreatedBy is the declarer, not the creator.",
     "SCAN, plus a new column.", ""),
    ("Total number of documents migrated", "Tile", "No", "s57",
     "Removed 17 Aug. s11 and s57 disagree on whether migration reporting "
     "is wanted at all.",
     "ASK. Is migration reporting in scope?", "Q9"),
    ("Total number of users migrating documents", "Tile", "No", "s57",
     "Removed 17 Aug. Same.", "ASK. Same.", "Q9"),
    ("Total migrated documents size (in GB)", "Tile", "No", "s57",
     "Removed 17 Aug. Same.", "ASK. Same.", "Q9"),

    ("BAND", "SLIDE 58, RECORDS DECLARATION"),
    ("Total number of records declared", "Table column", "Yes", "s58", "", "", ""),
    ("Total number of users declaring records", "Table column", "Yes", "s58", "", "", ""),
    ("Total declared record size (in GB)", "Tile", "Yes", "s58", "", "", ""),
    ("Number of records declared per division", "Table column", "No", "s58",
     "Removed 17 Aug. Division is undefined.", "ASK. Division.", "Q5"),
    ("Number of users declaring records per division", "Table column", "No", "s58",
     "Removed 17 Aug. Division is undefined.", "ASK. Division.", "Q5"),

    ("BAND", "SLIDE 59, PHYSICAL COUNTERPART"),
    ("Records with a counterpart, by site", "Table", "Yes", "s59", "", "", ""),
    ("Share of records with a counterpart", "Table column", "Yes", "s59",
     "Built as a derived column in place of the removed completion rate.", "", ""),
    ("Per division, two columns", "Table column", "No", "s59",
     "Removed 17 Aug. Division is undefined.", "ASK. Division.", "Q5"),
    ("Physical counterpart completion rate (turned over to RAC)", "Indicator", "No", "s59",
     "Removed 17 Aug. A custody event no reachable system records. Opus may.",
     "NONE today. Check Opus physical records.", "Q6"),

    ("BAND", "SLIDE 60, DISPOSAL"),
    ("Library name", "Table column", "Yes", "s60", "", "", ""),
    ("Number of records due for disposal", "Table column", "Yes", "s60", "", "", ""),
    ("Next due date for disposal", "Table column", "Yes", "s60", "", "", ""),
    ("With physical counterpart?", "Table column", "Yes", "s60", "", "", ""),
    ("Disposal approver", "Table column", "No", "s60",
     "Removed 17 Aug. No field in the application. Opus has an approver.",
     "APP, or Opus.", "Q7"),
    ("Status (Approved)", "Table column", "No", "s60",
     "Removed 17 Aug. Opus carries Approved.", "APP, or Opus.", "Q7"),
    ("Status (Declined)", "Table column", "No", "s60",
     "Removed 17 Aug. Opus carries Rejected, the same outcome renamed.",
     "APP, or Opus. A mapping decision.", "Q7"),
    ("Status (Extended)", "Table column", "No", "s60",
     "Removed 17 Aug. Opus does not carry Extended either.",
     "APP. Or confirm it is not a process and drop it.", "Q7"),
    ("Total number of records disposed, per year and month", "Tile", "No", "s60",
     "Removed 17 Aug. Needs a disposal date. Opus reports destroyed records.",
     "APP, or Opus.", "Q7"),
    ("Total disposed record size (in GB)", "Tile", "No", "s60",
     "Removed 17 Aug. Same blocker plus file size, which is Gap 2.",
     "APP, plus the scan for size.", "Q7"),

    ("BAND", "SLIDES 61 TO 66, DEPARTMENTAL LIBRARY USAGE, ONE PER FILE PLAN CATEGORY"),
    ("Library names", "Table column", "Yes", "s61 to s66", "", "", ""),
    ("Number of documents", "Table column", "Yes", "s61 to s66", "", "", ""),
    ("Number of records declared", "Table column", "Yes", "s61 to s66", "", "", ""),
    ("Number of physical counterparts", "Table column", "Yes", "s61 to s66", "", "", ""),
    ("Category picker across the six categories", "Behaviour", "Yes", "s61 to s66", "", "", ""),
    ("Number of users per library", "Table column", "No", "s61 to s66",
     "Removed 17 Aug. Microsoft 365 reports activity per SITE and never per "
     "LIBRARY. This is a grain problem, not a missing source.",
     "NONE. No source will ever fill this. Recommend dropping the column.", ""),
    ("Indicators per library, per division", "Indicator", "No", "s61 to s66",
     "Removed 17 Aug. Division is undefined, and per library user counts "
     "are impossible.", "ASK. Division. The per library half is impossible.", "Q5"),

    ("BAND", "OTHER PANELS ON THIS DASHBOARD"),
    ("Cumulative records declaration trend", "Panel", "Yes", "s53", "", "", ""),
    ("Programme dates", "Panel", "Yes", "s26, s28", "", "", ""),
]

# =========================================================================
# 3. PROJECT INSIGHTS, THE MISSING DASHBOARD
pj = [
    ("BAND", "THIS ENTIRE DASHBOARD IS MISSING FROM THE PROTOTYPE"),
    ("Project Insights, as one of the six required key views", "Dashboard", "No", "s13, s14",
     "Not in the prototype. The navigation carries five dashboards, not six. "
     "Slides 13 and 14 both name Project Insights among the six key views.",
     "Blocked on two separate sources, below. Neither has an owner yet.", "Q2, Q16"),

    ("BAND", "SLIDE 38, THE EIGHT PROFILE FIELDS"),
    ("Facility type", "Profile field", "No", "s38",
     "Comes from an ADB project system that has never been named in this project.",
     "ASK. Which system holds project attributes, and who owns it.", "Q16"),
    ("Modality", "Profile field", "No", "s38", "Same system, never named.", "ASK. Same.", "Q16"),
    ("Country", "Profile field", "No", "s38", "Same system, never named.", "ASK. Same.", "Q16"),
    ("Status", "Profile field", "No", "s38", "Same system, never named.", "ASK. Same.", "Q16"),
    ("Effectivity date", "Profile field", "No", "s38", "Same system, never named.", "ASK. Same.", "Q16"),
    ("Closing date", "Profile field", "No", "s38", "Same system, never named.", "ASK. Same.", "Q16"),
    ("Project number", "Profile field", "No", "s38", "Same system, never named.", "ASK. Same.", "Q16"),
    ("Project name", "Profile field", "No", "s38", "Same system, never named.", "ASK. Same.", "Q16"),

    ("BAND", "SLIDE 38, THE SEVEN CLICKABLE TILES AND THEIR DRILLS"),
    ("Sites for this project", "Tile and drill", "No", "s38",
     "Blocked on the site to project register. Nothing says a site belongs to a project.",
     "ASK. A site to project register.", "Q2"),
    ("Users for this project", "Tile and drill", "No", "s38",
     "Blocked on the register, and on the user register for the split.",
     "ASK. Both registers.", "Q2, Q3"),
    ("Site visits for this project", "Tile and drill", "No", "s38",
     "Blocked on the register.", "ASK. The register.", "Q2"),
    ("Documents for this project", "Tile and drill", "No", "s38",
     "Blocked on the register, and the scan for the figure.", "ASK. The register.", "Q2"),
    ("Records declared for this project", "Tile and drill", "No", "s38",
     "Blocked on the register.", "ASK. The register.", "Q2"),
    ("Physical counterparts for this project", "Tile and drill", "No", "s38",
     "Blocked on the register.", "ASK. The register.", "Q2"),
    ("Records due for disposal for this project", "Tile and drill", "No", "s38",
     "Blocked on the register.", "ASK. The register.", "Q2"),

    ("BAND", "SLIDE 38, THE THREE CHARTS"),
    ("Users pie, split staff / consultants / contractors", "Chart", "No", "s38",
     "Blocked on the register AND the user register.", "ASK. Both.", "Q2, Q3"),
    ("Cumulative declaration trend for this project", "Chart", "No", "s38",
     "Blocked on the register.", "ASK. The register.", "Q2"),
    ("Documents against records declared, site by site, with declaration rate",
     "Chart", "No", "s38",
     "Blocked on the register, and the scan for the denominator.",
     "ASK. The register. Then SCAN.", "Q2"),

    ("BAND", "SLIDES 36 AND 37, THE PROJECT LISTS THAT REACH THIS DASHBOARD"),
    ("Sovereign project site list", "Table", "No", "s36",
     "Lives on Bank-wide as the drill behind a tile. Both are missing.",
     "ASK. The register.", "Q2"),
    ("Nonsovereign project site list", "Table", "No", "s37",
     "Lives on Bank-wide as the drill behind a tile. Both are missing.",
     "ASK. The register.", "Q2"),
]

# =========================================================================
# 4. INSTITUTIONAL FILE PLAN
fp = [
    ("BAND", "SLIDE 47, THE ROLLUP ACROSS FIVE CATEGORIES"),
    ("Total Institutional Management terms", "Table row", "Yes", "s47", "", "", ""),
    ("Total Administration terms", "Table row", "Yes", "s47", "", "", ""),
    ("Total People management terms", "Table row", "Yes", "s47", "", "", ""),
    ("Total Programs and operations terms", "Table row", "Yes", "s47", "", "", ""),
    ("Total other terms", "Table row", "Yes", "s47", "", "", ""),
    ("Total terms", "Table row", "Yes", "s47", "", "", ""),
    ("Departments / Offices / RMs provisioned", "Table column", "No", "s47",
     "Needs a term to document join. The file plan table has no key that reaches "
     "the document table, and the document table carries no TermId.",
     "Design change plus a scan change, on top of a missing source.", "Q12, Q13"),
    ("Number of libraries provisioned", "Table column", "No", "s47",
     "Same missing join.", "Same.", "Q12, Q13"),
    ("Number of documents", "Table column", "No", "s47",
     "Same missing join, plus the scan.", "Same, plus SCAN.", "Q12, Q13"),
    ("Number of records declared", "Table column", "No", "s47",
     "Same missing join.", "Same.", "Q12, Q13"),
    ("Number of physical counterparts", "Table column", "No", "s47",
     "Same missing join.", "Same.", "Q12, Q13"),

    ("BAND", "SLIDES 48 TO 52, ONE SCREEN PER CATEGORY"),
    ("Term (Top level) list per category", "Table column", "Yes", "s48 to s52", "", "", ""),
    ("Category screens, all five present", "Panel", "Yes", "s48 to s52", "", "", ""),
    ("A notice saying what is missing behind this dashboard", "Panel", "Yes", "Ours",
     "Not a client requirement. Added so a reader sees the two gaps.", "", ""),
    ("Departments / Offices / RMs provisioned, per term", "Table column", "No", "s48 to s52",
     "Same missing join.", "Design change plus a source.", "Q12, Q13"),
    ("Number of libraries provisioned, per term", "Table column", "No", "s48 to s52",
     "Same missing join.", "Same.", "Q12, Q13"),
    ("Number of documents, per term", "Table column", "No", "s48 to s52",
     "Same missing join, plus the scan.", "Same, plus SCAN.", "Q12, Q13"),
    ("Number of records declared, per term", "Table column", "No", "s48 to s52",
     "Same missing join.", "Same.", "Q12, Q13"),
    ("Number of physical counterparts, per term", "Table column", "No", "s48 to s52",
     "Same missing join.", "Same.", "Q12, Q13"),
    ("Indicator: top most used terms", "Indicator", "No", "s48 to s52",
     "Same missing join.", "Same.", "Q12, Q13"),
    ("Indicator: least used and unused terms", "Indicator", "No", "s48 to s52",
     "Same missing join.", "Same.", "Q12, Q13"),
    ("Indicator: new libraries created outside convention", "Indicator", "No", "s48 to s52",
     "Never built. Needs the approved convention as a machine readable rule, "
     "not a document.",
     "ASK. The convention in a form a query can test.", ""),
    ("Request for new library creation and/or deletion", "Indicator", "No", "s48 to s52",
     "Never built. This is a governance request queue, not a measurement. "
     "Cloud Governance may hold it.",
     "ASK. Whether these requests are raised in Cloud Governance.", ""),
    ("Request for new term", "Indicator", "No", "s48 to s52",
     "Never built. Same. Opus term requests may also be relevant.",
     "ASK. Where term requests are raised.", ""),

    ("BAND", "THE TWO GAPS THAT GOVERN THIS WHOLE DASHBOARD"),
    ("Where the Institutional File Plan actually lives", "Source", "No", "s47 to s52",
     "Never answered. Purview holds 53 flat retention labels. Opus holds 65 terms "
     "in groups named Department and ADB. Neither has been confirmed as the file plan.",
     "ASK. Which system is the institutional file plan.", "Q12"),
    ("How a library or document is known to use a term", "Join", "No", "s47 to s52",
     "No key exists anywhere. This is the single blocker behind every column above.",
     "Design change. A TermId on the document table, and a scan that fills it.", "Q13"),
]

# =========================================================================
# 5. RETENTION AND DISPOSAL
rdz = [
    ("BAND", "SLIDE 44, THE ROLLUP. Permanent, Temporary, Total"),
    ("Departments / Offices / RMs provisioned", "Table column", "Yes", "s44", "", "", ""),
    ("Number of libraries provisioned", "Table column", "Yes", "s44", "", "", ""),
    ("Number of records declared", "Table column", "Yes", "s44", "", "", ""),
    ("Number of physical counterparts", "Table column", "Yes", "s44", "", "", ""),
    ("Number of records due for disposal", "Table column", "Yes", "s44", "", "", ""),
    ("Number of records disposed", "Table column", "No", "s44",
     "Removed 17 Aug. No disposal date or status in the EDRMS application. "
     "Opus reports destroyed records, which may fill this.",
     "APP, or Opus. Check the Opus approval and destruction exports.", "Q7"),

    ("BAND", "SLIDE 45, PERMANENT RETENTION"),
    ("Term (Top level)", "Table column", "Yes", "s45", "", "", ""),
    ("Departments / Offices / RMs provisioned", "Table column", "Yes", "s45", "", "", ""),
    ("Number of libraries provisioned", "Table column", "Yes", "s45", "", "", ""),
    ("Number of documents", "Table column", "Yes", "s45", "", "", ""),
    ("Number of records declared", "Table column", "Yes", "s45", "", "", ""),
    ("Number of physical counterparts", "Table column", "Yes", "s45", "", "", ""),

    ("BAND", "SLIDE 46, TEMPORARY RETENTION"),
    ("Term (Top level)", "Table column", "Yes", "s46", "", "", ""),
    ("Departments / Offices / RMs provisioned", "Table column", "Yes", "s46", "", "", ""),
    ("Number of libraries provisioned", "Table column", "Yes", "s46", "", "", ""),
    ("Number of records declared", "Table column", "Yes", "s46", "", "", ""),
    ("Number of physical counterparts", "Table column", "Yes", "s46", "", "", ""),
    ("Retention label, 5 / 10 / 20 years", "Table column", "Yes", "s46", "", "", ""),
    ("Number of records due for disposal", "Table column", "Yes", "s46", "", "", ""),
    ("Number of records disposed", "Table column", "No", "s46",
     "Removed 17 Aug. Same blocker as s44.", "APP, or Opus.", "Q7"),

    ("BAND", "THE RULE THIS WHOLE DASHBOARD RESTS ON, AND IT IS NOT SETTLED"),
    ("Which retention labels are permanent and which are temporary", "Rule", "No", "s44 to s46",
     "Never confirmed. The Purview file plan has never been exported, so whether "
     "all 53 labels map cleanly to one side of the split is unknown. Opus splits "
     "THREE ways, Long Term / Short Term / Permanent, not two.",
     "ASK. The mapping, and whether Long Term and Short Term both roll into Temporary.",
     "Q11"),
    ("Declared records by retention label", "Panel", "Yes", "s46",
     "Built, and it is the panel most exposed to the rule above.", "", ""),
]

# =========================================================================
# 6. RECORDS AND ARCHIVE HOLDINGS
ra = [
    ("BAND", "THE DASHBOARD ITSELF"),
    ("Records and Archive Holdings, as a key view", "Dashboard", "Yes", "s13, s14",
     "Present, but as a SPECIFICATION not a report. Every measure below is absent, "
     "and a notice on the screen says the dashboard is not yet specified.", "", "Q17"),
    ("The two tables drawn with the client's own column headings", "Panel", "Yes", "s68, s69",
     "The shape is present. The content is not.", "", "Q17"),

    ("BAND", "SLIDE 68, STORAGE"),
    ("Location, as Archives Room / Records Center / Offsite Storage", "Table column", "No", "s68",
     "No source held. Opus has a Locations node under Physical records.",
     "Check the Opus Locations export. Confirm it is a picklist, not free text.", "Q17"),
    ("No. of requests", "Table column", "No", "s68",
     "No source held. Opus reports physical records requests.",
     "Check the Opus requests export.", "Q17"),
    ("Total number of requestors (departments)", "Table column", "No", "s68",
     "No source held. Opus may carry a requestor but perhaps not their department.",
     "Check whether the Opus requestor carries a department.", "Q17"),
    ("Total number of requestors (RMs)", "Table column", "No", "s68",
     "No source held.", "Check the Opus requests export.", "Q17"),
    ("Total number of boxes stored", "Table column", "No", "s68",
     "No source held. Opus Physical records Explorer holds physical items.",
     "Check the Opus Explorer export for a box type.", "Q17"),
    ("Total number of folders stored", "Table column", "No", "s68",
     "No source held. Same export.", "Check the Opus Explorer export.", "Q17"),
    ("Remarks", "Table column", "No", "s68",
     "No source held. Free text, may not exist in any system.",
     "ASK. Whether Remarks is captured anywhere.", "Q17"),
    ("Indicator: new boxes and folders month on month per department",
     "Indicator", "No", "s68",
     "No source held. Needs an accession date and a department per item.",
     "Check the Opus Explorer export for both.", "Q17"),
    ("No. and list of outstanding / pending requests", "Indicator", "No", "s68",
     "No source held. Opus does show pending loan and overdue loan counts.",
     "Check the Opus requests export.", "Q17"),
    ("Physical counterparts identified but not transferred", "Indicator", "No", "s68",
     "No source held. This is the same custody event as 'turned over to RAC'.",
     "Check whether an Opus physical item links to a declared record.", "Q6, Q17"),
    ("Room capacity and % available storage capacity", "Chart", "No", "s68",
     "No source held. Note the slide asks US whether it CAN be included. "
     "An earlier prototype invented this chart and it was removed for that reason.",
     "Check the Opus Locations export for a capacity field. Then answer the client.",
     "Q17"),

    ("BAND", "SLIDE 69, RETRIEVAL"),
    ("Location", "Table column", "No", "s69", "No source held.", "Opus Locations.", "Q17"),
    ("No. of requests", "Table column", "No", "s69",
     "No source held. Opus reports loan requests.", "Opus requests export.", "Q17"),
    ("Total number of requestors (departments / RMs)", "Table column", "No", "s69",
     "No source held.", "Opus requests export.", "Q17"),
    ("Total number of boxes retrieved", "Table column", "No", "s69",
     "No source held. A retrieval is a loan in Opus terms.",
     "Opus requests export, loan type.", "Q17"),
    ("Total number of folders retrieved", "Table column", "No", "s69",
     "No source held.", "Opus requests export.", "Q17"),
    ("Status: Loan, Return to owner, For Disposal", "Table column", "No", "s69",
     "No source held. Opus carries a status but its vocabulary may differ.",
     "Check the Opus status values against the client's three words.", "Q17"),
    ("Remarks", "Table column", "No", "s69",
     "No source held.", "ASK.", "Q17"),
    ("Indicator: retrieval requests and boxes / folders retrieved month on month per department",
     "Indicator", "No", "s69",
     "No source held.", "Opus requests export, with dates and department.", "Q17"),
    ("No. and list of outstanding / pending requests", "Indicator", "No", "s69",
     "No source held. Opus shows pending and overdue loans.",
     "Opus requests export.", "Q17"),
    ("No. of boxes / folders disposed", "Indicator", "No", "s69",
     "No source held. Opus shows pending destruction.",
     "Opus destruction export.", "Q17"),
    ("Can room activities update capacity and freed up storage?", "Question", "No", "s69",
     "An open question the client put to US, not a requirement to build.",
     "Answer it once the Opus Locations export is read.", "Q17"),

    ("BAND", "SLIDE 67, WHAT THE CLIENT SAID ABOUT THIS DASHBOARD"),
    ("Retrieval as processed from eServe", "Source", "No", "s67",
     "eServe has never been connected to this project and nobody has named an owner.",
     "ASK. Is eServe the retrieval system of record, and who owns it.", "Q17"),
    ("What is available in Opus and how we can apply it", "Source", "No", "s67",
     "The client asked this on s67. It was closed on 17 Aug as 'nothing from Opus' "
     "because Opus was not a source we could see. Access was granted 19 Aug, so the "
     "basis for that decision has changed.",
     "Run the seven exports in OPUS_EXPORT_INSTRUCTIONS.md, then put it back to the client.",
     "Q17"),
]

s_bw, hr_bw, last_bw = sheet(wb, "1 Bank-wide Oversight", "Bank-wide Oversight",
    "Prototype key 'bw'. Deck slides 16, 17, 18, 34 to 44 and 47.", bw)
s_dp, hr_dp, last_dp = sheet(wb, "2 Department Insights", "Department Insights",
    "Prototype key 'dp'. Deck slides 26, 28 and 53 to 66.", dp)
s_pj, hr_pj, last_pj = sheet(wb, "3 Project Insights", "Project Insights: THE MISSING DASHBOARD",
    "No prototype key. The client requires six key views, s13 and s14. The prototype has five. "
    "Deck slides 36, 37 and 38.", pj)
s_fp, hr_fp, last_fp = sheet(wb, "4 Institutional File Plan", "Institutional File Plan",
    "Prototype key 'fp'. Deck slides 47 to 52.", fp)
s_rd, hr_rd, last_rd = sheet(wb, "5 Retention and Disposal", "Retention and Disposal",
    "Prototype key 'rd'. Deck slides 44, 45 and 46.", rdz)
s_ra, hr_ra, last_ra = sheet(wb, "6 Records and Archive Holdings", "Records and Archive Holdings",
    "Prototype key 'ra'. Deck slides 67, 68 and 69. Present as a specification, not a report.", ra)

# ---------------------------------------------------------------- SUMMARY
sm = wb.create_sheet("Summary")
sm.sheet_view.showGridLines = False
sm["A1"] = "Summary: what is built against what the client asked for"
sm["A1"].font = Font(name=FONT, size=14, bold=True, color="10243E")
sm["A2"] = "Counts are formulas over each dashboard sheet. Edit a sheet and this updates."
sm["A2"].font = NOTE_FT

heads = ["Dashboard", "In the prototype", "Not in the prototype", "Total asked for",
         "% built", "Chief blocker"]
for i, h in enumerate(heads, start=1):
    c = sm.cell(row=4, column=i, value=h)
    c.fill, c.font, c.border = H_FILL, H_FONT, BORD
    c.alignment = Alignment(vertical="center", wrap_text=True)
for i, w in enumerate([32, 18, 22, 18, 12, 46], start=1):
    sm.column_dimensions[get_column_letter(i)].width = w
sm.row_dimensions[4].height = 30

specs = [
    ("Bank-wide Oversight", "'1 Bank-wide Oversight'", hr_bw + 1, last_bw,
     "The two project tiles, division, the disposal change request"),
    ("Department Insights", "'2 Department Insights'", hr_dp + 1, last_dp,
     "The user register, division, the disposal change request"),
    ("Project Insights", "'3 Project Insights'", hr_pj + 1, last_pj,
     "MISSING DASHBOARD. A site to project register, and an unnamed ADB project system"),
    ("Institutional File Plan", "'4 Institutional File Plan'", hr_fp + 1, last_fp,
     "No key joins a term to a library or a document"),
    ("Retention and Disposal", "'5 Retention and Disposal'", hr_rd + 1, last_rd,
     "The permanent against temporary rule, and records disposed"),
    ("Records and Archive Holdings", "'6 Records and Archive Holdings'", hr_ra + 1, last_ra,
     "No source at all today. AvePoint Opus is the open lead"),
]

r = 5
for name, ref, first, last, blocker in specs:
    rng = f"{ref}!$D${first}:$D${last}"
    sm.cell(row=r, column=1, value=name)
    sm.cell(row=r, column=2, value=f'=COUNTIF({rng},"Yes")')
    sm.cell(row=r, column=3, value=f'=COUNTIF({rng},"No")')
    sm.cell(row=r, column=4, value=f"=B{r}+C{r}")
    sm.cell(row=r, column=5, value=f'=IFERROR(B{r}/D{r},0)')
    sm.cell(row=r, column=6, value=blocker)
    for i in range(1, 7):
        cell = sm.cell(row=r, column=i)
        cell.border = BORD
        cell.alignment = TOP
        cell.font = OUT_FONT if name == "Project Insights" else IN_FONT
        if name == "Project Insights":
            cell.fill = OUT_FILL
    sm.cell(row=r, column=5).number_format = "0.0%"
    r += 1

# total row
sm.cell(row=r, column=1, value="TOTAL")
sm.cell(row=r, column=2, value=f"=SUM(B5:B{r-1})")
sm.cell(row=r, column=3, value=f"=SUM(C5:C{r-1})")
sm.cell(row=r, column=4, value=f"=SUM(D5:D{r-1})")
sm.cell(row=r, column=5, value=f"=IFERROR(B{r}/D{r},0)")
for i in range(1, 7):
    cell = sm.cell(row=r, column=i)
    cell.font = Font(name=FONT, size=10, bold=True)
    cell.border = BORD
    cell.fill = BAND
sm.cell(row=r, column=5).number_format = "0.0%"

# notes under the table
r += 2
notes = [
    "HOW TO USE THIS WITH THE CLIENT",
    "1. Every red row on every sheet is a thing the client drew and we cannot build. Each carries its slide number.",
    "2. The four kinds of blocker are ASK, APP, SCAN and NONE. Only ASK needs the client. The rest need a decision here or a change request.",
    "3. Project Insights is the one to raise first. It is a whole key view, named on s13 and s14, and it is not in the prototype at all.",
    "4. AvePoint Opus was opened on 19 August and may close the disposal columns and the whole of Records and Archive Holdings. Nothing in this workbook assumes it will.",
    "",
    "This workbook checks PRESENCE against the deck. It does not verify a single figure. Numbers in the prototype are test tenant.",
]
for n in notes:
    c = sm.cell(row=r, column=1, value=n)
    c.font = Font(name=FONT, size=10, bold=n.startswith("HOW TO"))
    c.alignment = TOP
    r += 1

wb.save("/home/user/Jim/EDRMS_Dashboard_Gap_Checker_2026-08-19.xlsx")
print("saved")
