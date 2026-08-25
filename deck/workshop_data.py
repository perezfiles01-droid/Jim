#!/usr/bin/env python3
"""
Read EDRMS_Util_Dashboard_Gap_Checker_2026-08-21.xlsx and write workshop.json,
the single content model behind both the navigation deck and the script. Both
build from this, so a count on a slide and the same count in the script cannot
drift apart.

The workbook the client will have on screen is the authority for every number
quoted in either file. Nothing is typed by hand.
"""
import json, os
from openpyxl import load_workbook

D = os.path.dirname(os.path.abspath(__file__))
XL = os.path.join(D, "..", "EDRMS_Util_Dashboard_Gap_Checker_2026-08-21.xlsx")

wb = load_workbook(XL)
sheets = []
for idx, name in enumerate(wb.sheetnames, start=1):
    ws = wb[name]
    rows, section = [], ""
    for r in range(5, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a is not None and b is None:          # a section band
            section = str(a).strip()
            continue
        if not isinstance(a, int):
            continue
        rows.append({
            "n": a, "row": r, "section": section,
            "item": str(b).strip() if b else "",
            "type": str(ws.cell(row=r, column=3).value or "").strip(),
            "in": str(ws.cell(row=r, column=4).value or "").strip(),
            "slide": str(ws.cell(row=r, column=5).value or "").strip(),
            "why": str(ws.cell(row=r, column=6).value or "").strip(),
            "needs": str(ws.cell(row=r, column=7).value or "").strip(),
            "q": str(ws.cell(row=r, column=8).value or "").strip(),
        })
    yes = sum(1 for x in rows if x["in"] == "Yes")
    no = sum(1 for x in rows if x["in"] == "No")
    sheets.append({"tab": idx, "name": name, "rows": rows,
                   "yes": yes, "no": no, "tot": len(rows),
                   "first_row": rows[0]["row"] if rows else 5,
                   "last_row": rows[-1]["row"] if rows else 5})

out = {
    "sheets": sheets,
    "tot": {"yes": sum(s["yes"] for s in sheets),
            "no": sum(s["no"] for s in sheets),
            "tot": sum(s["tot"] for s in sheets)},
}
json.dump(out, open(os.path.join(D, "workshop.json"), "w"), indent=1)
print("tabs", len(sheets), out["tot"])
for s in sheets:
    print(f'  {s["tab"]} {s["name"][:34]:36} rows {s["tot"]:3}  in {s["yes"]:3}  out {s["no"]:3}'
          f'  xl rows {s["first_row"]}-{s["last_row"]}')
